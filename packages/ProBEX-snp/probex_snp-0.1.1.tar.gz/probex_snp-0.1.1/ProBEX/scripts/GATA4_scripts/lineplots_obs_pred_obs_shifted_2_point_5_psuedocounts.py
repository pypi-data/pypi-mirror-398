import os
import time, psutil, gc
import logging
import pandas as pd
import re
import numpy as np
import matplotlib.pyplot as plt
from matplotlib import ticker
import warnings
import traceback
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

'''
Author: [Shreya Sharm]
Date: [March 28, 2025]
Description:
    This script processes observed and predicted enrichment data, generates combined plots,
    and saves all plotting data to an Excel file for further analysis.
'''

warnings.filterwarnings("ignore", category=Warning)
logging.basicConfig(level=logging.INFO, 
                    format="%(asctime)s - %(levelname)s - %(message)s", 
                    handlers=[logging.FileHandler("debug.log"), logging.StreamHandler()])

class CombinedEnrichmentPlotter:
    def __init__(self, output_folder, csv_files_order, concentration_map, 
                 merged_df_path, denominators_lambda_unb_bi_path, 
                 k_values_lambda_bi_path, denominators_lambda_b_bi_path):
        """
        Initialize the plotter with all necessary paths and settings
        """
        self.output_folder = output_folder
        self.csv_files_order = csv_files_order
        self.concentration_map = concentration_map
        self.rsid_pattern = re.compile(r'^rs') 
        
        self.merged_df_path = merged_df_path
        self.denominators_lambda_unb_bi_path = denominators_lambda_unb_bi_path
        self.k_values_lambda_bi_path = k_values_lambda_bi_path
        self.denominators_lambda_b_bi_path = denominators_lambda_b_bi_path
        self.excel_output_path = os.path.join(output_folder, "all_enrichment_data.xlsx")
        
        os.makedirs(output_folder, exist_ok=True)
        self.enrich_dict = {}

    def generate_alt_alleles(self, ref, alt):
        """
        Generate possible alternative alleles given reference and alternate alleles
        """
        possible_alleles = ['A', 'T', 'G', 'C']
        if ref in possible_alleles:
            possible_alleles.remove(ref)
        if alt in possible_alleles:
            possible_alleles.remove(alt)

        return possible_alleles[:2] if len(possible_alleles) >= 2 else (possible_alleles[0], None)
    
    def process_observed_files(self):
        """
        Process enrichment values from the merged_df DataFrame with improved allele handling
        """
        try:
            logging.info(f"Loading merged_df from {self.merged_df_path}")
            merged_df = pd.read_csv(self.merged_df_path)
            
            logging.info(f"Merged_df shape: {merged_df.shape}")
            #logging.info(f"Sample rsIDs: {merged_df['rsID'].dropna().head(5).tolist()}")

            processed_rsids = 0
            valid_enrichments = 0
            
            for _, row in merged_df.iterrows():
                if not isinstance(row['rsID'], str):
                    continue

                try:
                    rsID_allele = row['rsID']  # e.g. "rs72775262_G"
                    rsID, allele = rsID_allele.split('_', 1)  # Split on first underscore only
                except ValueError:
                    continue

                if rsID not in self.enrich_dict:
                    self.enrich_dict[rsID] = {
                        'observed': {},
                        'observed_lambda_withbi': {},
                        'predicted_lambda': {},
                    }
                    processed_rsids += 1
                    
                # Process each concentration and replicate
                concentrations = ['100', '500', '1000', '1500', '2000', '3000']
                replicates = ['R1', 'R2']
                
                for conc in concentrations:
                    conc_label = f"{conc} nM"
                    
                    for rep in replicates:
                        # Try both possible column patterns
                        col_name = None
                        for suffix in ['b_unb']:
                            test_col = f"Enrich_{conc}_{rep}_{suffix}"
                            if test_col in merged_df.columns:
                                col_name = test_col
                                break
                        
                        if not col_name:
                            continue
                            
                        value = row[col_name]
                        if pd.isna(value):
                            continue
                            
                        # Create allele key (rsID_allele format)
                        allele_key = f"{rsID}_{allele}"
                        
                        # Initialize nested dictionaries if needed
                        if conc_label not in self.enrich_dict[rsID]['observed']:
                            self.enrich_dict[rsID]['observed'][conc_label] = {}
                            
                        if allele_key not in self.enrich_dict[rsID]['observed'][conc_label]:
                            self.enrich_dict[rsID]['observed'][conc_label][allele_key] = []
                            
                        # Store the value
                        self.enrich_dict[rsID]['observed'][conc_label][allele_key].append(value)
                        valid_enrichments += 1
                        
        except Exception as e:
            logging.error(f"Error processing observed files: {str(e)}")
            logging.error(traceback.format_exc())
            raise

                
    def calculate_observed_enrich_lambda_withbi(self):
        """
        Calculate observed enrichment using lambda values with proper mapping of bi per rsID.
        """
        try:
            merged_df = pd.read_csv(self.merged_df_path)
            concentrations = np.array([100, 500, 1000, 1500, 2000, 3000])

            # Total unbound and bound counts for each concentration and replicate
            T_Uj_R1 = np.array([np.sum(merged_df[f"{conc}_count_R1_unb"]) for conc in concentrations])
            T_Uj_R2 = np.array([np.sum(merged_df[f"{conc}_count_R2_unb"]) for conc in concentrations])
            T_cj_R1 = np.array([np.sum(merged_df[f"{conc}_count_R1_b"]) for conc in concentrations])
            T_cj_R2 = np.array([np.sum(merged_df[f"{conc}_count_R2_b"]) for conc in concentrations])

            # Pseudo counts for unbound
            pseudo_b_R1 = 0
            pseudo_b_R2 = 0
            pseudo_unb_R1 = (2.5 * T_Uj_R1) / 1e6
            pseudo_unb_R2 = (2.5 * T_Uj_R2) / 1e6

            # Stack counts arrays
            n_sij_R1 = np.stack([merged_df[f"{conc}_count_R1_b"].values for conc in concentrations], axis=1)
            n_sij_R2 = np.stack([merged_df[f"{conc}_count_R2_b"].values for conc in concentrations], axis=1)
            n_Uij_R1 = np.stack([merged_df[f"{conc}_count_R1_unb"].values for conc in concentrations], axis=1)
            n_Uij_R2 = np.stack([merged_df[f"{conc}_count_R2_unb"].values for conc in concentrations], axis=1)

            global T_L_R1, T_L_R2
            T_L_R1 = np.sum(merged_df["Library_1"].values)
            T_L_R2 = np.sum(merged_df["Library_2"].values)

            # Read lambda values with bi
            lambda_df = pd.read_csv(self.k_values_lambda_bi_path)

            # Initialize updated counts arrays as copies of original + pseudo counts
            
            n_sij_R1_updated = n_sij_R1 + pseudo_b_R1
            n_sij_R2_updated = n_sij_R2 + pseudo_b_R2

            n_Uij_R1_updated = n_Uij_R1 + pseudo_unb_R1
            n_Uij_R2_updated = n_Uij_R2 + pseudo_unb_R2

            # Apply bi correction per rsID by updating only the corresponding row
            for _, row in lambda_df.iterrows():
                full_rsid = row['rsID']
                bi = row['Fitted_bi_iter10']

                # Find index in merged_df where rsID matches exactly
                idxs = merged_df.index[merged_df['rsID'] == full_rsid].tolist()
                if not idxs:
                    # If no match, skip this rsID
                    continue
                idx = idxs[0]

                # Safely apply bi correction only for that SNP's counts across all concentrations
                n_sij_R1_updated[idx, :] = n_sij_R1[idx, :] / (1.0 - bi)
                n_sij_R2_updated[idx, :] = n_sij_R2[idx, :] / (1.0 - bi)
                # Note: pseudo unbound counts are added but not corrected by bi, as per original logic

            # Sum total updated counts
            T_cj_R1_updated = np.sum(n_sij_R1_updated, axis=0)
            T_cj_R2_updated = np.sum(n_sij_R2_updated, axis=0)
            T_Uj_R1_updated = np.sum(n_Uij_R1_updated, axis=0)
            T_Uj_R2_updated = np.sum(n_Uij_R2_updated, axis=0)

            # Now compute enrichment per rsID and concentration
            for _, row in lambda_df.iterrows():
                try:
                    full_rsid = row['rsID']
                    rsID, allele = full_rsid.split("_")
                    key_name = f"{rsID}_{allele}"

                    if rsID not in self.enrich_dict:
                        self.enrich_dict[rsID] = {
                            'observed': {}, 
                            'observed_lambda_withbi': {},
                            'predicted_lambda': {},
                        }

                    # Find index of this rsID in merged_df
                    idxs = merged_df.index[merged_df['rsID'] == full_rsid].tolist()
                    if not idxs:
                        continue
                    idx = idxs[0]

                    enrich_R1 = ((n_sij_R1_updated[idx] / T_cj_R1_updated) / (n_Uij_R1_updated[idx] / T_Uj_R1_updated))
                    enrich_R2 = ((n_sij_R2_updated[idx] / T_cj_R2_updated) / (n_Uij_R2_updated[idx] / T_Uj_R2_updated))

                    for i, conc in enumerate(concentrations):
                        conc_str = f"{conc} nM"
                        if conc_str not in self.enrich_dict[rsID]['observed_lambda_withbi']:
                            self.enrich_dict[rsID]['observed_lambda_withbi'][conc_str] = {}

                        if key_name not in self.enrich_dict[rsID]['observed_lambda_withbi'][conc_str]:
                            self.enrich_dict[rsID]['observed_lambda_withbi'][conc_str][key_name] = []

                        self.enrich_dict[rsID]['observed_lambda_withbi'][conc_str][key_name].extend([enrich_R1[i], enrich_R2[i]])

                except Exception as e:
                    logging.error(f"Error processing rsID {row.get('rsID', 'unknown')}: {str(e)}")
                    logging.debug(traceback.format_exc())
                    continue

        except Exception as e:
            logging.error(f"Error in calculate_observed_enrich_lambda_withbi: {str(e)}")
            logging.error(traceback.format_exc())
            raise
        
    def process_predicted_data(self):
        """
        Process predicted data from C values and denominators with improved rsID matching
        """
        logging.info("Processing predicted data")
        
        try:
            # Load data for bi model
            K_values = pd.read_csv(self.k_values_lambda_bi_path)
            logging.info(f"Total K values loaded: {len(K_values)}")
            logging.debug(f"Sample rsIDs in K_values: {K_values['rsID'].head(10).tolist()}")
            
            denominators_unb = pd.read_csv(self.denominators_lambda_unb_bi_path, header=None) 
            denominators_b = pd.read_csv(self.denominators_lambda_b_bi_path, header=None) 
            
            # Create a mapping of rsID_allele to K value for faster lookup
            k_value_map = {}
            missing_in_enrich_dict = []
            
            for _, row in K_values.iterrows():
                rsID_allele = row['rsID']
                k_value_map[rsID_allele] = row['Fitted_K_iter10']
                
                # Check if this rsID exists in enrich_dict
                rsID = rsID_allele.split('_')[0]
                if rsID not in self.enrich_dict:
                    missing_in_enrich_dict.append(rsID_allele)
            
            if missing_in_enrich_dict:
                logging.warning(f"{len(missing_in_enrich_dict)} rsIDs in K_values not found in enrich_dict. Sample: {missing_in_enrich_dict[:5]}")
            
            # Prepare denominators for bi model
            deno_lambda_b_R1 = {
                '100': denominators_b.iloc[0, 0],
                '500': denominators_b.iloc[1, 0],
                '1000': denominators_b.iloc[2, 0],
                '1500': denominators_b.iloc[3, 0],
                '2000': denominators_b.iloc[4, 0],
                '3000': denominators_b.iloc[5, 0]
            }

            deno_lambda_b_R2 = {
                '100': denominators_b.iloc[6, 0],
                '500': denominators_b.iloc[7, 0],
                '1000': denominators_b.iloc[8, 0],
                '1500': denominators_b.iloc[9, 0],
                '2000': denominators_b.iloc[10, 0],
                '3000': denominators_b.iloc[11, 0]
            }
            
            deno_lambda_unb_R1 = {
                '100': denominators_unb.iloc[0, 0],
                '500': denominators_unb.iloc[1, 0],
                '1000': denominators_unb.iloc[2, 0],
                '1500': denominators_unb.iloc[3, 0],
                '2000': denominators_unb.iloc[4, 0],
                '3000': denominators_unb.iloc[5, 0]
            }

            deno_lambda_unb_R2 = {
                '100': denominators_unb.iloc[6, 0],
                '500': denominators_unb.iloc[7, 0],
                '1000': denominators_unb.iloc[8, 0],
                '1500': denominators_unb.iloc[9, 0],
                '2000': denominators_unb.iloc[10, 0],
                '3000': denominators_unb.iloc[11, 0]
            }        
                
            def enrichment_equation_new(K, concentrations, deno_lambda_b, deno_lambda_unb):
                """
                Calculate predicted enrichment using the enrichment equation (with bi)
                """
                predicted_enrich = []
                for conc in concentrations:
                    numerator = ((K * conc) * (deno_lambda_unb[str(conc)]))
                    denominator = deno_lambda_b[str(conc)]
                    predicted_enrich.append(numerator / denominator)
                return predicted_enrich
        
            concentrations = np.array([100, 500, 1000, 1500, 2000, 3000])
            
            # Process all rsID_allele combinations that have K values
            batch_size = 1000
            for i in range(0, len(k_value_map.items()), batch_size):
                batch = list(k_value_map.items())[i:i+batch_size]
                
                for rsID_allele, K in batch:
                    try:
                        # Split rsID and allele more carefully
                        parts = rsID_allele.split('_')
                        if len(parts) < 2:
                            logging.warning(f"Invalid rsID_allele format: {rsID_allele}")
                            continue
                            
                        rsID = '_'.join(parts[:-1])  # Handle cases where rsID might contain underscores
                        allele = parts[-1]
                        
                        # Initialize structure if not exists
                        if rsID not in self.enrich_dict:
                            self.enrich_dict[rsID] = {
                                'observed': {}, 
                                'observed_lambda_withbi': {},
                                'predicted_lambda': {},
                            }
                        elif 'predicted_lambda' not in self.enrich_dict[rsID]:
                            self.enrich_dict[rsID]['predicted_lambda'] = {}
                            
                        # Calculate predicted enrichments
                        enrich_lambda_R1 = enrichment_equation_new(K, concentrations, deno_lambda_b_R1, deno_lambda_unb_R1)
                        enrich_lambda_R2 = enrichment_equation_new(K, concentrations, deno_lambda_b_R2, deno_lambda_unb_R2)

                        for conc, val_lambda1, val_lambda2 in zip(concentrations, enrich_lambda_R1, enrich_lambda_R2):
                            conc_str = f"{conc} nM"
                            if conc_str not in self.enrich_dict[rsID]['predicted_lambda']:
                                self.enrich_dict[rsID]['predicted_lambda'][conc_str] = {}
                                
                            if rsID_allele not in self.enrich_dict[rsID]['predicted_lambda'][conc_str]:
                                self.enrich_dict[rsID]['predicted_lambda'][conc_str][rsID_allele] = []
                                
                            self.enrich_dict[rsID]['predicted_lambda'][conc_str][rsID_allele].extend([val_lambda1, val_lambda2])
                            
                    except Exception as e:
                        logging.error(f"Error processing {rsID_allele}: {str(e)}")
                        continue
                
                gc.collect()
                    
        except Exception as e:
            logging.error(f"Error in process_predicted_data: {str(e)}")
            logging.error(traceback.format_exc())
            raise
    
    def save_all_plotting_data_to_excel(self):
        """
        Save all plotting data (observed, observed_with_bi, predicted) to an Excel file
        with separate sheets for each data type.
        """
        
        if os.path.exists(self.excel_output_path):
            logging.info(f"Excel file already exists. Returning existing file.")
            return self.excel_output_path
        
        logging.info("Saving all plotting data to Excel")
        
        # Create a new Excel workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Create sheets for each data type
        observed_sheet = wb.create_sheet("Observed")
        observed_bi_sheet = wb.create_sheet("Observed_with_BI")
        predicted_sheet = wb.create_sheet("Predicted")
        
        # Define concentration order
        concentration_order = ['100 nM', '500 nM', '1000 nM', '1500 nM', '2000 nM', '3000 nM']
        
        # Prepare headers
        headers = ['rsID_Allele'] + concentration_order
        observed_sheet.append(headers)
        observed_bi_sheet.append(headers)
        predicted_sheet.append(headers)
        
        # Collect all unique rsID_allele combinations from all data types
        all_alleles = set()
        
        # First collect from observed data
        for rsID, data in self.enrich_dict.items():
            if 'observed' in data:
                for conc in data['observed']:
                    for allele_key in data['observed'][conc]:
                        all_alleles.add(allele_key)
        
        # Then from other data types
        for rsID, data in self.enrich_dict.items():
            for data_type in ['observed_lambda_withbi', 'predicted_lambda']:
                if data_type in data:
                    for conc in data[data_type]:
                        for allele_key in data[data_type][conc]:
                            all_alleles.add(allele_key)
        
        # Now populate data for each allele
        for allele_key in sorted(all_alleles):
            try:
                rsID, allele = allele_key.split('_', 1)  # Split only on first underscore
            except ValueError:
                logging.warning(f"Invalid allele_key format: {allele_key}")
                continue
            
            # Process observed data
            obs_row = [allele_key]
            for conc in concentration_order:
                if rsID in self.enrich_dict and 'observed' in self.enrich_dict[rsID]:
                    if conc in self.enrich_dict[rsID]['observed'] and allele_key in self.enrich_dict[rsID]['observed'][conc]:
                        values = self.enrich_dict[rsID]['observed'][conc][allele_key]
                        # Filter out NaN values
                        valid_values = [v for v in values if not np.isnan(v)]
                        mean_val = np.mean(valid_values) if valid_values else np.nan
                        obs_row.append(mean_val)
                    else:
                        obs_row.append(np.nan)
                else:
                    obs_row.append(np.nan)
            observed_sheet.append(obs_row)
            
            # Process observed with bi data
            obs_bi_row = [allele_key]
            for conc in concentration_order:
                if rsID in self.enrich_dict and 'observed_lambda_withbi' in self.enrich_dict[rsID]:
                    if conc in self.enrich_dict[rsID]['observed_lambda_withbi'] and allele_key in self.enrich_dict[rsID]['observed_lambda_withbi'][conc]:
                        values = self.enrich_dict[rsID]['observed_lambda_withbi'][conc][allele_key]
                        valid_values = [v for v in values if not np.isnan(v)]
                        mean_val = np.mean(valid_values) if valid_values else np.nan
                        obs_bi_row.append(mean_val)
                    else:
                        obs_bi_row.append(np.nan)
                else:
                    obs_bi_row.append(np.nan)
            observed_bi_sheet.append(obs_bi_row)
            
            # Process predicted data
            pred_row = [allele_key]
            for conc in concentration_order:
                if rsID in self.enrich_dict and 'predicted_lambda' in self.enrich_dict[rsID]:
                    if conc in self.enrich_dict[rsID]['predicted_lambda'] and allele_key in self.enrich_dict[rsID]['predicted_lambda'][conc]:
                        values = self.enrich_dict[rsID]['predicted_lambda'][conc][allele_key]
                        valid_values = [v for v in values if not np.isnan(v)]
                        mean_val = np.mean(valid_values) if valid_values else np.nan
                        pred_row.append(mean_val)
                    else:
                        pred_row.append(np.nan)
                else:
                    pred_row.append(np.nan)
            predicted_sheet.append(pred_row)
        
        # Save the workbook
        wb.save(self.excel_output_path)
        logging.info(f"All plotting data saved to {self.excel_output_path}")
        
    def plot_combined_enrichment(self):
        """
        Create plots with proper error bars for all data types
        """
        concentration_order = ['100 nM', '500 nM', '1000 nM', '1500 nM', '2000 nM', '3000 nM']
        x_positions = [0, 1, 2, 3, 4, 6]
        
        COLORS = {
    		'observed': '#1D6F3D',                
    		'observed_lambda_withbi': '#254B96',  
    		'predicted_lambda': '#BF212F'         
	}

        for rsID, data in self.enrich_dict.items():
            if 'observed' not in data:
                continue

            # Collect all alleles present in any data type
            alleles = set()
            for data_type in ['observed', 'observed_lambda_withbi', 'predicted_lambda']:
                if data_type in data:
                    for conc in data[data_type]:
                        alleles.update(key.split('_')[1] for key in data[data_type][conc].keys())

            for allele in sorted(alleles):
                allele_key = f"{rsID}_{allele}"
                output_path = os.path.join(self.output_folder, f"combined_{allele_key}.png")
                
                if os.path.exists(output_path):
                    logging.info(f"Plot already exists for {allele_key}. Skipping.")
                    continue
                    
                skip_plot = False
                required_data_types = ['observed', 'observed_lambda_withbi', 'predicted_lambda']
                
                for data_type in required_data_types:
                    if data_type not in data:
                        logging.info(f"Skipping {allele_key} - missing {data_type} data")
                        skip_plot = True
                        break
                    
                    for conc in concentration_order:
                        if conc not in data[data_type] or allele_key not in data[data_type][conc]:
                            logging.info(f"Skipping {allele_key} - missing {data_type} data for {conc}")
                            skip_plot = True
                            break
                        
                        values = data[data_type][conc][allele_key]
                        if len(values) == 0 or any(np.isnan(v) or np.isinf(v) for v in values):
                            logging.info(f"Skipping {allele_key} - invalid values in {data_type} for {conc}")
                            skip_plot = True
                            break
                        
                    if skip_plot:
                        break
                    
                if skip_plot:
                    continue

                plt.figure(figsize=(10, 6), dpi=120)
                ax = plt.gca()
                ax.set_facecolor('#f8f8f8')
                plt.grid(color='white', linestyle='-', linewidth=1.5)

                label_map = {
                    'observed': "Observed Enrichment",
                    'observed_lambda_withbi': "Binding-Shift corrected",
                    'predicted_lambda': "Binding-Adjusted Predicted"
                }

                for data_type, marker, linestyle in [
                    ('observed', 'o', '--'),
                    ('observed_lambda_withbi', '^', '-.'),
                    ('predicted_lambda', 's', ':')
                ]:
                    if data_type not in data:
                        continue

                    means = []
                    stds = []
                    valid_positions = []

                    for i, conc in enumerate(concentration_order):
                        if conc in data[data_type] and allele_key in data[data_type][conc]:
                            values = [v for v in data[data_type][conc][allele_key] if not np.isnan(v)]
                            if values:
                                means.append(np.mean(values))
                                stds.append(np.std(values) if len(values) > 1 else 0)
                                valid_positions.append(i)

                    if not valid_positions:
                        continue

                    color = COLORS[data_type]  # <-- Fix: get the color here

                    plt.errorbar(
                        x_positions, means, yerr=stds,
                        label=label_map.get(data_type, data_type),
                        color=color,
                        marker=marker,
                        linestyle=linestyle,
                        markerfacecolor='white',
                        markeredgecolor=color,
                        markersize=8,
                        linewidth=2.5,
                        alpha=0.9,
                        capsize=5,
                        markeredgewidth=1.5,
                        elinewidth=1.5
                    )

                plt.title(f"Enrichment Profile: {allele_key}", fontsize=14, fontweight='bold')
                plt.xlabel("Concentration (nM)", fontsize=12)
                plt.ylabel("Enrichment Ratio", fontsize=12)
                plt.xticks(x_positions, [c.split()[0] for c in concentration_order])
                plt.legend()
                plt.tight_layout()
                plt.savefig(output_path, dpi=350)
                plt.close()
                    
    def run_processing(self):
        """
        Run the complete processing and plotting pipeline
        """
        start_time = time.time()
        try:
            # Process all data first
            logging.info("Processing observed data")
            self.process_observed_files()
            
            logging.info("Calculating observed enrichment with binding shift")
            self.calculate_observed_enrich_lambda_withbi()
            
            logging.info("Processing predicted data")
            self.process_predicted_data()
            
            # Save all data to Excel
            logging.info("Saving all data to Excel")
            self.save_all_plotting_data_to_excel()
            
            # Generate plots
            logging.info("Generating enrichment plots")
            self.plot_combined_enrichment()
                
        except Exception as e:
            logging.error(f"Error during processing: {str(e)}")
            logging.error(traceback.format_exc())
            raise
        finally:
            plt.close('all')
            gc.collect()      
            end_time = time.time()
            logging.info(f"Completed processing and plotting in {end_time - start_time:.2f} seconds")

if __name__ == "__main__":
    start_time = time.time()
    logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
    paths = {
        "output_folder": "./output/GATA4/fitting/linePlots_fittedK",
        "merged_df_path": "./output/GATA4/merged_df_GATA4_with_pseudocounts_2point5.csv",
        "denominators_lambda_unb_bi_path": "./output/GATA4/fitting/iteration_results_withunb_with_bi/updated_unbound_denominators_iteration_10.csv",
        "k_values_lambda_bi_path": "./output/GATA4/fitting/iteration_results_withunb_with_bi/fitted_K_results_iteration_10.csv",
        "denominators_lambda_b_bi_path": "./output/GATA4/fitting/iteration_results_withunb_with_bi/updated_bound_denominators_iteration_10.csv"
    }
    csv_files_order = [
        "R52_46_Vs_R52_47.csv", "R52_44_Vs_R52_45.csv", "R52_42_Vs_R52_43.csv",
        "R52_40_Vs_R52_41.csv", "R52_38_Vs_R52_39.csv", "R52_36_Vs_R52_37.csv",
        "R52_34_Vs_R52_35.csv", "R52_50_Vs_R52_51.csv", "R52_52_Vs_R52_53.csv",
        "R52_54_Vs_R52_55.csv", "R52_56_Vs_R52_57.csv", "R52_58_Vs_R52_59.csv",
        "R52_60_Vs_R52_61.csv", "R52_62_Vs_R52_63.csv"
    ]
    concentration_map = {
        file: f"{conc} nM" for file, conc in zip(csv_files_order, 
        [3000, 2000, 1500, 1000, 500, 100, 0, 0, 100, 500, 1000, 1500, 2000, 3000])
    }
    
    logging.info("Initializing CombinedEnrichmentPlotter")

    plotter = CombinedEnrichmentPlotter(
            output_folder=paths["output_folder"],
            csv_files_order=csv_files_order,
            concentration_map=concentration_map,
            merged_df_path=paths["merged_df_path"],
            denominators_lambda_unb_bi_path=paths["denominators_lambda_unb_bi_path"],
            k_values_lambda_bi_path=paths["k_values_lambda_bi_path"],
            denominators_lambda_b_bi_path=paths["denominators_lambda_b_bi_path"]
        )

    # Run the complete processing pipeline
    logging.info("Starting the combined enrichment processing")
    plotter.run_processing()
    logging.info("All processing completed successfully")
    end_time = time.time()
    total_time = end_time - start_time
    print(f"Total time taken: {total_time:.2f} seconds")
