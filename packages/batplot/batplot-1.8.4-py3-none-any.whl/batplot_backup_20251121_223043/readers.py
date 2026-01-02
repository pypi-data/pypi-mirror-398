"""Readers for various battery cycler data formats.

This module provides parsers for different battery testing equipment file formats:

Supported Formats:
    - BioLogic .mpt: Native binary format from BioLogic potentiostats
    - BioLogic .txt: Exported text format from EC-Lab software
    - Neware .csv: CSV export from Neware battery testers
    - Landt/Lanhe .xlsx: Excel files with Chinese column headers
    - Generic .csv: Generic CSV with standard battery cycling columns

Key Functions:
    - read_mpt_file(): Parse BioLogic .mpt files for CV, GC, CPC modes
    - read_biologic_txt_file(): Parse BioLogic .txt exports
    - read_ec_csv_file(): Parse Neware CSV and Excel files, handles half-cycles
    - read_ec_csv_dqdv_file(): Parse CSV for differential capacity analysis

Data Return Formats:
    CV mode: (voltage, current, cycles)
    GC mode: (capacity, voltage, cycles, charge_mask, discharge_mask)
    CPC mode: (cycle_nums, cap_charge, cap_discharge, efficiency)
    dQ/dV mode: (voltage, dqdv, cycles)

Special Features:
    - Half-cycle detection and merging (Neware compatibility)
    - Automatic column detection with fuzzy matching
    - Chinese column name support (Landt/Lanhe cyclers)
    - Specific capacity calculation for .mpt files
"""

from __future__ import annotations

import numpy as np
from typing import Tuple, List


def _infer_cycles_from_masks(charge_mask: np.ndarray, discharge_mask: np.ndarray, n_points: int) -> np.ndarray:
    """Infer full-cycle numbers by pairing alternating charge/discharge segments.

    Each contiguous charge or discharge block becomes a "run". Runs are ordered by their
    starting index and paired sequentially so that run 0 and 1 form Cycle 1, run 2 and 3
    form Cycle 2, etc. If a final run has no pair (e.g., data ends mid-cycle), it is
    assigned to the previous cycle. Rest/CV intervals inherit the cycle number of the
    preceding segment so filters cover both halves seamlessly.
    """

    segments: List[Tuple[int, int, bool]] = []  # (start, end_exclusive, is_charge)

    def _append_segments(mask: np.ndarray, is_charge_segment: bool):
        idx = np.where(mask)[0]
        if idx.size == 0:
            return
        start = int(idx[0])
        prev = int(idx[0])
        for cur in idx[1:]:
            if cur != prev + 1:
                segments.append((start, prev + 1, is_charge_segment))
                start = int(cur)
            prev = int(cur)
        segments.append((start, prev + 1, is_charge_segment))

    _append_segments(charge_mask, True)
    _append_segments(discharge_mask, False)
    segments.sort(key=lambda seg: seg[0])

    cycles = np.zeros(n_points, dtype=int)
    if not segments:
        cycles.fill(1)
        return cycles

    current_cycle = 1
    half_index = 0  # 0 = first half, 1 = second half of current cycle
    for start, end, _flag in segments:
        cycles[start:end] = current_cycle
        half_index += 1
        if half_index == 2:
            current_cycle += 1
            half_index = 0

    last_cycle = 1
    for i in range(n_points):
        if cycles[i] == 0:
            cycles[i] = last_cycle
        else:
            last_cycle = cycles[i]

    return cycles


def read_excel_to_csv_like(fname: str, header_row: int = 2, data_start_row: int = 3) -> Tuple[list, list]:
    """Read Excel file and convert to CSV-like structure for batplot.
    
    This is designed for Chinese cycler data Excel files where:
    - Row 1: File/sample name
    - Row 2: Column headers
    - Row 3+: Data
    
    Args:
        fname: Path to Excel file (.xlsx)
        header_row: Row number containing headers (1-indexed, default=2)
        data_start_row: First row containing data (1-indexed, default=3)
    
    Returns:
        Tuple of (header_list, rows_list) compatible with CSV processing
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl is required to read Excel files. Install with: pip install openpyxl")
    
    wb = openpyxl.load_workbook(fname, read_only=True, data_only=True)
    ws = wb.active
    
    # Read header row
    header = []
    for cell in ws[header_row]:
        header.append(str(cell.value) if cell.value is not None else '')
    
    # Read data rows
    rows = []
    for row in ws.iter_rows(min_row=data_start_row, values_only=True):
        # Convert row to list of strings (handle None, datetime, etc.)
        row_data = []
        for val in row:
            if val is None:
                row_data.append('')
            elif isinstance(val, (int, float)):
                row_data.append(str(val))
            else:
                row_data.append(str(val))
        rows.append(row_data)
    
    wb.close()
    return header, rows


def read_csv_file(fname: str):
    for delim in [",", ";", "\t"]:
        try:
            data = np.genfromtxt(fname, delimiter=delim, comments="#")
            if data.ndim == 1:
                data = data.reshape(1, -1)
            if data.shape[1] >= 2:
                return data
        except Exception:
            continue
    raise ValueError(f"Invalid CSV format in {fname}, need at least 2 columns (x,y).")


def read_gr_file(fname: str):
    """Read a PDF .gr file (r, G(r))."""
    r_vals = []
    g_vals = []
    with open(fname, "r") as f:
        for line in f:
            ls = line.strip()
            if not ls or ls.startswith("#"):
                continue
            parts = ls.replace(",", " ").split()
            floats = []
            for p in parts:
                try:
                    floats.append(float(p))
                except ValueError:
                    break
            if len(floats) >= 2:
                r_vals.append(floats[0])
                g_vals.append(floats[1])
    if not r_vals:
        raise ValueError(f"No numeric data found in {fname}")
    return np.array(r_vals, dtype=float), np.array(g_vals, dtype=float)


def read_fullprof_rowwise(fname: str):
    with open(fname, "r") as f:
        lines = f.readlines()[1:]
    y_rows = []
    for line in lines:
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        y_rows.extend([float(val) for val in line.split()])
    y = np.array(y_rows)
    return y, len(lines)


def robust_loadtxt_skipheader(fname: str):
    """Skip comments/non-numeric lines and load at least 2-column numeric data.
    
    Flexibly handles comma, space, tab, or mixed delimiters.
    """
    data_lines = []
    with open(fname, "r") as f:
        for line in f:
            ls = line.strip()
            if not ls or ls.startswith("#"):
                continue
            # Normalize delimiters: replace commas and tabs with spaces
            # This handles CSV (comma), TSV (tab), space-separated, and mixed formats
            ls_normalized = ls.replace(",", " ").replace("\t", " ")
            floats = []
            for p in ls_normalized.split():
                try:
                    floats.append(float(p))
                except ValueError:
                    break
            if len(floats) >= 2:
                # Store the normalized line (with all delimiters converted to spaces)
                data_lines.append(ls_normalized)
    if not data_lines:
        raise ValueError(f"No numeric data found in {fname}")
    from io import StringIO
    return np.loadtxt(StringIO("\n".join(data_lines)))


def read_mpt_file(fname: str, mode: str = 'gc', mass_mg: float = None):
    """Read BioLogic .mpt file in various modes.
    
    BioLogic .mpt files come in two formats:
    1. Full EC-Lab format: Complete header with metadata and column names
    2. Simple export: Just 2-3 columns with minimal/no header
    
    This function automatically detects the format and parses accordingly.
    
    Modes Explained:
        - 'gc' (Galvanostatic Cycling): Returns capacity vs voltage curves
          Calculates specific capacity from Q(discharge) and active material mass
          Identifies charge/discharge segments from current sign
          
        - 'cv' (Cyclic Voltammetry): Returns voltage vs current curves
          Used for electrochemical characterization
          
        - 'cpc' (Capacity Per Cycle): Returns cycle statistics
          Extracts max charge/discharge capacity for each cycle
          Calculates coulombic efficiency = Q_discharge / Q_charge * 100
          
        - 'time': Returns time-series data (for operando plots)
          Simple x-y format without cycle processing
    
    Args:
        fname: Path to .mpt file
        mode: Operating mode - 'gc', 'cv', 'cpc', or 'time'
        mass_mg: Active material mass in milligrams
                Required for 'gc' and 'cpc' modes to calculate specific capacity
                Units: mAh/g = (mAh) / (mg / 1000)
    
    Returns:
        Depends on mode:
        
        'gc' mode: Tuple[np.ndarray, np.ndarray, np.ndarray, np.ndarray, np.ndarray]
            (specific_capacity, voltage, cycles, charge_mask, discharge_mask)
            - specific_capacity: Specific capacity in mAh/g
            - voltage: Voltage in V
            - cycles: Cycle number for each data point
            - charge_mask: Boolean array, True for charging points
            - discharge_mask: Boolean array, True for discharging points
            
        'cv' mode: Tuple[np.ndarray, np.ndarray, np.ndarray]
            (voltage, current, cycles)
            
        'cpc' mode: Tuple[np.ndarray, np.ndarray, np.ndarray, np.ndarray]
            (cycle_nums, cap_charge, cap_discharge, efficiency)
            - cycle_nums: Array of cycle numbers [1, 2, 3, ...]
            - cap_charge: Max specific capacity for each cycle (charge)
            - cap_discharge: Max specific capacity for each cycle (discharge)
            - efficiency: Coulombic efficiency % for each cycle
            
        'time' mode: Tuple[np.ndarray, np.ndarray, np.ndarray, str, str]
            (time, voltage, current, x_label, y_label)
            For simple 2-column files: returns (x, y, empty_array, 'x', 'y')
    
    Raises:
        ValueError: If mass_mg not provided for 'gc' or 'cpc' mode
        FileNotFoundError: If file doesn't exist
        Exception: If file format is invalid or columns not found
    
    File Format Notes:
        - EC-Lab format starts with "EC-Lab ASCII FILE"
        - Header contains "Nb header lines" to skip
        - Column names may be in French or English
        - Simple exports are just tab/space-separated numbers
    """
    import re
    
    # === STEP 1: Detect file format ===
    # EC-Lab files start with specific marker, simple exports don't
    is_eclab_format = False
    with open(fname, 'r', encoding='utf-8', errors='ignore') as f:
        first_line = f.readline().strip()
        if first_line.startswith('EC-Lab ASCII FILE'):
            is_eclab_format = True
    
    # Handle simple 2-column time/voltage export format (for operando time mode)
    if not is_eclab_format and mode == 'time':
        try:
            # Read with flexible delimiter (tab or whitespace) and handle European comma decimal separator
            with open(fname, 'r', encoding='utf-8', errors='ignore') as f:
                x_vals = []
                y_vals = []
                x_label = 'x'
                y_label = 'y'
                first_line_processed = False
                
                for line in f:
                    line = line.strip()
                    if not line:
                        continue
                    
                    # Try to parse as numeric data
                    # First try tab-separated, then space-separated
                    parts = line.split('\t') if '\t' in line else line.split()
                    
                    if len(parts) >= 2:
                        try:
                            # Try to parse as numbers (replace comma with period for European locale)
                            x_val = float(parts[0].replace(',', '.'))
                            y_val = float(parts[1].replace(',', '.'))
                            x_vals.append(x_val)
                            y_vals.append(y_val)
                            first_line_processed = True
                        except ValueError:
                            # This line contains non-numeric data (likely header)
                            if not first_line_processed:
                                # This is a header line, extract column names
                                x_label = parts[0].strip()
                                y_label = parts[1].strip() if len(parts) > 1 else 'y'
                                continue
                            else:
                                # Stop reading if we hit non-numeric after data started
                                break
                
                if not x_vals:
                    raise ValueError("No data found in file")
                
                x_data = np.array(x_vals)
                y_data = np.array(y_vals)
                current_mA = np.zeros_like(x_data)  # No current data in simple format
                
                # Return raw data without conversion, and include column labels
                return x_data, y_data, current_mA, x_label, y_label
        except Exception as e:
            raise ValueError(f"Failed to read simple .mpt format: {e}")
    
    # For non-time modes or EC-Lab format, require full EC-Lab format
    if not is_eclab_format:
        raise ValueError(f"Not a valid EC-Lab .mpt file: {fname}")
    
    # Read header to find number of header lines
    header_lines = 0
    with open(fname, 'r', encoding='utf-8', errors='ignore') as f:
        first_line = f.readline().strip()
        
        # Find header lines count
        for line in f:
            if line.startswith('Nb header lines'):
                match = re.search(r'Nb header lines\s*:\s*(\d+)', line)
                if match:
                    header_lines = int(match.group(1))
                    break
        if header_lines == 0:
            raise ValueError(f"Could not find header line count in {fname}")
    
    # Read the data
    data_lines = []
    column_names = []
    
    with open(fname, 'r', encoding='utf-8', errors='ignore') as f:
        # Skip header lines
        for i in range(header_lines - 1):
            f.readline()
        
        # Read column names (should be at header_lines - 1)
        header_line = f.readline().strip()
        column_names = [col.strip() for col in header_line.split('\t')]
        
        # Read data lines
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                # Replace comma decimal separator with period (European locale support)
                values = [float(val.replace(',', '.')) for val in line.split('\t')]
                if len(values) == len(column_names):
                    data_lines.append(values)
            except ValueError:
                continue
    
    if not data_lines:
        raise ValueError(f"No valid data found in {fname}")
    
    # Convert to numpy array
    data = np.array(data_lines)
    
    # Create column index mapping
    col_map = {name: i for i, name in enumerate(column_names)}
    
    if mode == 'gc':
        # Galvanostatic cycling: use BioLogic's Q charge and Q discharge columns
        if mass_mg is None or mass_mg <= 0:
            raise ValueError("Mass loading (in mg) is required and must be positive for GC mode. Use --mass parameter.")

        mass_g = float(mass_mg) / 1000.0

        # Skip first line of data as requested
        data = data[1:]

        # Required columns - try common variations
        voltage_col = col_map.get('Ewe/V', None)
        if voltage_col is None:
            voltage_col = col_map.get('Ewe', None)
        
        q_charge_col = col_map.get('Q charge/mA.h', None)
        if q_charge_col is None:
            q_charge_col = col_map.get('Q charge/mAh', None)
        
        q_discharge_col = col_map.get('Q discharge/mA.h', None)
        if q_discharge_col is None:
            q_discharge_col = col_map.get('Q discharge/mAh', None)
        
        if voltage_col is None:
            available = ', '.join(f"'{c}'" for c in column_names)
            raise ValueError(f"Could not find 'Ewe/V' or 'Ewe' column for voltage.\nAvailable columns: {available}")
        if q_charge_col is None:
            available = ', '.join(f"'{c}'" for c in column_names)
            raise ValueError(f"Could not find 'Q charge/mA.h' or 'Q charge/mAh' column.\nAvailable columns: {available}")
        if q_discharge_col is None:
            available = ', '.join(f"'{c}'" for c in column_names)
            raise ValueError(f"Could not find 'Q discharge/mA.h' or 'Q discharge/mAh' column.\nAvailable columns: {available}")

        voltage = data[:, voltage_col]
        q_charge = data[:, q_charge_col]
        q_discharge = data[:, q_discharge_col]
        
        n = len(voltage)
        
        # Determine if experiment starts with charge or discharge
        # by checking which Q column increases first
        starts_with_charge = None
        for i in range(min(100, n - 1)):
            if q_charge[i+1] > q_charge[i] + 1e-6:
                starts_with_charge = True
                break
            elif q_discharge[i+1] > q_discharge[i] + 1e-6:
                starts_with_charge = False
                break
        
        if starts_with_charge is None:
            # Default to charge if no clear increase detected
            starts_with_charge = True
        
        # Detect charge/discharge segments based on when Q values drop to 0
        # The end of charge is when Q charge drops to ~0
        # The end of discharge is when Q discharge drops to ~0
        is_charge = np.zeros(n, dtype=bool)
        
        # Set initial state
        current_is_charge = starts_with_charge
        is_charge[0] = current_is_charge
        
        # Detect segment boundaries by finding where Q values reset to ~0
        for i in range(1, n):
            if current_is_charge:
                # We're in a charge segment
                # End of charge is when Q charge drops to near 0
                if q_charge[i] < 1e-10 and q_charge[i-1] > 1e-6:
                    # Q charge just dropped to 0, switch to discharge
                    current_is_charge = False
            else:
                # We're in a discharge segment
                # End of discharge is when Q discharge drops to near 0
                if q_discharge[i] < 1e-10 and q_discharge[i-1] > 1e-6:
                    # Q discharge just dropped to 0, switch to charge
                    current_is_charge = True
            
            is_charge[i] = current_is_charge
        
        # Find charge/discharge segment boundaries
        run_starts = [0]
        for k in range(1, n):
            if is_charge[k] != is_charge[k-1]:
                run_starts.append(k)
        run_starts.append(n)
        
        # Create masks
        charge_mask = is_charge
        discharge_mask = ~is_charge
        
        # Calculate specific capacity for each segment, starting from 0
        specific_capacity = np.zeros(n, dtype=float)
        
        for r in range(len(run_starts) - 1):
            start_idx = run_starts[r]
            end_idx = run_starts[r + 1]
            
            if is_charge[start_idx]:
                # Use Q charge column
                q_values = q_charge[start_idx:end_idx]
            else:
                # Use Q discharge column
                q_values = q_discharge[start_idx:end_idx]
            
            # Reset capacity to start from 0 for this segment
            q_start = q_values[0]
            specific_capacity[start_idx:end_idx] = (q_values - q_start) / mass_g
        
        # Assign cycle numbers: each full charge-discharge or discharge-charge pair is one cycle
        cycle_numbers = np.zeros(n, dtype=int)
        current_cycle = 1
        half_cycle = 0  # Track if we're on first or second half of cycle
        
        for r in range(len(run_starts) - 1):
            start_idx = run_starts[r]
            end_idx = run_starts[r + 1]
            
            cycle_numbers[start_idx:end_idx] = current_cycle
            
            half_cycle += 1
            if half_cycle == 2:
                # Completed one full cycle (charge+discharge or discharge+charge)
                current_cycle += 1
                half_cycle = 0

        return (specific_capacity, voltage, cycle_numbers, charge_mask, discharge_mask)
    
    elif mode == 'time':
        # Time series: time vs voltage/current
        time_col = col_map.get('time/s', None)
        voltage_col = col_map.get('Ewe/V', None)
        if voltage_col is None:
            voltage_col = col_map.get('Ewe', None)
        current_col = col_map.get('<I>/mA', None)
        
        if time_col is None:
            available = ', '.join(f"'{c}'" for c in column_names)
            raise ValueError(f"Could not find 'time/s' column.\nAvailable columns: {available}")
        if voltage_col is None:
            available = ', '.join(f"'{c}'" for c in column_names)
            raise ValueError(f"Could not find 'Ewe/V' or 'Ewe' column.\nAvailable columns: {available}")
        
        time_data = data[:, time_col]
        voltage_data = data[:, voltage_col]
        
        # Current column is optional (only needed for advanced features like ion counting)
        if current_col is None:
            # Return None for current_data if column not found
            current_data = None
        else:
            current_data = data[:, current_col]
        
        # For EC-Lab files, return standard labels
        return (time_data, voltage_data, current_data, 'Time (h)', 'Voltage (V)')
    
    elif mode == 'cv':
        # Cyclic voltammetry: voltage vs current, split by cycle
        voltage_col = col_map.get('Ewe/V', None)
        if voltage_col is None:
            voltage_col = col_map.get('Ewe', None)
        current_col = col_map.get('<I>/mA', None)
        cycle_col = col_map.get('cycle number', None)
        
        if voltage_col is None:
            available = ', '.join(f"'{c}'" for c in column_names)
            raise ValueError(f"Could not find 'Ewe/V' or 'Ewe' column for voltage.\nAvailable columns: {available}")
        if current_col is None:
            available = ', '.join(f"'{c}'" for c in column_names)
            raise ValueError(f"Could not find '<I>/mA' column for current.\nAvailable columns: {available}")
        
        voltage = data[:, voltage_col]
        current = data[:, current_col]
        if cycle_col is not None:
            cycles = data[:, cycle_col].astype(int)
        else:
            cycles = np.ones(len(voltage), dtype=int)
        return voltage, current, cycles
    elif mode == 'cpc':
        # Capacity-per-cycle: extract end-of-segment charge/discharge capacities and efficiency
        if mass_mg is None or mass_mg <= 0:
            raise ValueError("Mass loading (mg) is required and must be positive for CPC mode. Use --mass.")

        mass_g = float(mass_mg) / 1000.0

        # Skip first line of data
        data = data[1:]

        # Required columns - try common variations
        q_charge_col = col_map.get('Q charge/mA.h', None)
        if q_charge_col is None:
            q_charge_col = col_map.get('Q charge/mAh', None)
        
        q_discharge_col = col_map.get('Q discharge/mA.h', None)
        if q_discharge_col is None:
            q_discharge_col = col_map.get('Q discharge/mAh', None)
        
        if q_charge_col is None:
            available = ', '.join(f"'{c}'" for c in column_names)
            raise ValueError(f"Could not find 'Q charge/mA.h' or 'Q charge/mAh' column.\nAvailable columns: {available}")
        if q_discharge_col is None:
            available = ', '.join(f"'{c}'" for c in column_names)
            raise ValueError(f"Could not find 'Q discharge/mA.h' or 'Q discharge/mAh' column.\nAvailable columns: {available}")

        q_charge = data[:, q_charge_col]
        q_discharge = data[:, q_discharge_col]
        
        n = len(q_charge)
        
        # Determine if experiment starts with charge or discharge
        starts_with_charge = None
        for i in range(min(100, n - 1)):
            if q_charge[i+1] > q_charge[i] + 1e-6:
                starts_with_charge = True
                break
            elif q_discharge[i+1] > q_discharge[i] + 1e-6:
                starts_with_charge = False
                break
        
        if starts_with_charge is None:
            starts_with_charge = True
        
        # Detect segment boundaries by finding where Q values reset to ~0
        is_charge = np.zeros(n, dtype=bool)
        current_is_charge = starts_with_charge
        is_charge[0] = current_is_charge
        
        for i in range(1, n):
            if current_is_charge:
                if q_charge[i] < 1e-10 and q_charge[i-1] > 1e-6:
                    current_is_charge = False
            else:
                if q_discharge[i] < 1e-10 and q_discharge[i-1] > 1e-6:
                    current_is_charge = True
            is_charge[i] = current_is_charge
        
        # Find segment boundaries
        run_starts = [0]
        for k in range(1, n):
            if is_charge[k] != is_charge[k-1]:
                run_starts.append(k)
        run_starts.append(n)
        
        # Extract end-of-segment capacities
        cyc_nums = []
        cap_charge_spec = []
        cap_discharge_spec = []
        eff_percent = []
        
        current_cycle = 1
        half_cycle = 0
        cycle_charge_cap = np.nan
        cycle_discharge_cap = np.nan
        
        for r in range(len(run_starts) - 1):
            start_idx = run_starts[r]
            end_idx = run_starts[r + 1]
            
            if is_charge[start_idx]:
                # Charge segment: get capacity at end (just before it resets)
                # Use the last valid value before segment ends
                end_cap = q_charge[end_idx - 1] if end_idx > start_idx else 0.0
                cycle_charge_cap = end_cap / mass_g
            else:
                # Discharge segment: get capacity at end
                end_cap = q_discharge[end_idx - 1] if end_idx > start_idx else 0.0
                cycle_discharge_cap = end_cap / mass_g
            
            half_cycle += 1
            if half_cycle == 2:
                # Completed one full cycle
                cyc_nums.append(current_cycle)
                cap_charge_spec.append(cycle_charge_cap)
                cap_discharge_spec.append(cycle_discharge_cap)
                
                # Calculate efficiency
                if np.isfinite(cycle_charge_cap) and cycle_charge_cap > 0 and np.isfinite(cycle_discharge_cap):
                    eff = (cycle_discharge_cap / cycle_charge_cap) * 100.0
                else:
                    eff = np.nan
                eff_percent.append(eff)
                
                # Reset for next cycle
                current_cycle += 1
                half_cycle = 0
                cycle_charge_cap = np.nan
                cycle_discharge_cap = np.nan

        return (np.array(cyc_nums, dtype=float),
                np.array(cap_charge_spec, dtype=float),
                np.array(cap_discharge_spec, dtype=float),
                np.array(eff_percent, dtype=float))

    else:
        raise ValueError(f"Unknown mode '{mode}'. Use 'gc', 'time', or 'cpc'.")


def read_biologic_txt_file(fname: str, mode: str = 'cv') -> Tuple[np.ndarray, np.ndarray, np.ndarray]:
    """Read BioLogic tab-separated text export (simplified format without EC-Lab header).
    
    These .txt files have a single header line with tab-separated column names,
    followed by tab-separated data rows. Common format from BioLogic EC-Lab exports.
    
    Args:
        fname: Path to .txt file
        mode: Currently only 'cv' is supported (cyclic voltammetry)
    
    Returns:
        For 'cv' mode: (voltage, current, cycles)
    """
    data_lines = []
    column_names = []
    
    with open(fname, 'r', encoding='utf-8', errors='ignore') as f:
        # First line is the header
        header_line = f.readline().strip()
        column_names = [col.strip() for col in header_line.split('\t')]
        
        # Read data lines
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                # Replace comma decimal separator with period (European locale support)
                values = [float(val.replace(',', '.')) for val in line.split('\t')]
                if len(values) == len(column_names):
                    data_lines.append(values)
            except ValueError:
                continue
    
    if not data_lines:
        raise ValueError(f"No valid data found in {fname}")
    
    # Convert to numpy array
    data = np.array(data_lines)
    
    # Create column index mapping
    col_map = {name: i for i, name in enumerate(column_names)}
    
    if mode == 'cv':
        # Cyclic voltammetry: voltage vs current, split by cycle
        voltage_col = col_map.get('Ewe/V', None)
        if voltage_col is None:
            voltage_col = col_map.get('Ewe', None)
        current_col = col_map.get('<I>/mA', None)
        cycle_col = col_map.get('cycle number', None)
        
        if voltage_col is None:
            available = ', '.join(f"'{c}'" for c in column_names)
            raise ValueError(f"Could not find 'Ewe/V' or 'Ewe' column for voltage.\nAvailable columns: {available}")
        if current_col is None:
            available = ', '.join(f"'{c}'" for c in column_names)
            raise ValueError(f"Could not find '<I>/mA' column for current.\nAvailable columns: {available}")
        
        voltage = data[:, voltage_col]
        current = data[:, current_col]
        if cycle_col is not None:
            cycles = data[:, cycle_col].astype(int)
        else:
            cycles = np.ones(len(voltage), dtype=int)
        return voltage, current, cycles
    else:
        raise ValueError(f"Unknown mode '{mode}' for .txt file. Currently only 'cv' is supported.")


def read_ec_csv_file(fname: str, prefer_specific: bool = True) -> Tuple[np.ndarray, np.ndarray, np.ndarray, np.ndarray, np.ndarray]:
    """Read battery cycler CSV or Excel file and extract galvanostatic cycling data.
    
    This function handles data files exported from battery cyclers (primarily Neware format),
    supporting both CSV and Excel formats with optional Chinese column names. It implements
    sophisticated half-cycle merging logic to properly pair charge/discharge segments into
    complete cycles.
    
    Supported File Formats:
    -----------------------
    CSV Files:
        - Two-line header format (Neware standard):
            * Line 1: Main column names
            * Line 2: Continues with additional columns (first cell empty)
        - Single-line header format (simplified exports)
        - UTF-8 encoding with error tolerance for malformed characters
        
    Excel Files (.xlsx, .xls):
        - Row 1: File/sample metadata (ignored)
        - Row 2: Column headers (Chinese or English)
        - Row 3+: Data rows
        - Automatically converted to CSV-like format internally
    
    Column Detection (Flexible Fuzzy Matching):
    -------------------------------------------
    Required columns (at least one voltage/current variant):
        - Voltage: 'Voltage(V)' or Chinese equivalents ('充电中压/V', '放电中压/V', etc.)
        - Current: 'Current(mA)' (optional for summary files)
        - Cycle Index: 'Cycle Index' or '循环序号' (used for summary exports; per-point data
          infers cycles from the charge/discharge order)
    
    Capacity columns (prioritized by prefer_specific flag):
        Specific capacity (mAh/g):
            - 'Spec. Cap.(mAh/g)' - combined capacity
            - 'Chg. Spec. Cap.(mAh/g)' or '充电比容量/mAh/g' - charge only
            - 'DChg. Spec. Cap.(mAh/g)' or '放电比容量/mAh/g' - discharge only
        
        Absolute capacity (mAh):
            - 'Capacity(mAh)' - combined capacity
            - 'Chg. Cap.(mAh)' or '充电容量/mAh' - charge only
            - 'DChg. Cap.(mAh)' or '放电容量/mAh' - discharge only
    
    Optional columns:
        - 'Step Type': Explicit charge/discharge indicator ('CC Chg', 'CC DChg', etc.)
          Used to determine charge_mask if present, otherwise inferred from capacity columns
    
    Half-Cycle Merging Algorithm:
    ------------------------------
    Many cyclers export data where each charge and discharge is a separate segment.
    This function merges them into complete cycles:
    
    1. Detect segment boundaries:
        - Use 'Step Type' column if available
        - Otherwise, infer from split capacity columns (charge-only vs discharge-only)
        - Fallback: detect by voltage trend (increasing = charge, decreasing = discharge)
    
    2. Merge logic:
        - Pair consecutive segments as complete cycles regardless of the cycler's own
          numbering (makes Cycle 1 = first charge+discharge, or discharge+charge)
        - Handles odd number of segments (incomplete final cycle) by keeping the lone
          half-cycle with the previous cycle
    
    3. Capacity continuity:
        - Charge segments: capacity increases from 0 → max_charge
        - Discharge segments: capacity continues from max_charge → max_charge + discharge_capacity
        - Ensures continuous X-axis for plotting (no gaps between charge/discharge)
    
    Summary File Detection:
    -----------------------
    Automatically detects summary files (cycle statistics without per-point data):
        - Has capacity columns but missing 'Voltage(V)' or 'Current(mA)'
        - Returns aggregated cycle data instead of per-point measurements
        - Used for cycle life plots (capacity retention vs cycle number)
    
    Args:
        fname: Path to CSV or Excel file (.csv, .xlsx, .xls extensions supported).
        prefer_specific: Capacity unit preference flag (default=True).
            - True: Prioritize specific capacity (mAh/g) if available, fallback to absolute (mAh)
            - False: Prioritize absolute capacity (mAh) if available, fallback to specific (mAh/g)
            - Useful when files contain both units and you want to standardize plots
    
    Returns:
        Tuple of 5 numpy arrays for galvanostatic cycling plots:
            capacity_x (np.ndarray): X-axis capacity values for plotting
                - Units: mAh/g (specific) or mAh (absolute) depending on file and prefer_specific flag
                - Length: N data points
                - Continuous across charge/discharge (discharge starts where charge ends)
                - Example: [0, 50, 100, 150, 100, 50, 0] for one cycle (charge 0→150, discharge 150→0)
            
            voltage (np.ndarray): Y-axis voltage values in V
                - Length: N data points (matches capacity_x)
                - Typical range: 2.5-4.2 V for Li-ion cells
            
            cycle_numbers (np.ndarray): Cycle index for each data point
                - Length: N data points
                - dtype: int
                - Values: 1, 2, 3, ... (1-indexed)
                - Always inferred by pairing alternating charge/discharge runs in
                  chronological order (ignoring half-cycle numbering in the file)
            
            charge_mask (np.ndarray): Boolean mask indicating charging data points
                - Length: N data points
                - dtype: bool
                - True where current > 0 (charging), False otherwise
                - Determined by 'Step Type' column if present, otherwise inferred from:
                    * Split capacity columns (Chg. Cap. vs DChg. Cap.)
                    * Voltage trend (increasing = charge, decreasing = discharge)
                - Used to apply different colors/markers to charge vs discharge in plots
            
            discharge_mask (np.ndarray): Boolean mask indicating discharging data points
                - Length: N data points
                - dtype: bool
                - Exactly inverse of charge_mask (True where charge_mask is False)
                - Used for styling discharge curves differently in plots
    
    Raises:
        ValueError: If CSV/Excel file is malformed or missing required columns:
            - Empty file or missing header rows
            - No 'Voltage(V)' or 'Current(mA)' columns (unless summary file)
            - No valid capacity columns found
            - File encoding errors (though UTF-8 errors='ignore' provides tolerance)
        FileNotFoundError: If fname path does not exist
        openpyxl errors: If Excel file is corrupted or unsupported format
    
    Examples:
        >>> # Read Neware CSV with specific capacity preference
        >>> cap, v, cyc, chg_mask, dchg_mask = read_ec_csv_file('neware_export.csv', prefer_specific=True)
        >>> print(f"Loaded {len(cap)} points across {cyc.max()} cycles")
        >>> print(f"Capacity range: {cap.min():.1f} - {cap.max():.1f} mAh/g")
        >>> 
        >>> # Plot charge and discharge with different colors
        >>> import matplotlib.pyplot as plt
        >>> plt.plot(cap[chg_mask], v[chg_mask], 'r-', label='Charge')
        >>> plt.plot(cap[dchg_mask], v[dchg_mask], 'b-', label='Discharge')
        >>> 
        >>> # Read Excel file with Chinese headers, prefer absolute capacity
        >>> cap, v, cyc, _, _ = read_ec_csv_file('循环数据.xlsx', prefer_specific=False)
        >>> # cap will be in mAh if '充电容量/mAh' columns present
    
    Notes:
        - Chinese column name support: Automatically translates common Chinese headers
          (循环序号, 充电比容量/mAh/g, etc.) to English equivalents for processing
        - Memory efficiency: Uses numpy arrays for fast vectorized operations on large datasets
        - Fuzzy matching: Column detection is case-sensitive but exact match on stripped header text
        - Half-cycle files: Many cyclers export charge and discharge as separate segments;
          this function intelligently merges them into complete cycles for proper visualization
        - Summary vs detail: Automatically detects summary files (one row per cycle) vs
          detailed files (many points per cycle) and handles appropriately
    """
    import csv
    import os
    
    # Check if file is Excel
    _, ext = os.path.splitext(fname)
    if ext.lower() in ['.xlsx', '.xls']:
        # Read Excel file
        header, rows = read_excel_to_csv_like(fname)
    else:
        # Read first two rows to compose header from CSV
        with open(fname, newline='', encoding='utf-8', errors='ignore') as f:
            r = csv.reader(f)
            try:
                r1 = next(r)
                r2 = next(r)
            except StopIteration:
                raise ValueError(f"CSV '{fname}' is empty or missing header rows")
            # If second line begins with an empty first cell, treat it as a continuation of header.
            if len(r2) > 0 and (r2[0] == '' or str(r2[0]).strip() == ''):
                header = [c.strip() for c in r1] + [c.strip() for c in r2[1:]]
                rows = list(r)
            else:
                # Single-line header: r2 is the first data row; include it back in rows
                header = [c.strip() for c in r1]
                rows = [r2] + list(r)

    # Build fast name->index map (case-insensitive match on exact header text)
    name_to_idx = {h: i for i, h in enumerate(header)}
    
    # Chinese to English column name mappings
    chinese_mappings = {
        '循环序号': 'Cycle Index',
        '充电比容量/mAh/g': 'Chg. Spec. Cap.(mAh/g)',
        '放电比容量/mAh/g': 'DChg. Spec. Cap.(mAh/g)',
        '充电容量/mAh': 'Chg. Cap.(mAh)',
        '放电容量/mAh': 'DChg. Cap.(mAh)',
        '效率/%': 'Efficiency(%)',
        '充电中压/V': 'Voltage(V)',
        '放电中压/V': 'Voltage(V)',
        '充电均压/V': 'Voltage(V)',
        '放电均压/V': 'Voltage(V)',
    }
    
    # Add Chinese mappings to name_to_idx
    for i, h in enumerate(header):
        if h in chinese_mappings:
            eng_name = chinese_mappings[h]
            if eng_name not in name_to_idx:
                name_to_idx[eng_name] = i

    def _find(name: str):
        return name_to_idx.get(name, None)

    # Required columns
    v_idx = _find('Voltage(V)')
    i_idx = _find('Current(mA)')
    cyc_idx = _find('Cycle Index')
    step_type_idx = _find('Step Type')  # Optional: explicitly indicates charge/discharge
    
    # Capacity columns (absolute preferred unless prefer_specific True)
    cap_abs_idx = _find('Capacity(mAh)')
    cap_abs_chg_idx = _find('Chg. Cap.(mAh)')
    cap_abs_dch_idx = _find('DChg. Cap.(mAh)')
    cap_spec_idx = _find('Spec. Cap.(mAh/g)')
    cap_spec_chg_idx = _find('Chg. Spec. Cap.(mAh/g)')
    cap_spec_dch_idx = _find('DChg. Spec. Cap.(mAh/g)')
    
    # Check if this is a summary file (has capacity columns but no voltage/current)
    has_capacity_cols = any([cap_abs_chg_idx, cap_abs_dch_idx, cap_spec_chg_idx, cap_spec_dch_idx])
    is_summary_file = has_capacity_cols and (v_idx is None or i_idx is None)
    
    if not is_summary_file and (v_idx is None or i_idx is None):
        raise ValueError("CSV missing required 'Voltage(V)' or 'Current(mA)' columns")

    use_specific = False
    # Decide which flavor to use
    if prefer_specific and (cap_spec_chg_idx is not None or cap_spec_idx is not None):
        use_specific = True
    elif not prefer_specific and (cap_abs_chg_idx is not None or cap_abs_idx is not None):
        use_specific = False
    elif cap_abs_chg_idx is None and cap_abs_idx is None and (cap_spec_idx is not None or cap_spec_chg_idx is not None):
        use_specific = True
    # else: fallback stays False (absolute) if both missing we'll error later

    # Prepare arrays
    n = len(rows)
    
    def _to_float(val: str) -> float:
        try:
            return float(val.strip()) if isinstance(val, str) else float(val)
        except Exception:
            return np.nan
    
    # Special handling for summary files (charge/discharge capacities per cycle, no point-by-point data)
    if is_summary_file:
        # For summary files, create synthetic points: one charge point and one discharge point per cycle
        voltage = []
        current = []
        cycles = []
        cap_x = []
        is_charge_list = []
        
        for k, row in enumerate(rows):
            if len(row) < len(header):
                row = row + [''] * (len(header) - len(row))
            
            # Get cycle number
            cycle_num = 1
            if cyc_idx is not None:
                cval = _to_float(row[cyc_idx])
                cycle_num = int(cval) if not np.isnan(cval) and cval > 0 else 1
            
            # Get charge and discharge capacities
            if use_specific:
                cap_chg = _to_float(row[cap_spec_chg_idx]) if cap_spec_chg_idx is not None else 0
                cap_dch = _to_float(row[cap_spec_dch_idx]) if cap_spec_dch_idx is not None else 0
            else:
                cap_chg = _to_float(row[cap_abs_chg_idx]) if cap_abs_chg_idx is not None else 0
                cap_dch = _to_float(row[cap_abs_dch_idx]) if cap_abs_dch_idx is not None else 0
            
            # Create charge point
            if cap_chg > 0 and not np.isnan(cap_chg):
                voltage.append(3.5)  # Synthetic voltage
                current.append(0.1)  # Synthetic current
                cycles.append(cycle_num)
                cap_x.append(cap_chg)
                is_charge_list.append(True)
            
            # Create discharge point
            if cap_dch > 0 and not np.isnan(cap_dch):
                voltage.append(2.5)  # Synthetic voltage
                current.append(-0.1)  # Synthetic current
                cycles.append(cycle_num)
                cap_x.append(cap_dch)
                is_charge_list.append(False)
        
        voltage = np.array(voltage, dtype=float)
        current = np.array(current, dtype=float)
        cycles = np.array(cycles, dtype=int)
        cap_x = np.array(cap_x, dtype=float)
        is_charge = np.array(is_charge_list, dtype=bool)
        
        charge_mask = is_charge
        discharge_mask = ~is_charge
        
        return (cap_x, voltage, cycles, charge_mask, discharge_mask)
    
    # Normal processing for point-by-point data
    voltage = np.empty(n, dtype=float)
    current = np.empty(n, dtype=float)
    cap_x = np.full(n, np.nan, dtype=float)

    for k, row in enumerate(rows):
        # Ensure row has enough columns
        if len(row) < len(header):
            row = row + [''] * (len(header) - len(row))
        v = _to_float(row[v_idx])
        i = _to_float(row[i_idx])
        voltage[k] = v
        current[k] = i
        # Don't decide chg/dchg capacity here; we will assign after deriving direction
        # Fill combined capacity columns if present (used when split columns missing)
        if use_specific and cap_spec_idx is not None:
            cap_x[k] = _to_float(row[cap_spec_idx])
        elif (not use_specific) and cap_abs_idx is not None:
            cap_x[k] = _to_float(row[cap_abs_idx])

    # --- Derive charge/discharge from Step Type column, capacity columns, or voltage trend ---
    # Priority 1: Use explicit Step Type column (e.g., "CC Chg", "CC DChg", "Charge", "Discharge")
    is_charge = np.zeros(n, dtype=bool)
    is_rest_segment = np.zeros(n, dtype=bool)
    is_rest_segment = np.zeros(n, dtype=bool)
    used_step_type = False
    used_capacity_columns = False
    
    if step_type_idx is not None:
        # Parse Step Type to determine charge/discharge
        # We'll also track which rows are Rest/CV/other non-charge/discharge steps
        is_rest_or_other = np.zeros(n, dtype=bool)
        
        for k, row in enumerate(rows):
            if len(row) < len(header):
                row = row + [''] * (len(header) - len(row))
            step_type = str(row[step_type_idx]).strip().lower()
            
            # Check for Rest or other non-active steps
            is_rest = 'rest' in step_type or 'pause' in step_type or 'cv' in step_type
            
            # Check for discharge first (since "discharge" contains "charge")
            # Discharge indicators: "dchg", "dischg", "discharge", "cc dchg", etc.
            is_dchg = 'dchg' in step_type or 'dischg' in step_type or step_type.startswith('dis')
            # Then check for charge indicators: "chg", "charge", "cc chg", etc.
            is_chg = (not is_dchg) and (not is_rest) and (('chg' in step_type) or ('charge' in step_type))
            
            # Mark Rest/CV steps to exclude them from masks
            if is_rest:
                is_rest_or_other[k] = True
                is_charge[k] = False  # Will be excluded from both masks
            elif is_chg:
                is_charge[k] = True
            elif is_dchg:
                is_charge[k] = False
            else:
                # For unknown types, inherit from previous row (or default to False)
                is_charge[k] = is_charge[k-1] if k > 0 else False
        used_step_type = True
    
    # Priority 2: Use split charge/discharge capacity columns if available
    elif (use_specific and cap_spec_chg_idx is not None and cap_spec_dch_idx is not None) or \
         (not use_specific and cap_abs_chg_idx is not None and cap_abs_dch_idx is not None):
        # Read capacity columns to determine charge/discharge
        # Whichever column has non-zero value indicates the segment type
        cap_chg_vals = np.empty(n, dtype=float)
        cap_dch_vals = np.empty(n, dtype=float)
        
        if use_specific:
            chg_col_idx = cap_spec_chg_idx
            dch_col_idx = cap_spec_dch_idx
        else:
            chg_col_idx = cap_abs_chg_idx
            dch_col_idx = cap_abs_dch_idx
        
        for k, row in enumerate(rows):
            if len(row) < len(header):
                row = row + [''] * (len(header) - len(row))
            cap_chg_vals[k] = _to_float(row[chg_col_idx])
            cap_dch_vals[k] = _to_float(row[dch_col_idx])
        
        # A point is charge if its charge capacity > threshold AND discharge capacity ≈ 0
        # A point is discharge if its discharge capacity > threshold AND charge capacity ≈ 0
        threshold = 1e-6
        for k in range(n):
            chg_val = cap_chg_vals[k] if not np.isnan(cap_chg_vals[k]) else 0.0
            dch_val = cap_dch_vals[k] if not np.isnan(cap_dch_vals[k]) else 0.0
            
            if chg_val > threshold and dch_val <= threshold:
                is_charge[k] = True
            elif dch_val > threshold and chg_val <= threshold:
                is_charge[k] = False
            else:
                # Both zero or both non-zero: inherit from previous or default
                is_charge[k] = is_charge[k-1] if k > 0 else True
        used_capacity_columns = True
    
    # Priority 3: Fallback to voltage trend (robust to current sign conventions)
    else:
        # Compute a tolerant derivative and propagate direction over plateaus (|dv| <= eps)
        # eps based on dynamic range to avoid noise flips
        v_clean = np.array(voltage, dtype=float)
        v_min = np.nanmin(v_clean) if np.isfinite(v_clean).any() else 0.0
        v_max = np.nanmax(v_clean) if np.isfinite(v_clean).any() else 1.0
        v_span = max(1e-6, float(v_max - v_min))
        eps = max(1e-6, 1e-4 * v_span)
        dv = np.diff(v_clean)
        dv = np.nan_to_num(dv, nan=0.0, posinf=0.0, neginf=0.0)
        # Initial direction: first significant dv sets the sign; fallback to current sign if needed
        init_dir = None
        for d in dv[: min(500, dv.size)]:
            if abs(d) > eps:
                init_dir = (d > 0)
                break
        if init_dir is None:
            # Fallback: use sign of first non-zero current; else assume charge
            nz = None
            for i_val in current:
                if abs(i_val) > 1e-12 and np.isfinite(i_val):
                    nz = (i_val >= 0)
                    break
            init_dir = True if nz is None else bool(nz)
        prev_dir = init_dir
        for k in range(n):
            dir_set = None
            # Prefer backward-looking difference to keep the last sample of a run with its run
            if k > 0:
                db = dv[k-1]
                if abs(db) > eps:
                    dir_set = (db > 0)
            # Fallback: look forward to the next informative change (for first sample of a run)
            if dir_set is None:
                j = k
                while j < n-1:
                    d = dv[j]
                    if abs(d) > eps:
                        dir_set = (d > 0)
                        break
                    j += 1
            # If still None (flat series), keep previous
            if dir_set is None:
                dir_set = prev_dir
            is_charge[k] = dir_set
            prev_dir = dir_set

    # Build run-length encoding and optionally merge very short flicker runs
    # (Only apply smoothing when using voltage trend detection, not when using explicit methods)
    if not used_step_type and not used_capacity_columns:
        # Smoothing logic for voltage-trend-based detection
        run_starts = [0]
        for k in range(1, n):
            if is_charge[k] != is_charge[k-1]:
                run_starts.append(k)
        run_starts.append(n)
        # Merge runs shorter than 3 samples (or 0.2% of data length, whichever larger)
        min_len = max(3, int(0.002 * n))
        if len(run_starts) >= 3:
            keep_mask = is_charge.copy()
            new_is_charge = is_charge.copy()
            for r in range(len(run_starts)-1):
                a = run_starts[r]
                b = run_starts[r+1]
                if (b - a) < min_len:
                    # Prefer to merge into previous run if exists; else next
                    if r > 0:
                        new_is_charge[a:b] = new_is_charge[a-1]
                    elif r+1 < len(run_starts)-1:
                        new_is_charge[a:b] = new_is_charge[b]
            is_charge = new_is_charge

    # Compute final run starts for cycle inference
    run_starts = [0]
    for k in range(1, n):
        if is_charge[k] != is_charge[k-1]:
            run_starts.append(k)
    run_starts.append(n)

    # Build masks from voltage trend
    # Exclude Rest/CV steps if they were identified
    if used_step_type and 'is_rest_or_other' in locals():
        charge_mask = is_charge & ~is_rest_or_other
        discharge_mask = ~is_charge & ~is_rest_or_other
    else:
        charge_mask = is_charge
        discharge_mask = ~is_charge

    # Assign capacity per-point when split chg/dchg columns exist, using derived direction
    if use_specific and (cap_spec_chg_idx is not None and cap_spec_dch_idx is not None):
        for k, row in enumerate(rows):
            # Ensure row length
            if len(row) < len(header):
                row = row + [''] * (len(header) - len(row))
            cap_chg = _to_float(row[cap_spec_chg_idx])
            cap_dch = _to_float(row[cap_spec_dch_idx])
            cap_x[k] = cap_chg if is_charge[k] else cap_dch
    elif (not use_specific) and (cap_abs_chg_idx is not None and cap_abs_dch_idx is not None):
        for k, row in enumerate(rows):
            if len(row) < len(header):
                row = row + [''] * (len(header) - len(row))
            cap_chg = _to_float(row[cap_abs_chg_idx])
            cap_dch = _to_float(row[cap_abs_dch_idx])
            cap_x[k] = cap_chg if is_charge[k] else cap_dch

    # If capacity column was missing entirely, raise
    if np.all(np.isnan(cap_x)):
        raise ValueError("No usable capacity columns found in CSV (looked for 'Capacity(mAh)' or 'Spec. Cap.(mAh/g)')")

    # Replace NaNs in capacity by 0 to avoid plotting gaps within valid segments
    # but keep masks to split charge/discharge and cycles (NaN voltage gets dropped later by plotting logic)
    cap_x = np.nan_to_num(cap_x, nan=0.0)

    cycles = _infer_cycles_from_masks(charge_mask, discharge_mask, n)

    return cap_x, voltage, cycles, charge_mask, discharge_mask


def read_ec_csv_dqdv_file(fname: str, prefer_specific: bool = True) -> Tuple[np.ndarray, np.ndarray, np.ndarray, np.ndarray, np.ndarray, str]:
    """Read differential capacity (dQ/dV) data from battery cycler CSV for phase transition analysis.
    
    Differential capacity analysis (dQ/dV vs V plots) is used to identify electrochemical phase
    transitions and reaction mechanisms in battery materials. Peaks in dQ/dV correspond to flat
    voltage plateaus in the galvanostatic profile, revealing redox reactions and structural changes.
    
    This function extracts pre-calculated dQ/dV data from cycler software exports (e.g., Neware),
    supporting both absolute (dQ/dV in mAh/V) and specific (dQm/dV in mAh g⁻¹ V⁻¹) units.
    
    Differential Capacity Theory:
    -----------------------------
    dQ/dV = dQ/dt × dt/dV = I / (dV/dt)
    
    Where:
        - Q: Capacity (charge passed, mAh or mAh/g)
        - V: Voltage (V)
        - I: Current (mA)
        - t: Time (hours)
    
    Physical interpretation:
        - High dQ/dV: Voltage changes slowly with capacity (flat plateau) → sharp peak in plot
        - Low dQ/dV: Voltage changes rapidly with capacity (sloped region) → baseline in plot
        - Peaks identify specific phase transitions, intercalation stages, or side reactions
    
    Common applications:
        - Identifying Li+ intercalation stages in graphite anodes (peaks around 0.1-0.2 V)
        - Detecting phase transitions in cathode materials (NMC, LFP, etc.)
        - Monitoring electrode degradation (peak shift/broadening over cycles)
        - Comparing electrode materials (peak positions reveal thermodynamics)
    
    Expected CSV Format:
    --------------------
    Uses same two-line header format as read_ec_csv_file():
        - Line 1: Main column names
        - Line 2: Additional columns (first cell empty, merged with Line 1)
    
    Required columns:
        - 'Voltage(V)': X-axis for dQ/dV plot (typical range 2.5-4.2 V for Li-ion)
        - At least one dQ/dV column:
            * 'dQm/dV(mAh/V.g)': Specific differential capacity (preferred for comparing materials)
            * 'dQ/dV(mAh/V)': Absolute differential capacity (for single cell analysis)
    
    Optional columns (for charge/discharge classification):
        - 'Step Type': Explicit step identifier ('CC Chg', 'CC DChg', etc.)
        - 'Chg. Spec. Cap.(mAh/g)', 'DChg. Spec. Cap.(mAh/g)': Used to infer direction
        - 'Chg. Cap.(mAh)', 'DChg. Cap.(mAh)': Alternative capacity columns
        - 'Current(mA)': Used as fallback for voltage trend analysis
    
    Charge/Discharge Detection Logic (Priority Order):
    ---------------------------------------------------
    Uses same robust detection algorithm as read_ec_csv_file():
    
    1. **Step Type column** (highest priority, most reliable):
        - Looks for keywords: 'Chg' or 'DChg' in 'Step Type' column
        - Example: 'CC Chg' → charge, 'CC DChg' → discharge
    
    2. **Split capacity columns** (medium priority, reliable for well-formatted exports):
        - If 'Chg. Spec. Cap.(mAh/g)' and 'DChg. Spec. Cap.(mAh/g)' both exist:
            * Non-zero Chg. Cap. → charge segment
            * Non-zero DChg. Cap. → discharge segment
        - Same logic for absolute capacity columns ('Chg. Cap.(mAh)', 'DChg. Cap.(mAh)')
    
    3. **Voltage trend analysis** (fallback, less reliable):
        - Calculate voltage differences: dV[i] = V[i+1] - V[i]
        - dV > 0 (voltage increasing) → charge
        - dV < 0 (voltage decreasing) → discharge
        - Includes smoothing to remove flicker from measurement noise
        - Merges runs shorter than 3 samples (or 0.2% of data) to avoid false transitions
    
    Cycle Inference:
    ----------------
    Cycles are always inferred by pairing alternating charge/discharge segments in
    chronological order:
        - Segment 1 + Segment 2 = Cycle 1 (charge→discharge or discharge→charge)
        - Segment 3 + Segment 4 = Cycle 2
        - etc.

    This guarantees Cycle 1 always contains the first two electrochemical halves regardless
    of how the cycler numbered its rows.
    
    Args:
        fname: Path to CSV file containing differential capacity data.
        prefer_specific: Unit preference flag (default=True).
            - True: Use specific differential capacity (dQm/dV in mAh g⁻¹ V⁻¹) if available
            - False: Use absolute differential capacity (dQ/dV in mAh V⁻¹) if available
            - Fallback: Use whichever unit is present if preferred unit missing
            - Note: Specific capacity allows fair comparison between different active mass loadings
    
    Returns:
        Tuple of 6 elements for differential capacity plotting:
            voltage (np.ndarray): X-axis voltage values in V
                - Length: N data points
                - Typical range: 2.5-4.2 V for Li-ion, 1.0-2.5 V for Li-S, 0-3.5 V for supercapacitors
                - Used as X-axis for dQ/dV plots (dQ/dV vs V)
            
            dqdv (np.ndarray): Y-axis differential capacity values
                - Length: N data points
                - Units: mAh g⁻¹ V⁻¹ (specific) or mAh V⁻¹ (absolute) depending on prefer_specific
                - Typical values: 0-1000 mAh g⁻¹ V⁻¹ for graphite, 0-5000 for some cathodes
                - Peaks correspond to electrochemical reactions / phase transitions
                - May contain NaN values if cycler software couldn't calculate derivative
            
            cycles (np.ndarray): Cycle number for each data point
                - Length: N data points
                - dtype: int
                - Values: 1, 2, 3, ... (1-indexed)
                - Used to separate and color-code different cycles in overlay plots
                - Always inferred by pairing alternating charge/discharge segments so Cycle 1
                  contains the first two electrochemical halves
            
            charge_mask (np.ndarray): Boolean mask indicating charging data points
                - Length: N data points
                - dtype: bool
                - True during charge (positive current), False during discharge
                - Determined by Step Type → split capacity columns → voltage trend (priority order)
                - Used to plot charge/discharge with different styles (e.g., solid vs dashed lines)
            
            discharge_mask (np.ndarray): Boolean mask indicating discharging data points
                - Length: N data points
                - dtype: bool
                - Exactly inverse of charge_mask
                - Used for discharge-specific styling in plots
            
            y_label (str): Formatted axis label for Y-axis with proper LaTeX notation
                - Value: 'dQm/dV (mAh g$^{-1}$ V$^{-1}$)' for specific capacity
                - Value: 'dQ/dV (mAh V$^{-1}$)' for absolute capacity
                - Includes proper superscript formatting for matplotlib rendering
                - Can be used directly as plt.ylabel() argument
    
    Raises:
        ValueError: If CSV file is malformed or missing required data:
            - Empty file or missing header rows
            - No 'Voltage(V)' column found
            - No dQ/dV columns ('dQ/dV(mAh/V)' or 'dQm/dV(mAh/V.g)') found
            - All dQ/dV values are NaN (calculation failed in cycler software)
        FileNotFoundError: If fname path does not exist
        UnicodeDecodeError: Rarely raised due to errors='ignore' flag in file reading
    
    Examples:
        >>> # Read specific differential capacity for material comparison
        >>> v, dqdv, cyc, chg, dchg, ylabel = read_ec_csv_dqdv_file('neware_dqdv.csv', prefer_specific=True)
        >>> print(f"Loaded {len(v)} points, {cyc.max()} cycles")
        >>> print(f"Y-axis label: {ylabel}")  # 'dQm/dV (mAh g$^{-1}$ V$^{-1}$)'
        >>> 
        >>> # Plot charge and discharge dQ/dV curves separately
        >>> import matplotlib.pyplot as plt
        >>> plt.figure()
        >>> plt.plot(v[chg], dqdv[chg], 'r-', label='Charge')
        >>> plt.plot(v[dchg], dqdv[dchg], 'b-', label='Discharge')
        >>> plt.xlabel('Voltage (V)')
        >>> plt.ylabel(ylabel)
        >>> plt.legend()
        >>> 
        >>> # Overlay multiple cycles to show degradation
        >>> for cycle_num in range(1, 6):  # First 5 cycles
        >>>     mask = (cyc == cycle_num) & chg  # Charge only
        >>>     plt.plot(v[mask], dqdv[mask], label=f'Cycle {cycle_num}')
        >>> # Peak shifts/broadening indicate structural changes
        >>> 
        >>> # Identify peaks (phase transitions)
        >>> from scipy.signal import find_peaks
        >>> chg_data = dqdv[chg & (cyc == 1)]  # First charge cycle
        >>> chg_v = v[chg & (cyc == 1)]
        >>> peaks, _ = find_peaks(chg_data, height=100, distance=10)
        >>> print(f"Phase transition voltages: {chg_v[peaks]}")
    
    Notes:
        - dQ/dV calculation quality depends on cycler settings:
            * Smaller voltage steps → better dV resolution → smoother dQ/dV
            * Typical recommendation: 5 mV voltage steps for high-quality dQ/dV
            * GITT or PITT techniques provide best dQ/dV resolution (equilibrium data)
        - Peaks in dQ/dV correspond to inflection points in Q vs V curve (second derivative = 0)
        - Charge and discharge dQ/dV often show hysteresis due to kinetic limitations
        - NaN handling: This function preserves NaN values from cycler software (bad derivatives)
          Plotting code should use `plt.plot(..., 'o-', markevery=lambda x: ~np.isnan(dqdv[x]))`
        - Memory: Uses numpy arrays for efficient handling of large datasets (100k+ points)
        - Cycle inference: Assumes alternating charge/discharge; unusual protocols may need manual cycle assignment
    """
    import csv

    # Read two header rows logic identical to read_ec_csv_file
    with open(fname, newline='', encoding='utf-8', errors='ignore') as f:
        r = csv.reader(f)
        try:
            r1 = next(r)
            r2 = next(r)
        except StopIteration:
            raise ValueError(f"CSV '{fname}' is empty or missing header rows")
        if len(r2) > 0 and (r2[0] == '' or str(r2[0]).strip() == ''):
            header = [c.strip() for c in r1] + [c.strip() for c in r2[1:]]
            rows = list(r)
        else:
            header = [c.strip() for c in r1]
            rows = [r2] + list(r)

    name_to_idx = {h: i for i, h in enumerate(header)}
    def _find(name: str):
        return name_to_idx.get(name, None)

    v_idx = _find('Voltage(V)')
    i_idx = _find('Current(mA)')
    dq_abs_idx = _find('dQ/dV(mAh/V)')
    dq_spec_idx = _find('dQm/dV(mAh/V.g)')
    step_type_idx = _find('Step Type')  # Optional: explicitly indicates charge/discharge
    
    # Also look for capacity columns to help determine charge/discharge
    cap_spec_chg_idx = _find('Chg. Spec. Cap.(mAh/g)')
    cap_spec_dch_idx = _find('DChg. Spec. Cap.(mAh/g)')
    cap_abs_chg_idx = _find('Chg. Cap.(mAh)')
    cap_abs_dch_idx = _find('DChg. Cap.(mAh)')
    
    if v_idx is None:
        raise ValueError("CSV missing required 'Voltage(V)' column for dQ/dV plot")
    if dq_abs_idx is None and dq_spec_idx is None:
        raise ValueError("CSV missing dQ/dV columns: need 'dQ/dV(mAh/V)' or 'dQm/dV(mAh/V.g)'")

    use_spec = False
    if prefer_specific and dq_spec_idx is not None:
        use_spec = True
    elif dq_abs_idx is not None:
        use_spec = False
    elif dq_spec_idx is not None:
        use_spec = True

    y_label = r'dQm/dV (mAh g$^{-1}$ V$^{-1}$)' if use_spec else r'dQ/dV (mAh V$^{-1}$)'
    n = len(rows)
    voltage = np.empty(n, dtype=float)
    dqdv = np.empty(n, dtype=float)
    current = np.zeros(n, dtype=float)
    def _to_float(val: str) -> float:
        try:
            return float(val.strip()) if isinstance(val, str) else float(val)
        except Exception:
            return np.nan

    for k, row in enumerate(rows):
        if len(row) < len(header):
            row = row + [''] * (len(header) - len(row))
        voltage[k] = _to_float(row[v_idx])
        if use_spec:
            dqdv[k] = _to_float(row[dq_spec_idx])
        else:
            dqdv[k] = _to_float(row[dq_abs_idx])
        if i_idx is not None:
            current[k] = _to_float(row[i_idx])

    # --- Derive charge/discharge using same logic as GC mode ---
    # Priority 1: Use explicit Step Type column
    is_charge = np.zeros(n, dtype=bool)
    is_rest_segment = np.zeros(n, dtype=bool)
    used_step_type = False
    used_capacity_columns = False
    
    if step_type_idx is not None:
        for k, row in enumerate(rows):
            if len(row) < len(header):
                row = row + [''] * (len(header) - len(row))
            step_type = str(row[step_type_idx]).strip().lower()
            is_rest = 'rest' in step_type or 'pause' in step_type or 'wait' in step_type or step_type in {'cv', 'idle'}
            if is_rest:
                is_rest_segment[k] = True
                is_charge[k] = is_charge[k-1] if k > 0 else True
                continue
            is_rest = ('rest' in step_type) or ('wait' in step_type) or step_type in {'pause', 'idle'}
            if is_rest:
                is_rest_segment[k] = True
                is_charge[k] = is_charge[k-1] if k > 0 else True
                continue
            is_dchg = 'dchg' in step_type or 'dischg' in step_type or step_type.startswith('dis')
            is_chg = (not is_dchg) and (('chg' in step_type) or ('charge' in step_type))
            if is_chg:
                is_charge[k] = True
            elif is_dchg:
                is_charge[k] = False
            else:
                is_charge[k] = is_charge[k-1] if k > 0 else True
        used_step_type = True
    
    # Priority 2: Use split charge/discharge capacity columns if available
    elif (cap_spec_chg_idx is not None and cap_spec_dch_idx is not None) or \
         (cap_abs_chg_idx is not None and cap_abs_dch_idx is not None):
        # Prefer specific capacity columns if they exist
        if cap_spec_chg_idx is not None and cap_spec_dch_idx is not None:
            chg_col_idx = cap_spec_chg_idx
            dch_col_idx = cap_spec_dch_idx
        else:
            chg_col_idx = cap_abs_chg_idx
            dch_col_idx = cap_abs_dch_idx
        
        cap_chg_vals = np.empty(n, dtype=float)
        cap_dch_vals = np.empty(n, dtype=float)
        
        for k, row in enumerate(rows):
            if len(row) < len(header):
                row = row + [''] * (len(header) - len(row))
            cap_chg_vals[k] = _to_float(row[chg_col_idx])
            cap_dch_vals[k] = _to_float(row[dch_col_idx])
        
        # Determine charge/discharge based on which capacity is non-zero
        threshold = 1e-6
        for k in range(n):
            chg_val = cap_chg_vals[k] if not np.isnan(cap_chg_vals[k]) else 0.0
            dch_val = cap_dch_vals[k] if not np.isnan(cap_dch_vals[k]) else 0.0
            
            if chg_val > threshold and dch_val <= threshold:
                is_charge[k] = True
            elif dch_val > threshold and chg_val <= threshold:
                is_charge[k] = False
            else:
                is_charge[k] = is_charge[k-1] if k > 0 else True
        used_capacity_columns = True
    
    # Priority 3: Fallback to voltage trend
    else:
        v_clean = np.array(voltage, dtype=float)
        v_min = np.nanmin(v_clean) if np.isfinite(v_clean).any() else 0.0
        v_max = np.nanmax(v_clean) if np.isfinite(v_clean).any() else 1.0
        v_span = max(1e-6, float(v_max - v_min))
        eps = max(1e-6, 1e-4 * v_span)
        dv = np.diff(v_clean)
        dv = np.nan_to_num(dv, nan=0.0, posinf=0.0, neginf=0.0)

        init_dir = None
        for d in dv[: min(500, dv.size)]:
            if abs(d) > eps:
                init_dir = (d > 0)
                break
        if init_dir is None:
            nz = None
            for i_val in current:
                if abs(i_val) > 1e-12 and np.isfinite(i_val):
                    nz = (i_val >= 0)
                    break
            init_dir = True if nz is None else bool(nz)
        
        prev_dir = init_dir
        for k in range(n):
            dir_set = None
            # Prefer backward-looking difference to keep the last sample of a run with its run
            if k > 0:
                db = dv[k-1]
                if abs(db) > eps:
                    dir_set = (db > 0)
            # Fallback: look forward to the next informative change
            if dir_set is None:
                j = k
                while j < n-1:
                    d = dv[j]
                    if abs(d) > eps:
                        dir_set = (d > 0)
                        break
                    j += 1
            if dir_set is None:
                dir_set = prev_dir
            is_charge[k] = dir_set
            prev_dir = dir_set

    charge_mask = is_charge & ~is_rest_segment
    discharge_mask = (~is_charge) & ~is_rest_segment
    inferred_cycles = _infer_cycles_from_masks(charge_mask, discharge_mask, n)

    return voltage, dqdv, inferred_cycles, charge_mask, discharge_mask, y_label


def read_csv_time_voltage(fname: str) -> Tuple[np.ndarray, np.ndarray]:
    """Read time (in hours) and voltage from a cycler CSV file.
    
    Args:
        fname: Path to CSV file with columns like 'Total Time' and 'Voltage(V)'
        
    Returns:
        (time_h, voltage) where time_h is in hours and voltage in volts
    """
    import csv
    
    # Read first two rows to compose header (same logic as read_ec_csv_file)
    with open(fname, newline='', encoding='utf-8', errors='ignore') as f:
        r = csv.reader(f)
        try:
            r1 = next(r)
            r2 = next(r)
        except StopIteration:
            raise ValueError(f"CSV '{fname}' is empty or missing header rows")
        
        # If second line begins with an empty first cell, treat it as continuation of header
        if len(r2) > 0 and (r2[0] == '' or str(r2[0]).strip() == ''):
            header = [c.strip() for c in r1] + [c.strip() for c in r2[1:]]
            rows = list(r)
        else:
            header = [c.strip() for c in r1]
            rows = [r2] + list(r)
    
    # Build column index map
    name_to_idx = {h: i for i, h in enumerate(header)}
    
    # Look for time and voltage columns (try multiple common names)
    time_idx = None
    for name in ['Total Time', 'Time', 'time/s', 'Time(s)', 'Test Time(s)']:
        if name in name_to_idx:
            time_idx = name_to_idx[name]
            break
    
    voltage_idx = None
    for name in ['Voltage(V)', 'Voltage', 'Ewe/V', 'Voltage/V']:
        if name in name_to_idx:
            voltage_idx = name_to_idx[name]
            break
    
    if time_idx is None:
        raise ValueError(f"CSV '{fname}' missing time column. Expected 'Total Time', 'Time', 'time/s', etc.")
    if voltage_idx is None:
        raise ValueError(f"CSV '{fname}' missing voltage column. Expected 'Voltage(V)', 'Voltage', 'Ewe/V', etc.")
    
    # Parse data
    n = len(rows)
    time_data = np.empty(n, dtype=float)
    voltage_data = np.empty(n, dtype=float)
    
    def _parse_time(val: str) -> float:
        """Parse time from string, handling HH:MM:SS format and numeric seconds."""
        if isinstance(val, (int, float)):
            return float(val)
        val = str(val).strip()
        # Try HH:MM:SS format
        if ':' in val:
            parts = val.split(':')
            try:
                if len(parts) == 3:  # HH:MM:SS
                    h, m, s = float(parts[0]), float(parts[1]), float(parts[2])
                    return h * 3600 + m * 60 + s
                elif len(parts) == 2:  # MM:SS
                    m, s = float(parts[0]), float(parts[1])
                    return m * 60 + s
            except:
                pass
        # Try as plain number (seconds)
        try:
            return float(val)
        except:
            return np.nan
    
    def _to_float(val: str) -> float:
        try:
            return float(str(val).strip()) if val else np.nan
        except:
            return np.nan
    
    for k, row in enumerate(rows):
        if len(row) < len(header):
            row = row + [''] * (len(header) - len(row))
        time_data[k] = _parse_time(row[time_idx])
        voltage_data[k] = _to_float(row[voltage_idx])
    
    # Convert time from seconds to hours
    time_h = time_data / 3600.0
    
    # Remove NaN values
    mask = ~(np.isnan(time_h) | np.isnan(voltage_data))
    return time_h[mask], voltage_data[mask]


def read_mpt_time_voltage(fname: str) -> Tuple[np.ndarray, np.ndarray]:
    """Read time (in hours) and voltage from a BioLogic .mpt file.
    
    Args:
        fname: Path to .mpt file
        
    Returns:
        (time_h, voltage) where time_h is in hours and voltage in volts
    """
    import re
    
    # Read header to find number of header lines
    header_lines = 0
    with open(fname, 'r', encoding='utf-8', errors='ignore') as f:
        first_line = f.readline().strip()
        if not first_line.startswith('EC-Lab ASCII FILE'):
            raise ValueError(f"Not a valid EC-Lab .mpt file: {fname}")
        
        for line in f:
            if line.startswith('Nb header lines'):
                match = re.search(r'Nb header lines\s*:\s*(\d+)', line)
                if match:
                    header_lines = int(match.group(1))
                    break
        if header_lines == 0:
            raise ValueError(f"Could not find header line count in {fname}")
    
    # Read data
    data_lines = []
    column_names = []
    
    with open(fname, 'r', encoding='utf-8', errors='ignore') as f:
        # Skip header lines
        for i in range(header_lines - 1):
            f.readline()
        
        # Read column names
        header_line = f.readline().strip()
        column_names = [col.strip() for col in header_line.split('\t')]
        
        # Read data lines
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                # Replace comma decimal separator with period
                values = [float(val.replace(',', '.')) for val in line.split('\t')]
                if len(values) == len(column_names):
                    data_lines.append(values)
            except ValueError:
                continue
    
    if not data_lines:
        raise ValueError(f"No valid data found in {fname}")
    
    # Convert to numpy array
    data = np.array(data_lines)
    col_map = {name: i for i, name in enumerate(column_names)}
    
    # Look for time column (try multiple names)
    time_idx = None
    for name in ['time/s', 'Time/s', 'time', 'Time']:
        if name in col_map:
            time_idx = col_map[name]
            break
    
    # Look for voltage column
    voltage_idx = None
    for name in ['Ewe/V', 'Voltage/V', 'Voltage', 'Ewe']:
        if name in col_map:
            voltage_idx = col_map[name]
            break
    
    if time_idx is None:
        raise ValueError(f"MPT file '{fname}' missing time column")
    if voltage_idx is None:
        raise ValueError(f"MPT file '{fname}' missing voltage column")
    
    time_s = data[:, time_idx]
    voltage = data[:, voltage_idx]
    
    # Convert time from seconds to hours
    time_h = time_s / 3600.0
    
    # Remove NaN values
    mask = ~(np.isnan(time_h) | np.isnan(voltage))
    return time_h[mask], voltage[mask]
