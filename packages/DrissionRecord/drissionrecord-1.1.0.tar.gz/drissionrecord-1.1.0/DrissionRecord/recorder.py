# -*- coding:utf-8 -*-
from csv import reader as csv_reader, writer as csv_writer
from json import loads, load, dump, dumps
from pathlib import Path
from time import sleep

from openpyxl.reader.excel import load_workbook

from .base import BaseRecorder
from .setter import RecorderSetter, set_csv_header
from .tools import (ok_list_str, process_content_json, get_key_cols, img2ws, link2ws, height2ws, width2ws,
                    get_csv, parse_coord, do_nothing, Header, get_wb, get_ws, is_single_data,
                    is_1D_data, data2ws, styles2ws, get_real_row, get_ws_real_coord, RowData, RowText)


class Recorder(BaseRecorder):
    def __init__(self, path=None, cache_size=1000):
        self._header = {None: None}
        self._methods = {'xlsx': self._to_xlsx_fast,
                         'csv': self._to_csv_fast,
                         'txt': self._to_txt_fast,
                         'jsonl': self._to_jsonl_fast,
                         'json': self._to_json_fast,
                         'addData': not_type,
                         # fast模式下非添加data使用的方法
                         'addImg': do_nothing,
                         'addLink': do_nothing,
                         'addStyle': do_nothing,
                         'addHeight': do_nothing,
                         'addWidth': do_nothing,
                         # 写入不同类型数据时使用的方法
                         'img': img2ws,
                         'link': link2ws,
                         'style': styles2ws,
                         'height': height2ws,
                         'width': width2ws,
                         'data': data2ws}
        super().__init__(path=path, cache_size=cache_size)
        self._data = {}
        self._delimiter = ','  # csv文件分隔符
        self._quote_char = '"'  # csv文件引用符
        self._follow_styles = False
        self._row_height = None
        self._styles = None
        self._header_row = {None: 1}
        self._fast = True
        self._link_style = None
        self._None_header_is_newest = None
        self._None_header_row_is_newest = None
        self.data_col = 1

    def _set_methods(self, file_type):
        self._methods[file_type] = getattr(self, f'_to_{file_type}_fast')
        if file_type == 'xlsx':
            self._methods['addImg'] = self._add_img
            self._methods['addLink'] = self._add_link
            self._methods['addStyle'] = self._add_styles
            self._methods['addHeight'] = self._add_rows_height
            self._methods['addWidth'] = self._add_cols_width
            self._methods['addData'] = self._add_data_any
        else:
            self._methods['addImg'] = do_nothing
            self._methods['addLink'] = do_nothing
            self._methods['addStyle'] = do_nothing
            self._methods['addHeight'] = do_nothing
            self._methods['addWidth'] = do_nothing
            self._methods['addData'] = self._add_data_txt

    @property
    def set(self):
        if self._setter is None:
            self._setter = RecorderSetter(self)
        return self._setter

    @property
    def delimiter(self):
        return self._delimiter

    @property
    def quote_char(self):
        return self._quote_char

    @property
    def header(self):
        return get_header(self)

    def add_data(self, data, coord=None, table=None):
        coord = parse_coord(coord, self.data_col)
        data, data_num = self._handle_data(data, coord)
        self._add(data, table,
                  True if self._fast and coord[0] else False,
                  data_num, self._methods['addData'])

    def _handle_data(self, data, coord):
        if is_single_data(data):
            data = {'type': 'data', 'data': [self._make_final_data(self, (data,))], 'coord': coord}
            data_num = 1
        elif not data:
            data = {'type': 'data', 'data': [self._make_final_data(self, tuple())], 'coord': coord}
            data_num = 1
        elif is_1D_data(data):
            data = {'type': 'data', 'data': [self._make_final_data(self, data)], 'coord': coord}
            data_num = 1
        else:  # 二维数组
            data = {'type': 'data', 'coord': coord,
                    'data': [self._make_final_data(self, (d,)) if is_single_data(d)
                             else self._make_final_data(self, d) for d in data]}
            data_num = len(data)
        return data, data_num

    def _add(self, data, table, to_slow, num, add_method):
        while self._pause_add:  # 等待其它线程写入结束
            sleep(.02)

        if to_slow:
            self._slow_mode()

        if table is None:
            table = self._table
        elif table is True:
            table = None

        add_method(data, table)

        self._data_count += num
        if 0 < self.cache_size <= self._data_count:
            self.record()

    def _add_data_any(self, data, table):
        if (self._data.get(table, 0) != 0 and self._data[table]
                and data['coord'] == self._data[table][-1]['coord']
                and self._data[table][-1]['type'] == 'data'):
            self._data[table][-1]['data'].extend(data['data'])
        else:
            self._data.setdefault(table, []).append(data)

    def _add_data_txt(self, data, table):
        if self._data.get(None, None) and data['coord'] == self._data[None][-1]['coord']:
            self._data[table][-1]['data'].extend(data['data'])
        else:
            self._data.setdefault(None, []).append(data)

    def _add_others(self, data, table):
        self._data.setdefault(table, []).append(data)

    def _add_link(self, coord, link, content=None, table=None):
        self._add({'type': 'link', 'link': link, 'content': content,
                   'coord': parse_coord(coord, self.data_col)}, table, self._fast, 1, self._add_others)

    def _add_img(self, coord, img_path, width=None, height=None, table=None):
        self._add({'type': 'img', 'imgPath': img_path, 'width': width, 'height': height,
                   'coord': parse_coord(coord, self.data_col)}, table, self._fast, 1, self._add_others)

    def _add_styles(self, styles, coord, rows, cols, replace=True, table=None):
        self._add({'type': 'style', 'mode': 'replace' if replace else 'cover',
                   'styles': styles, 'coord': (1, 1), 'real_coord': coord, 'rows': rows, 'cols': cols},
                  table, self._fast, 1, self._add_others)

    def _add_rows_height(self, rows, height, table=None):
        self._add({'type': 'height', 'rows': rows, 'height': height}, table, self._fast, 1, self._add_others)

    def _add_cols_width(self, cols, width, table=None):
        self._add({'type': 'width', 'cols': cols, 'width': width}, table, self._fast, 1,
                  self._add_others)

    def add_link(self, link, coord, content=None, table=None):
        self._methods['addLink'](coord, link, content, table)

    def add_img(self, img_path, coord, width=None, height=None, table=None):
        self._methods['addImg'](coord, img_path, width, height, table)

    def add_styles(self, styles, coord=None, rows=None, cols=None, replace=True, table=None):
        self._methods['addStyle'](styles, coord, rows, cols, replace, table)

    def add_rows_height(self, height, rows=True, table=None):
        self._methods['addHeight'](rows, height, table)

    def add_cols_width(self, width, cols=True, table=None):
        self._methods['addWidth'](cols, width, table)

    def rows(self, cols=True, sign_col=True,
             signs=None, deny_sign=False, count=None, begin_row=None, end_row=None):
        if not self._path or not Path(self._path).exists():
            raise RuntimeError('未指定文件路径或文件不存在。')

        if self.type == 'xlsx':
            wb = load_workbook(self.path, data_only=True, read_only=True)
            if self.table and self.table not in [i.title for i in wb.worksheets]:
                raise RuntimeError(f'xlsx文件未包含指定工作表：{self.table}')
            ws = wb[self.table] if self.table else wb.active
            if ws.max_column is None:  # 遇到过read_only时无法获取列数的文件
                wb.close()
                wb = load_workbook(self.path, data_only=True)
                ws = wb[self.table] if self.table else wb.active
            method = get_xlsx_rows

        elif self.type == 'csv':
            ws = None
            method = get_csv_rows

        elif self.type == 'jsonl':
            ws = None
            method = get_jsonl_rows

        elif self.type == 'json':
            ws = None
            method = get_json_rows

        elif self.type == 'txt':
            ws = None
            method = get_txt_rows

        else:
            raise RuntimeError('不支持的文件格式。')

        header = get_header(self, ws)

        if not isinstance(signs, (list, tuple, set)):
            signs = (signs,)
        if self.type in ('csv', 'xlsx'):
            if sign_col is not True:
                sign_col = header.get_num(sign_col) or 1
            cols = get_key_cols(cols, header)
            if not begin_row:
                begin_row = self._header_row.get(self.table, self._header_row[None]) + 1
        elif not begin_row:
            begin_row = 1

        return method(self, header=header, key_cols=cols, begin_row=begin_row, end_row=end_row or 0,
                      sign_col=sign_col, sign=signs, deny_sign=deny_sign, count=count, ws=ws)

    def _record(self):
        self._methods[self.type]()
        if not self._fast:
            self._fast_mode()

    def _fast_mode(self):
        self._methods['csv'] = self._to_csv_fast
        self._methods['txt'] = self._to_txt_fast
        self._methods['json'] = self._to_json_fast
        self._methods['jsonl'] = self._to_jsonl_fast
        self._fast = True

    def _slow_mode(self):
        self._methods['csv'] = self._to_csv_slow
        self._methods['txt'] = self._to_txt_slow
        self._methods['json'] = self._to_json_slow
        self._methods['jsonl'] = self._to_jsonl_slow
        self._fast = False

    def _to_xlsx_fast(self):
        wb, new_file = get_wb(self)
        tables = wb.sheetnames
        rewrite_method = 'make_num_dict_rewrite' if self._auto_new_header else 'make_num_dict'

        for table, data in self._data.items():
            ws, new_sheet = get_ws(wb, table, tables, new_file)
            new_file = False
            if table is None:
                if self._None_header_is_newest or ws.title not in self._header:
                    self._header[ws.title] = self._header[None]
                    self._None_header_is_newest = None
                if self._None_header_row_is_newest or ws.title not in self._header_row:
                    self._header_row[ws.title] = self._header_row[None]

            begin_row = True
            if new_sheet:
                begin_row = handle_new_sheet(self, ws, data)
            elif self._header.get(ws.title, None) is None:
                self._header[ws.title] = (Header([c.value for c in ws[self._header_row[ws.title]]])
                                          if ws.title in self._header_row else Header())

            header = self._header[ws.title]
            rewrite = False
            if not begin_row and not data[0]['coord'][0]:  # 首行为空，将数据填入首行
                cur = data[0]
                rewrite = self._methods[cur['type']](
                    **{'recorder': self,
                       'ws': ws,
                       'data': cur,
                       'coord': (1, header._get_num(cur.get('coord', (1, 1))[1])),
                       'new_row': not cur.get('coord', (1, 1))[0],
                       'header': header,
                       'rewrite': rewrite,
                       'rewrite_method': rewrite_method})
                data = data[1:]

            for cur in data:
                rewrite = self._methods[cur['type']](
                    **{'recorder': self,
                       'ws': ws,
                       'data': cur,
                       'coord': get_ws_real_coord(cur.get('coord', (1, 1)), ws, header),
                       'new_row': not cur.get('coord', (1, 1))[0],
                       'header': header,
                       'rewrite': rewrite,
                       'rewrite_method': rewrite_method})

            if rewrite:
                for c in range(1, ws.max_column + 1):
                    ws.cell(self._header_row[ws.title], c, value=header[c])

        wb.save(self.path)
        wb.close()

    def _to_csv_fast(self):
        file, new_csv = get_csv(self)
        writer = csv_writer(file, delimiter=self.delimiter, quotechar=self.quote_char)
        get_and_set_csv_header(self, new_csv, file, writer)
        rewrite_method = 'make_insert_list_rewrite' if self._auto_new_header else 'make_insert_list'

        rewrite = False
        header = self._header[None]
        for d in self._data[None]:
            col = header._get_num(d['coord'][1])
            for data in d['data']:
                data, rewrite = header.__getattribute__(rewrite_method)(data, 'csv', rewrite)
                data = [None] * (col - 1) + data
                writer.writerow(data)
        file.close()

        if rewrite:
            set_csv_header(self, self._header[None], self._header_row[None])

    def _to_csv_slow(self):
        file, new_csv = get_csv(self)
        writer = csv_writer(file, delimiter=self.delimiter, quotechar=self.quote_char)
        get_and_set_csv_header(self, new_csv, file, writer)
        file.seek(0)
        reader = csv_reader(file, delimiter=self.delimiter, quotechar=self.quote_char)
        lines = list(reader)
        lines_count = len(lines)
        header = self._header[None]

        rewrite = False
        method = 'make_change_list_rewrite' if self._auto_new_header else 'make_change_list'
        for i in self._data[None]:
            data = i['data']
            row = get_real_row(i['coord'][0], lines_count)
            col = header._get_num(i['coord'][1])
            for r, da in enumerate(data, row):
                add_rows = r - lines_count
                if add_rows > 0:  # 若行数不够，填充行数
                    [lines.append([]) for _ in range(add_rows)]
                    lines_count += add_rows
                row_num = r - 1
                lines[row_num], rewrite = self._header[None].__getattribute__(method)(lines[row_num], da, col,
                                                                                      'csv', rewrite)

        if rewrite:
            [lines.append([]) for _ in range(self._header_row[None] - lines_count)]  # 若行数不够，填充行数
            lines[self._header_row[None] - 1] = list(header.num_key.values())

        file.close()
        writer = csv_writer(open(self.path, 'w', encoding=self.encoding, newline=''),
                            delimiter=self.delimiter, quotechar=self.quote_char)
        writer.writerows(lines)

    def _to_txt_fast(self):
        with open(self.path, 'a+', encoding=self.encoding) as f:
            all_data = []
            for data in self._data[None]:
                for d in data['data']:
                    all_data.append(' '.join(ok_list_str(d)))
            f.write('\n'.join(all_data) + '\n')

    def _to_jsonl_fast(self):
        with open(self.path, 'a+', encoding=self.encoding) as f:
            all_data = []
            for data in self._data[None]:
                for d in data['data']:
                    all_data.append(d if isinstance(d, str) else dumps(d))
            f.write('\n'.join(all_data) + '\n')

    def _to_json_fast(self):
        if self._file_exists or Path(self.path).exists():
            with open(self.path, 'r', encoding=self.encoding) as f:
                json_data = load(f)
        else:
            json_data = []

        for i in self._data[None]:
            for data in i['data']:
                if isinstance(data, dict):
                    for k, d in data.items():
                        data[k] = process_content_json(d)
                    json_data.append(data)
                else:
                    json_data.append([process_content_json(d) for d in i])

        with open(self.path, 'w', encoding=self.encoding) as f:
            dump(json_data, f)

    def _to_txt_slow(self):
        if not self._file_exists and not Path(self.path).exists():
            with open(self.path, 'w', encoding=self.encoding):
                pass
        with open(self.path, 'r+', encoding=self.encoding) as f:
            lines = f.readlines()
            handle_txt_lines(self._data[None], lines, '\n', handle_txt_data)
            f.seek(0)
            f.writelines(lines)

    def _to_jsonl_slow(self):
        if not self._file_exists and not Path(self.path).exists():
            with open(self.path, 'w', encoding=self.encoding):
                pass
        with open(self.path, 'r+', encoding=self.encoding) as f:
            lines = f.readlines()
            handle_txt_lines(self._data[None], lines, '[]\n', handle_jsonl_data)
            f.seek(0)
            f.writelines(lines)

    def _to_json_slow(self):
        if self._file_exists or Path(self.path).exists():
            with open(self.path, 'r', encoding=self.encoding) as f:
                lines = load(f)
        else:
            lines = []
        handle_txt_lines(self._data[None], lines, None, handle_json_data)
        with open(self.path, 'w', encoding=self.encoding) as f:
            dump(lines, f)


def handle_txt_lines(data_lst, lines, val, method):
    lines_len = len(lines)
    for data in data_lst:
        num = get_real_row(data['coord'][0], lines_len)
        data_end = num + len(data['data'])
        if lines_len < data_end:
            diff = data_end - lines_len - 1
            [lines.append(val) for _ in range(diff)]
            lines_len += diff
        for num, i in enumerate(data['data'], num - 1):
            method(lines, num, i)


def handle_txt_data(lines, num, data):
    lines[num] = ' '.join(ok_list_str(data)) + '\n'


def handle_jsonl_data(lines, num, data):
    lines[num] = data if isinstance(data, str) else dumps(data) + '\n'


def handle_json_data(lines, num, data):
    if isinstance(data, dict):
        for k, d in data.items():
            data[k] = process_content_json(d)
        lines[num] = data
    else:
        lines[num] = [process_content_json(d) for d in data]


def get_header(recorder, ws=None):
    header = recorder._header.get(recorder._table, None)
    if header is not None:
        return header
    if not recorder.path or not Path(recorder.path).exists():
        return None

    if recorder.type == 'xlsx':
        if not ws:
            wb = load_workbook(recorder.path)
            if not recorder.table:
                ws = wb.active
            elif recorder.table not in wb.sheetnames:
                wb.close()
                return Header()
            else:
                ws = wb[recorder.table]
        header_row = recorder._header_row.get(recorder.table, recorder._header_row[None])
        if header_row > ws.max_row:
            recorder._header[recorder.table] = Header()
        else:
            recorder._header[recorder.table] = Header(
                [i.value for i in ws[recorder._header_row.get(recorder.table, recorder._header_row[None])]])

        if not ws:
            wb.close()
        return recorder._header[recorder.table]

    elif recorder.type == 'csv':
        from csv import reader
        with open(recorder.path, 'r', newline='', encoding=recorder.encoding) as f:
            u = reader(f, delimiter=recorder.delimiter, quotechar=recorder.quote_char)
            try:
                for _ in range(recorder._header_row[None]):
                    header = next(u)
            except StopIteration:  # 文件是空的
                header = []
        recorder._header[None] = Header(header)
        return recorder._header[None]

    elif recorder.type == 'jsonl':
        with open(recorder.path, 'r', newline='', encoding=recorder.encoding) as f:
            try:
                for _ in range(recorder._header_row[None]):
                    header = next(f)
            except StopIteration:  # 文件是空的
                header = '[]'
            header = loads(header.strip())
            if isinstance(header, dict):
                header = header.keys()
            elif not isinstance(header, list):
                header = [str(header)]
        recorder._header[None] = Header(header)
        return recorder._header[None]

    elif recorder.type == 'json':
        with open(recorder.path, 'r', newline='', encoding=recorder.encoding) as f:
            j = load(f)
            if recorder._header_row[None] > len(j):
                header = []
            else:
                header = j[recorder._header_row[None] - 1]
                if isinstance(header, dict):
                    header = header.keys()
                elif not isinstance(header, list):
                    header = [str(header)]
        recorder._header[None] = Header(header)
        return recorder._header[None]


def handle_new_sheet(recorder, ws, data):
    if not recorder._header_row:
        return 0

    if recorder._header.get(ws.title, None) is not None:
        for c, h in recorder._header[ws.title].items():
            ws.cell(row=recorder._header_row[ws.title], column=c, value=h)
        begin_row = recorder._header_row

    else:
        data = get_first_dict(data)
        if data:
            header = Header([h for h in data.keys() if isinstance(h, str)])
            recorder._header[ws.title] = header
            for c, h in header.items():
                ws.cell(row=recorder._header_row[ws.title], column=c, value=h)
            begin_row = recorder._header_row

        else:
            recorder._header[ws.title] = Header()
            begin_row = 0

    return begin_row


def get_first_dict(data):
    if not data:
        return False
    elif data[0]['type'] == 'data' and data[0]['data'] and isinstance(data[0]['data'][0], dict):
        return data[0]['data'][0]


def get_xlsx_rows(recorder, header, key_cols, begin_row, end_row, sign_col, sign, deny_sign, count, ws):
    rows = ws.rows
    try:
        for _ in range(begin_row - 1):
            next(rows)
    except StopIteration:
        return []

    if sign_col is True or sign_col > ws.max_column:  # 获取所有行
        if count or end_row:
            rows = list(rows)[:(min(count, end_row - begin_row + 1)
                                if count and end_row else (count or end_row - begin_row + 1))]

        if key_cols is True:  # 获取整行
            res = [header.make_row_data(ind, {col: cell.value for col, cell in enumerate(row, 1)})
                   for ind, row in enumerate(rows, begin_row)]
        else:  # 只获取对应的列
            res = [header.make_row_data(ind, {col: row[col - 1].value for col in key_cols})
                   for ind, row in enumerate(rows, begin_row)]

    else:  # 获取符合条件的行
        if count:
            res = get_xlsx_rows_with_count(key_cols, deny_sign, header, rows,
                                           begin_row, end_row, sign_col, sign, count)
        else:
            res = get_xlsx_rows_without_count(key_cols, deny_sign, header, rows, begin_row, end_row,
                                              sign_col, sign)

    ws.parent.close()
    return res


def get_xlsx_rows_with_count(key_cols, deny_sign, header, rows, begin_row, end_row, sign_col, sign, count):
    got = 0
    res = []
    if key_cols is True:  # 获取整行
        if deny_sign:
            for ind, row in enumerate(rows, begin_row):
                if got == count or (end_row and ind > end_row):
                    break
                if row[sign_col - 1].value not in sign:
                    res.append(header.make_row_data(ind, {col: cell.value for col, cell in enumerate(row, 1)}))
                    got += 1
        else:
            for ind, row in enumerate(rows, begin_row):
                if got == count or (end_row and ind > end_row):
                    break
                if row[sign_col - 1].value in sign:
                    res.append(header.make_row_data(ind, {col: cell.value for col, cell in enumerate(row, 1)}))
                    got += 1

    else:  # 只获取对应的列
        if deny_sign:
            for ind, row in enumerate(rows, begin_row):
                if got == count or (end_row and ind > end_row):
                    break
                if row[sign_col - 1].value not in sign:
                    res.append(header.make_row_data(ind, {col: row[col - 1].value for col in key_cols}))
                    got += 1
        else:
            for ind, row in enumerate(rows, begin_row):
                if got == count or (end_row and ind > end_row):
                    break
                if row[sign_col - 1].value in sign:
                    res.append(header.make_row_data(ind, {col: row[col - 1].value for col in key_cols}))
                    got += 1
    return res


def get_xlsx_rows_without_count(key_cols, deny_sign, header, rows, begin_row, end_row, sign_col, sign):
    if end_row:
        if end_row < begin_row:
            return []
        rows = list(rows)[:end_row - begin_row + 1]
    if key_cols is True:  # 获取整行
        if deny_sign:
            return [header.make_row_data(ind, {col: cell.value for col, cell in enumerate(row, 1)})
                    for ind, row in enumerate(rows, begin_row)
                    if row[sign_col - 1].value not in sign]
        else:
            return [header.make_row_data(ind, {col: cell.value for col, cell in enumerate(row, 1)})
                    for ind, row in enumerate(rows, begin_row)
                    if row[sign_col - 1].value in sign]

    else:  # 只获取对应的列
        if deny_sign:
            return [header.make_row_data(ind, {col: row[col - 1].value for col in key_cols})
                    for ind, row in enumerate(rows, begin_row)
                    if row[sign_col - 1].value not in sign]
        else:
            return [header.make_row_data(ind, {col: row[col - 1].value for col in key_cols})
                    for ind, row in enumerate(rows, begin_row)
                    if row[sign_col - 1].value in sign]


def get_csv_rows(recorder, header, key_cols, begin_row, end_row, sign_col, sign, deny_sign, count, ws):
    sign = ['' if i is None else str(i) for i in sign]
    begin_row -= 1
    res = []
    with open(recorder.path, 'r', encoding=recorder.encoding) as f:
        try:
            for i in range(begin_row):
                next(f)
        except StopIteration:
            return res
        reader = csv_reader(f, delimiter=recorder.delimiter, quotechar=recorder.quote_char)

        if sign_col is True:  # 获取所有行
            header_len = len(header)
            if count or end_row:
                end = min(count + begin_row, end_row) if count and end_row else (end_row or count + begin_row)
            else:
                end = False
            method = get_csv_rows_key_is_True if key_cols is True else get_csv_rows_key_not_True
            for ind, line in enumerate(reader, begin_row + 1):
                if end and ind > end:
                    break
                method(line, res, header, ind, key_cols, header_len)

        else:  # 获取符合条件的行
            sign_col -= 1
            get_csv_rows_with_count(reader, begin_row, end_row, sign_col, sign, deny_sign,
                                    key_cols, res, header, count)

    return res


def get_csv_rows_key_is_True(line, res, header, ind, key_cols, header_len):
    if not line:
        res.append(header.make_row_data(ind, {col: '' for col in range(1, header_len + 1)}))
    else:
        line_len = len(line)
        x = max(header_len, line_len)
        res.append(header.make_row_data(ind, {col: line[col - 1] if col <= line_len else ''
                                              for col in range(1, x + 1)}))


def get_csv_rows_key_not_True(line, res, header, ind, key_cols, header_len):
    x = len(line) + 1
    res.append(header.make_row_data(ind, {col: line[col - 1] if col < x else '' for col in key_cols}))


def get_csv_rows_with_count(lines, begin_row, end_row, sign_col, sign, deny_sign, key_cols, res, header, count):
    got = 0
    header_len = len(header)
    for ind, line in enumerate(lines, begin_row + 1):
        if (end_row and ind > end_row) or (count and got == count):
            break
        row_sign = '' if sign_col > len(line) - 1 else line[sign_col]
        if (row_sign not in sign) if deny_sign else (row_sign in sign):
            if key_cols is True:  # 获取整行
                if not line:
                    res.append(header.make_row_data(ind, {col: '' for col in range(1, header_len + 1)}))
                else:
                    line_len = len(line)
                    x = max(header_len, line_len)
                    res.append(header.make_row_data(ind, {col: line[col - 1] if col <= line_len else ''
                                                          for col in range(1, x + 1)}))
            else:  # 只获取对应的列
                x = len(line) + 1
                res.append(header.make_row_data(ind, {col: line[col - 1] if col < x else '' for col in key_cols}))
            got += 1


def get_jsonl_rows(recorder, header, key_cols, begin_row, end_row, sign_col, sign, deny_sign, count, ws):
    sign = ['' if i is None else str(i) for i in sign]
    begin_row -= 1
    res = []
    with open(recorder.path, 'r', encoding=recorder.encoding) as f:
        try:
            for i in range(begin_row):
                next(f)
        except StopIteration:
            return res

        if sign_col is True:  # 获取所有行
            header_len = len(header)
            if count or end_row:
                end = min(count + begin_row, end_row) if count and end_row else (end_row or count + begin_row)
            else:
                end = False
            method = get_jsonl_rows_key_is_True if key_cols is True else get_jsonl_rows_key_not_True
            for ind, line in enumerate(f, begin_row + 1):
                if end and ind > end:
                    break
                line = loads(line.strip())
                method(line, res, header, ind, key_cols, header_len)

        else:  # 获取符合条件的行
            get_jsonl_rows_with_count(f, begin_row, end_row, sign_col, sign, deny_sign,
                                      key_cols, res, header, count)

    return res


def get_jsonl_rows_key_is_True(line, res, header, ind, key_cols, header_len):
    if isinstance(line, dict):
        res.append(RowData(ind, header, None, line))
    elif isinstance(line, list):
        if not line:
            res.append(header.make_row_data(ind, {col: None for col in range(1, header_len + 1)}))
        else:
            line_len = len(line)
            x = max(header_len, line_len)
            res.append(header.make_row_data(ind, {col: line[col - 1] if col <= line_len else None
                                                  for col in range(1, x + 1)}))


def get_jsonl_rows_key_not_True(line, res, header, ind, key_cols, header_len):
    if isinstance(line, dict):
        header = Header(line.keys())
        key_cols = get_key_cols(key_cols, header)
        res.append(RowData(ind, header, None, {header[c]: line[header[c]] for c in key_cols}))
    else:
        x = len(line) + 1
        key_cols = get_key_cols(key_cols, header)
        res.append(header.make_row_data(ind, {col: line[col - 1] if col < x else None for col in key_cols}))


def get_jsonl_rows_with_count(lines, begin_row, end_row, sign_col, sign, deny_sign, key_cols, res, header, count):
    got = 0
    header_len = len(header)
    for ind, line in enumerate(lines, begin_row + 1):
        if (end_row and ind > end_row) or (count and got == count):
            break
        line = loads(line.strip())
        if isinstance(sign_col, str):
            if isinstance(line, dict):
                row_sign = line[sign_col]
            else:  # list
                sign_col = header[sign_col]
                row_sign = None if sign_col > len(line) else line[sign_col - 1]
        else:  # int
            if isinstance(line, dict):
                row_sign = None if sign_col > len(line) else line[list(line.keys())[sign_col - 1]]
            else:
                row_sign = None if sign_col > len(line) else line[sign_col - 1]

        if (row_sign not in sign) if deny_sign else (row_sign in sign):
            if key_cols is True:  # 获取整行
                get_jsonl_rows_key_is_True(line, res, header, ind, key_cols, header_len)
            else:  # 只获取对应的列
                get_jsonl_rows_key_not_True(line, res, header, ind, key_cols, header_len)
            got += 1


def get_json_rows(recorder, header, key_cols, begin_row, end_row, sign_col, sign, deny_sign, count, ws):
    sign = ['' if i is None else str(i) for i in sign]
    begin_row -= 1
    res = []
    with open(recorder.path, 'r', encoding=recorder.encoding) as f:
        lines = load(f)
        if sign_col is True:  # 获取所有行
            header_len = len(header)
            if count or end_row:
                end = min(count + begin_row, end_row) if count and end_row else (end_row or count + begin_row)
            else:
                end = None
            method = get_jsonl_rows_key_is_True if key_cols is True else get_jsonl_rows_key_not_True
            for ind, line in enumerate(lines[begin_row:end], begin_row + 1):
                if not isinstance(line, (dict, list)):
                    line = [line]
                method(line, res, header, ind, key_cols, header_len)

        else:  # 获取符合条件的行
            get_json_rows_with_count(lines, begin_row, end_row, sign_col, sign, deny_sign,
                                     key_cols, res, header, count)
    return res


def get_json_rows_with_count(lines, begin_row, end_row, sign_col, sign, deny_sign, key_cols, res, header, count):
    got = 0
    header_len = len(header)
    for ind, line in enumerate(lines, begin_row + 1):
        if (end_row and ind > end_row) or (count and got == count):
            break
        if not isinstance(line, (dict, list)):
            line = [line]
        if isinstance(sign_col, str):
            if isinstance(line, dict):
                row_sign = line[sign_col]
            else:  # list
                sign_col = header[sign_col]
                row_sign = None if sign_col > len(line) else line[sign_col - 1]
        else:  # int
            if isinstance(line, dict):
                row_sign = None if sign_col > len(line) else line[list(line.keys())[sign_col - 1]]
            else:
                row_sign = None if sign_col > len(line) else line[sign_col - 1]

        if (row_sign not in sign) if deny_sign else (row_sign in sign):
            if key_cols is True:  # 获取整行
                get_jsonl_rows_key_is_True(line, res, header, ind, key_cols, header_len)
            else:  # 只获取对应的列
                get_jsonl_rows_key_not_True(line, res, header, ind, key_cols, header_len)
            got += 1


def get_txt_rows(recorder, header, key_cols, begin_row, end_row, sign_col, sign, deny_sign, count, ws):
    begin_row -= 1
    res = []
    with open(recorder.path, 'r', encoding=recorder.encoding) as f:
        try:
            for i in range(begin_row):
                next(f)
        except StopIteration:
            return res

        got = 0
        for ind, line in enumerate(f, begin_row + 1):
            if (end_row and ind > end_row) or (count and got == count):
                break
            t = RowText(line.strip())
            t.row = ind
            res.append(t)
            got += 1

    return res


def get_and_set_csv_header(recorder, new_csv, file, writer):
    if not recorder._header_row:
        return

    if new_csv:
        if recorder._header[None]:
            for _ in range(recorder._header_row[None] - 1):
                writer.writerow([])
            writer.writerow(ok_list_str(recorder._header[None]))

        if recorder._header[None] is None and recorder._data_count:
            data = get_first_dict(recorder._data[None])
            if data:
                recorder._header[None] = Header([h for h in data.keys() if isinstance(h, str)])
            else:
                recorder._header[None] = Header()
            if recorder._header[None]:
                writer.writerow(ok_list_str(recorder._header[None]))
        else:
            recorder._header[None] = Header()

    elif recorder._header[None] is None:  # 从文件读取表头
        file.seek(0)
        reader = csv_reader(file, delimiter=recorder.delimiter, quotechar=recorder.quote_char)
        header = []
        try:
            for _ in range(recorder._header_row[None]):
                header = next(reader)
        except StopIteration:
            pass
        recorder._header[None] = Header(header)
        file.seek(2)


def not_type(*keys):
    raise RuntimeError('添加数据前请先指定文件路径。')
