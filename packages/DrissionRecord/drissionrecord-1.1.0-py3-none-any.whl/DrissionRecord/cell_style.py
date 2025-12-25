# -*- coding:utf-8 -*-
from copy import copy
from threading import Lock

from openpyxl.styles import Alignment, Font, Side, Border, Protection, GradientFill, PatternFill, Color
from openpyxl.utils import get_column_letter


class CellStyle(object):
    font_args = ('name', 'size', 'charset', 'underline', 'color', 'scheme', 'vertAlign',
                 'bold', 'italic', 'strike', 'outline', 'shadow', 'condense', 'extend')
    border_args = ('start', 'end', 'left', 'right', 'top', 'bottom', 'diagonal', 'vertical', 'horizontal',
                   'horizontal', 'outline', 'diagonalUp', 'diagonalDown')
    alignment_args = ('horizontal', 'vertical', 'indent', 'relativeIndent', 'justifyLastLine', 'readingOrder',
                      'textRotation', 'wrapText', 'shrinkToFit')
    protection_args = ('locked', 'hidden')
    gradient_fill_args = ('type', 'degree', 'left', 'right', 'top', 'bottom', 'stop')
    pattern_fill_args = ('patternType', 'fgColor', 'bgColor')

    def __init__(self):
        self._font = None
        self._border = None
        self._alignment = None
        self._pattern_fill = None
        self._gradient_fill = None
        self._number_format = None
        self._protection = None

        # 用于覆盖目标单元格的对象
        self._Font = None
        self._Border = None
        self._Alignment = None
        self._Fill = None
        self._Protection = None

        self.height = None
        self.width = None

    @property
    def font(self):
        if self._font is None:
            self._font = CellFont()
        return self._font

    @property
    def border(self):
        if self._border is None:
            self._border = CellBorder()
        return self._border

    @property
    def alignment(self):
        if self._alignment is None:
            self._alignment = CellAlignment()
        return self._alignment

    @property
    def pattern_fill(self):
        self._gradient_fill = None
        if self._pattern_fill is None:
            self._pattern_fill = CellPatternFill()
        return self._pattern_fill

    @property
    def gradient_fill(self):
        self._pattern_fill = None
        if self._gradient_fill is None:
            self._gradient_fill = CellGradientFill()
        return self._gradient_fill

    @property
    def number_format(self):
        if self._number_format is None:
            self._number_format = CellNumberFormat()
        return self._number_format

    @property
    def protection(self):
        if self._protection is None:
            self._protection = CellProtection()
        return self._protection

    def to_cell(self, cell, replace=True):
        if replace:
            self._replace_to_cell(cell)
        else:
            self._cover_to_cell(cell)
        if self.height is not None:
            cell.parent.row_dimensions[cell.row].height = self.height
        if self.width is not None:
            cell.parent.column_dimensions[get_column_letter(cell.column)].width = self.width

    def set_size(self, width=None, height=None):
        if width == height is None:
            self.width = self.height = None
            return self
        if width is not None:
            self.width = width
        if height is not None:
            self.height = height
        return self

    def set_bgColor(self, color):
        self.pattern_fill.set_fgColor(color)
        return self

    def set_txtColor(self, color):
        self.font.set_color(color)
        return self

    def set_txtSize(self, size=None, bold=None):
        if size is not None:
            self.font.set_size(size)
        if bold is not None:
            self.font.set_bold(bold)
        return self

    def set_bold(self, on_off=True):
        self.font.set_bold(on_off)
        return self

    def set_delLine(self, on_off=True):
        self.font.set_strike(on_off)
        return self

    def set_underLine(self, on_off=True):
        self.font.set_underline('single' if on_off else None)
        return self

    def set_center(self):
        self.alignment.set_horizontal('center')
        self.alignment.set_vertical('center')
        return self

    def set_border(self, on_off=True):
        if on_off:
            self.border.set_quad('thin', 'black')
        else:
            self.border.set_quad(None, None)
        return self

    def _cover_to_cell(self, cell):
        if self._font:
            d = _handle_args(self.font_args, self._font, cell.font)
            d['family'] = cell.font.family
            cell.font = Font(**d)

        if self._border:
            d = _handle_args(self.border_args, self._border, cell.border)
            cell.border = Border(**d)

        if self._alignment:
            d = _handle_args(self.alignment_args, self._alignment, cell.alignment)
            cell.alignment = Alignment(**d)

        if self._pattern_fill:
            f = None if 'fills.GradientFill' in str(cell.fill) else cell.fill
            d = _handle_args(self.pattern_fill_args, self._pattern_fill, f)
            cell.fill = PatternFill(**d)

        elif self._gradient_fill:
            f = None if 'fills.PatternFill' in str(cell.fill) else cell.fill
            d = _handle_args(self.gradient_fill_args, self._gradient_fill, f)
            cell.fill = GradientFill(**d)

        if self._number_format and self._number_format.format != 'notSet':
            cell.number_format = self._number_format.format

        if self._protection:
            d = _handle_args(self.protection_args, self._protection, cell.protection)
            cell.protection = Protection(**d)

    def _replace_to_cell(self, cell):
        if self._font:
            if self._Font is None:
                d = _handle_args(self.font_args, self._font, None)
                self._Font = Font(**d)
            cell.font = self._Font

        if self._border:
            if self._Border is None:
                d = _handle_args(self.border_args, self._border, None)
                self._Border = Border(**d)
            cell.border = self._Border

        if self._alignment:
            if self._Alignment is None:
                d = _handle_args(self.alignment_args, self._alignment, None)
                self._Alignment = Alignment(**d)
            cell.alignment = self._Alignment

        if self._pattern_fill:
            if not isinstance(self._Fill, PatternFill):
                d = _handle_args(self.pattern_fill_args, self._pattern_fill, None)
                self._Fill = PatternFill(**d)
            cell.fill = self._Fill

        elif self._gradient_fill:
            if not isinstance(self._Fill, GradientFill):
                d = _handle_args(self.gradient_fill_args, self._gradient_fill, None)
                self._Fill = GradientFill(**d)
            cell.fill = self._Fill

        if self._number_format and self._number_format.format != 'notSet':
            cell.number_format = self._number_format.format

        if self._protection:
            if self._Protection is None:
                d = _handle_args(self.protection_args, self._protection, None)
                self._Protection = Protection(**d)
            cell.protection = self._Protection


def _handle_args(args, src, target=None):
    d = {}
    for arg in args:
        tmp = getattr(src, arg)
        if tmp != 'notSet':
            d[arg] = tmp
        elif target:
            d[arg] = getattr(target, arg)
    return d


class CellFont(object):
    _LINE_STYLES = ('single', 'double', 'singleAccounting', 'doubleAccounting', None)
    _SCHEMES = ('major', 'minor', None)
    _VERT_ALIGNS = ('superscript', 'subscript', 'baseline', None)

    def __init__(self):
        self.name = 'notSet'
        self.charset = 'notSet'
        self.size = 'notSet'
        self.bold = 'notSet'
        self.italic = 'notSet'
        self.strike = 'notSet'
        self.outline = 'notSet'
        self.shadow = 'notSet'
        self.condense = 'notSet'
        self.extend = 'notSet'
        self.underline = 'notSet'
        self.vertAlign = 'notSet'
        self.color = 'notSet'
        self.scheme = 'notSet'

    def set_name(self, name):
        self.name = name

    def set_charset(self, charset):
        if not isinstance(charset, int):
            raise TypeError('charset参数只能接收int类型。')
        self.charset = charset

    def set_size(self, size):
        self.size = size

    def set_bold(self, on_off):
        self.bold = on_off

    def set_italic(self, on_off):
        self.italic = on_off

    def set_strike(self, on_off):
        self.strike = on_off

    def set_outline(self, on_off):
        self.outline = on_off

    def set_shadow(self, on_off):
        self.shadow = on_off

    def set_condense(self, on_off):
        self.condense = on_off

    def set_extend(self, on_off):
        self.extend = on_off

    def set_color(self, color):
        self.color = get_color_code(color)

    def set_underline(self, option):
        if option not in self._LINE_STYLES:
            raise ValueError(f'option参数只能是{self._LINE_STYLES}其中之一。')
        self.underline = option

    def set_vertAlign(self, option):
        if option not in self._VERT_ALIGNS:
            raise ValueError(f'option参数只能是{self._VERT_ALIGNS}其中之一。')
        self.vertAlign = option

    def set_scheme(self, option):
        if option not in self._SCHEMES:
            raise ValueError(f'option参数只能是{self._SCHEMES}其中之一。')
        self.scheme = option


class CellBorder(object):
    _LINE_STYLES = ('dashDot', 'dashDotDot', 'dashed', 'dotted', 'double', 'hair', 'medium', 'mediumDashDot',
                    'mediumDashDotDot', 'mediumDashed', 'slantDashDot', 'thick', 'thin', None)

    def __init__(self):
        self.start = 'notSet'
        self.end = 'notSet'
        self.left = 'notSet'
        self.right = 'notSet'
        self.top = 'notSet'
        self.bottom = 'notSet'
        self.diagonal = 'notSet'
        self.vertical = 'notSet'
        self.horizontal = 'notSet'
        self.horizontal = 'notSet'
        self.outline = 'notSet'
        self.diagonalUp = 'notSet'
        self.diagonalDown = 'notSet'

    def set_start(self, style, color):
        if style not in self._LINE_STYLES:
            raise ValueError(f'style参数只能是{self._LINE_STYLES}之一。')
        self.start = Side(style=style, color=get_color_code(color))

    def set_end(self, style, color):
        if style not in self._LINE_STYLES:
            raise ValueError(f'style参数只能是{self._LINE_STYLES}之一。')
        self.end = Side(style=style, color=get_color_code(color))

    def set_left(self, style, color):
        if style not in self._LINE_STYLES:
            raise ValueError(f'style参数只能是{self._LINE_STYLES}之一。')
        self.left = Side(style=style, color=get_color_code(color))

    def set_right(self, style, color):
        if style not in self._LINE_STYLES:
            raise ValueError(f'style参数只能是{self._LINE_STYLES}之一。')
        self.right = Side(style=style, color=get_color_code(color))

    def set_top(self, style, color):
        if style not in self._LINE_STYLES:
            raise ValueError(f'style参数只能是{self._LINE_STYLES}之一。')
        self.top = Side(style=style, color=get_color_code(color))

    def set_bottom(self, style, color):
        if style not in self._LINE_STYLES:
            raise ValueError(f'style参数只能是{self._LINE_STYLES}之一。')
        self.bottom = Side(style=style, color=get_color_code(color))

    def set_quad(self, style, color):
        if style not in self._LINE_STYLES:
            raise ValueError(f'style参数只能是{self._LINE_STYLES}之一。')
        self.bottom = self.top = self.right = self.left = Side(style=style, color=get_color_code(color))

    def set_diagonal(self, style, color):
        if style not in self._LINE_STYLES:
            raise ValueError(f'style参数只能是{self._LINE_STYLES}之一。')
        self.diagonal = Side(style=style, color=get_color_code(color))

    def set_vertical(self, style, color):
        if style not in self._LINE_STYLES:
            raise ValueError(f'style参数只能是{self._LINE_STYLES}之一。')
        self.vertical = Side(style=style, color=get_color_code(color))

    def set_horizontal(self, style, color):
        if style not in self._LINE_STYLES:
            raise ValueError(f'style参数只能是{self._LINE_STYLES}之一。')
        self.horizontal = Side(style=style, color=get_color_code(color))

    def set_outline(self, on_off):
        self.outline = on_off

    def set_diagonalDown(self, on_off):
        self.diagonalDown = on_off

    def set_diagonalUp(self, on_off):
        self.diagonalUp = on_off


class CellAlignment(object):
    _horizontal_alignments = ('general', 'left', 'center', 'right', 'fill', 'justify', 'centerContinuous',
                              'distributed', None)
    _vertical_alignments = ('top', 'center', 'bottom', 'justify', 'distributed', None)

    def __init__(self):
        self.horizontal = 'notSet'
        self.vertical = 'notSet'
        self.indent = 'notSet'
        self.relativeIndent = 'notSet'
        self.justifyLastLine = 'notSet'
        self.readingOrder = 'notSet'
        self.textRotation = 'notSet'
        self.wrapText = 'notSet'
        self.shrinkToFit = 'notSet'

    def set_horizontal(self, horizontal):
        if horizontal not in self._horizontal_alignments:
            raise ValueError(f'horizontal参数必须是{self._horizontal_alignments}其中之一。')
        self.horizontal = horizontal

    def set_vertical(self, vertical):
        if vertical not in self._vertical_alignments:
            raise ValueError(f'horizontal参数必须是{self._vertical_alignments}其中之一。')
        self.vertical = vertical

    def set_indent(self, indent):
        if not (isinstance(indent, int) and 0 <= indent <= 255):
            raise ValueError('value参数必须在0到255之间。')
        self.indent = indent

    def set_relativeIndent(self, indent):
        if not (isinstance(indent, int) and -255 <= indent <= 255):
            raise ValueError('value参数必须在-255到255之间。')
        self.relativeIndent = indent

    def set_justifyLastLine(self, on_off):
        self.justifyLastLine = on_off

    def set_readingOrder(self, value):
        if not (isinstance(value, int) and value >= 0):
            raise ValueError('value参数必须不小于0。')
        self.readingOrder = value

    def set_textRotation(self, value):
        if not (0 <= value <= 180 or value == 255):
            raise ValueError('value必须在0到180之间。')
        self.textRotation = value

    def set_wrapText(self, on_off):
        self.wrapText = on_off

    def set_shrinkToFit(self, on_off):
        self.shrinkToFit = on_off


class CellGradientFill(object):
    def __init__(self):
        self.type = 'notSet'
        self.degree = 'notSet'
        self.left = 'notSet'
        self.right = 'notSet'
        self.top = 'notSet'
        self.bottom = 'notSet'
        self.stop = 'notSet'

    def set_type(self, name):
        if name not in ('linear', 'path'):
            raise ValueError("name参数只能是 'linear', 'path' 之一。")
        self.type = name

    def set_degree(self, value):
        self.degree = value

    def set_left(self, value):
        self.left = value

    def set_right(self, value):
        self.right = value

    def set_top(self, value):
        self.top = value

    def set_bottom(self, value):
        self.bottom = value

    def set_stop(self, values):
        self.stop = values


class CellPatternFill(object):
    _FILES = ('none', 'solid', 'darkDown', 'darkGray', 'darkGrid', 'darkHorizontal', 'darkTrellis', 'darkUp',
              'darkVertical', 'gray0625', 'gray125', 'lightDown', 'lightGray', 'lightGrid', 'lightHorizontal',
              'lightTrellis', 'lightUp', 'lightVertical', 'mediumGray', None)

    def __init__(self):
        self.patternType = 'notSet'
        self.fgColor = 'notSet'
        self.bgColor = 'notSet'

    def set_patternType(self, name):
        if name not in self._FILES:
            raise ValueError(f'name参数只能是{self._FILES}其中之一。')
        self.patternType = name

    def set_fgColor(self, color):
        self.fgColor = get_color_code(color)
        if self.patternType == 'notSet':
            self.patternType = 'solid'

    def set_bgColor(self, color):
        self.bgColor = get_color_code(color)
        if self.patternType == 'notSet':
            self.patternType = 'solid'


class CellNumberFormat(object):
    def __init__(self):
        self.format = 'notSet'

    def set_format(self, string):
        if string is None:
            string = 'General'
        self.format = string


class CellProtection(object):
    def __init__(self):
        self.hidden = 'notSet'
        self.locked = 'notSet'

    def set_hidden(self, on_off):
        self.hidden = on_off

    def set_locked(self, on_off):
        self.locked = on_off


class CellStyleCopier(object):
    def __init__(self, from_cell):
        """
        :param from_cell: 被复制单元格对象
        """
        self._style = copy(from_cell._style)
        self._font = copy(from_cell.font)
        self._border = copy(from_cell.border)
        self._fill = copy(from_cell.fill)
        self._number_format = copy(from_cell.number_format)
        self._protection = copy(from_cell.protection)
        self._alignment = copy(from_cell.alignment)

    def to_cell(self, cell):
        cell._style = self._style
        cell.alignment = self._alignment
        cell.font = self._font
        cell.border = self._border
        cell.fill = self._fill
        cell.number_format = self._number_format
        cell.protection = self._protection


def get_color_code(color):
    if color is None:
        return '000000'
    if isinstance(color, Color):
        return color
    __COLORS__ = {
        'white': 'FFFFFF',
        'black': '000000',
        'red': 'FF0000',
        'green': '7FB80E',
        'blue': '009AD6',
        'purple': '8552A1',
        'yellow': 'FFFF00',
        'orange': 'F58220'
    }
    color = str(color)
    if ',' in color:
        color = color.replace(' ', '').lstrip('(').rstrip(')')
        RGB = color.split(',')
        color = ''
        for i in RGB:
            num = int(i)
            color += str(hex(num))[-2:].replace('x', '0').upper()
        return color

    return __COLORS__.get(color, color).lstrip('#')


class NoneStyle(object):
    _instance_lock = Lock()

    def __init__(self):
        self._font = Font()
        self._border = Border()
        self._alignment = Alignment()
        self._fill = PatternFill()
        self._number_format = 'General'
        self._protection = Protection()

    def __new__(cls, *args, **kwargs):
        if not hasattr(NoneStyle, "_instance"):
            with NoneStyle._instance_lock:
                if not hasattr(NoneStyle, "_instance"):
                    NoneStyle._instance = object.__new__(cls)
        return NoneStyle._instance

    def to_cell(self, cell, replace=True):
        cell.font = self._font
        cell.border = self._border
        cell.alignment = self._alignment
        cell.fill = self._fill
        cell.protection = self._protection
        cell.number_format = 'General'
