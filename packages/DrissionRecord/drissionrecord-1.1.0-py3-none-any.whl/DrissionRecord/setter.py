# -*- coding:utf-8 -*-
from pathlib import Path

from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook

from .cell_style import CellStyle
from .tools import (make_valid_name, make_final_data_simplify, make_final_data,
                    Header, ZeroHeader, process_content_xlsx, ok_list_str, data2ws_follow, data2ws, data2ws_style)


class OriginalSetter(object):
    def __init__(self, recorder):
        self._recorder = recorder

    def cache_size(self, size):
        if not isinstance(size, int) or size < 0:
            raise TypeError('cache_size值只能是int，且必须>=0')
        self._recorder._cache = size
        return self

    def path(self, path):
        if self._recorder._path:
            self._recorder.record()
        p = Path(path)
        self._recorder._path = str((p.parent / make_valid_name(p.name)).absolute())
        self._recorder._file_exists = False
        return self

    def show_msg(self, on_off):
        self._recorder.show_msg = on_off
        return self

    def auto_backup(self, interval=None, folder=None, overwrite=None):
        if folder is not None:
            self._recorder._backup_path = folder
        if isinstance(overwrite, bool):
            self._recorder._backup_overwrite = overwrite
        if interval is not None:
            self._recorder._backup_interval = interval
        return self


class BaseSetter(OriginalSetter):
    def table(self, name):
        self._recorder._table = name
        return self

    def auto_new_header(self, on_off=True):
        self._recorder.record()
        self._recorder._auto_new_header = on_off
        return self

    def before(self, data):
        return self._set_after_before(True, data)

    def after(self, data):
        return self._set_after_before(False, data)

    def _set_after_before(self, before, data):
        if isinstance(data, (list, dict)):
            data = data
        elif isinstance(data, tuple):
            data = list(data)
        elif data is not None:
            data = [data]
        setattr(self._recorder, '_before' if before else '_after', data)
        if self._recorder._after or self._recorder._before:
            self._recorder._make_final_data = make_final_data
        else:
            self._recorder._make_final_data = make_final_data_simplify
        return self


class RecorderSetter(BaseSetter):
    def encoding(self, encoding):
        self._recorder._encoding = encoding
        return self

    def header(self, header, table=None, to_file=True, row=None):
        if not header or not isinstance(header, (list, tuple)):
            raise ValueError('header不能为空且必须为list或tuple格式。')

        self._recorder.record()
        row = row or self._recorder._header_row.get(table, 1)
        with self._recorder._lock:
            header = Header(header)
            if self._recorder.type == 'xlsx':
                if table is None:
                    table = self._recorder.table
                elif table is True:
                    table = None
                elif not isinstance(table, str):
                    raise ValueError('table只能是None、True或str。')
                self._recorder._None_is_newest = table is None
                self._recorder._header[table] = header
                if to_file:
                    set_xlsx_header(self._recorder, header, table, row)
            elif self._recorder.type == 'csv':
                self._recorder._header[None] = header
                if to_file:
                    set_csv_header(self._recorder, header, row)
            else:
                self._recorder._header[None] = header

        return self

    def header_row(self, num, table=None):
        if num < 0:
            raise ValueError('num不能小于0。')
        self._recorder.record()
        with self._recorder._lock:
            if table is None:
                table = self._recorder.table
            elif table is True:
                table = None
            elif not isinstance(table, str):
                raise ValueError('table只能是None、True或str。')
            self._recorder._header_row[table] = num
            self._recorder._header[table] = ZeroHeader() if num == 0 else None
            self._recorder._None_header_is_newest = table is None
            self._recorder._None_header_row_is_newest = table is None
        return self

    def delimiter(self, delimiter):
        self._recorder._delimiter = delimiter
        return self

    def quote_char(self, quote_char):
        self._recorder._quote_char = quote_char
        return self

    def path(self, path, file_type=None):
        super().path(path)
        if not file_type:
            suffix = Path(path).suffix.lower()
            if suffix:
                file_type = suffix[1:]
        self.file_type(file_type)
        self._recorder._header = {None: None}
        self._recorder._header_row = {None: 1}
        self._recorder._None_header_is_newest = None
        self._recorder._None_header_row_is_newest = None
        return self

    def file_type(self, file_type):
        if file_type not in ('csv', 'xlsx', 'txt', 'jsonl', 'json'):
            file_type = 'txt'
        self._recorder._type = file_type
        self._recorder._set_methods(file_type)
        if file_type != 'xlsx':
            self._recorder._table = None
        return self

    def table(self, name):
        self._recorder._table = name if name is not True else None
        return self

    def follow_styles(self, on_off=True):
        self._recorder._follow_styles = on_off
        if on_off:
            self._recorder._styles = None
            self._recorder._row_height = None
            self._recorder._methods['data'] = data2ws_follow
        else:
            self._recorder._methods['data'] = data2ws
        return self

    def new_row_height(self, height):
        self._recorder._row_height = height
        if height is not None:
            self._recorder._follow_styles = False
            self._recorder._methods['data'] = data2ws_style
        else:
            self._recorder._methods['data'] = data2ws
        return self

    def new_row_styles(self, styles):
        self._recorder.record()
        self._recorder._styles = styles
        if styles is not None:
            self._recorder._follow_styles = False
            self._recorder._methods['data'] = data2ws_style
        else:
            self._recorder._methods['data'] = data2ws
        return self

    def data_col(self, col):
        if col is None:
            self._recorder.data_col = 0
        elif not isinstance(col, (int, str)):
            raise TypeError('col值只能是int、str或None。')
        else:
            self._recorder.data_col = col
        return self

    def link_style(self, style=True):
        if style is True:
            style = CellStyle()
            style.font.set_color("0000FF")
            style.font.set_underline('single')
        self._recorder._link_style = style
        return self


class DBSetter(BaseSetter):
    def path(self, path, table=None):
        with self._recorder._lock:
            super().path(path)
            if self._recorder._conn is not None:
                self._recorder._close_connection()
            self._recorder._connect()

            if table:
                self.table(table)
            else:
                r = self._recorder.run_sql("select name from sqlite_master where type='table'")
                self._recorder._table = r[0] if r else None

            self._recorder._data = {}
            self._recorder._close_connection()
        return self

    def table(self, name):
        if '`' in name:
            raise ValueError('table名称不能包含字符"`"。')
        self._recorder._table = name
        return self


def set_csv_header(recorder, header, row):
    if not recorder.path:
        raise FileNotFoundError('未指定文件。')
    from csv import writer
    if recorder._file_exists or Path(recorder.path).exists():
        with open(recorder.path, 'r', newline='', encoding=recorder._encoding) as f:
            lines = f.readlines()
            content1 = lines[:row - 1]
            content2 = lines[row:]

        with open(recorder.path, 'w', newline='', encoding=recorder._encoding) as f:
            f.write("".join(content1))
            csv_write = writer(f, delimiter=recorder._delimiter, quotechar=recorder._quote_char)
            con_len = len(content1)
            if con_len < row - 1:
                for _ in range(row - con_len - 1):
                    csv_write.writerow([])
            csv_write.writerow(ok_list_str(header))

        with open(recorder.path, 'a+', newline='', encoding=recorder._encoding) as f:
            f.write("".join(content2))

    else:
        Path(recorder.path).parent.mkdir(parents=True, exist_ok=True)
        with open(recorder.path, 'w', newline='', encoding=recorder._encoding) as f:
            csv_write = writer(f, delimiter=recorder._delimiter, quotechar=recorder._quote_char)
            for _ in range(row - 1):
                csv_write.writerow([])
            csv_write.writerow(ok_list_str(header))

    recorder._file_exists = True


def set_xlsx_header(recorder, header, table, row):
    if not recorder.path:
        raise FileNotFoundError('未指定文件。')
    if recorder._file_exists or Path(recorder.path).exists():
        wb = load_workbook(recorder.path)
        if table:
            ws = wb[table] if table in [i.title for i in wb.worksheets] else wb.create_sheet(title=table)
        else:
            ws = wb.active

    else:
        Path(recorder.path).parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        ws = wb.active
        if table:
            ws.title = table

    for c, i in header.items():
        ws.cell(row, c, value=process_content_xlsx(i))
    len_row = len(ws[row])
    len_header = len(header)
    if len_row > len_header:
        for c in range(len_header + 1, len_row + 1):
            ws.cell(row, c, value=None)

    wb.save(recorder.path)
    wb.close()
    recorder._file_exists = True
