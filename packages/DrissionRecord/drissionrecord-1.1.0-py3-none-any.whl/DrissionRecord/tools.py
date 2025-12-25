# -*- coding:utf-8 -*-
from collections.abc import Iterable
from csv import reader as csv_reader, writer as csv_writer
from pathlib import Path
from re import search, sub, match

from openpyxl.cell import Cell, ReadOnlyCell
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook

from .cell_style import CellStyle, CellStyleCopier, NoneStyle


def line2ws(ws, header, row, col, data, rewrite_method, rewrite):
    if isinstance(data, dict):
        data, rewrite, header_len = header.__getattribute__(rewrite_method)(data, 'xlsx', rewrite)
        for c, val in data.items():
            ws.cell(row, c, value=process_content_xlsx(val))
    else:
        for key, val in enumerate(data):
            ws.cell(row, col + key, value=process_content_xlsx(val))
    return rewrite


def line2ws_follow(ws, header, row, col, data, rewrite_method, rewrite, styles, height, new_row):
    if new_row:
        styles2new_row(ws, styles.values(), height, row)

    if isinstance(data, dict):
        data, rewrite, header_len = header.__getattribute__(rewrite_method)(data, 'xlsx', rewrite)
        if new_row:
            for c, val in data.items():
                ws.cell(row, c, value=process_content_xlsx(val))
        else:
            for c, val in data.items():
                styles.get(c, NoneStyle()).to_cell(ws.cell(row, c, value=process_content_xlsx(val)))
    else:
        if new_row:
            for key, val in enumerate(data):
                ws.cell(row, col + key, value=process_content_xlsx(val))
        else:
            for key, val in enumerate(data):
                col1 = col + key
                styles.get(col1, NoneStyle()).to_cell(ws.cell(row, col1, value=process_content_xlsx(val)))
    return rewrite


def data2ws(recorder, ws, data, coord, header, rewrite, rewrite_method, new_row):
    row, col = coord
    for r, d in enumerate(data['data'], row):
        rewrite = line2ws(ws, header, r, col, d, rewrite_method, rewrite)
    return rewrite


def data2ws_follow(recorder, ws, data, coord, header, rewrite, rewrite_method, new_row):
    row, col = coord
    if row > 1:
        styles = {ind: CellStyleCopier(cell) for ind, cell in enumerate(ws[row - 1], 1)}
        height = ws.row_dimensions[row - 1].height
        for r, d in enumerate(data['data'], row):
            rewrite = line2ws_follow(ws, header, r, col, d, rewrite_method, rewrite, styles, height, new_row)

    else:
        for r, d in enumerate(data['data'], row):
            rewrite = line2ws(ws, header, r, col, d, rewrite_method, rewrite)

    return rewrite


def data2ws_style(recorder, ws, data, coord, header, rewrite, rewrite_method, new_row):
    row, col = coord
    if new_row:
        styles = recorder._styles
        if isinstance(styles, dict):
            styles = header.make_num_dict(styles, None)[0]
            styles = [styles.get(c, None) for c in range(1, ws.max_column + 1)]
        elif isinstance(styles, CellStyle):
            styles = [styles] * ws.max_column
        height = ws.row_dimensions[row].height

        for r, d in enumerate(data['data'], row):
            rewrite = line2ws(ws, header, r, col, d, rewrite_method, rewrite)
            styles2new_row(ws, styles, height, row)

    else:
        for r, d in enumerate(data['data'], row):
            rewrite = line2ws(ws, header, r, col, d, rewrite_method, rewrite)

    return rewrite


def styles2new_row(ws, styles, height, row):
    if height is not None:
        ws.row_dimensions[row].height = height
    if styles:
        for c, s in enumerate(styles, start=1):
            if s:
                s.to_cell(ws.cell(row=row, column=c))


def styles2ws(**kwargs):
    ws = kwargs['ws']
    header = kwargs['header']
    data = kwargs['data']
    styles = data['styles']
    coord = data['real_coord']  # 'A3'、'A1:C3'、(1, 3)、['A1', 'B2', 'C3']
    rows = data['rows']  # 3、'1:3'、[1, 2, 3]
    cols = data['cols']
    mode = data['mode'] == 'replace'
    if not styles:
        styles = [NoneStyle()]
    elif isinstance(styles, CellStyle):
        styles = [styles]

    if isinstance(styles, dict):
        for coord, val in styles.items():
            styles2ws(ws=ws, header=header,
                      data={'styles': val, 'real_coord': coord, 'rows': None, 'cols': None, 'mode': mode})
        return

    if rows:
        if isinstance(rows, int):
            styles_len = len(styles)
            for col, cell in enumerate(ws[header.get_num(rows)]):
                styles[col % styles_len].to_cell(cell, replace=mode)

        elif isinstance(rows, str) and ':' in rows:
            begin, end = rows.split(':', 1)
            try:
                begin = header.get_num(int(begin))
                end = header.get_num(int(end))
            except ValueError:
                raise ValueError('行号必须是数字，现在是：', rows)
            if begin > end:
                begin, end = end, begin
            for i in range(begin, end + 1):
                styles2ws(ws=ws, header=header,
                          data={'styles': styles, 'real_coord': None, 'rows': i, 'cols': None, 'mode': mode})

        elif isinstance(rows, (tuple, list)):
            for i in rows:
                styles2ws(ws=ws, header=header,
                          data={'styles': styles, 'real_coord': None, 'rows': i, 'cols': None, 'mode': mode})

    if cols:
        if isinstance(cols, int):  # 列序号
            styles_len = len(styles)
            for col, cell in enumerate(ws[header.get_col(cols)]):
                styles[col % styles_len].to_cell(cell, replace=mode)

        elif isinstance(cols, str):  # 表头值
            cols = header.get_num(cols)
            if cols:
                styles2ws(ws=ws, header=header,
                          data={'styles': styles, 'real_coord': None, 'rows': None, 'cols': cols, 'mode': mode})

        elif isinstance(cols, tuple) and len(cols) == 2:
            begin, end = cols
            begin = header.get_num(begin)
            end = header.get_num(end)
            if begin > end:
                begin, end = end, begin
            for i in range(begin, end + 1):
                styles2ws(ws=ws, header=header,
                          data={'styles': styles, 'real_coord': None, 'rows': None, 'cols': i, 'mode': mode})

        elif isinstance(cols, (tuple, list)):
            for i in cols:
                styles2ws(ws=ws, header=header,
                          data={'styles': styles, 'real_coord': None, 'rows': None, 'cols': i, 'mode': mode})

    if coord:
        if isinstance(coord, str):
            if ':' in coord:
                begin, end = coord.split(':', 1)
                begin = parse_coord(begin)
                end = parse_coord(end)
                begin = f'{header.get_col(begin[1])}{begin[0]}'
                end = f'{header.get_col(end[1])}{end[0]}'
                styles_len = len(styles)
                for row in ws[f'{begin}:{end}']:
                    for col, cell in enumerate(row):
                        styles[col % styles_len].to_cell(cell, replace=mode)
            else:
                coord = parse_coord(coord)
                coord = f'{header.get_col(coord[1])}{coord[0]}'
                styles[0].to_cell(ws[coord], replace=mode)

        elif isinstance(coord, tuple) and len(coord) == 2:
            coord = parse_coord(coord)
            coord = f'{header.get_col(coord[1])}{coord[0]}'
            styles[0].to_cell(ws[coord], replace=mode)

        elif isinstance(coord, (tuple, list)):
            for i in coord:
                styles2ws(ws=ws, header=header,
                          data={'styles': styles, 'real_coord': i, 'rows': None, 'cols': None, 'mode': mode})


def link2ws(**kwargs):
    recorder = kwargs['recorder']
    data = kwargs['data']
    cell = kwargs['ws'].cell(*kwargs['coord'])
    has_link = bool(cell.hyperlink)
    cell.hyperlink = data['link']
    if data['content'] is not None:
        cell.value = process_content_xlsx(data['content'])
    if data['link']:
        if recorder._link_style:
            recorder._link_style.to_cell(cell, replace=False)
    elif has_link:
        NoneStyle().to_cell(cell, replace=False)


def img2ws(**kwargs):
    row, col = kwargs['coord']
    data = kwargs['data']
    ws = kwargs['ws']
    from openpyxl.drawing.image import Image
    img = Image(data['imgPath'])
    width, height = data['width'], data['height']
    if width and height:
        img.width = width
        img.height = height
    elif width:
        img.height = int(img.height * (width / img.width))
        img.width = width
    elif height:
        img.width = int(img.width * (height / img.height))
        img.height = height
    # ws.add_image(img, (row, Header._NUM_KEY[col]))
    ws.add_image(img, f'{Header._NUM_KEY[col]}{row}')


def width2ws(**kwargs):
    # 用int表示列序号，str表示表头值，用tuple设置某列到某列，用list指定每一列，为Ture设置所有列
    cols = kwargs['data']['cols']
    width = kwargs['data']['width']
    ws = kwargs['ws']
    header = kwargs['header']

    if isinstance(width, dict):
        for col, val in width.items():
            width2ws(ws=ws, header=header, data={'cols': col, 'width': val})

    elif isinstance(cols, (int, str)):  # 表头值或列序号
        cols = header.get_col(cols)
        if cols:
            ws.column_dimensions[cols].width = width

    elif isinstance(cols, tuple) and len(cols) == 2:  # 连续多列
        beg, end = cols
        if not beg:
            beg = 1
        if not end:
            end = -1

        beg = header.get_num(beg)
        end = header.get_num(end)

        if beg and end:
            if beg > end:
                beg, end = end, beg
            for c in range(beg, end + 1):
                ws.column_dimensions[header.get_col(c)].width = width

    elif isinstance(cols, list):
        for col in cols:
            width2ws(ws=ws, header=header, data={'cols': col, 'width': width})

    elif cols is True:
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[ZeroHeader()[col]].width = width


def height2ws(**kwargs):
    # int表示行号，str为'1:3'格式，dict可指定行独立设置高度，list或tuple指定多行设置同一个高度。True设置所有行
    rows = kwargs['data']['rows']
    height = kwargs['data']['height']
    ws = kwargs['ws']

    if isinstance(height, dict):
        for row, val in height.items():
            height2ws(ws=ws, data={'rows': row, 'height': val})

    elif isinstance(rows, int):
        if rows < 1:
            rows = get_real_row(rows, ws.max_row)
        ws.row_dimensions[rows].height = height

    elif isinstance(rows, str) and ':' in rows:
        beg, end = rows.split(':', 1)
        if beg == '':
            beg = 1
        if end == '':
            end = -1
        beg = int(beg)
        end = int(end)
        if beg < 1 or end < 1:
            max_row = ws.max_row
            beg = get_real_row(beg, max_row)
            end = get_real_row(end, max_row)

        if beg > end:
            beg, end = end, beg
        for c in range(beg, end + 1):
            ws.row_dimensions[c].height = height

    elif isinstance(rows, (list, tuple)):
        for row in rows:
            height2ws(ws=ws, data={'rows': row, 'height': height})

    elif rows is True:
        for i in range(1, ws.max_row + 1):
            ws.row_dimensions[i].height = height


def is_single_data(data):
    return not isinstance(data, Iterable) or isinstance(data, str)


def is_1D_data(data):
    if isinstance(data, dict):
        return True
    for i in data:
        return is_single_data(i)


def remove_end_Nones(in_list):
    h = []
    flag = True
    for i in in_list[::-1]:
        if flag:
            if i in (None, ''):
                continue
            else:
                flag = False
        h.append(i)
    return h[::-1]


def _get_column_letter(col_idx):
    letters = []
    while col_idx > 0:
        col_idx, remainder = divmod(col_idx, 26)
        if remainder == 0:
            remainder = 26
            col_idx -= 1
        letters.append(chr(remainder + 64))
    return ''.join(reversed(letters))


def process_content_xlsx(content):
    if isinstance(content, (str, int, float, type(None))):
        data = content
    elif isinstance(content, (Cell, ReadOnlyCell)):
        data = content.value
    else:
        data = str(content)

    if isinstance(data, str):
        data = sub(r'[\000-\010]|[\013-\014]|[\016-\037]', '', data)

    return data


def process_content_json(content):
    if isinstance(content, (str, int, float, type(None))):
        return content
    elif isinstance(content, (Cell, ReadOnlyCell)):
        return content.value
    else:
        return str(content)


def process_content_str(content):
    if isinstance(content, str):
        return content
    elif content is None:
        return ''
    elif isinstance(content, (Cell, ReadOnlyCell)):
        return str(content.value)
    else:
        return str(content)


def process_nothing(content):
    return content


def do_nothing(*args, **kwargs):
    return


class BaseHeader(object):
    _NUM_KEY = {}
    _KEY_NUM = {}
    _CONTENT_FUNCS = {'csv': process_content_str,
                      'xlsx': process_content_xlsx,
                      None: process_nothing}

    def __new__(cls, header=None):
        if not cls._NUM_KEY:
            for i in range(1, 18279):
                col = _get_column_letter(i)
                cls._NUM_KEY[i] = col
                cls._KEY_NUM[col] = i
        return object.__new__(cls)

    @property
    def _str_num(self):
        return Header._KEY_NUM

    @property
    def _num_str(self):
        return Header._NUM_KEY

    def __iter__(self):
        return iter(self.key_num)


class Header(BaseHeader):
    def __init__(self, header=None):
        if isinstance(header, dict):
            self._NUM_KEY = {c: str(v) if v not in ('', None) else None for c, v in header.items()}
        elif isinstance(header, (list, tuple)):
            self._NUM_KEY = {c: str(i) if i not in ('', None) else c
                             for c, i in enumerate(remove_end_Nones(header), start=1)}
        elif isinstance(header, Iterable):
            self._NUM_KEY = {c: str(i) if i not in ('', None) else c
                             for c, i in enumerate(remove_end_Nones(list(header)), start=1)}
        else:
            self._NUM_KEY = {}
            self._KEY_NUM = {}
            return
        self._KEY_NUM = {c: h for h, c in self._NUM_KEY.items()} if self._NUM_KEY else {}

    @property
    def key_num(self):
        return self._KEY_NUM

    @property
    def num_key(self):
        return self._NUM_KEY

    def values(self):
        return self.num_key.values()

    def items(self):
        return self.num_key.items()

    def make_row_data(self, row, row_values, None_val=None):
        data = {self.get_key(col): val for col, val in row_values.items()}
        return RowData(row, self, None_val, data)

    def make_insert_list(self, data, file_type, rewrite):  # 修改时记得ZeroHeader对应方法
        if isinstance(data, dict):
            data = self.make_num_dict(data, file_type)[0]
            data = [data.get(i, None) for i in range(1, max(max(data), len(self.num_key)) + 1)] if data else []
        else:
            data = [self._CONTENT_FUNCS[file_type](v) for v in data]
        return data, False

    def make_insert_list_rewrite(self, data, file_type, rewrite):
        if isinstance(data, dict):
            data, rewrite, header_len = self.make_num_dict_rewrite(data, file_type, rewrite)
            data = [data.get(i, None) for i in range(1, max(max(data), header_len) + 1)]
        else:
            data = [self._CONTENT_FUNCS[file_type](v) for v in data]
        return data, rewrite

    def make_change_list(self, line_data, data, col, file_type, rewrite):
        if isinstance(data, dict):
            data = self.make_num_dict(data, file_type)[0]
            raw_data = {c: v for c, v in enumerate(line_data, 1)}
            raw_data = {**raw_data, **data}
            line_data = [raw_data.get(c, None) for c in range(1, max(raw_data) + 1)]
        else:
            line_data.extend([''] * (col - len(line_data) + len(data) - 1))  # 若列数不够，填充空列
            for k, j in enumerate(data):  # 填充数据
                line_data[col + k - 1] = self._CONTENT_FUNCS[file_type](j)
        return line_data, False

    def make_change_list_rewrite(self, line_data, data, col, file_type, rewrite):
        if isinstance(data, dict):
            data, rewrite, header_len = self.make_num_dict_rewrite(data, file_type, rewrite)
            raw_data = {c: v for c, v in enumerate(line_data, 1)}
            raw_data = {**raw_data, **data}
            line_data = [raw_data.get(c, None) for c in range(1, max(raw_data) + 1)]
        else:
            line_data.extend([''] * (col - len(line_data) + len(data) - 1))  # 若列数不够，填充空列
            for k, j in enumerate(data):  # 填充数据
                line_data[col + k - 1] = self._CONTENT_FUNCS[file_type](j)
        return line_data, rewrite

    def make_num_dict(self, *keys):
        data = keys[0]
        file_type = keys[1]
        val = {}
        for k, v in data.items():
            num = self.get_num(k)
            if num:
                val[num] = self._CONTENT_FUNCS[file_type](v)
        return val, False, 0

    def make_num_dict_rewrite(self, *keys):
        data, file_type, rewrite = keys
        val = {}
        header_len = len(self.num_key)
        for k, v in data.items():
            if isinstance(k, str) and k not in self.key_num:
                header_len += 1
                self.key_num[k] = header_len
                self.num_key[header_len] = k
                rewrite = True
            num = self.get_num(k)
            if num:
                val[num] = self._CONTENT_FUNCS[file_type](v)
        return val, rewrite, header_len

    def get_key(self, num):
        key = self[num]
        return num if key is None else key

    def get_col(self, header_or_num):
        num = self.get_num(header_or_num)
        return ZeroHeader()[num] if num else None

    def get_num(self, header_or_num):  # 修改时记得ZeroHeader
        if isinstance(header_or_num, int):
            return self._num2num(header_or_num)
        elif isinstance(header_or_num, str):
            return self.key_num.get(header_or_num, None)
        else:
            raise TypeError(f'col值只能是int或str。当前值：{header_or_num}')

    def _get_num(self, header_or_num):
        return self.get_num(header_or_num) or len(self) + 1

    def _num2num(self, num):
        if num > 0:
            return num
        elif num < 0:
            l = len(self)
            return num % l + 1 if -num <= l else None
        else:
            return len(self) + 1

    def __getitem__(self, item):
        if isinstance(item, str):
            return self.key_num.get(item)
        elif isinstance(item, int) and item != 0:
            return self.num_key.get(self._num2num(item), None)
        else:
            raise ValueError('值只能时str或int，且不能为0。')

    def __len__(self):
        return len(self.num_key)

    def __repr__(self):
        return str(self.num_key)

    def __bool__(self):
        return True if self.num_key else False


class ZeroHeader(Header):
    _OBJ = None

    def __new__(cls):
        super().__new__(cls)
        if cls._OBJ is None:
            cls._OBJ = object.__new__(cls)
        return cls._OBJ

    def __init__(self):
        return

    def get_num(self, col):
        if isinstance(col, int) and col > 0:
            return col
        elif isinstance(col, str):
            return self.key_num.get(col.upper(), None)
        else:
            raise TypeError(f'表头行为0时，col值只能str或大于0的int。当前值：{col}')

    def make_insert_list(self, data, file_type, rewrite):
        if isinstance(data, dict):
            val = self.make_num_dict(data, file_type)[0]
            data = [val.get(c, None) for c in range(1, max(val) + 1)] if val else []
        else:
            data = [self._CONTENT_FUNCS[file_type](v) for v in data]
        return data, False

    def make_insert_list_rewrite(self, data, file_type, rewrite):
        return self.make_insert_list(data, file_type, rewrite)

    def make_num_dict_rewrite(self, *keys):
        data, file_type, rewrite = keys
        return self.make_num_dict(data, file_type)

    def get_col(self, header_or_num):
        return self[header_or_num] if isinstance(header_or_num, int) else header_or_num

    def _num2num(self, num):
        if num > 0:
            return num if num <= len(self) else None
        elif num < 0:
            return num % len(self) + 1 if -num <= len(self) else None
        else:
            raise ValueError('列序号不能为0。')

    def _get_num(self, header_or_num):
        return self.get_num(header_or_num) or 1

    def __getitem__(self, item):
        return self.num_key.get(item, None) if isinstance(item, int) else self.key_num.get(item.upper(), None)

    def __len__(self):
        return 0


class RowData(dict):
    def __init__(self, row, header, None_val, seq):
        self.header = header
        self.row = row
        self._None_val = None_val
        super().__init__(seq)

    def __getitem__(self, item):
        ite = self.header[item] if isinstance(item, int) else item
        if ite is None:
            raise RuntimeError(f'header中无{item}项。\nheader：{self.header.values()}')
        return self.get(ite, self._None_val)

    def col(self, key_or_num, as_num=True):
        return self.header.get_num(key_or_num) if as_num else self.header.get_col(key_or_num)

    def coord(self, key_or_num, col_num=False):
        return self.row, self.col(key_or_num, col_num)


class RowText(str):
    def __init__(self, value):
        super().__init__()
        self.value = value
        self.row = None


def Col(key):
    return ZeroHeader().key_num[key.upper()]


def align_csv(path, encoding='utf-8', delimiter=',', quotechar='"'):
    with open(path, 'r', encoding=encoding) as f:
        reader = csv_reader(f, delimiter=delimiter, quotechar=quotechar)
        lines = list(reader)
        lines_data = {}
        max_len = 0

        # 把每行列数用字典记录，并找到最长的一行
        for k, i in enumerate(lines):
            line_len = len(i)
            if line_len > max_len:
                max_len = line_len
            lines_data[k] = line_len

        # 把所有行用空值补全到和最长一行一样
        for i in lines_data:
            lines[i].extend([None] * (max_len - lines_data[i]))

        writer = csv_writer(open(path, 'w', encoding=encoding, newline=''), delimiter=delimiter, quotechar=quotechar)
        writer.writerows(lines)


def get_usable_path(path, is_file=True, parents=True):
    path = Path(path)
    parent = path.parent
    if parents:
        parent.mkdir(parents=True, exist_ok=True)
    path = parent / make_valid_name(path.name)
    name = path.stem if path.is_file() else path.name
    ext = path.suffix if path.is_file() else ''

    first_time = True

    while path.exists() and path.is_file() == is_file:
        r = search(r'(.*)_(\d+)$', name)

        if not r or (r and first_time):
            src_name, num = name, '1'
        else:
            src_name, num = r.group(1), int(r.group(2)) + 1

        name = f'{src_name}_{num}'
        path = parent / f'{name}{ext}'
        first_time = None

    return path


def make_valid_name(full_name):
    # ----------------去除前后空格----------------
    full_name = full_name.strip()

    # ----------------去除不允许存在的字符----------------
    if search(r'[<>/\\|:*?\n"]', full_name):
        full_name = sub(r'<', '＜', full_name)
        full_name = sub(r'>', '＞', full_name)
        full_name = sub(r'/', '／', full_name)
        full_name = sub(r'\\', '＼', full_name)
        full_name = sub(r'\|', '｜', full_name)
        full_name = sub(r':', '：', full_name)
        full_name = sub(r'\*', '＊', full_name)
        full_name = sub(r'\?', '？', full_name)
        full_name = sub(r'\n', '', full_name)
        full_name = sub(r'"(.*?)"', r'“\1”', full_name)
        full_name = sub(r'"', '“', full_name)

    # ----------------使总长度不大于255个字符（一个汉字是2个字符）----------------
    r = search(r'(.*)(\.[^.]+$)', full_name)  # 拆分文件名和后缀名
    if r:
        name, ext = r.group(1), r.group(2)
        ext_long = len(ext)
    else:
        name, ext = full_name, ''
        ext_long = 0

    while get_long(name) > 255 - ext_long:
        name = name[:-1]

    return f'{name}{ext}'.rstrip('.')


def get_long(txt):
    txt_len = len(txt)
    return int((len(txt.encode('utf-8')) - txt_len) / 2 + txt_len)


def parse_coord(coord, data_col=1):
    if not coord:  # 新增一行，列为data_col
        return_coord = 0, data_col

    elif isinstance(coord, int):
        return_coord = coord, data_col

    elif isinstance(coord, str):  # 'A3'格式
        m = match(r'^[$]?([A-Za-z]{1,3})[$]?(-?\d+)$', coord)
        if not m:
            raise ValueError(f'{coord} 坐标格式不正确。')
        y, x = m.groups()
        return_coord = int(x), ZeroHeader()[y] or 1

    elif isinstance(coord, (tuple, list)) and len(coord) == 2:
        if isinstance(coord[0], int):
            x = int(coord[0])
        elif coord[0] is None:
            x = 0
        else:
            raise TypeError('行格式不正确。')

        if isinstance(coord[1], (str, int)):
            y = coord[1]
        elif coord[1] is None:
            y = 0
        else:
            raise TypeError('列格式不正确。')

        return_coord = x, y

    else:
        raise ValueError(f'{coord} 坐标格式不正确。')

    return return_coord


def ok_list_xlsx(data_list):
    if isinstance(data_list, (dict, Header)):
        data_list = data_list.values()
    return [process_content_xlsx(i) for i in data_list]


def ok_list_str(data_list):
    if isinstance(data_list, (dict, Header)):
        data_list = data_list.values()
    return [process_content_str(i) for i in data_list]


def ok_list_db(data_list):
    if isinstance(data_list, (dict, Header)):
        data_list = data_list.values()
    return [process_content_json(i) for i in data_list]


def get_real_row(row, max_row):
    if row <= 0:
        row = max_row + row + 1
    return 1 if row < 1 else row


def get_real_coord(coord, max_row, header):
    row, col = coord
    return get_real_row(row, max_row), header._get_num(col)


def get_ws_real_coord(coord, ws, header):
    row, col = coord
    if row <= 0:
        row = ws.max_row + row + 1
    return 1 if row < 1 else row, header._get_num(col)


def make_final_data_simplify(recorder, data):
    return data if isinstance(data, (dict, list, tuple)) else list(data)


def make_final_data(recorder, data):
    if isinstance(data, dict):
        if isinstance(recorder.before, dict):
            data = {**recorder.before, **data}
        if isinstance(recorder.after, dict):
            data = {**data, **recorder.after}
        return data

    else:
        return_list = []
        for i in (recorder.before, data, recorder.after):
            if isinstance(i, dict):
                return_list.extend(list(i.values()))
            elif not i:
                pass
            else:
                return_list.extend(list(i))
        return return_list


def _set_style(height, styles, ws, row):
    if height is not None:
        ws.row_dimensions[row].height = height

    if styles:
        if isinstance(styles, CellStyle):
            for c in ws[row]:
                styles.to_cell(c)
        else:
            for k, s in enumerate(styles, start=1):
                if s:
                    s.to_cell(ws.cell(row=row, column=k))


def get_csv(recorder):
    new_csv = not recorder._file_exists and not Path(recorder.path).exists()
    return open(recorder.path, 'a+', newline='', encoding=recorder.encoding), new_csv


def get_wb(recorder):
    if recorder._file_exists or Path(recorder.path).exists():
        wb = load_workbook(recorder.path)
        new_file = False
    else:
        wb = Workbook()
        new_file = True
    return wb, new_file


def get_ws(wb, table, tables, new_file):
    new_sheet = new_file
    if table is None:
        ws = wb.active
        if ws.max_row == 1 and ws.max_column == 1 and ws.cell(row=1, column=1).value is None:
            new_sheet = True

    elif table in tables:
        ws = wb[table]
        if ws.max_row == 1 and ws.max_column == 1 and not ws.cell(row=1, column=1).value:
            new_sheet = True

    elif new_file is True:
        ws = wb.active
        tables.remove(ws.title)
        ws.title = table
        tables.append(table)
        new_sheet = True

    else:
        ws = wb.create_sheet(title=table)
        tables.append(table)
        new_sheet = True

    return ws, new_sheet


def get_tables(path):
    wb = load_workbook(path)
    tables = wb.sheetnames
    wb.close()
    return tables


def get_key_cols(cols, header):
    if cols is True:
        return True
    elif isinstance(cols, (int, str)):
        cols = header.get_num(cols)
        return [cols] if cols else []
    elif isinstance(cols, (list, tuple)):
        res = []
        for i in cols:
            i = header.get_num(i)
            if i:
                res.append(i)
        return res
    else:
        raise TypeError('col值只能是int或str。')
