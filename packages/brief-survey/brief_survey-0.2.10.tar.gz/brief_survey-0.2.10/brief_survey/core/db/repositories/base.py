import contextlib
import csv
from typing import (
    Any,
    AsyncGenerator,
    Dict,
    Generator,
    List,
    Optional,
    Sequence,
    Type,
    Union,
    cast, TypeVar, Generic,
)

import openpyxl
from sqlalchemy import delete, select, update
from sqlalchemy.ext.asyncio import (
    AsyncEngine,
    AsyncSession,
    create_async_engine,
)
from sqlalchemy.engine import Engine
from sqlalchemy.orm import (
    Session as SQLSession,
    scoped_session,
    sessionmaker,
)

from openpyxl.utils import get_column_letter

try:
    from sqlalchemy.ext.asyncio import async_sessionmaker
except ImportError:
    async_sessionmaker = type("async_sessionmaker", (type,), {})  # type: ignore[assignment,misc]

DBConnection = Union[AsyncEngine, Engine, str]

from ..db_models import Base

T = TypeVar("T", bound=Base)

class BaseRepository(Generic[T]):
    """
    Универсальный репозиторий для SQLAlchemy моделей с произвольными CRUD операциями.

    Все фильтры для get/update/delete передаются через kwargs.
    Создание и обновление с передачей kwargs для полей.
    """

    @property
    def Session(self) -> Union[scoped_session, async_sessionmaker]:
        return self.session_maker

    def __init__(
        self,
        sql_model_class: Type[T],
        connection_string: Optional[str] = 'sqlite+aiosqlite:///brief.db',
        connection: Optional[DBConnection] = None,
        engine_args: Optional[Dict[str, Any]] = None,
        async_mode: Optional[bool] = None,
        table_name: Optional[str] = None,

    ):


        if connection_string and connection:
            raise ValueError("connection_string and connection are mutually exclusive")

        if connection_string:
            connection = connection_string



        if isinstance(connection, str):
            self.async_mode = async_mode if async_mode is not None else False
            if self.async_mode:
                self.async_engine = create_async_engine(connection, **(engine_args or {}))
            else:
                from sqlalchemy import create_engine

                self.engine = create_engine(connection, **(engine_args or {}))
        elif isinstance(connection, Engine):
            self.async_mode = False
            self.engine = connection
        elif isinstance(connection, AsyncEngine):
            self.async_mode = True
            self.async_engine = connection
        else:
            raise ValueError(
                "connection should be connection string or instance of "
                "sqlalchemy.engine.Engine or sqlalchemy.ext.asyncio.engine.AsyncEngine"
            )

        self.session_maker: Union[scoped_session, async_sessionmaker]
        if self.async_mode:
            self.session_maker = async_sessionmaker(bind=self.async_engine)
        else:
            self.session_maker = scoped_session(sessionmaker(bind=self.engine))
        if table_name:

            DynamicModel = type(
                f"{sql_model_class.__name__}DynamicTable",
                (sql_model_class,),
                {"__tablename__": table_name}
            )
            self.sql_model_class = DynamicModel
        else:
            self.sql_model_class = sql_model_class

        if not self.async_mode:
            self._create_table_if_not_exists()

        self._table_created = False

    def _create_table_if_not_exists(self) -> None:
        self.sql_model_class.metadata.create_all(self.engine)
        self._table_created = True

    async def _acreate_table_if_not_exists(self) -> None:
        if not self._table_created:
            assert self.async_mode, "Async only"
            async with self.async_engine.begin() as conn:
                await conn.run_sync(self.sql_model_class.metadata.create_all)
            self._table_created = True

    def get_all(self, *args: Any, **filters: Any) -> List[T]:
        """
            Получить все записи, соответствующие переданным фильтрам.

            Поддерживаемые фильтры:
                - Обычное равенство: field=value
                - "LIKE" с подстрокой: field__like=value

            Аргументы:
                *args: позиционные условия (обычно не используются).
                **filters: фильтры по полям модели.
                    Ключи с обычными именами фильтруют по равенству.
                    Ключи с суффиксом '__like' фильтруют по оператору LIKE.
                    Например:
                        get_all(status='active', title__like='пример')

            Возвращает:
                Список объектов модели, где все условия выполнены (AND).

            Исключения:
                AttributeError — при отсутствии поля в модели.
        """
        with self._make_sync_session() as session:
            query = session.query(self.sql_model_class)

            if filters:
                # Собираем условия
                conditions = []
                for key, value in filters.items():
                    if key.endswith("__like"):
                        field_name = key[:-6]
                        field = getattr(self.sql_model_class, field_name, None)
                        if field is None:
                            raise AttributeError(f"Model has no field '{field_name}'")
                        conditions.append(field.like(f"%{value}%"))
                    else:
                        field = getattr(self.sql_model_class, key, None)
                        if field is None:
                            raise AttributeError(f"Model has no field '{key}'")
                        conditions.append(field == value)
                query = query.filter(*conditions)

            query = query.order_by(self.sql_model_class.id.asc())
            return query.all()

    async def aget_all(self, *args: Any, **filters: Any) -> List[T]:
        """
        Асинхронно получить все записи, удовлетворяющие фильтрам.

        Поддерживаемые фильтры:
            - Обычное равенство: field=value
            - "LIKE" с подстрокой: field__like=value

        Аргументы:
            *args: позиционные условия (обычно не используются).
            **filters: фильтры по полям модели.
                Ключи с суффиксом '__like' используются для фильтрации через LIKE.
                Пример:
                    await aget_all(status='active', description__like='транспорт')

        Возвращает:
            Список объектов модели, удовлетворяющих всем фильтрам (логика AND).

        Исключения:
            AttributeError — если указано поле, отсутствующее в модели.
        """
        await self._acreate_table_if_not_exists()
        async with self._make_async_session() as session:
            stmt = select(self.sql_model_class)

            if filters:
                from sqlalchemy import and_
                conditions = []
                for key, value in filters.items():
                    if key.endswith("__like"):
                        field_name = key[:-6]
                        field = getattr(self.sql_model_class, field_name, None)
                        if field is None:
                            raise AttributeError(f"Model has no field '{field_name}'")
                        conditions.append(field.like(f"%{value}%"))
                    else:
                        field = getattr(self.sql_model_class, key, None)
                        if field is None:
                            raise AttributeError(f"Model has no field '{key}'")
                        conditions.append(field == value)
                stmt = stmt.where(and_(*conditions))

            stmt = stmt.order_by(self.sql_model_class.id.asc())
            result = await session.execute(stmt)
            return result.scalars().all()


    def get_one(self, **filters: Any) -> Optional[T]:
        """
        Получить одну запись по фильтрам.

        Возвращает None, если не найдено.
        """
        with self._make_sync_session() as session:
            query = session.query(self.sql_model_class)
            if filters:
                query = query.filter_by(**filters)
            res:T = query.first()
            return res

    async def aget_one(self, **filters: Any) -> Optional[T]:
        """Асинхронный аналог get_one."""
        await self._acreate_table_if_not_exists()
        async with self._make_async_session() as session:
            stmt = select(self.sql_model_class)
            if filters:
                stmt = stmt.filter_by(**filters)
            result = await session.execute(stmt)
            return result.scalars().first()

    def create(self, **fields: Any) -> T:
        """
        Создать новую запись с переданными полями.

        Возвращает созданный объект (но ещё не сохранён в БД, если требуется можно вызвать add_item).
        """
        obj = self.sql_model_class(**fields)
        self.add_item(obj)
        return obj

    async def acreate(self, **fields: Any) -> T:
        """
        Асинхронно создать новую запись и сохранить в БД.
        """
        obj = self.sql_model_class(**fields)
        await self.aadd_item(obj)
        return obj

    def update(self, values: Dict[str, Any], **filters: Any) -> int:
        """
        Обновить записи, удовлетворяющие фильтрам, новыми значениями.

        Args:
            values: словарь поле->значение для обновления.
            **filters: фильтры для выбора записей.

        Returns:
            Количество обновлённых записей.
        """
        with self._make_sync_session() as session:
            stmt = update(self.sql_model_class).where(
                *[getattr(self.sql_model_class, k) == v for k, v in filters.items()]
            ).values(**values)
            result = session.execute(stmt)
            session.commit()
            return result.rowcount or 0

    async def aupdate(self, values: Dict[str, Any], **filters: Any) -> int:
        """
        Асинхронное обновление записей по фильтрам.
        """
        await self._acreate_table_if_not_exists()
        async with self._make_async_session() as session:
            stmt = update(self.sql_model_class).where(
                *[getattr(self.sql_model_class, k) == v for k, v in filters.items()]
            ).values(**values)
            result = await session.execute(stmt)
            await session.commit()
            return result.rowcount or 0

    def delete(self, **filters: Any) -> int:
        """
        Удалить записи, удовлетворяющие фильтрам.

        Returns:
            Количество удалённых записей.
        """
        with self._make_sync_session() as session:
            stmt = delete(self.sql_model_class).where(
                *[getattr(self.sql_model_class, k) == v for k, v in filters.items()]
            )
            result = session.execute(stmt)
            session.commit()
            return result.rowcount or 0

    async def adelete(self, **filters: Any) -> int:
        """Асинхронное удаление по фильтрам."""
        await self._acreate_table_if_not_exists()
        async with self._make_async_session() as session:
            stmt = delete(self.sql_model_class).where(
                *[getattr(self.sql_model_class, k) == v for k, v in filters.items()]
            )
            result = await session.execute(stmt)
            await session.commit()
            return result.rowcount or 0

    def add_item(self, item: Any) -> None:
        """Сохранить один объект (sync)."""
        with self._make_sync_session() as session:
            session.add(item)
            session.commit()

    async def aadd_item(self, item: Any) -> None:
        """Сохранить один объект (async)."""
        await self._acreate_table_if_not_exists()
        async with self._make_async_session() as session:
            session.add(item)
            await session.commit()

    @contextlib.contextmanager
    def _make_sync_session(self) -> Generator[SQLSession, None, None]:
        if self.async_mode:
            raise ValueError(
                "sync session requested, but async_mode=True, use async methods"
            )
        with self.session_maker() as session:
            yield cast(SQLSession, session)

    @contextlib.asynccontextmanager
    async def _make_async_session(self) -> AsyncGenerator[AsyncSession, None]:
        if not self.async_mode:
            raise ValueError(
                "async session requested, but async_mode=False, use sync methods"
            )
        async with self.session_maker() as session:
            yield cast(AsyncSession, session)

    def export_to_csv(
            self,
            items: List[Type[T]],
            filepath: str,
            fields: Optional[List[str]] = None,
            delimiter: str = ",",
            encoding: str = "utf-8"
    ) -> None:
        """
        Экспортирует список ORM объектов или схем в CSV файл.

        Args:
            items: Список объектов self.sql_model_class (или Pydantic моделей)
            filepath: Путь к файлу для записи
            fields: Список полей для экспорта. Если None — берутся все колонки модели.
            delimiter: Разделитель. По умолчанию запятая.
            encoding: Кодировка файла.
        """
        if not items:
            raise ValueError("Нет данных для экспорта")

        if fields is None:
            # Получаем все столбцы SQLAlchemy модели (если это ORM объекты)
            if hasattr(self.sql_model_class, "__table__"):
                fields = [col.name for col in self.sql_model_class.__table__.columns]
            else:
                # Если объекты не ORM, пробуем взять атрибуты первого объекта
                fields = list(vars(items[0]).keys())

        with open(filepath, mode="w", encoding=encoding, newline="") as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=fields, delimiter=delimiter)
            writer.writeheader()
            for item in items:
                row = {}
                for f in fields:
                    val = getattr(item, f, None)
                    # Если поле — datetime/date, конвертируем в строку
                    if hasattr(val, "isoformat"):
                        val = val.isoformat()
                    row[f] = val
                writer.writerow(row)

        return filepath

    def export_to_excel(
            self,
            items: List[Type[T]],
            filepath: str,
            fields: Optional[List[str]] = None,
    ) -> str:
        """
        Экспортирует список ORM объектов или схем в Excel файл (xlsx).

        Args:
            items: Список объектов self.sql_model_class или Pydantic моделей
            filepath: Путь к результату, например 'export.xlsx'
            fields: Поля для экспорта. Если None — все столбцы модели.
        Return:
            str - filepath
        """
        if not items:
            raise ValueError("Нет данных для экспорта")

        if fields is None:
            if hasattr(self.sql_model_class, "__table__"):
                fields = [col.name for col in self.sql_model_class.__table__.columns]
            else:
                fields = list(vars(items[0]).keys())

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Export"

        # Запишем заголовки
        for col_idx, field in enumerate(fields, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = field

        # Запишем данные
        for row_idx, item in enumerate(items, start=2):
            for col_idx, field in enumerate(fields, start=1):
                val = getattr(item, field, None)
                if hasattr(val, "isoformat"):
                    val = val.isoformat()
                ws.cell(row=row_idx, column=col_idx, value=val)

        # Можно немного подстроить ширину столбцов
        for idx, col in enumerate(fields, start=1):
            max_length = max(
                len(str(getattr(item, col, "")) or "") for item in items
            )
            max_length = max(max_length, len(col))
            ws.column_dimensions[get_column_letter(idx)].width = max_length + 2

        wb.save(filepath)
        return filepath