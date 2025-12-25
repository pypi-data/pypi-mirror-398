import json
import os
from openpyxl import Workbook

def save_table_json_to_xlsx(table_json, output_dir="."):
    """
    将数据库表定义的 JSON 数据写入 .xlsx 文件
    :param table_json: 字典，符合数据库表定义规范的 JSON 数据
    :param output_dir: 输出目录
    """
    # 提取表名作为文件名
    table_name = table_json.get("name")
    if not table_name:
        raise ValueError("JSON 中缺少 'name' 字段")

    filename = f"{table_name}.xlsx"
    filepath = os.path.join(output_dir, filename)

    # 创建工作簿
    wb = Workbook()

    # --- summary sheet ---
    ws_summary = wb.active
    ws_summary.title = "summary"
    summary_headers = ["name", "title", "primary", "catelog"]
    ws_summary.append(summary_headers)
    ws_summary.append([
        table_json.get("name", ""),
        table_json.get("title", ""),
        table_json.get("primary", ""),
        table_json.get("catelog", "")
    ])

    # --- fields sheet ---
    ws_fields = wb.create_sheet("fields")
    field_headers = ["name", "title", "type", "length", "dec", "nullable", "default", "comments"]
    ws_fields.append(field_headers)
    for field in table_json.get("fields", []):
        row = [
            field.get("name", ""),
            field.get("title", ""),
            field.get("type", ""),
            field.get("length", ""),
            field.get("dec", ""),
            field.get("nullable", ""),
            field.get("default", ""),
            field.get("comments", "")
        ]
        ws_fields.append(row)

    # --- indexes sheet ---
    ws_indexes = wb.create_sheet("indexes")
    index_headers = ["name", "idxtype", "idxfields"]
    ws_indexes.append(index_headers)
    for idx in table_json.get("indexes", []):
        idxfields = idx.get("idxfields")
        # 如果 idxfields 是列表，转成字符串形式 f1,f2；否则保持原样
        idxfields_str = ",".join(idxfields) if isinstance(idxfields, list) else str(idxfields)
        ws_indexes.append([
            idx.get("name", ""),
            idx.get("idxtype", ""),
            idxfields_str
        ])

    # --- codes sheet ---
    ws_codes = wb.create_sheet("codes")
    code_headers = ["field", "table", "valuefield", "textfield", "cond"]
    ws_codes.append(code_headers)
    for code in table_json.get("codes", []):
        ws_codes.append([
            code.get("field", ""),
            code.get("table", ""),
            code.get("valuefield", ""),
            code.get("textfield", ""),
            code.get("cond", "")
        ])

    # 保存文件
    wb.save(filepath)
    print(f"✅ 已生成文件: {filepath}")


# ======================
# 使用示例
# ======================
def main():
	import os, sys
	import codecs
	import json
	if len(sys.argv) < 2:
		print(f'{sys.argv[0]} dbtable_json_file ...')
		sys.exit(1)
	for a in sys.argv[1:]:
		with open(a, 'r', encoding='utf-8') as file:
			content = file.read()
			data = json.loads(content)
			if not isinstance(data, list):
				data = [data]
			for d in data:
				save_table_json_to_xlsx(d)

if __name__ == "__main__":
	main()
	"""
    # 示例 JSON 数据
    sample_json = {
        "name": "user_info",
        "title": "用户信息表",
        "primary": "id",
        "catelog": "entity",
        "fields": [
            {
                "name": "id",
                "title": "用户ID",
                "type": "str",
                "length": 32,
                "nullable": "no",
                "comments": "主键，32位字符串"
            },
            {
                "name": "age",
                "title": "年龄",
                "type": "short",
                "nullable": "yes",
                "default": 0,
                "comments": "用户年龄"
            },
            {
                "name": "salary",
                "title": "月薪",
                "type": "decimal",
                "length": 10,
                "dec": 2,
                "nullable": "yes",
                "default": "0.00",
                "comments": "精确到分"
            }
        ],
        "indexes": [
            {
                "name": "idx_user_age",
                "idxtype": "index",
                "idxfields": ["age"]
            },
            {
                "name": "idx_user_salary",
                "idxtype": "unique",
                "idxfields": ["id", "salary"]
            }
        ],
        "codes": [
            {
                "field": "dept_id",
                "table": "department",
                "valuefield": "id",
                "textfield": "name",
                "cond": "status=1"
            }
        ]
    }

    # 调用函数生成 xlsx
    save_table_json_to_xlsx(sample_json)
	"""
