"""
TLF 目录占位模板与填充
---------------------------------

目标：
- 以现有 `/src/triclick_doc_toolset/toolset/TLF table of content.xlsx` 为样板，生成占位符模板；
- 提供将数据写入模板并输出最终 Excel 的简洁 API。

设计要点：
- 保留原表格样式与筛选、列宽；
- 仅将数据区域替换为占位符（单行），便于复制样式批量填充；
- 数据写入支持按列名匹配（不强依赖列顺序）。

占位符命名：根据表头规范化为 snake_case，例如：
Type -> type；Number -> number；Unique/Repeat -> unique_repeat；Flag 1 -> flag_1。

使用示例：
>>> from triclick_doc_toolset.toolset.gen_tlf_header_excel import write_tlf_toc_file
>>> rows = [
...   {"type":"Table","number":"X.1","unique_repeat":"Unique","title":"Analysis Populations","population":"ITT Population"},
... ]
>>> out = write_tlf_toc_file(rows=rows, output_path="TLF_TOC_filled.xlsx")

"""

from __future__ import annotations

from pathlib import Path
from typing import List, Mapping, Optional, Sequence, Union
from copy import copy
from typing import cast
from importlib import resources as importlib_resources

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import MergedCell
from openpyxl.cell.rich_text import CellRichText
from decimal import Decimal
from datetime import datetime, date, time
import io


# openpyxl 的单元格可写入类型集合
ExcelValue = Union[str, int, float, Decimal, bool, datetime, date, time, CellRichText]
ExcelSetValue = Optional[ExcelValue]


def _normalize_header_name(name: str) -> str:
    s = (name or "").strip().lower()
    # 将非字母数字转为下划线，并压缩重复下划线
    import re
    s = re.sub(r"[^a-z0-9]+", "_", s)
    s = re.sub(r"_+", "_", s).strip("_")
    return s


def _find_header_row(ws: Worksheet, max_scan_rows: int = 10) -> Optional[int]:
    """在前若干行中寻找包含关键列名的表头行。"""
    targets = {"type", "number", "title"}
    for r in range(1, max_scan_rows + 1):
        values = [str(c.value).strip() if c.value is not None else "" for c in ws[r]]
        norm = {_normalize_header_name(v) for v in values if v}
        if targets.issubset(norm):
            return r
    return None


def _header_key_order(ws: Worksheet, header_row: int) -> List[str]:
    """按当前列顺序返回规范化后的列键列表。"""
    order: List[str] = []
    for c in ws[header_row]:
        key = _normalize_header_name(str(c.value or ""))
        if key:
            order.append(key)
        else:
            order.append("")
    return order


def _placeholder_for(key: str) -> str:
    return f"{{{{{key}}}}}"

def _copy_cell_style(src_cell, dst_cell):
    dst_cell.font = copy(src_cell.font)
    dst_cell.fill = copy(src_cell.fill)
    dst_cell.border = copy(src_cell.border)
    dst_cell.alignment = copy(src_cell.alignment)
    dst_cell.number_format = src_cell.number_format
    dst_cell.protection = copy(src_cell.protection)


def _collect_placeholder_row(ws: Worksheet, header_row: int) -> Optional[int]:
    """返回占位符行索引（紧随表头且含 {{...}}），不存在则返回 None。"""
    r = header_row + 1
    vals = [str(c.value) for c in ws[r] if c.value is not None]
    if any("{{" in v and "}}" in v for v in vals):
        return r
    return None


def _next_writable_row(ws: Worksheet, start_row: int) -> int:
    """在存在合并单元的情况下，返回从 start_row 开始的首个可写行。"""
    r = start_row
    ranges = list(ws.merged_cells.ranges)
    while any(rr.min_row <= r <= rr.max_row for rr in ranges):
        max_cover = max((rr.max_row for rr in ranges if rr.min_row <= r <= rr.max_row), default=r)
        r = max_cover + 1
    return r


def _select_sheet(wb, sheet_name: Optional[str]) -> Worksheet:
    """选择 TOC 工作表或首个工作表，保证类型为 Worksheet。"""
    if sheet_name and sheet_name in wb.sheetnames:
        return cast(Worksheet, wb[sheet_name])
    return cast(Worksheet, wb["TOC"] if "TOC" in wb.sheetnames else wb.worksheets[0])


def _fill_toc_sheet(ws: Worksheet, rows: Sequence[Mapping[str, ExcelSetValue]]) -> None:
    """核心填充逻辑：发现表头、确保占位符行、复制样式并写入数据。"""
    header_row = _find_header_row(ws)
    if not header_row:
        raise ValueError("模板中未找到表头行")
    keys = _header_key_order(ws, header_row)

    ph_row = _collect_placeholder_row(ws, header_row)
    if not ph_row:
        ph_row = _next_writable_row(ws, header_row + 1)
        ws.insert_rows(ph_row)
        for col_idx, key in enumerate(keys, start=1):
            cell = ws.cell(row=ph_row, column=col_idx)
            if isinstance(cell, MergedCell):
                continue
            cell.value = _placeholder_for(key) if key else None

    current = ph_row
    for i, row in enumerate(rows):
        if i > 0:
            current += 1
            ws.insert_rows(current)
            for col_idx in range(1, len(keys) + 1):
                src = ws.cell(row=ph_row, column=col_idx)
                dst = ws.cell(row=current, column=col_idx)
                if isinstance(dst, MergedCell):
                    continue
                _copy_cell_style(src, dst)
        for col_idx, key in enumerate(keys, start=1):
            dst = ws.cell(row=current, column=col_idx)
            if isinstance(dst, MergedCell):
                continue
            value: ExcelSetValue = row.get(key)
            dst.value = value if value is not None else None

    if not rows:
        ws.delete_rows(ph_row)


def _open_template_workbook(template_path: Optional[str]):
    """根据优先级加载模板 Workbook：
    1) 显式传入的 `template_path`
    2) 打包资源 `triclick_doc_toolset/resources/templates/TLF_TOC_heading_template.xlsx`（优先）
    3) 开发模式回退：仓库根目录 `resources/templates/TLF_TOC_heading_template.xlsx`
    """
    if template_path:
        return load_workbook(Path(template_path))

    # 2) 打包资源路径（wheel 内部）
    try:
        resource = importlib_resources.files("triclick_doc_toolset").joinpath(
            "resources/templates/TLF_TOC_heading_template.xlsx"
        )
        # 将资源解析为临时文件路径（适配 zip/目录两种形式）
        with importlib_resources.as_file(resource) as fp:
            return load_workbook(fp)
    except Exception:
        pass

    # 3) 开发模式回退：仓库根目录 resources/templates
    repo_root = Path(__file__).resolve().parents[3]
    dev_tpl = repo_root / "resources" / "templates" / "TLF_TOC_heading_template.xlsx"
    return load_workbook(dev_tpl)


def write_tlf_toc_bytes(
    *,
    rows: Sequence[Mapping[str, ExcelSetValue]],
    sheet_name: Optional[str] = None,
) -> bytes:
    """
    与 `write_tlf_toc_file` 行为一致，但返回 Excel 文件字节流（bytes）。

    适用于 HTTP 下载接口：将返回的 bytes 直接作为响应体，并设置
    `Content-Type: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet`。

    参数：
    - rows: 需要写入的行数据（列名为规范化后的键）。
    - sheet_name: 指定工作表名，默认自动选择 TOC 或首个工作表。
    """
    wb = _open_template_workbook(None)
    ws = _select_sheet(wb, sheet_name)
    _fill_toc_sheet(ws, rows)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def write_tlf_toc_file(
    *,
    rows: Sequence[Mapping[str, ExcelSetValue]],
    output_path: str,
    sheet_name: Optional[str] = None,
) -> str:
    """
    将 `rows` 按模板的列映射写入并输出最终 Excel。

    - rows: 每行数据为 dict，key 使用规范化后的列名，如：
      type/number/unique_repeat/title/population/flag_1/flag_2/flag_3。
    - output_path: 必填，输出文件路径（绝对或相对）。
    - 返回输出文件的绝对路径。
    """
    wb = _open_template_workbook(None)
    ws = _select_sheet(wb, sheet_name)
    _fill_toc_sheet(ws, rows)

    out = Path(output_path)
    wb.save(out)
    return str(out)