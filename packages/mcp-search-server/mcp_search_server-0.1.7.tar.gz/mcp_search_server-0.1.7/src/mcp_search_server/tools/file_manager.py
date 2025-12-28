"""
File management tool for reading and writing text files.
Supports reading/writing text files, PDFs, Word documents, Excel files, and images.
"""

import os
import asyncio
from pathlib import Path
from typing import Dict, Any
import base64

# Document processing imports
try:
    from pypdf import PdfReader

    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

try:
    from docx import Document

    DOCX_AVAILABLE = True
except ImportError:
    DOCX_AVAILABLE = False

try:
    import openpyxl
    import xlrd

    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    from PIL import Image
    import io

    IMAGE_AVAILABLE = True
except ImportError:
    IMAGE_AVAILABLE = False


class FileManager:
    """File management tool for basic file operations with MCP support."""

    def __init__(self):
        """Initialize file manager."""
        self.max_file_size = 10 * 1024 * 1024  # 10MB

        # Setup default files directory
        project_root = Path(__file__).parent.parent.parent.parent
        self.default_files_dir = project_root / "data" / "files"
        self.default_files_dir.mkdir(parents=True, exist_ok=True)

    def _normalize_path(self, file_path: str) -> Path:
        """
        Normalize file path - use default directory if relative path is given.

        Args:
            file_path: Original file path

        Returns:
            Normalized Path object
        """
        # Empty string means default files directory
        if not file_path or file_path.strip() == "":
            return self.default_files_dir

        path = Path(file_path)

        # Forbid absolute paths; treat them as filenames under default dir
        if path.is_absolute():
            return self.default_files_dir / path.name

        # Avoid double prefix if user already starts with data/files
        parts = path.parts
        if len(parts) >= 2 and parts[0] == "data" and parts[1] == "files":
            project_root = Path(__file__).parent.parent.parent.parent
            return project_root / path

        # Treat all other paths as relative to default_files_dir
        return self.default_files_dir / path

    def _validate_path(self, file_path: str) -> bool:
        """
        Validate file path for security.

        Args:
            file_path: File path to validate

        Returns:
            True if path is safe
        """
        # Convert to absolute path
        abs_path = os.path.abspath(file_path)

        # Check for dangerous patterns
        dangerous_patterns = [
            "..",  # Parent directory access
            "/etc",  # System configuration
            "/root",  # Root directory
            "/usr/bin",  # System binaries
            "/sys",  # System directory
            "/proc",  # Process directory
        ]

        for pattern in dangerous_patterns:
            if pattern in abs_path.lower():
                return False
        # Enforce that writes happen under default_files_dir
        try:
            base = str(self.default_files_dir.resolve())
            return abs_path.startswith(base)
        except Exception:
            return False
        return True

    def _read_pdf(self, path: Path) -> str:
        """Read PDF file content."""
        if not PDF_AVAILABLE:
            return "Error: The pypdf library is not installed"

        try:
            reader = PdfReader(str(path))
            text = ""
            for page_num, page in enumerate(reader.pages, 1):
                text += f"\n--- Page {page_num} ---\n"
                text += page.extract_text()
            return text
        except Exception as e:
            return f"Error reading PDF: {str(e)}"

    def _read_docx(self, path: Path) -> str:
        """Read Word document content."""
        if not DOCX_AVAILABLE:
            return "Error: The python-docx library is not installed"

        try:
            doc = Document(str(path))
            text = ""
            for paragraph in doc.paragraphs:
                text += paragraph.text + "\n"
            return text
        except Exception as e:
            return f"Error reading Word document: {str(e)}"

    def _read_excel(self, path: Path) -> str:
        """Read Excel file content."""
        if not EXCEL_AVAILABLE:
            return "Error: The openpyxl/xlrd libraries are not installed"

        try:
            if path.suffix.lower() == ".xlsx":
                wb = openpyxl.load_workbook(str(path), data_only=True)
                text = ""
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    text += f"\n--- Sheet '{sheet_name}' ---\n"
                    for row in ws.iter_rows(values_only=True):
                        row_text = "\t".join(
                            [str(cell) if cell is not None else "" for cell in row]
                        )
                        if row_text.strip():
                            text += row_text + "\n"
                return text
            else:
                # For .xls files
                wb = xlrd.open_workbook(str(path))
                text = ""
                for sheet in wb.sheets():
                    text += f"\n--- Sheet '{sheet.name}' ---\n"
                    for row_idx in range(sheet.nrows):
                        row_values = sheet.row_values(row_idx)
                        row_text = "\t".join([str(cell) for cell in row_values])
                        if row_text.strip():
                            text += row_text + "\n"
                return text
        except Exception as e:
            return f"Error reading Excel file: {str(e)}"

    def _read_image(self, path: Path) -> str:
        """Read and analyze image file."""
        if not IMAGE_AVAILABLE:
            return "Error: The Pillow library is not installed"

        try:
            with Image.open(str(path)) as img:
                # Get basic image info
                info = {
                    "format": img.format,
                    "mode": img.mode,
                    "size": img.size,
                    "width": img.width,
                    "height": img.height,
                }

                # Convert image to base64 for display (optional, for small images)
                if img.width * img.height < 1000000:  # Only for images < 1MP
                    buffer = io.BytesIO()
                    img.save(buffer, format=img.format or "PNG")
                    img_data = buffer.getvalue()
                    b64_data = base64.b64encode(img_data).decode("utf-8")

                    return (
                        f"Image '{path}':\n"
                        + f"Format: {info['format']}\n"
                        + f"Size: {info['width']}x{info['height']} pixels\n"
                        + f"Mode: {info['mode']}\n"
                        + f"Base64 data: {b64_data[:100]}..."
                        + f" (total {len(b64_data)} characters)"
                    )
                else:
                    return (
                        f"Image '{path}':\n"
                        + f"Format: {info['format']}\n"
                        + f"Size: {info['width']}x{info['height']} pixels\n"
                        + f"Mode: {info['mode']}\n"
                        + "(Image too large to display in base64)"
                    )

        except Exception as e:
            return f"Error reading image: {str(e)}"

    def _read_file(self, file_path: str) -> str:
        """
        Read content from file.

        Args:
            file_path: Path to file

        Returns:
            File content or error message
        """
        try:
            # Normalize path to use default directory if needed
            path = self._normalize_path(file_path)

            if not self._validate_path(str(path)):
                return "Error: Unsafe file path"

            if not path.exists():
                return f"Error: File not found: {file_path}"

            if not path.is_file():
                return f"Error: The specified path is not a file: {file_path}"

            # Check file size
            if path.stat().st_size > self.max_file_size:
                return f"Error: File is too large (max {self.max_file_size // 1024 // 1024}MB)"

            # Determine file type and read accordingly
            suffix = path.suffix.lower()

            if suffix == ".pdf":
                content = self._read_pdf(path)
            elif suffix == ".docx":
                content = self._read_docx(path)
            elif suffix in [".xlsx", ".xls"]:
                content = self._read_excel(path)
            elif suffix in [".jpg", ".jpeg", ".png", ".gif", ".bmp", ".webp", ".tiff"]:
                content = self._read_image(path)
            else:
                # Try to read as text file
                try:
                    with open(path, "r", encoding="utf-8") as f:
                        content = f.read()
                except UnicodeDecodeError:
                    return f"Error: The file is not text or has an unsupported encoding: {path}"

            return content

        except PermissionError:
            return f"Error: No permission to access file: {file_path}"
        except Exception as e:
            return f"Error reading file: {str(e)}"

    def _write_file(self, file_path: str, content: str, mode: str = "w") -> str:
        """
        Write content to file.

        Args:
            file_path: Path to file
            content: Content to write
            mode: Write mode ('w' for write, 'a' for append)

        Returns:
            Success message or error
        """
        try:
            # Normalize path to use default directory if needed
            path = self._normalize_path(file_path)

            if not self._validate_path(str(path)):
                return "Error: Unsafe file path"

            # Create directory if it doesn't exist
            path.parent.mkdir(parents=True, exist_ok=True)

            # Check content size
            content_size = len(content.encode("utf-8"))
            if content_size > self.max_file_size:
                return f"Error: Content too large (max {self.max_file_size // 1024 // 1024}MB)"

            with open(path, mode, encoding="utf-8") as f:
                f.write(content)

            action = "written to" if mode == "w" else "appended to"
            return f"Content successfully {action} file: {path}"

        except PermissionError:
            return f"Error: No write permission for file: {path}"
        except Exception as e:
            return f"Error writing file: {str(e)}"

    def _list_directory(self, dir_path: str) -> str:
        """
        List directory contents.

        Args:
            dir_path: Path to directory

        Returns:
            Directory listing or error message
        """
        try:
            # Normalize path to use default directory if needed
            path = self._normalize_path(dir_path)

            if not self._validate_path(str(path)):
                return "Error: Unsafe directory path"

            if not path.exists():
                return f"Error: Directory not found: {path}"

            if not path.is_dir():
                return f"Error: The specified path is not a directory: {path}"

            items = []
            for item in sorted(path.iterdir()):
                item_type = "directory" if item.is_dir() else "file"
                size = ""
                if item.is_file():
                    size_bytes = item.stat().st_size
                    if size_bytes < 1024:
                        size = f" ({size_bytes} bytes)"
                    elif size_bytes < 1024 * 1024:
                        size = f" ({size_bytes // 1024} KB)"
                    else:
                        size = f" ({size_bytes // 1024 // 1024} MB)"

                items.append(f"  {item.name} ({item_type}){size}")

            if not items:
                return f"Directory is empty: {path}"

            return f"Contents of directory '{path}':\n\n" + "\n".join(items)

        except PermissionError:
            return f"Error: No permission to access directory: {path}"
        except Exception as e:
            return f"Error reading directory: {str(e)}"

    async def read_file(self, file_path: str) -> Dict[str, Any]:
        """
        Async wrapper for reading files.

        Args:
            file_path: Path to file to read

        Returns:
            Dictionary with result and metadata
        """
        content = await asyncio.to_thread(self._read_file, file_path)
        path = self._normalize_path(file_path)

        return {
            "path": str(path),
            "content": content,
            "size": path.stat().st_size if path.exists() else 0,
            "exists": path.exists(),
        }

    async def write_file(
        self, file_path: str, content: str, append: bool = False
    ) -> Dict[str, Any]:
        """
        Async wrapper for writing files.

        Args:
            file_path: Path to file to write
            content: Content to write
            append: If True, append to file; if False, overwrite

        Returns:
            Dictionary with result and metadata
        """
        mode = "a" if append else "w"
        result = await asyncio.to_thread(self._write_file, file_path, content, mode)
        path = self._normalize_path(file_path)

        return {
            "path": str(path),
            "message": result,
            "size": path.stat().st_size if path.exists() else 0,
            "exists": path.exists(),
        }

    async def list_directory(self, dir_path: str = "") -> Dict[str, Any]:
        """
        Async wrapper for listing directories.

        Args:
            dir_path: Path to directory (empty for default files directory)

        Returns:
            Dictionary with directory listing
        """
        path = self._normalize_path(dir_path)
        content = await asyncio.to_thread(self._list_directory, dir_path)

        # Parse directory contents
        items = []
        if path.exists() and path.is_dir():
            for item in sorted(path.iterdir()):
                item_info = {
                    "name": item.name,
                    "type": "directory" if item.is_dir() else "file",
                    "size": item.stat().st_size if item.is_file() else 0,
                }
                items.append(item_info)

        return {"path": str(path), "items": items, "count": len(items), "listing": content}

    async def delete_file(self, file_path: str) -> Dict[str, Any]:
        """
        Delete a file.

        Args:
            file_path: Path to file to delete

        Returns:
            Dictionary with result
        """
        try:
            path = self._normalize_path(file_path)

            if not self._validate_path(str(path)):
                return {"success": False, "message": "Error: Unsafe file path", "path": str(path)}

            if not path.exists():
                return {
                    "success": False,
                    "message": f"Error: File not found: {file_path}",
                    "path": str(path),
                }

            if not path.is_file():
                return {
                    "success": False,
                    "message": f"Error: Not a file: {file_path}",
                    "path": str(path),
                }

            await asyncio.to_thread(path.unlink)

            return {
                "success": True,
                "message": f"File deleted successfully: {path}",
                "path": str(path),
            }
        except PermissionError:
            return {
                "success": False,
                "message": f"Error: No permission to delete file: {file_path}",
                "path": str(path),
            }
        except Exception as e:
            return {
                "success": False,
                "message": f"Error deleting file: {str(e)}",
                "path": str(path),
            }

    def manage_file(self, command: str) -> str:
        """
        Execute file management command.

        Args:
            command: Command string in format 'action:path:content'

        Returns:
            Result of operation
        """
        try:
            if not command or not isinstance(command, str):
                return "Error: Empty command"

            parts = command.split(":", 2)  # Split into max 3 parts

            if len(parts) < 2:
                return (
                    "Command format error. Use:\n"
                    "- read:/path/to/file\n"
                    "- write:/path/to/file:content\n"
                    "- append:/path/to/file:content\n"
                    "- list:/path/to/directory"
                )

            action = parts[0].lower().strip()
            path = parts[1].strip()
            content = parts[2] if len(parts) > 2 else ""

            if action == "read":
                return self._read_file(path)
            elif action == "write":
                if not content:
                    return "Error: No content specified for writing"
                return self._write_file(path, content, "w")
            elif action == "append":
                if not content:
                    return "Error: No content specified for appending"
                return self._write_file(path, content, "a")
            elif action == "list":
                return self._list_directory(path)
            else:
                return f"Error: Unknown command '{action}'. Available commands: read, write, append, list"

        except Exception as e:
            return f"Error executing command: {str(e)}"


# Create a global instance
file_manager = FileManager()
