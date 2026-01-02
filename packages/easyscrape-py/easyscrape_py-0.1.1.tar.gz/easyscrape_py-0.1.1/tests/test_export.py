"""Comprehensive tests for easyscrape.export module."""

import pytest
import json
import os
from easyscrape.export import (
    to_json, to_csv, to_jsonl, to_excel, to_parquet, to_sqlite,
    Exporter, ExportError, _validate_export_path,
)


# =============================================================================
# Path Validation Tests
# =============================================================================

class TestValidateExportPath:
    """Tests for path validation."""

    def test_valid_path(self):
        path = _validate_export_path("output.json")
        assert path is not None

    def test_nested_path(self):
        path = _validate_export_path("subdir/output.json")
        assert path is not None

    def test_creates_parent_dirs(self):
        path = _validate_export_path("test_export_dir/nested/output.json")
        assert path.parent.exists()
        # Cleanup
        import shutil
        shutil.rmtree("test_export_dir", ignore_errors=True)

    def test_path_traversal_blocked(self):
        with pytest.raises(ExportError, match="Path traversal"):
            _validate_export_path("../../../etc/passwd")

    def test_path_traversal_hidden(self):
        with pytest.raises(ExportError, match="Path traversal"):
            _validate_export_path("subdir/../../outside.txt")

    def test_extension_validation(self):
        path = _validate_export_path("file.json", allowed_extensions=[".json"])
        assert str(path).endswith(".json")

    def test_extension_not_allowed(self):
        with pytest.raises(ExportError, match="extension"):
            _validate_export_path("file.xyz", allowed_extensions=[".json", ".csv"])


# =============================================================================
# ExportError Tests
# =============================================================================

class TestExportError:
    """Tests for ExportError."""

    def test_is_exception(self):
        assert issubclass(ExportError, Exception)

    def test_can_raise(self):
        with pytest.raises(ExportError):
            raise ExportError("test error")

    def test_with_path(self):
        err = ExportError("Failed", path="/some/path.json")
        assert err.path == "/some/path.json"

    def test_with_format(self):
        err = ExportError("Failed", format="json")
        assert err.format == "json"

    def test_with_all_attrs(self):
        err = ExportError("Failed", path="/path.csv", format="csv")
        assert err.path == "/path.csv"
        assert err.format == "csv"


# =============================================================================
# JSON Export Tests
# =============================================================================

class TestToJson:
    """Tests for JSON export."""

    def test_dict_to_json(self):
        data = {"name": "test", "value": 123}
        path = "test_out.json"
        to_json(data, path)
        assert os.path.exists(path)
        with open(path) as f:
            parsed = json.load(f)
        assert parsed["name"] == "test"
        os.remove(path)

    def test_list_to_json(self):
        data = [{"a": 1}, {"a": 2}]
        path = "test_list.json"
        to_json(data, path)
        assert os.path.exists(path)
        os.remove(path)

    def test_pretty_print(self):
        data = {"name": "test"}
        path = "test_pretty.json"
        to_json(data, path, pretty=True)
        assert os.path.exists(path)
        with open(path) as f:
            content = f.read()
        assert "\n" in content
        os.remove(path)

    def test_compact_print(self):
        data = {"name": "test"}
        path = "test_compact.json"
        to_json(data, path, pretty=False)
        assert os.path.exists(path)
        with open(path) as f:
            content = f.read()
        assert "\n" not in content.strip()
        os.remove(path)

    def test_unicode(self):
        data = {"text": "Hello \u4e16\u754c"}
        path = "test_unicode.json"
        to_json(data, path)
        with open(path, encoding="utf-8") as f:
            parsed = json.load(f)
        assert parsed["text"] == "Hello \u4e16\u754c"
        os.remove(path)

    def test_nested_data(self):
        data = {"user": {"name": "Alice", "roles": ["admin", "user"]}}
        path = "test_nested.json"
        to_json(data, path)
        with open(path) as f:
            parsed = json.load(f)
        assert parsed["user"]["roles"][0] == "admin"
        os.remove(path)


# =============================================================================
# CSV Export Tests
# =============================================================================

class TestToCsv:
    """Tests for CSV export."""

    def test_list_of_dicts(self):
        data = [{"name": "Alice", "age": 30}]
        path = "test_csv.csv"
        to_csv(data, path)
        assert os.path.exists(path)
        with open(path) as f:
            content = f.read()
        assert "name" in content
        assert "Alice" in content
        os.remove(path)

    def test_custom_fieldnames(self):
        data = [{"name": "Bob", "age": 25}]
        path = "test_fields.csv"
        to_csv(data, path, fieldnames=["age", "name"])
        with open(path) as f:
            first_line = f.readline()
        assert first_line.startswith("age")
        os.remove(path)

    def test_multiple_rows(self):
        data = [
            {"name": "Alice", "age": 30},
            {"name": "Bob", "age": 25},
        ]
        path = "test_multi.csv"
        to_csv(data, path)
        with open(path) as f:
            lines = f.readlines()
        assert len(lines) == 3  # Header + 2 rows
        os.remove(path)

    def test_empty_list(self):
        data = []
        path = "test_empty.csv"
        to_csv(data, path)
        assert os.path.exists(path)
        os.remove(path)

    def test_unicode_csv(self):
        data = [{"name": "Alice \u4e16\u754c", "value": 1}]
        path = "test_unicode.csv"
        to_csv(data, path)
        with open(path, encoding="utf-8") as f:
            content = f.read()
        assert "\u4e16\u754c" in content
        os.remove(path)


# =============================================================================
# JSONL Export Tests
# =============================================================================

class TestToJsonl:
    """Tests for JSONL export."""

    def test_list_to_jsonl(self):
        data = [{"a": 1}, {"a": 2}]
        path = "test_data.jsonl"
        to_jsonl(data, path)
        assert os.path.exists(path)
        with open(path) as f:
            lines = f.readlines()
        assert len(lines) == 2
        os.remove(path)

    def test_each_line_valid_json(self):
        data = [{"x": 1}, {"y": 2}]
        path = "test_jsonl.jsonl"
        to_jsonl(data, path)
        with open(path) as f:
            for line in f:
                parsed = json.loads(line)
                assert isinstance(parsed, dict)
        os.remove(path)

    def test_empty_list(self):
        data = []
        path = "test_empty.jsonl"
        to_jsonl(data, path)
        assert os.path.exists(path)
        with open(path) as f:
            content = f.read()
        assert content == ""
        os.remove(path)

    def test_unicode_jsonl(self):
        data = [{"text": "Hello \u4e16\u754c"}]
        path = "test_unicode.jsonl"
        to_jsonl(data, path)
        with open(path, encoding="utf-8") as f:
            line = f.readline()
        parsed = json.loads(line)
        assert parsed["text"] == "Hello \u4e16\u754c"
        os.remove(path)


# =============================================================================
# Excel Export Tests
# =============================================================================

class TestToExcel:
    """Tests for Excel export."""

    def test_basic_export(self):
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed")
        
        data = [{"name": "Alice", "age": 30}]
        path = "test_excel.xlsx"
        to_excel(data, path)
        assert os.path.exists(path)
        os.remove(path)

    def test_custom_sheet_name(self):
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed")
        
        data = [{"x": 1}]
        path = "test_sheet.xlsx"
        to_excel(data, path, sheet_name="MyData")
        
        wb = openpyxl.load_workbook(path)
        assert "MyData" in wb.sheetnames
        wb.close()
        os.remove(path)


# =============================================================================
# Parquet Export Tests
# =============================================================================

class TestToParquet:
    """Tests for Parquet export."""

    def test_basic_export(self):
        try:
            import pandas
            import pyarrow
        except ImportError:
            pytest.skip("pandas/pyarrow not installed")
        
        data = [{"name": "Alice", "age": 30}]
        path = "test_parquet.parquet"
        to_parquet(data, path)
        assert os.path.exists(path)
        os.remove(path)


# =============================================================================
# SQLite Export Tests
# =============================================================================

class TestToSqlite:
    """Tests for SQLite export."""

    def test_basic_export(self):
        try:
            import pandas
        except ImportError:
            pytest.skip("pandas not installed")
        
        data = [{"name": "Alice", "age": 30}]
        path = "test_sqlite.db"
        to_sqlite(data, path, table_name="users")
        assert os.path.exists(path)
        os.remove(path)

    def test_custom_table_name(self):
        try:
            import pandas
            import sqlite3
        except ImportError:
            pytest.skip("pandas not installed")
        
        data = [{"id": 1, "value": "test"}]
        path = "test_table.db"
        to_sqlite(data, path, table_name="custom_table")
        
        conn = sqlite3.connect(path)
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
        tables = [row[0] for row in cursor.fetchall()]
        conn.close()
        
        assert "custom_table" in tables
        os.remove(path)


# =============================================================================
# Exporter Class Tests
# =============================================================================

class TestExporter:
    """Tests for Exporter class."""

    def test_creation(self):
        data = [{"a": 1}]
        exporter = Exporter(data)
        assert exporter is not None

    def test_export_json(self):
        data = [{"x": 1}]
        path = "exporter_test.json"
        Exporter(data).json(path)
        assert os.path.exists(path)
        os.remove(path)

    def test_export_csv(self):
        data = [{"x": 1}]
        path = "exporter_test.csv"
        Exporter(data).csv(path)
        assert os.path.exists(path)
        os.remove(path)

    def test_export_jsonl(self):
        data = [{"x": 1}]
        path = "exporter_test.jsonl"
        Exporter(data).jsonl(path)
        assert os.path.exists(path)
        os.remove(path)

    def test_chaining(self):
        data = [{"z": 1}]
        exporter = Exporter(data)
        result = exporter.json("chain.json")
        assert result is exporter
        os.remove("chain.json")

    def test_multiple_exports(self):
        data = [{"m": 1}]
        Exporter(data).json("multi.json").csv("multi.csv")
        assert os.path.exists("multi.json")
        assert os.path.exists("multi.csv")
        os.remove("multi.json")
        os.remove("multi.csv")

    def test_export_excel(self):
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed")
        
        data = [{"x": 1}]
        path = "exporter_test.xlsx"
        Exporter(data).excel(path)
        assert os.path.exists(path)
        os.remove(path)

    def test_data_attribute(self):
        data = [{"a": 1}, {"a": 2}]
        exporter = Exporter(data)
        assert exporter._data == data

    def test_empty_data(self):
        data = []
        exporter = Exporter(data)
        path = "empty.json"
        exporter.json(path)
        assert os.path.exists(path)
        os.remove(path)




# =============================================================================
# Additional Coverage Tests
# =============================================================================

class TestExportErrorPaths:
    """Tests for error handling in export functions."""

    def test_invalid_path_oserror(self):
        # Path with null bytes causes OSError
        with pytest.raises(ExportError, match="Invalid path"):
            _validate_export_path("file\x00name.json")

    def test_json_write_oserror(self, mocker):
        mocker.patch("pathlib.Path.write_text", side_effect=OSError("disk full"))
        with pytest.raises(ExportError, match="Failed to write JSON"):
            to_json({"a": 1}, "test.json")

    def test_csv_write_oserror(self, mocker):
        mocker.patch("builtins.open", side_effect=OSError("permission denied"))
        with pytest.raises(ExportError, match="Failed to write CSV"):
            to_csv([{"a": 1}], "test.csv")

    def test_jsonl_write_oserror(self, mocker):
        mocker.patch("builtins.open", side_effect=OSError("permission denied"))
        with pytest.raises(ExportError, match="Failed to write JSONL"):
            to_jsonl([{"a": 1}], "test.jsonl")


class TestExporterExtendedMethods:
    """Tests for extended Exporter methods."""

    def test_exporter_parquet(self):
        try:
            import pandas
            import pyarrow
        except ImportError:
            pytest.skip("pandas/pyarrow not installed")
        
        data = [{"x": 1}]
        path = "exporter_parquet.parquet"
        result = Exporter(data).parquet(path)
        assert os.path.exists(path)
        assert result._data == data  # Test chaining
        os.remove(path)

    def test_exporter_sqlite(self):
        try:
            import pandas
        except ImportError:
            pytest.skip("pandas not installed")
        
        data = [{"x": 1}]
        path = "exporter_sqlite.db"
        result = Exporter(data).sqlite(path, table="test_table")
        assert os.path.exists(path)
        assert result._data == data  # Test chaining
        os.remove(path)

    def test_parquet_error_handling(self, mocker):
        try:
            import pandas
        except ImportError:
            pytest.skip("pandas not installed")
        
        mocker.patch("pandas.DataFrame.to_parquet", side_effect=ValueError("bad data"))
        with pytest.raises(ExportError, match="Failed to write Parquet"):
            to_parquet([{"a": 1}], "test.parquet")

    def test_sqlite_error_handling(self, mocker):
        try:
            import pandas
        except ImportError:
            pytest.skip("pandas not installed")
        
        mocker.patch("pandas.DataFrame.to_sql", side_effect=OSError("db error"))
        with pytest.raises(ExportError, match="Failed to write SQLite"):
            to_sqlite([{"a": 1}], "test.db", table_name="t")



class TestExcelEdgeCases:
    """Tests for Excel edge cases."""

    def test_excel_with_nested_data(self):
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed")
        
        # Test converting nested data (list/dict) to JSON strings
        data = [{"name": "test", "tags": ["a", "b"], "meta": {"x": 1}}]
        path = "test_nested_excel.xlsx"
        to_excel(data, path)
        
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        # Check that nested data was converted to JSON string
        assert ws.cell(row=2, column=2).value is not None  # tags column
        wb.close()
        os.remove(path)

    def test_excel_empty_data(self):
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed")
        
        data = []
        path = "test_empty_excel.xlsx"
        to_excel(data, path)
        assert os.path.exists(path)
        os.remove(path)

    def test_excel_write_oserror(self, mocker):
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed")
        
        mocker.patch("openpyxl.Workbook.save", side_effect=OSError("disk error"))
        with pytest.raises(ExportError, match="Failed to write Excel"):
            to_excel([{"a": 1}], "test.xlsx")

    def test_openpyxl_import_error(self, mocker):
        import builtins
        original_import = builtins.__import__
        
        def mock_import(name, *args, **kwargs):
            if name == "openpyxl":
                raise ImportError("No module named 'openpyxl'")
            return original_import(name, *args, **kwargs)
        
        mocker.patch("builtins.__import__", side_effect=mock_import)
        with pytest.raises(ImportError, match="openpyxl required"):
            to_excel([{"a": 1}], "test.xlsx")


class TestImportErrorHandling:
    """Tests for import error handling."""

    def test_pandas_import_error_parquet(self, mocker):
        import builtins
        original_import = builtins.__import__
        
        def mock_import(name, *args, **kwargs):
            if name == "pandas":
                raise ImportError("No module named 'pandas'")
            return original_import(name, *args, **kwargs)
        
        mocker.patch("builtins.__import__", side_effect=mock_import)
        with pytest.raises(ImportError, match="pandas required for Parquet"):
            to_parquet([{"a": 1}], "test.parquet")

    def test_pandas_import_error_sqlite(self, mocker):
        import builtins
        original_import = builtins.__import__
        
        def mock_import(name, *args, **kwargs):
            if name == "pandas":
                raise ImportError("No module named 'pandas'")
            return original_import(name, *args, **kwargs)
        
        mocker.patch("builtins.__import__", side_effect=mock_import)
        with pytest.raises(ImportError, match="pandas required for SQLite"):
            to_sqlite([{"a": 1}], "test.db", table_name="t")



# =============================================================================
# to_dataframe Tests
# =============================================================================

class TestToDataframe:
    """Tests for to_dataframe function."""

    def test_basic_conversion(self):
        """Test basic list of dicts to DataFrame conversion."""
        from easyscrape.export import to_dataframe
        data = [{"a": 1, "b": 2}, {"a": 3, "b": 4}]
        df = to_dataframe(data)
        assert len(df) == 2
        assert list(df.columns) == ["a", "b"]
        assert df["a"].tolist() == [1, 3]

    def test_empty_list(self):
        """Test empty list returns empty DataFrame."""
        from easyscrape.export import to_dataframe
        df = to_dataframe([])
        assert len(df) == 0

    def test_single_row(self):
        """Test single row conversion."""
        from easyscrape.export import to_dataframe
        df = to_dataframe([{"name": "test", "value": 42}])
        assert len(df) == 1
        assert df.iloc[0]["name"] == "test"
        assert df.iloc[0]["value"] == 42

    def test_mixed_types(self):
        """Test mixed data types in values."""
        from easyscrape.export import to_dataframe
        data = [{"str": "hello", "int": 1, "float": 1.5, "none": None}]
        df = to_dataframe(data)
        assert df.iloc[0]["str"] == "hello"
        assert df.iloc[0]["int"] == 1
        assert df.iloc[0]["float"] == 1.5

    def test_nested_data(self):
        """Test nested dict/list values."""
        from easyscrape.export import to_dataframe
        data = [{"nested": {"a": 1}, "list": [1, 2, 3]}]
        df = to_dataframe(data)
        assert df.iloc[0]["nested"] == {"a": 1}
        assert df.iloc[0]["list"] == [1, 2, 3]
