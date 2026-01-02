"""Data Export Module - Multi-Format Output for Scraped Data.

This module provides a unified interface for exporting scraped data to various
file formats: JSON, CSV, JSON Lines, Excel, Parquet, and SQLite.
"""
from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any, Final

from .exceptions import EasyScrapeError

__all__: Final[tuple[str, ...]] = (
    "to_json",
    "to_csv",
    "to_jsonl",
    "to_excel",
    "to_parquet",
    "to_sqlite",
    "to_dataframe",
    "Exporter",
)


class ExportError(EasyScrapeError):
    """Raised when an export operation fails."""

    def __init__(
        self,
        message: str,
        path: str | None = None,
        format: str | None = None
    ) -> None:
        super().__init__(message)
        self.path = path
        self.format = format


def _validate_export_path(
    path: str | Path,
    allowed_extensions: list[str] | None = None,
    create_parents: bool = True,
) -> Path:
    """Validate and sanitise an export file path."""
    try:
        resolved = Path(path).resolve()
    except (OSError, ValueError) as exc:
        raise ExportError(f"Invalid path: {exc}", path=str(path)) from exc

    # Security check: ensure path is within current working directory
    cwd = Path.cwd().resolve()

    try:
        resolved.relative_to(cwd)
    except ValueError:
        raise ExportError(
            f"Path traversal detected: '{path}' resolves outside working directory. "
            f"All exports must stay within: {cwd}",
            path=str(path),
        ) from None

    # Validate extension if restrictions specified
    if allowed_extensions:
        suffix = resolved.suffix.lower()
        if suffix not in [ext.lower() for ext in allowed_extensions]:
            raise ExportError(
                f"Invalid file extension '{suffix}'. Allowed: {allowed_extensions}",
                path=str(path),
            )

    # Create parent directories if requested
    if create_parents:
        resolved.parent.mkdir(parents=True, exist_ok=True)

    return resolved


def to_json(
    data: list[dict[str, Any]] | dict[str, Any],
    path: str | Path,
    pretty: bool = True,
) -> None:
    """Export data to a JSON file."""
    validated_path = _validate_export_path(path, [".json"])

    try:
        content = json.dumps(
            data,
            indent=2 if pretty else None,
            ensure_ascii=False,
            default=str,
        )
        validated_path.write_text(content, encoding="utf-8")
    except (OSError, TypeError) as exc:
        raise ExportError(f"Failed to write JSON: {exc}", path=str(path), format="json") from exc


def to_csv(
    data: list[dict[str, Any]],
    path: str | Path,
    fieldnames: list[str] | None = None,
) -> None:
    """Export data to a CSV file."""
    validated_path = _validate_export_path(path, [".csv"])

    try:
        if not data:
            validated_path.write_text("", encoding="utf-8")
            return

        fields = fieldnames or list(data[0].keys())

        with open(validated_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
            writer.writeheader()
            writer.writerows(data)
    except (OSError, csv.Error) as exc:
        raise ExportError(f"Failed to write CSV: {exc}", path=str(path), format="csv") from exc


def to_jsonl(
    data: list[dict[str, Any]],
    path: str | Path,
) -> None:
    """Export data as JSON Lines."""
    validated_path = _validate_export_path(path, [".jsonl", ".ndjson"])

    try:
        with open(validated_path, "w", encoding="utf-8") as f:
            for row in data:
                f.write(json.dumps(row, ensure_ascii=False, default=str))
                f.write("\n")
    except (OSError, TypeError) as exc:
        raise ExportError(f"Failed to write JSONL: {exc}", path=str(path), format="jsonl") from exc


def to_excel(
    data: list[dict[str, Any]],
    path: str | Path,
    sheet_name: str = "Sheet1",
) -> None:
    """Export data to an Excel workbook."""
    try:
        import openpyxl
    except ImportError as exc:
        raise ImportError(
            "openpyxl required for Excel export. Install with: pip install openpyxl"
        ) from exc

    validated_path = _validate_export_path(path, [".xlsx", ".xls"])

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        if not data:
            wb.save(validated_path)
            return

        headers = list(data[0].keys())
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        for row_idx, row_data in enumerate(data, 2):
            for col_idx, header in enumerate(headers, 1):
                value = row_data.get(header)
                if isinstance(value, (list, dict)):
                    value = json.dumps(value, ensure_ascii=False)
                ws.cell(row=row_idx, column=col_idx, value=value)

        wb.save(validated_path)
    except OSError as exc:
        raise ExportError(f"Failed to write Excel: {exc}", path=str(path), format="excel") from exc


def to_parquet(
    data: list[dict[str, Any]],
    path: str | Path,
) -> None:
    """Export data to Apache Parquet format."""
    try:
        import pandas as pd
    except ImportError as exc:
        raise ImportError(
            "pandas required for Parquet export. Install with: pip install pandas pyarrow"
        ) from exc

    validated_path = _validate_export_path(path, [".parquet"])

    try:
        df = pd.DataFrame(data)
        df.to_parquet(validated_path, index=False)
    except (OSError, ValueError) as exc:
        raise ExportError(f"Failed to write Parquet: {exc}", path=str(path), format="parquet") from exc


def to_sqlite(
    data: list[dict[str, Any]],
    path: str | Path,
    table_name: str = "scraped_data",
    if_exists: str = "replace",
) -> None:
    """Export data to a SQLite database."""
    try:
        import pandas as pd
    except ImportError as exc:
        raise ImportError(
            "pandas required for SQLite export. Install with: pip install pandas"
        ) from exc

    import sqlite3

    validated_path = _validate_export_path(path, [".db", ".sqlite", ".sqlite3"])

    conn = None
    try:
        df = pd.DataFrame(data)
        conn = sqlite3.connect(str(validated_path))
        df.to_sql(table_name, conn, if_exists=if_exists, index=False)
    except (OSError, sqlite3.Error) as exc:
        raise ExportError(f"Failed to write SQLite: {exc}", path=str(path), format="sqlite") from exc
    finally:
        if conn is not None:
            conn.close()


def to_dataframe(data: list[dict[str, Any]]) -> Any:
    """Convert scraped data to a pandas DataFrame."""
    try:
        import pandas as pd
    except ImportError as exc:
        raise ImportError("pandas required: pip install pandas") from exc
    return pd.DataFrame(data)


class Exporter:
    """Fluent interface for exporting data to multiple formats."""

    __slots__ = ("_data",)

    def __init__(self, data: list[dict[str, Any]]) -> None:
        self._data = data

    def json(self, path: str | Path, pretty: bool = True) -> Exporter:
        """Export to JSON format."""
        to_json(self._data, path, pretty)
        return self

    def csv(self, path: str | Path, fieldnames: list[str] | None = None) -> Exporter:
        """Export to CSV format."""
        to_csv(self._data, path, fieldnames)
        return self

    def jsonl(self, path: str | Path) -> Exporter:
        """Export to JSON Lines format."""
        to_jsonl(self._data, path)
        return self

    def excel(self, path: str | Path, sheet_name: str = "Sheet1") -> Exporter:
        """Export to Excel format."""
        to_excel(self._data, path, sheet_name)
        return self

    def parquet(self, path: str | Path) -> Exporter:
        """Export to Parquet format."""
        to_parquet(self._data, path)
        return self

    def sqlite(self, path: str | Path, table: str = "scraped_data") -> Exporter:
        """Export to SQLite database."""
        to_sqlite(self._data, path, table)
        return self
