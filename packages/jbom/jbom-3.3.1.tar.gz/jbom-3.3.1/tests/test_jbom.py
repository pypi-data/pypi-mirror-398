#!/usr/bin/env python3
"""
Unit tests for jBOM

Updated to reflect current design (Dec 2025):
- Priority-based inventory ranking (removed ACTIVE/Reorder dependencies)
- Simplified tie-handling (Priority column encodes all business logic)
- BOM sorting by component category and natural numbering
- Removed stocking information columns from test data
"""

import unittest
import tempfile
import csv
from pathlib import Path
import sys

# Add src directory to path for imports during testing
sys.path.insert(0, str(Path(__file__).parent.parent / "src"))

# Import the classes and functions to test
from jbom.common.types import Component, InventoryItem, BOMEntry
from jbom.common.constants import (
    ComponentType,
    DiagnosticIssue,
    CommonFields,
    ScoreWeights,
    CATEGORY_FIELDS,
    VALUE_INTERPRETATION,
)
from jbom.generators.bom import BOMGenerator
from jbom.processors.component_types import (
    get_category_fields,
    get_value_interpretation,
)
from jbom.processors.inventory_matcher import InventoryMatcher
from jbom import EXCEL_SUPPORT, NUMBERS_SUPPORT


class TestResistorParsing(unittest.TestCase):
    """Test resistor value parsing and EIA formatting"""

    def setUp(self):
        # Create a temporary inventory file for matcher
        self.temp_inv = tempfile.NamedTemporaryFile(
            mode="w", delete=False, suffix=".csv"
        )
        self.temp_inv.write("IPN,Name,Category,Value,LCSC\ntest,test,RES,330R,C123\n")
        self.temp_inv.close()
        self.matcher = InventoryMatcher(Path(self.temp_inv.name))

    def tearDown(self):
        Path(self.temp_inv.name).unlink()

    def test_parse_res_to_ohms(self):
        """Test parsing various resistor value formats to ohms"""
        test_cases = [
            ("330", 330.0),
            ("330R", 330.0),
            ("330Ω", 330.0),
            ("3R3", 3.3),
            ("22k", 22000.0),
            ("22K", 22000.0),
            ("22k0", 22000.0),
            ("22K0", 22000.0),
            ("2M2", 2200000.0),
            ("1M", 1000000.0),
            ("0R22", 0.22),
            ("10K1", 10100.0),
            ("47K5", 47500.0),
            ("2M7", 2700000.0),
            ("1R2", 1.2),
            ("", None),
            ("invalid", None),
        ]

        for input_val, expected in test_cases:
            with self.subTest(input_val=input_val):
                result = self.matcher._parse_res_to_ohms(input_val)
                if expected is None:
                    self.assertIsNone(result)
                else:
                    self.assertAlmostEqual(result, expected, places=6)

    def test_ohms_to_eia_basic(self):
        """Test formatting ohms to EIA format"""
        test_cases = [
            (330.0, False, "330R"),
            (3300.0, False, "3K3"),
            (10000.0, False, "10K"),
            (22000.0, False, "22K"),
            (1000000.0, False, "1M"),
            (2200000.0, False, "2M2"),
            (3.3, False, "3R3"),
            (0.22, False, "0R22"),
        ]

        for ohms, force_zero, expected in test_cases:
            with self.subTest(ohms=ohms, force_zero=force_zero):
                result = self.matcher._ohms_to_eia(ohms, force_trailing_zero=force_zero)
                self.assertEqual(result, expected)

    def test_ohms_to_eia_precision(self):
        """Test EIA formatting with precision trailing zeros"""
        test_cases = [
            (10000.0, True, "10K0"),
            (22000.0, True, "22K0"),
            (1000000.0, True, "1M0"),
            (330.0, True, "330R"),  # No trailing zero for R values
        ]

        for ohms, force_zero, expected in test_cases:
            with self.subTest(ohms=ohms, force_zero=force_zero):
                result = self.matcher._ohms_to_eia(ohms, force_trailing_zero=force_zero)
                self.assertEqual(result, expected)


class TestCapacitorParsing(unittest.TestCase):
    """Test capacitor value parsing and formatting"""

    def setUp(self):
        self.temp_inv = tempfile.NamedTemporaryFile(
            mode="w", delete=False, suffix=".csv"
        )
        self.temp_inv.write("IPN,Name,Category,Value,LCSC\ntest,test,CAP,100nF,C123\n")
        self.temp_inv.close()
        self.matcher = InventoryMatcher(Path(self.temp_inv.name))

    def tearDown(self):
        Path(self.temp_inv.name).unlink()

    def test_parse_cap_to_farad(self):
        """Test parsing capacitor values to farads"""
        test_cases = [
            ("100nF", 100e-9),
            ("100n", 100e-9),
            ("0.1uF", 0.1e-6),
            ("1uF", 1e-6),
            ("1u", 1e-6),
            ("220pF", 220e-12),
            ("220p", 220e-12),
            ("1n0", 1e-9),
            ("1u0", 1e-6),
            ("", None),
            ("invalid", None),
        ]

        for input_val, expected in test_cases:
            with self.subTest(input_val=input_val):
                result = self.matcher._parse_cap_to_farad(input_val)
                if expected is None:
                    self.assertIsNone(result)
                else:
                    self.assertAlmostEqual(result, expected, places=12)

    def test_farad_to_eia(self):
        """Test formatting farads to EIA format"""
        test_cases = [
            (100e-9, "100nF"),
            (1e-6, "1uF"),
            (0.1e-6, "100nF"),
            (220e-12, "220pF"),
            (2.2e-6, "2u2F"),
            (4.7e-9, "4n7F"),
        ]

        for farads, expected in test_cases:
            with self.subTest(farads=farads):
                result = self.matcher._farad_to_eia(farads)
                self.assertEqual(result, expected)


class TestInductorParsing(unittest.TestCase):
    """Test inductor value parsing and formatting"""

    def setUp(self):
        self.temp_inv = tempfile.NamedTemporaryFile(
            mode="w", delete=False, suffix=".csv"
        )
        self.temp_inv.write("IPN,Name,Category,Value,LCSC\ntest,test,IND,10uH,C123\n")
        self.temp_inv.close()
        self.matcher = InventoryMatcher(Path(self.temp_inv.name))

    def tearDown(self):
        Path(self.temp_inv.name).unlink()

    def test_parse_ind_to_henry(self):
        """Test parsing inductor values to henrys"""
        test_cases = [
            ("10uH", 10e-6),
            ("10u", 10e-6),
            ("100nH", 100e-9),
            ("100n", 100e-9),
            ("2m2H", 2.2e-3),
            ("2m2", 2.2e-3),
            ("1mH", 1e-3),
            ("", None),
            ("invalid", None),
        ]

        for input_val, expected in test_cases:
            with self.subTest(input_val=input_val):
                result = self.matcher._parse_ind_to_henry(input_val)
                if expected is None:
                    self.assertIsNone(result)
                else:
                    self.assertAlmostEqual(result, expected, places=9)

    def test_henry_to_eia(self):
        """Test formatting henrys to EIA format"""
        test_cases = [
            (10e-6, "10uH"),
            (100e-9, "100nH"),
            (2.2e-3, "2m2H"),
            (1e-3, "1mH"),
            (4.7e-6, "4u7H"),
        ]

        for henrys, expected in test_cases:
            with self.subTest(henrys=henrys):
                result = self.matcher._henry_to_eia(henrys)
                self.assertEqual(result, expected)


class TestCategorySpecificFields(unittest.TestCase):
    """Test category-specific field mappings and value interpretation"""

    def test_get_category_fields(self):
        """Test category-specific field retrieval"""
        # Test resistor fields
        res_fields = get_category_fields("RES")
        self.assertIn("Tolerance", res_fields)
        self.assertIn("V", res_fields)
        self.assertIn("W", res_fields)
        self.assertNotIn("Capacitance", res_fields)  # Should not have capacitor fields

        # Test capacitor fields
        cap_fields = get_category_fields("CAP")
        self.assertIn("Voltage", cap_fields)
        self.assertIn("Type", cap_fields)
        self.assertIn("Tolerance", cap_fields)
        self.assertNotIn("Resistance", cap_fields)  # Should not have resistor fields

        # Test LED fields
        led_fields = get_category_fields("LED")
        self.assertIn("mcd", led_fields)
        self.assertIn("Wavelength", led_fields)
        self.assertIn("Angle", led_fields)

        # Test unknown category (should get default fields)
        unknown_fields = get_category_fields("UNKNOWN")
        self.assertIn("Tolerance", unknown_fields)  # Should include common fields
        self.assertIn("Temperature Coefficient", unknown_fields)

    def test_get_value_interpretation(self):
        """Test value interpretation mapping"""
        # Test known interpretations
        self.assertEqual(get_value_interpretation("RES"), "Resistance")
        self.assertEqual(get_value_interpretation("CAP"), "Capacitance")
        self.assertEqual(get_value_interpretation("IND"), "Inductance")
        self.assertEqual(get_value_interpretation("LED"), "Color")

        # Test alternative naming
        self.assertEqual(get_value_interpretation("RESISTOR"), "Resistance")
        self.assertEqual(get_value_interpretation("CAPACITOR"), "Capacitance")

        # Test unknown types
        self.assertIsNone(get_value_interpretation("UNKNOWN"))
        self.assertIsNone(get_value_interpretation(""))

    def test_category_field_constants(self):
        """Test that the category field constants are properly defined"""
        # Test that main categories exist
        self.assertIn("RES", CATEGORY_FIELDS)
        self.assertIn("CAP", CATEGORY_FIELDS)
        self.assertIn("LED", CATEGORY_FIELDS)
        self.assertIn("IND", CATEGORY_FIELDS)

        # Test that value interpretation constants exist
        self.assertIn("RES", VALUE_INTERPRETATION)
        self.assertIn("CAP", VALUE_INTERPRETATION)
        self.assertIn("IND", VALUE_INTERPRETATION)
        self.assertIn("LED", VALUE_INTERPRETATION)


class TestComponentTypeDetection(unittest.TestCase):
    """Test component type detection from lib_id and footprint"""

    def setUp(self):
        self.temp_inv = tempfile.NamedTemporaryFile(
            mode="w", delete=False, suffix=".csv"
        )
        self.temp_inv.write("IPN,Name,Category,Value,LCSC\ntest,test,RES,330R,C123\n")
        self.temp_inv.close()
        self.matcher = InventoryMatcher(Path(self.temp_inv.name))

    def tearDown(self):
        Path(self.temp_inv.name).unlink()

    def test_get_component_type(self):
        """Test component type detection"""
        test_cases = [
            ("Device:R", "RES"),
            ("Device:C", "CAP"),
            ("Device:L", "IND"),
            ("Device:LED", "LED"),
            ("Device:Q_NPN_BCE", "Q"),  # Fixed: Q components are transistors
            ("Connector:Conn_01x02", "CON"),
            ("Switch:SW_Push", "SWI"),
            ("MCU:ESP32", "IC"),  # Fixed: MCU components are ICs
            ("IC:74HC595", "IC"),  # Additional test for IC prefix
            ("SPCoast:resistor", "RES"),
            ("SPCoast:capacitor", "CAP"),
            ("unknown:part", None),
        ]

        for lib_id, expected in test_cases:
            with self.subTest(lib_id=lib_id):
                component = Component(
                    reference="R1", lib_id=lib_id, value="", footprint=""
                )
                result = self.matcher._get_component_type(component)
                self.assertEqual(result, expected)


class TestPrecisionResistorDetection(unittest.TestCase):
    """Test precision resistor detection logic"""

    def setUp(self):
        # Create comprehensive inventory with standard decade values
        self.temp_inv = tempfile.NamedTemporaryFile(
            mode="w", delete=False, suffix=".csv"
        )

        # Standard E6/E12/E24 decade values with both 1% and 5% tolerances
        csv_content = [
            "IPN,Name,Category,Generic,Package,Value,Tolerance,LCSC,Priority"
        ]

        # E12 series standard values (5% tolerance)
        e12_values = [
            "10R",
            "12R",
            "15R",
            "18R",
            "22R",
            "27R",
            "33R",
            "39R",
            "47R",
            "56R",
            "68R",
            "82R",
        ]
        for i, val in enumerate(e12_values):
            csv_content.append(f"R{100+i},{val} 5%,RES,0603,0603,{val},5%,C{100+i},1")
            # Add K and M variants
            k_val = val.replace("R", "K")
            m_val = (
                val.replace("R", "M") if val != "10R" else None
            )  # Skip 10M (too large)
            csv_content.append(
                f"R{200+i},{k_val} 5%,RES,0603,0603,{k_val},5%,C{200+i},1"
            )
            if m_val:
                csv_content.append(
                    f"R{300+i},{m_val} 5%,RES,0603,0603,{m_val},5%,C{300+i},1"
                )

        # E24/E48 precision values (1% tolerance) - subset for testing
        precision_values = ["10K0", "10K1", "10K2", "47K5", "22K1", "33K2"]
        for i, val in enumerate(precision_values):
            csv_content.append(f"RP{i},{val} 1%,RES,0603,0603,{val},1%,CP{i},1")

        # Some standard values with only 5% available (for warning tests)
        warning_test_values = ["91K", "1M3"]
        for i, val in enumerate(warning_test_values):
            csv_content.append(f"RW{i},{val} 5%,RES,0603,0603,{val},5%,CW{i},2")

        self.temp_inv.write("\n".join(csv_content))
        self.temp_inv.close()

        self.matcher = InventoryMatcher(Path(self.temp_inv.name))

        # Test components covering various scenarios
        self.components = [
            # Precision components that HAVE 1% inventory matches
            Component("R1", "Device:R", "10K0", "PCM_SPCoast:0603-RES"),  # Has 1% match
            Component("R2", "Device:R", "47K5", "PCM_SPCoast:0603-RES"),  # Has 1% match
            # Precision components that DON'T have 1% matches (should warn)
            Component(
                "R3", "Device:R", "91K0", "PCM_SPCoast:0603-RES"
            ),  # Only 5% available
            Component(
                "R4", "Device:R", "1M30", "PCM_SPCoast:0603-RES"
            ),  # Only 5% available
            # Standard components (should not warn)
            Component("R5", "Device:R", "10K", "PCM_SPCoast:0603-RES"),  # Standard
            Component("R6", "Device:R", "47R", "PCM_SPCoast:0603-RES"),  # Standard
            # Component with no inventory match at all
            Component(
                "R7", "Device:R", "999K9", "PCM_SPCoast:0603-RES"
            ),  # Not in inventory
        ]

    def tearDown(self):
        Path(self.temp_inv.name).unlink()

    def test_precision_detection_pattern(self):
        """Test that precision resistor patterns are correctly detected"""
        import re

        precision_pattern = r"^\s*\d+[kKmMrR]\d+\s*"

        test_cases = [
            # Single trailing digit (precision)
            ("10K0", True),  # Precision - trailing 0
            ("10K1", True),  # Precision - trailing 1
            ("47K5", True),  # Precision - trailing 5
            ("2M7", True),  # Precision - trailing 7
            ("1R2", True),  # Precision - trailing 2
            # Multi-digit trailing (precision)
            ("1K33", True),  # Precision - 1.33kΩ
            ("2K74", True),  # Precision - 2.74kΩ
            ("10K05", True),  # Precision - 10.05kΩ
            ("1M47", True),  # Precision - 1.47MΩ
            ("0R125", True),  # Precision - 0.125Ω
            # Standard values (no trailing digits)
            ("10K", False),  # Standard - no trailing digit
            ("330R", False),  # Standard - no trailing digit
            ("22k", False),  # Standard - lowercase, no trailing digit
            ("1M", False),  # Standard - no trailing digit
        ]

        for value, should_match in test_cases:
            with self.subTest(value=value):
                matches = bool(re.match(precision_pattern, value))
                self.assertEqual(
                    matches,
                    should_match,
                    (
                        f"Value '{value}' should "
                        f"{'match' if should_match else 'not match'} precision pattern"
                    ),
                )

    def test_bom_generation_precision_warnings(self):
        """Test that BOM generation includes precision warnings"""
        bom_gen = BOMGenerator(self.matcher)
        bom_gen.components = self.components
        bom_entries, _, _ = bom_gen.generate_bom()

        # Create lookup by reference
        entries_by_ref = {e.reference: e for e in bom_entries}

        # Test cases:
        # R1 (10K0): Precision format, has 1% match -> No warning
        self.assertIn("R1", entries_by_ref)
        self.assertNotIn("Warning", entries_by_ref["R1"].notes)

        # R2 (47K5): Precision format, has 1% match -> No warning
        self.assertIn("R2", entries_by_ref)
        self.assertNotIn("Warning", entries_by_ref["R2"].notes)

        # R3 (91K0): Precision format, only 5% available -> Warning
        self.assertIn("R3", entries_by_ref)
        self.assertIn("Warning", entries_by_ref["R3"].notes)

        # R4 (1M30): Precision format, only 5% available -> Warning
        self.assertIn("R4", entries_by_ref)
        self.assertIn("Warning", entries_by_ref["R4"].notes)

        # R5 (10K): Standard format -> No warning
        self.assertIn("R5", entries_by_ref)
        self.assertNotIn("Warning", entries_by_ref["R5"].notes)

        # R6 (47R): Standard format -> No warning
        self.assertIn("R6", entries_by_ref)
        self.assertNotIn("Warning", entries_by_ref["R6"].notes)

        # R7 (999K9): No match at all -> "No inventory match found"
        self.assertIn("R7", entries_by_ref)
        self.assertIn("No inventory match found", entries_by_ref["R7"].notes)


class TestInventoryMatching(unittest.TestCase):
    """Test inventory matching logic"""

    def setUp(self):
        # Create comprehensive test inventory with Priority column
        self.temp_inv = tempfile.NamedTemporaryFile(
            mode="w", delete=False, suffix=".csv"
        )
        csv_content = [
            (
                "IPN,Name,Category,Generic,Package,Value,Tolerance,LCSC,Manufacturer,"
                "MFGPN,Description,Datasheet,Priority"
            ),
            (
                "R001,330R 5%,RES,0603,0603,330R,5%,C25231,UNI-ROYAL,0603WAJ0331T5E,"
                "330Ω 5% 0603,https://example.com/r1,1"
            ),
            (
                "R002,10K 1%,RES,0603,0603,10K,1%,C25232,UNI-ROYAL,0603WAF1002T5E,"
                "10kΩ 1% 0603,https://example.com/r2,2"
            ),
            (
                "C001,100nF,CAP,0603,0603,100nF,10%,C14663,YAGEO,CC0603KRX7R9BB104,"
                "100nF X7R 0603,https://example.com/c1,1"
            ),
            (
                "L001,10uH,IND,0603,0603,10uH,20%,C1608,SUNLORD,SWPA3012S100MT,"
                "10µH 0603,https://example.com/l1,5"
            ),
        ]
        self.temp_inv.write("\n".join(csv_content))
        self.temp_inv.close()

        self.matcher = InventoryMatcher(Path(self.temp_inv.name))

    def tearDown(self):
        Path(self.temp_inv.name).unlink()

    def test_find_matches_resistor(self):
        """Test finding matches for resistors"""
        component = Component("R1", "Device:R", "330R", "PCM_SPCoast:0603-RES")
        matches = self.matcher.find_matches(component)

        self.assertGreater(len(matches), 0)
        best_item, best_score, _ = matches[0]
        self.assertEqual(best_item.value, "330R")
        self.assertEqual(best_item.category, "RES")

    def test_find_matches_capacitor(self):
        """Test finding matches for capacitors"""
        component = Component("C1", "Device:C", "100nF", "PCM_SPCoast:0603-CAP")
        matches = self.matcher.find_matches(component)

        self.assertGreater(len(matches), 0)
        best_item, best_score, _ = matches[0]
        self.assertEqual(best_item.value, "100nF")
        self.assertEqual(best_item.category, "CAP")

    def test_find_matches_inductor(self):
        """Test finding matches for inductors"""
        component = Component("L1", "Device:L", "10uH", "PCM_SPCoast:0603-IND")
        matches = self.matcher.find_matches(component)

        self.assertGreater(len(matches), 0)
        best_item, best_score, _ = matches[0]
        self.assertEqual(best_item.value, "10uH")
        self.assertEqual(best_item.category, "IND")

    def test_no_matches_found(self):
        """Test behavior when no matches are found"""
        component = Component(
            "U1", "MCU:Unknown", "ESP32-NONEXISTENT", "Package_QFP:LQFP-64"
        )
        matches = self.matcher.find_matches(component)

        self.assertEqual(len(matches), 0)

    def test_priority_ranking(self):
        """Test that Priority values are used for ranking (lower Priority = better)"""
        component = Component(
            "R1", "Device:R", "10K", "PCM_SPCoast:0603-RES"
        )  # Should match both R001 and R002
        matches = self.matcher.find_matches(component)

        if len(matches) > 1:
            # First match should have lower priority number (better)
            first_priority = matches[0][0].priority
            second_priority = matches[1][0].priority
            self.assertLessEqual(
                first_priority,
                second_priority,
                "Lower priority numbers should be ranked first",
            )


class TestBOMGeneration(unittest.TestCase):
    """Test BOM generation and CSV output"""

    def setUp(self):
        self.temp_inv = tempfile.NamedTemporaryFile(
            mode="w", delete=False, suffix=".csv"
        )
        csv_content = [
            (
                "IPN,Name,Category,Generic,Package,Value,LCSC,Manufacturer,MFGPN,"
                "Description,Datasheet,Priority"
            ),
            (
                "R001,330R,RES,0603,0603,330R,C25231,UNI-ROYAL,0603WAJ0331T5E,"
                "330Ω 5% 0603,https://example.com/r1,1"
            ),
            (
                "C001,100nF,CAP,0603,0603,100nF,C14663,YAGEO,CC0603KRX7R9BB104,"
                "100nF X7R 0603,https://example.com/c1,1"
            ),
        ]
        self.temp_inv.write("\n".join(csv_content))
        self.temp_inv.close()

        self.matcher = InventoryMatcher(Path(self.temp_inv.name))
        self.components = [
            Component("R1", "Device:R", "330R", "PCM_SPCoast:0603-RES"),
            Component(
                "R2", "Device:R", "330R", "PCM_SPCoast:0603-RES"
            ),  # Duplicate for grouping
            Component("C1", "Device:C", "100nF", "PCM_SPCoast:0603-CAP"),
        ]

    def tearDown(self):
        Path(self.temp_inv.name).unlink()

    def test_component_grouping(self):
        """Test that components are grouped by their matching inventory item"""
        bom_gen = BOMGenerator(self.matcher)
        bom_gen.components = self.components
        groups = bom_gen._group_components()

        # Should have 2 groups: one for R1,R2 (same inventory item) and one for C1
        self.assertEqual(len(groups), 2)

        # Find the resistor group (should be grouped by IPN, not value)
        resistor_group = None
        for key, comps in groups.items():
            if "R001" in key:  # IPN-based grouping
                resistor_group = comps
                break

        self.assertIsNotNone(resistor_group)
        self.assertEqual(len(resistor_group), 2)  # R1 and R2

    def test_bom_generation(self):
        """Test basic BOM generation"""
        bom_gen = BOMGenerator(self.matcher)
        bom_gen.components = self.components
        bom_entries, _, _ = bom_gen.generate_bom()

        # Should have entries for grouped components
        self.assertGreater(len(bom_entries), 0)

        # Find resistor entry (should be grouped)
        resistor_entry = None
        for entry in bom_entries:
            if "R1" in entry.reference and "R2" in entry.reference:
                resistor_entry = entry
                break

        self.assertIsNotNone(resistor_entry)
        self.assertEqual(resistor_entry.quantity, 2)
        self.assertEqual(resistor_entry.value, "330R")

    def test_csv_output_basic(self):
        """Test basic CSV output format"""
        bom_gen = BOMGenerator(self.matcher)
        bom_gen.components = self.components
        bom_entries, _, _ = bom_gen.generate_bom()

        with tempfile.NamedTemporaryFile(mode="w", delete=False, suffix=".csv") as f:
            temp_csv = Path(f.name)

        try:
            # Basic field list (non-verbose, no manufacturer) - normalized snake_case
            fields = [
                "reference",
                "quantity",
                "description",
                "value",
                "footprint",
                "lcsc",
                "datasheet",
                "smd",
            ]
            bom_gen.write_bom_csv(bom_entries, temp_csv, fields)

            # Read back and verify format
            with open(temp_csv, "r") as f:
                reader = csv.reader(f)
                header = next(reader)

                # Headers should be Title Case
                expected_header = [
                    "Reference",
                    "Quantity",
                    "Description",
                    "Value",
                    "Footprint",
                    "LCSC",
                    "Datasheet",
                    "SMD",
                ]
                self.assertEqual(header, expected_header)

                # Should have at least one data row
                rows = list(reader)
                self.assertGreater(len(rows), 0)
        finally:
            temp_csv.unlink()

    def test_csv_output_verbose(self):
        """Test verbose CSV output format"""
        from jbom.common.generator import GeneratorOptions

        bom_gen = BOMGenerator(self.matcher, GeneratorOptions(verbose=True))
        bom_gen.components = self.components
        bom_entries, _, _ = bom_gen.generate_bom(
            verbose=True
        )  # Pass verbose to generate_bom

        with tempfile.NamedTemporaryFile(mode="w", delete=False, suffix=".csv") as f:
            temp_csv = Path(f.name)

        try:
            # Verbose field list with manufacturer - normalized snake_case
            fields = [
                "reference",
                "quantity",
                "description",
                "value",
                "footprint",
                "lcsc",
                "manufacturer",
                "mfgpn",
                "datasheet",
                "smd",
                "match_quality",
                "priority",
            ]
            bom_gen.write_bom_csv(bom_entries, temp_csv, fields)

            # Read back and verify format includes extra columns
            with open(temp_csv, "r") as f:
                reader = csv.reader(f)
                header = next(reader)

                # Should include manufacturer and verbose columns (simplified) - Title Case headers
                self.assertIn("Manufacturer", header)
                self.assertIn("MFGPN", header)
                self.assertIn("Match Quality", header)
                self.assertIn("Priority", header)
        finally:
            temp_csv.unlink()


class TestSchematicLoader(unittest.TestCase):
    """Test KiCad schematic parsing (basic functionality)"""

    def test_component_creation(self):
        """Test Component dataclass creation"""
        comp = Component(
            reference="R1",
            lib_id="Device:R",
            value="330R",
            footprint="Resistor_SMD:R_0603_1608Metric",
            properties={"Tolerance": "5%", "W": "0.1W"},
        )

        self.assertEqual(comp.reference, "R1")
        self.assertEqual(comp.lib_id, "Device:R")
        self.assertEqual(comp.value, "330R")
        self.assertEqual(comp.properties["Tolerance"], "5%")
        self.assertTrue(comp.in_bom)
        self.assertFalse(comp.dnp)

    def test_inventory_item_creation(self):
        """Test InventoryItem dataclass creation"""
        item = InventoryItem(
            ipn="R001",
            keywords="resistor",
            category="RES",
            description="330Ω 5% 0603 resistor",
            smd="Yes",
            value="330R",
            type="Resistor",
            tolerance="5%",
            voltage="75V",
            amperage="",
            wattage="0.1W",
            lcsc="C25231",
            manufacturer="UNI-ROYAL",
            mfgpn="0603WAJ0331T5E",
            datasheet="https://example.com",
            package="0603",
            priority=1,
        )

        self.assertEqual(item.ipn, "R001")
        self.assertEqual(item.category, "RES")
        self.assertEqual(item.value, "330R")
        self.assertEqual(item.priority, 1)


class TestBOMSorting(unittest.TestCase):
    """Test BOM sorting by category and component numbering"""

    def setUp(self):
        self.temp_inv = tempfile.NamedTemporaryFile(
            mode="w", delete=False, suffix=".csv"
        )
        csv_content = [
            (
                "IPN,Name,Category,Generic,Package,Value,LCSC,Manufacturer,MFGPN,"
                "Description,Datasheet,Priority"
            ),
            (
                "R001,330R,RES,0603,0603,330R,C25231,UNI-ROYAL,0603WAJ0331T5E,"
                "330Ω 5% 0603,https://example.com/r1,1"
            ),
            (
                "R002,10K,RES,0603,0603,10K,C25232,UNI-ROYAL,0603WAF1002T5E,"
                "10kΩ 5% 0603,https://example.com/r2,1"
            ),
            (
                "C001,1uF,CAP,0603,0603,1uF,C14663,YAGEO,CC0603KRX7R9BB104,"
                "1uF X7R 0603,https://example.com/c1,1"
            ),
            (
                "D001,BAT54A,DIO,SOT-23,SOT-23,BAT54A,C12345,FOSAN,BAT54A,BAT54A Schottky,"
                "https://example.com/d1,1"
            ),
            (
                "LED001,WS2812B,LED,5050,5050,WS2812B,C54678,XINGLIGHT,WS2812B,RGB LED,"
                "https://example.com/led1,1"
            ),
        ]
        self.temp_inv.write("\n".join(csv_content))
        self.temp_inv.close()

        self.matcher = InventoryMatcher(Path(self.temp_inv.name))
        # Components in mixed order to test sorting
        self.components = [
            Component("R10", "Device:R", "330R", "PCM_SPCoast:0603-RES"),
            Component("LED3", "Device:LED", "WS2812B", "PCM_SPCoast:WS2812B"),
            Component("C2", "Device:C", "1uF", "PCM_SPCoast:0603-CAP"),
            Component("R1", "Device:R", "10K", "PCM_SPCoast:0603-RES"),
            Component("D5", "Device:D", "BAT54A", "Package_TO_SOT_SMD:SOT-23"),
            Component("R2", "Device:R", "10K", "PCM_SPCoast:0603-RES"),  # Same as R1
        ]

    def tearDown(self):
        Path(self.temp_inv.name).unlink()

    def test_bom_sort_key_parsing(self):
        """Test component reference parsing for sorting"""
        bom_gen = BOMGenerator(self.matcher)
        bom_gen.components = self.components

        test_cases = [
            ("R1", ("R", 1)),
            ("C10", ("C", 10)),
            ("LED4", ("LED", 4)),
            ("U100", ("U", 100)),
        ]

        for ref, expected in test_cases:
            with self.subTest(reference=ref):
                category, number = bom_gen._parse_reference(ref)
                self.assertEqual(category, expected[0])
                self.assertEqual(number, expected[1])

    def test_bom_sorting_order(self):
        """Test that BOM entries are sorted correctly"""
        bom_gen = BOMGenerator(self.matcher)
        bom_gen.components = self.components
        bom_entries, _, _ = bom_gen.generate_bom()

        # Extract references to check order
        references = [entry.reference for entry in bom_entries]

        # The actual references might be grouped, so check the essential ordering
        self.assertEqual(len(references), 5)

        # Check that categories appear in correct order
        categories_found = []
        for ref in references:
            # Extract first letter/category from reference
            if ref.startswith("C"):
                categories_found.append("C")
            elif ref.startswith("D"):
                categories_found.append("D")
            elif ref.startswith("LED"):
                categories_found.append("LED")
            elif ref.startswith("R"):
                categories_found.append("R")

        # Should be in alphabetical order by category
        expected_categories = ["C", "D", "LED", "R", "R"]
        self.assertEqual(categories_found, expected_categories)


class TestFieldPrefixSystem(unittest.TestCase):
    """Test I:/C: prefix system for field disambiguation"""

    def setUp(self):
        # Create inventory with fields that might conflict with component properties
        self.temp_inv = tempfile.NamedTemporaryFile(
            mode="w", delete=False, suffix=".csv"
        )
        csv_content = [
            "IPN,Name,Category,Package,Value,Tolerance,Voltage,LCSC,Priority",
            "R001,330R 5%,RES,0603,330R,5%,75V,C25231,1",
            "C001,100nF,CAP,0603,100nF,10%,25V,C14663,1",
        ]
        self.temp_inv.write("\n".join(csv_content))
        self.temp_inv.close()

        self.matcher = InventoryMatcher(Path(self.temp_inv.name))

        # Components with properties that might conflict with inventory fields
        self.components = [
            Component(
                "R1",
                "Device:R",
                "330R",
                "PCM_SPCoast:0603-RES",
                properties={"Tolerance": "1%", "Voltage": "50V"},
            ),
            Component(
                "C1",
                "Device:C",
                "100nF",
                "PCM_SPCoast:0603-CAP",
                properties={"Voltage": "50V"},
            ),
        ]

    def tearDown(self):
        Path(self.temp_inv.name).unlink()

    def test_field_discovery(self):
        """Test that field discovery finds both inventory and component fields"""
        bom_gen = BOMGenerator(self.matcher)
        bom_gen.components = self.components
        available_fields = bom_gen.get_available_fields(self.components)

        # Should have standard BOM fields (normalized snake_case)
        self.assertIn("reference", available_fields)
        self.assertIn("quantity", available_fields)
        self.assertIn("value", available_fields)

        # Should have inventory fields (both prefixed and unprefixed where appropriate)
        self.assertIn("i:tolerance", available_fields)
        self.assertIn("i:package", available_fields)

        # Should have component property fields
        self.assertIn("c:tolerance", available_fields)
        self.assertIn("c:voltage", available_fields)

        # Should have ambiguous fields that exist in both sources
        self.assertIn("tolerance", available_fields)  # Ambiguous field
        self.assertIn("voltage", available_fields)  # Ambiguous field

    def test_field_value_extraction_prefixed(self):
        """Test field value extraction with explicit prefixes"""
        bom_gen = BOMGenerator(self.matcher)
        bom_gen.components = self.components
        bom_entries, _, _ = bom_gen.generate_bom()

        # Find resistor entry
        resistor_entry = None
        for entry in bom_entries:
            if "R1" in entry.reference:
                resistor_entry = entry
                break

        self.assertIsNotNone(resistor_entry)

        # Find corresponding component and inventory item
        component = self.components[0]  # R1
        inventory_item = None
        for item in self.matcher.inventory:
            if item.lcsc == resistor_entry.lcsc:
                inventory_item = item
                break

        self.assertIsNotNone(inventory_item)

        # Test explicit inventory field extraction
        inv_tolerance = bom_gen._get_field_value(
            "I:Tolerance", resistor_entry, component, inventory_item
        )
        self.assertEqual(inv_tolerance, "5%")  # From inventory

        # Test explicit component property extraction
        comp_tolerance = bom_gen._get_field_value(
            "C:Tolerance", resistor_entry, component, inventory_item
        )
        self.assertEqual(comp_tolerance, "1%")  # From component properties

    def test_ambiguous_field_handling(self):
        """Test that ambiguous fields return combined values"""
        bom_gen = BOMGenerator(self.matcher)
        bom_gen.components = self.components
        bom_entries, _, _ = bom_gen.generate_bom()

        # Find resistor entry
        resistor_entry = None
        for entry in bom_entries:
            if "R1" in entry.reference:
                resistor_entry = entry
                break

        component = self.components[0]
        inventory_item = None
        for item in self.matcher.inventory:
            if item.lcsc == resistor_entry.lcsc:
                inventory_item = item
                break

        # Test ambiguous field (should return combined value)
        tolerance_value = bom_gen._get_field_value(
            "Tolerance", resistor_entry, component, inventory_item
        )
        self.assertIn(
            "i:", tolerance_value
        )  # Should contain inventory marker (lowercase)
        self.assertIn(
            "c:", tolerance_value
        )  # Should contain component marker (lowercase)


class TestDebugFunctionality(unittest.TestCase):
    """Test debug functionality and alternative match display"""

    def setUp(self):
        # Create test inventory with multiple matching items
        self.temp_inv = tempfile.NamedTemporaryFile(
            mode="w", delete=False, suffix=".csv"
        )
        csv_content = [
            (
                "IPN,Name,Category,Package,Value,Tolerance,LCSC,Manufacturer,MFGPN,"
                "Description,Priority"
            ),
            (
                "R001,330R 5%,RES,0603,330R,5%,C25231,UNI-ROYAL,0603WAJ0331T5E,"
                "330Ω 5% 0603,1"
            ),
            (
                "R002,330R 1%,RES,0603,330R,1%,C25232,YAGEO,RC0603FR-07330RL,"
                "330Ω 1% 0603,2"
            ),
            (
                "R003,330R 10%,RES,0603,330R,10%,C25233,VISHAY,CRCW0603330RJNEA,"
                "330Ω 10% 0603,3"
            ),
        ]
        self.temp_inv.write("\n".join(csv_content))
        self.temp_inv.close()

        self.matcher = InventoryMatcher(Path(self.temp_inv.name))
        self.components = [
            Component("R1", "Device:R", "330R", "PCM_SPCoast:0603-RES"),
        ]

    def tearDown(self):
        Path(self.temp_inv.name).unlink()

    def test_debug_mode_enabled(self):
        """Test that debug mode works without polluting BOM notes"""
        bom_gen = BOMGenerator(self.matcher)
        bom_gen.components = self.components
        bom_entries, _, _ = bom_gen.generate_bom(debug=True)

        self.assertEqual(len(bom_entries), 1)
        entry = bom_entries[0]

        # Debug information should NOT be in the BOM notes - it's handled by console output
        # BOM files should remain clean and professional
        self.assertNotIn("DEBUG:", entry.notes or "")

        # Verify basic BOM entry structure is intact
        self.assertTrue(entry.reference)
        self.assertTrue(entry.lcsc)
        self.assertTrue(entry.description)

    def test_debug_alternatives_displayed(self):
        """Test that debug mode processes multiple matches correctly"""
        bom_gen = BOMGenerator(self.matcher)
        bom_gen.components = self.components
        bom_entries, _, _ = bom_gen.generate_bom(debug=True)

        entry = bom_entries[0]

        # Debug information should NOT pollute BOM notes
        self.assertNotIn("ALTERNATIVES:", entry.notes or "")
        self.assertNotIn("Alt1:", entry.notes or "")

        # However, the BOM should still contain the best match
        self.assertTrue(entry.lcsc)  # Should have matched a component
        self.assertIn(
            "C252", entry.lcsc
        )  # Should match one of the LCSC numbers (C25231, C25232, C25233)

    def test_debug_mode_disabled(self):
        """Test that normal mode doesn't show debug information"""
        bom_gen = BOMGenerator(self.matcher)
        bom_gen.components = self.components
        bom_entries, _, _ = bom_gen.generate_bom(debug=False)

        self.assertEqual(len(bom_entries), 1)
        entry = bom_entries[0]

        # Should not contain debug information
        self.assertNotIn("DEBUG:", entry.notes)
        self.assertNotIn("ALTERNATIVES:", entry.notes)
        self.assertNotIn("Component: R1", entry.notes)

    def test_find_matches_debug_signature(self):
        """Test that find_matches returns proper 3-tuple with debug info"""
        component = Component("R1", "Device:R", "330R", "PCM_SPCoast:0603-RES")

        # Test with debug enabled
        matches = self.matcher.find_matches(component, debug=True)

        self.assertGreater(len(matches), 0)

        # Each match should be a 3-tuple: (item, score, debug_info)
        for match in matches:
            self.assertEqual(len(match), 3)
            item, score, debug_info = match
            self.assertIsInstance(item, InventoryItem)
            self.assertIsInstance(score, int)
            # Debug info should be present for first match, may be None for others
            if debug_info is not None:
                self.assertIsInstance(debug_info, str)


class TestHierarchicalSupport(unittest.TestCase):
    """Test hierarchical schematic support functionality"""

    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.project_dir = Path(self.temp_dir.name) / "test_project"
        self.project_dir.mkdir()

        # Create test inventory
        self.temp_inv = tempfile.NamedTemporaryFile(
            mode="w", delete=False, suffix=".csv"
        )
        csv_content = [
            "IPN,Name,Category,Package,Value,LCSC,Manufacturer,MFGPN,Description,Priority",
            "R001,330R,RES,0603,330R,C25231,UNI-ROYAL,0603WAJ0331T5E,330Ω 5% 0603,1",
            "C001,100nF,CAP,0603,100nF,C14663,YAGEO,CC0603KRX7R9BB104,100nF X7R 0603,1",
        ]
        self.temp_inv.write("\n".join(csv_content))
        self.temp_inv.close()

    def tearDown(self):
        self.temp_dir.cleanup()
        Path(self.temp_inv.name).unlink()

    def create_hierarchical_root(self, filename="test_project.kicad_sch"):
        """Create a mock hierarchical root schematic file"""
        root_content = """(kicad_sch
    (version 20231120)
    (generator "eeschema")
    (generator_version "8.0")
    (uuid "test-root-uuid")
    (paper "A4")
    (lib_symbols)
    (sheet
        (at 25.4 25.4)
        (size 12.7 3.81)
        (property "Sheetname" "SubSheet1")
        (property "Sheetfile" "subsheet1.kicad_sch")
    )
    (sheet
        (at 45.4 25.4)
        (size 12.7 3.81)
        (property "Sheetname" "SubSheet2")
        (property "Sheetfile" "subsheet2.kicad_sch")
    )
)"""
        root_file = self.project_dir / filename
        root_file.write_text(root_content)
        return root_file

    def create_simple_schematic(
        self, filename="simple.kicad_sch", with_components=True
    ):
        """Create a mock simple schematic file"""
        if with_components:
            content = """(kicad_sch
    (version 20231120)
    (lib_symbols
        (symbol "Device:R" (properties...))
    )
    (symbol (lib_id "Device:R") (at 50 50 0) (unit 1)
        (in_bom yes) (on_board yes) (dnp no)
        (property "Reference" "R1" (at 52 50 0))
        (property "Value" "330R" (at 50 47 0))
        (property "Footprint" "PCM_SPCoast:0603-RES" (at 50 45 0))
    )
)"""
        else:
            content = """(kicad_sch
    (version 20231120)
    (lib_symbols)
)"""
        simple_file = self.project_dir / filename
        simple_file.write_text(content)
        return simple_file

    def test_is_hierarchical_schematic(self):
        """Test detection of hierarchical schematics"""
        from jbom.common.utils import is_hierarchical_schematic

        # Create hierarchical schematic
        hierarchical_file = self.create_hierarchical_root()
        self.assertTrue(is_hierarchical_schematic(hierarchical_file))

        # Create simple schematic
        simple_file = self.create_simple_schematic("simple.kicad_sch")
        self.assertFalse(is_hierarchical_schematic(simple_file))

    def test_extract_sheet_files(self):
        """Test extraction of sheet file references"""
        from jbom.common.utils import extract_sheet_files

        hierarchical_file = self.create_hierarchical_root()
        sheet_files = extract_sheet_files(hierarchical_file)

        self.assertEqual(len(sheet_files), 2)
        self.assertIn("subsheet1.kicad_sch", sheet_files)
        self.assertIn("subsheet2.kicad_sch", sheet_files)

    def test_find_best_schematic_normal_file(self):
        """Test finding best schematic with normal files"""
        from jbom.common.utils import find_best_schematic

        # Create files - should prefer directory-matching name
        self.create_simple_schematic("test_project.kicad_sch")
        self.create_simple_schematic("other.kicad_sch")

        best = find_best_schematic(self.project_dir)
        self.assertEqual(best.name, "test_project.kicad_sch")

    def test_find_best_schematic_hierarchical_priority(self):
        """Test that hierarchical schematics are preferred"""
        from jbom.common.utils import find_best_schematic

        # Create hierarchical root and simple schematic
        self.create_hierarchical_root("test_project.kicad_sch")
        self.create_simple_schematic("simple.kicad_sch")

        best = find_best_schematic(self.project_dir)
        self.assertEqual(
            best.name, "test_project.kicad_sch"
        )  # Should prefer hierarchical root

    def test_find_best_schematic_autosave_warning(self):
        """Test autosave file handling with warning"""
        from jbom.common.utils import find_best_schematic
        import io
        import sys

        # Create only autosave file
        self.create_hierarchical_root("_autosave-test_project.kicad_sch")

        # Capture stdout to check for warning
        captured_output = io.StringIO()
        sys.stdout = captured_output

        try:
            best = find_best_schematic(self.project_dir)
            self.assertEqual(best.name, "_autosave-test_project.kicad_sch")

            # Check that warning was issued
            output = captured_output.getvalue()
            self.assertIn("WARNING", output)
            self.assertIn("autosave", output.lower())
        finally:
            sys.stdout = sys.__stdout__

    def test_process_hierarchical_schematic(self):
        """Test processing of hierarchical schematic files"""
        from jbom.common.utils import process_hierarchical_schematic

        # Create hierarchical root
        root_file = self.create_hierarchical_root()

        # Create sub-sheets
        self.create_simple_schematic("subsheet1.kicad_sch")
        self.create_simple_schematic("subsheet2.kicad_sch")

        files_to_process = process_hierarchical_schematic(root_file, self.project_dir)

        # Should return root + 2 sub-sheets
        self.assertEqual(len(files_to_process), 3)
        file_names = [f.name for f in files_to_process]
        self.assertIn("test_project.kicad_sch", file_names)
        self.assertIn("subsheet1.kicad_sch", file_names)
        self.assertIn("subsheet2.kicad_sch", file_names)

    def test_process_hierarchical_missing_subsheet(self):
        """Test handling of missing sub-sheet files"""
        from jbom.common.utils import process_hierarchical_schematic
        import io
        import sys

        # Create hierarchical root but not sub-sheets
        root_file = self.create_hierarchical_root()

        # Capture stdout to check for warnings
        captured_output = io.StringIO()
        sys.stdout = captured_output

        try:
            files_to_process = process_hierarchical_schematic(
                root_file, self.project_dir
            )

            # Should still return root file even if sub-sheets missing
            self.assertEqual(len(files_to_process), 1)
            self.assertEqual(files_to_process[0].name, "test_project.kicad_sch")

            # Check for warning messages
            output = captured_output.getvalue()
            self.assertIn("Warning", output)
            self.assertIn("not found", output)
        finally:
            sys.stdout = sys.__stdout__

    def test_process_simple_schematic(self):
        """Test that simple schematics are processed normally"""
        from jbom.common.utils import process_hierarchical_schematic

        simple_file = self.create_simple_schematic()
        files_to_process = process_hierarchical_schematic(simple_file, self.project_dir)

        # Should return only the single file
        self.assertEqual(len(files_to_process), 1)
        self.assertEqual(files_to_process[0].name, "simple.kicad_sch")


class TestSMDFiltering(unittest.TestCase):
    """Test SMD component filtering functionality"""

    def setUp(self):
        # Create test inventory with both SMD and PTH components
        self.temp_inv = tempfile.NamedTemporaryFile(
            mode="w", delete=False, suffix=".csv"
        )
        csv_content = [
            "IPN,Name,Category,Package,Value,SMD,LCSC,Manufacturer,MFGPN,Description,Priority",
            "R001,330R SMD,RES,0603,330R,SMD,C25231,UNI-ROYAL,0603WAJ0331T5E,330Ω SMD 0603,1",
            "R002,330R PTH,RES,THT,330R,PTH,C25232,VISHAY,PTR0603330R,330Ω PTH,1",
            "C001,100nF SMD,CAP,0603,100nF,SMD,C14663,YAGEO,CC0603KRX7R9BB104,100nF SMD 0603,1",
            "C002,100nF PTH,CAP,THT,100nF,PTH,C14664,VISHAY,K104K15X7RF5TL2,100nF PTH,1",
        ]
        self.temp_inv.write("\n".join(csv_content))
        self.temp_inv.close()

        self.matcher = InventoryMatcher(Path(self.temp_inv.name))
        # Create mixed components - some that will match SMD, some PTH
        # Note: THT components use different footprints that won't match 0603 package
        self.components = [
            Component(
                "R1", "Device:R", "330R", "PCM_SPCoast:0603-RES"
            ),  # Should match SMD (R001)
            Component(
                "R2", "Device:R", "330R", "PCM_SPCoast:DIP-RES"
            ),  # Should match PTH (R002)
            Component(
                "C1", "Device:C", "100nF", "PCM_SPCoast:0603-CAP"
            ),  # Should match SMD (C001)
            Component(
                "C2", "Device:C", "100nF", "PCM_SPCoast:DIP-CAP"
            ),  # Should match PTH (C002)
        ]

    def tearDown(self):
        Path(self.temp_inv.name).unlink()

    def test_smd_filtering_enabled(self):
        """Test that SMD filtering works when enabled"""
        bom_gen = BOMGenerator(self.matcher)
        bom_gen.components = self.components

        # Generate BOM with SMD filtering
        bom_entries, _, _ = bom_gen.generate_bom(smd_only=True)

        # Should only include SMD components
        smd_entries = [entry for entry in bom_entries if entry.smd == "SMD"]
        pth_entries = [entry for entry in bom_entries if entry.smd == "PTH"]

        # All entries should be SMD when filtering is enabled
        self.assertEqual(len(smd_entries), len(bom_entries))
        self.assertEqual(len(pth_entries), 0)

    def test_smd_filtering_disabled(self):
        """Test that SMD filtering is off by default"""
        # Create a simple test with one SMD and one PTH component
        # that match different inventory items
        temp_inv2 = tempfile.NamedTemporaryFile(mode="w", delete=False, suffix=".csv")
        csv_content2 = [
            "IPN,Name,Category,Package,Value,SMD,LCSC,Manufacturer,MFGPN,Description,Priority",
            "R001,330R SMD,RES,0603,330R,SMD,C25231,UNI-ROYAL,0603WAJ0331T5E,330Ω SMD 0603,1",
            (
                "R003,1K PTH,RES,THT,1K,PTH,C25233,VISHAY,PTR06031K,1kΩ PTH,1"
            ),  # Different value to force different match
        ]
        temp_inv2.write("\n".join(csv_content2))
        temp_inv2.close()

        matcher2 = InventoryMatcher(Path(temp_inv2.name))
        components2 = [
            Component(
                "R1", "Device:R", "330R", "PCM_SPCoast:0603-RES"
            ),  # Should match SMD item
            Component(
                "R2", "Device:R", "1K", "PCM_SPCoast:DIP-RES"
            ),  # Should match PTH item
        ]

        bom_gen = BOMGenerator(matcher2)
        bom_gen.components = components2
        bom_entries, _, _ = bom_gen.generate_bom(smd_only=False)

        # Should include both SMD and PTH components
        smd_entries = [entry for entry in bom_entries if entry.smd == "SMD"]
        pth_entries = [entry for entry in bom_entries if entry.smd == "PTH"]

        # Should have both types
        self.assertGreater(
            len(smd_entries), 0, "Should have at least one SMD component"
        )
        self.assertGreater(
            len(pth_entries), 0, "Should have at least one PTH component"
        )

        # Clean up
        Path(temp_inv2.name).unlink()

    def test_is_smd_component_method(self):
        """Test the _is_smd_component helper method"""
        bom_gen = BOMGenerator([], self.matcher)

        # Test explicit SMD marking
        smd_entry = BOMEntry(
            reference="R1",
            quantity=1,
            value="330R",
            footprint="0603",
            lcsc="C123",
            manufacturer="TEST",
            mfgpn="TEST",
            description="Test",
            datasheet="",
            smd="SMD",
        )
        self.assertTrue(bom_gen._is_smd_component(smd_entry))

        # Test explicit PTH marking
        pth_entry = BOMEntry(
            reference="R2",
            quantity=1,
            value="330R",
            footprint="THT",
            lcsc="C124",
            manufacturer="TEST",
            mfgpn="TEST",
            description="Test",
            datasheet="",
            smd="PTH",
        )
        self.assertFalse(bom_gen._is_smd_component(pth_entry))

        # Test footprint-based inference for unclear SMD field
        unclear_smd_entry = BOMEntry(
            reference="R3",
            quantity=1,
            value="330R",
            footprint="PCM_SPCoast:0603-RES",
            lcsc="C125",
            manufacturer="TEST",
            mfgpn="TEST",
            description="Test",
            datasheet="",
            smd="",  # Unclear SMD field
        )
        self.assertTrue(
            bom_gen._is_smd_component(unclear_smd_entry)
        )  # Should infer SMD from 0603


class TestCustomFieldOutput(unittest.TestCase):
    """Test custom field selection in BOM output"""

    def setUp(self):
        self.temp_inv = tempfile.NamedTemporaryFile(
            mode="w", delete=False, suffix=".csv"
        )
        csv_content = [
            "IPN,Name,Category,Package,Value,Tolerance,LCSC,Manufacturer,MFGPN,Priority",
            "R001,330R 5%,RES,0603,330R,5%,C25231,UNI-ROYAL,0603WAJ0331T5E,1",
        ]
        self.temp_inv.write("\n".join(csv_content))
        self.temp_inv.close()

        self.matcher = InventoryMatcher(Path(self.temp_inv.name))
        self.components = [
            Component(
                "R1",
                "Device:R",
                "330R",
                "PCM_SPCoast:0603-RES",
                properties={"Tolerance": "1%", "Power": "0.1W"},
            ),
        ]

    def tearDown(self):
        Path(self.temp_inv.name).unlink()

    def test_custom_field_csv_output(self):
        """Test CSV output with custom field selection"""
        bom_gen = BOMGenerator(self.matcher)
        bom_gen.components = self.components
        bom_entries, _, _ = bom_gen.generate_bom()

        with tempfile.NamedTemporaryFile(mode="w", delete=False, suffix=".csv") as f:
            temp_csv = Path(f.name)

        try:
            # Test custom fields including prefixed ones (normalized snake_case)
            custom_fields = [
                "reference",
                "value",
                "i:package",
                "c:tolerance",
                "manufacturer",
            ]
            bom_gen.write_bom_csv(bom_entries, temp_csv, custom_fields)

            # Read back and verify
            with open(temp_csv, "r") as f:
                reader = csv.reader(f)
                header = next(reader)

                # Headers should be Title Case without spaces after prefixes
                expected_header = [
                    "Reference",
                    "Value",
                    "I:Package",
                    "C:Tolerance",
                    "Manufacturer",
                ]
                self.assertEqual(header, expected_header)

                # Check data row
                data_row = next(reader)
                self.assertEqual(data_row[0], "R1")  # Reference
                self.assertEqual(data_row[1], "330R")  # Value
                self.assertEqual(data_row[2], "0603")  # I:Package (from inventory)
                self.assertEqual(data_row[3], "1%")  # C:Tolerance (from component)
                self.assertEqual(data_row[4], "UNI-ROYAL")  # Manufacturer
        finally:
            temp_csv.unlink()

    def test_ambiguous_field_csv_output(self):
        """Test CSV output with ambiguous fields that auto-split into columns"""
        bom_gen = BOMGenerator(self.matcher)
        bom_gen.components = self.components
        bom_entries, _, _ = bom_gen.generate_bom()

        with tempfile.NamedTemporaryFile(mode="w", delete=False, suffix=".csv") as f:
            temp_csv = Path(f.name)

        try:
            # Use ambiguous field that should auto-split (normalized snake_case)
            custom_fields = ["reference", "value", "tolerance"]
            bom_gen.write_bom_csv(bom_entries, temp_csv, custom_fields)

            # Read back and verify
            with open(temp_csv, "r") as f:
                reader = csv.reader(f)
                header = next(reader)

                # Should auto-expand to separate I: and C: columns with Title Case headers
                expected_header = ["Reference", "Value", "I:Tolerance", "C:Tolerance"]
                self.assertEqual(header, expected_header)

                # Check data
                data_row = next(reader)
                self.assertEqual(data_row[0], "R1")  # Reference
                self.assertEqual(data_row[1], "330R")  # Value
                self.assertEqual(data_row[2], "5%")  # I:Tolerance (from inventory)
                self.assertEqual(data_row[3], "1%")  # C:Tolerance (from component)
        finally:
            temp_csv.unlink()


class TestNewConstants(unittest.TestCase):
    """Test the new constant classes and their values"""

    def test_component_type_constants(self):
        """Test ComponentType constants have expected values"""
        self.assertEqual(ComponentType.RESISTOR, "RES")
        self.assertEqual(ComponentType.CAPACITOR, "CAP")
        self.assertEqual(ComponentType.INDUCTOR, "IND")
        self.assertEqual(ComponentType.DIODE, "DIO")
        self.assertEqual(ComponentType.LED, "LED")
        self.assertEqual(ComponentType.TRANSISTOR, "Q")
        self.assertEqual(ComponentType.INTEGRATED_CIRCUIT, "IC")
        self.assertEqual(ComponentType.CONNECTOR, "CON")
        self.assertEqual(ComponentType.SWITCH, "SWI")
        self.assertEqual(ComponentType.MICROCONTROLLER, "MCU")
        self.assertEqual(ComponentType.REGULATOR, "REG")
        self.assertEqual(ComponentType.OSCILLATOR, "OSC")

    def test_diagnostic_issue_constants(self):
        """Test DiagnosticIssue constants have expected values"""
        self.assertEqual(DiagnosticIssue.TYPE_UNKNOWN, "type_unknown")
        self.assertEqual(DiagnosticIssue.NO_TYPE_MATCH, "no_type_match")
        self.assertEqual(DiagnosticIssue.NO_VALUE_MATCH, "no_value_match")
        self.assertEqual(DiagnosticIssue.PACKAGE_MISMATCH, "package_mismatch")
        self.assertEqual(
            DiagnosticIssue.PACKAGE_MISMATCH_GENERIC, "package_mismatch_generic"
        )
        self.assertEqual(DiagnosticIssue.NO_MATCH, "no_match")

    def test_common_fields_constants(self):
        """Test CommonFields constants have expected values"""
        self.assertEqual(CommonFields.VOLTAGE, "V")
        self.assertEqual(CommonFields.AMPERAGE, "A")
        self.assertEqual(CommonFields.WATTAGE, "W")
        self.assertEqual(CommonFields.TOLERANCE, "Tolerance")
        self.assertEqual(CommonFields.POWER, "Power")
        self.assertEqual(
            CommonFields.TEMPERATURE_COEFFICIENT, "Temperature Coefficient"
        )

    def test_score_weights_constants(self):
        """Test ScoreWeights constants have expected values"""
        self.assertEqual(ScoreWeights.TOLERANCE_EXACT, 15)
        self.assertEqual(ScoreWeights.TOLERANCE_BETTER, 12)
        self.assertEqual(ScoreWeights.VOLTAGE_MATCH, 10)
        self.assertEqual(ScoreWeights.CURRENT_MATCH, 10)
        self.assertEqual(ScoreWeights.POWER_MATCH, 10)
        self.assertEqual(ScoreWeights.LED_WAVELENGTH, 8)
        self.assertEqual(ScoreWeights.LED_INTENSITY, 8)
        self.assertEqual(ScoreWeights.OSC_FREQUENCY, 12)
        self.assertEqual(ScoreWeights.OSC_STABILITY, 8)
        self.assertEqual(ScoreWeights.LED_ANGLE, 5)
        self.assertEqual(ScoreWeights.OSC_LOAD, 5)
        self.assertEqual(ScoreWeights.CON_PITCH, 10)
        self.assertEqual(ScoreWeights.MCU_FAMILY, 8)
        self.assertEqual(ScoreWeights.GENERIC_PROPERTY, 3)


class TestNormalizeComponentType(unittest.TestCase):
    """Test the normalize_component_type function"""

    def test_normalize_component_type_mapping(self):
        """Test that normalize_component_type maps component types correctly"""
        from jbom import normalize_component_type

        # Test direct mapping to existing categories
        self.assertEqual(normalize_component_type("R"), "RES")
        self.assertEqual(normalize_component_type("RESISTOR"), "RES")
        self.assertEqual(normalize_component_type("C"), "CAP")
        self.assertEqual(normalize_component_type("CAPACITOR"), "CAP")
        self.assertEqual(normalize_component_type("D"), "DIO")
        self.assertEqual(normalize_component_type("DIODE"), "DIO")
        self.assertEqual(normalize_component_type("L"), "IND")
        self.assertEqual(normalize_component_type("INDUCTOR"), "IND")

        # Test direct category returns (case insensitive)
        self.assertEqual(normalize_component_type("RES"), "RES")
        self.assertEqual(normalize_component_type("res"), "RES")
        self.assertEqual(normalize_component_type("CAP"), "CAP")

        # Test transistor mapping
        self.assertEqual(normalize_component_type("TRANSISTOR"), "Q")

        # Test unknown component (returns as-is, uppercase)
        self.assertEqual(
            normalize_component_type("Device:Unknown_Part"), "DEVICE:UNKNOWN_PART"
        )
        self.assertEqual(normalize_component_type(""), "")


class TestDiagnosticWarning(unittest.TestCase):
    """Test diagnostic warning functionality"""

    def setUp(self):
        # Create a minimal test setup
        self.temp_inv = tempfile.NamedTemporaryFile(
            mode="w", delete=False, suffix=".csv"
        )
        csv_content = [
            "IPN,Name,Category,Package,Value,SMD,LCSC,Priority",
            "R001,330R,RES,0603,330R,SMD,C25231,1",
        ]
        self.temp_inv.write("\n".join(csv_content))
        self.temp_inv.close()

        self.matcher = InventoryMatcher(Path(self.temp_inv.name))
        self.bom_gen = BOMGenerator([], self.matcher)

    def tearDown(self):
        Path(self.temp_inv.name).unlink()

    def test_smd_warning_with_invalid_value(self):
        """Test that diagnostic warnings are generated for invalid SMD field values"""
        import sys
        from io import StringIO

        # Capture stderr to check for warnings
        old_stderr = sys.stderr
        captured_stderr = StringIO()
        sys.stderr = captured_stderr

        try:
            # Create an entry with invalid SMD field value
            entry = BOMEntry(
                reference="R1",
                quantity=1,
                value="330R",
                footprint="0603",
                lcsc="C25231",
                manufacturer="TEST",
                mfgpn="TEST",
                description="Test",
                datasheet="",
                smd="Q16",  # Invalid SMD value
            )

            # This should trigger the warning in _is_smd_component
            is_smd = self.bom_gen._is_smd_component(entry)

            # Check that warning was printed to stderr
            stderr_output = captured_stderr.getvalue()
            self.assertIn("Warning: Unexpected SMD field value", stderr_output)
            self.assertIn("Q16", stderr_output)
            self.assertIn("R1", stderr_output)

            # Should default to non-SMD
            self.assertFalse(is_smd)

        finally:
            sys.stderr = old_stderr


class TestSpreadsheetSupport(unittest.TestCase):
    """Test Excel and Numbers file support"""

    def setUp(self):
        # Create test data that matches expected inventory structure
        self.test_data = {
            "IPN": ["R001", "C001"],
            "Name": ["330R 5%", "100nF X7R"],
            "Category": ["RES", "CAP"],
            "Package": ["0603", "0603"],
            "Value": ["330R", "100nF"],
            "LCSC": ["C25231", "C14663"],
            "Manufacturer": ["UNI-ROYAL", "YAGEO"],
            "MFGPN": ["0603WAJ0331T5E", "CC0603KRX7R9BB104"],
            "Priority": [1, 1],
        }

    @unittest.skipUnless(EXCEL_SUPPORT, "openpyxl not available")
    def test_excel_file_loading(self):
        """Test loading inventory from Excel file"""
        import openpyxl

        # Create temporary Excel file
        temp_excel = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        temp_excel.close()

        try:
            # Create Excel workbook
            wb = openpyxl.Workbook()
            ws = wb.active

            # Write headers
            for col, header in enumerate(self.test_data.keys(), 1):
                ws.cell(row=1, column=col, value=header)

            # Write data
            for row_idx in range(len(self.test_data["IPN"])):
                for col, key in enumerate(self.test_data.keys(), 1):
                    ws.cell(
                        row=row_idx + 2, column=col, value=self.test_data[key][row_idx]
                    )

            wb.save(temp_excel.name)
            wb.close()

            # Test loading
            matcher = InventoryMatcher(Path(temp_excel.name))

            # Verify data was loaded correctly
            self.assertEqual(len(matcher.inventory), 2)

            # Check first item
            r_item = next(
                (item for item in matcher.inventory if item.ipn == "R001"), None
            )
            self.assertIsNotNone(r_item)
            self.assertEqual(r_item.category, "RES")
            self.assertEqual(r_item.value, "330R")
            self.assertEqual(r_item.lcsc, "C25231")
            self.assertEqual(r_item.manufacturer, "UNI-ROYAL")

            # Check second item
            c_item = next(
                (item for item in matcher.inventory if item.ipn == "C001"), None
            )
            self.assertIsNotNone(c_item)
            self.assertEqual(c_item.category, "CAP")
            self.assertEqual(c_item.value, "100nF")
            self.assertEqual(c_item.lcsc, "C14663")

        finally:
            Path(temp_excel.name).unlink()

    def test_unsupported_file_format(self):
        """Test error handling for unsupported file formats"""
        temp_file = tempfile.NamedTemporaryFile(suffix=".txt", delete=False)
        temp_file.write(b"Some text content")
        temp_file.close()

        try:
            with self.assertRaises(ValueError) as context:
                InventoryMatcher(Path(temp_file.name))

            self.assertIn("Unsupported inventory file format", str(context.exception))
            self.assertIn(".txt", str(context.exception))

        finally:
            Path(temp_file.name).unlink()

    @unittest.skipIf(
        EXCEL_SUPPORT, "Testing Excel import error when openpyxl not available"
    )
    def test_excel_import_error(self):
        """Test error handling when openpyxl is not available"""
        temp_excel = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        temp_excel.close()

        try:
            with self.assertRaises(ImportError) as context:
                InventoryMatcher(Path(temp_excel.name))

            self.assertIn("openpyxl package", str(context.exception))

        finally:
            Path(temp_excel.name).unlink()

    @unittest.skipIf(
        NUMBERS_SUPPORT,
        "Testing Numbers import error when numbers-parser not available",
    )
    def test_numbers_import_error(self):
        """Test error handling when numbers-parser is not available"""
        temp_numbers = tempfile.NamedTemporaryFile(suffix=".numbers", delete=False)
        temp_numbers.close()

        try:
            with self.assertRaises(ImportError) as context:
                InventoryMatcher(Path(temp_numbers.name))

            self.assertIn("numbers-parser package", str(context.exception))

        finally:
            Path(temp_numbers.name).unlink()

    def test_csv_still_works(self):
        """Test that CSV loading still works as before"""
        temp_csv = tempfile.NamedTemporaryFile(mode="w", suffix=".csv", delete=False)

        # Write CSV content
        csv_content = [
            "IPN,Name,Category,Package,Value,LCSC,Manufacturer,MFGPN,Priority",
            "R001,330R 5%,RES,0603,330R,C25231,UNI-ROYAL,0603WAJ0331T5E,1",
            "C001,100nF X7R,CAP,0603,100nF,C14663,YAGEO,CC0603KRX7R9BB104,1",
        ]
        temp_csv.write("\n".join(csv_content))
        temp_csv.close()

        try:
            # Test loading
            matcher = InventoryMatcher(Path(temp_csv.name))

            # Verify data was loaded correctly
            self.assertEqual(len(matcher.inventory), 2)

            # Check that items are correct
            r_item = next(
                (item for item in matcher.inventory if item.ipn == "R001"), None
            )
            self.assertIsNotNone(r_item)
            self.assertEqual(r_item.value, "330R")

        finally:
            Path(temp_csv.name).unlink()


class TestFieldArgumentParsing(unittest.TestCase):
    """Test --fields argument parsing with preset expansion"""

    def setUp(self):
        """Create a matcher with some test inventory for field validation"""
        self.temp_inv = tempfile.NamedTemporaryFile(
            mode="w", delete=False, suffix=".csv"
        )
        csv_writer = csv.DictWriter(
            self.temp_inv,
            fieldnames=["IPN", "Category", "Package", "Value", "LCSC", "Priority"],
        )
        csv_writer.writeheader()
        csv_writer.writerow(
            {
                "IPN": "R001",
                "Category": "RES",
                "Package": "0603",
                "Value": "330R",
                "LCSC": "C25231",
                "Priority": "1",
            }
        )
        self.temp_inv.close()
        self.matcher = InventoryMatcher(Path(self.temp_inv.name))

    def tearDown(self):
        Path(self.temp_inv.name).unlink()

    def test_parse_jlc_preset(self):
        """Test expanding +jlc preset"""
        from jbom.common.fields import parse_fields_argument as _parse_fields_argument

        available_fields = {
            "reference": "Standard BOM field",
            "quantity": "Standard BOM field",
            "lcsc": "Standard BOM field",
            "value": "Standard BOM field",
            "footprint": "Standard BOM field",
            "description": "Standard BOM field",
            "datasheet": "Standard BOM field",
            "smd": "Standard BOM field",
            "i:package": "Inventory field",
            "fabricator": "Standard BOM field",
            "fabricator_part_number": "Standard BOM field",
        }

        result = _parse_fields_argument("+jlc", available_fields, False, False)
        self.assertIn("reference", result)
        self.assertIn("quantity", result)
        self.assertIn("value", result)
        self.assertIn("i:package", result)
        self.assertIn("fabricator", result)
        self.assertIn("fabricator_part_number", result)

    def test_parse_standard_preset(self):
        """Test expanding +default preset"""
        from jbom.common.fields import parse_fields_argument as _parse_fields_argument

        available_fields = {
            "reference": "Standard",
            "quantity": "Standard",
            "description": "Standard",
            "value": "Standard",
            "footprint": "Standard",
            "lcsc": "Standard",
            "datasheet": "Standard",
            "smd": "Standard",
            "manufacturer": "Standard",
            "mfgpn": "Standard",
            "fabricator": "Standard",
            "fabricator_part_number": "Standard",
        }

        result = _parse_fields_argument("+default", available_fields, False, False)
        self.assertIn("reference", result)
        self.assertIn("quantity", result)
        self.assertIn("description", result)
        self.assertIn("manufacturer", result)
        self.assertIn("mfgpn", result)
        self.assertIn("fabricator", result)
        self.assertIn("fabricator_part_number", result)

    def test_parse_minimal_preset(self):
        """Test expanding +minimal preset"""
        from jbom.common.fields import parse_fields_argument as _parse_fields_argument

        available_fields = {
            "reference": "Standard",
            "quantity": "Standard",
            "value": "Standard",
            "lcsc": "Standard",
        }

        result = _parse_fields_argument("+minimal", available_fields, False, False)
        self.assertEqual(result, ["reference", "quantity", "value", "lcsc"])

    def test_parse_all_preset(self):
        """Test expanding +all preset to include all available fields"""
        from jbom.common.fields import parse_fields_argument as _parse_fields_argument

        available_fields = {
            "reference": "Standard",
            "value": "Standard",
            "customfield": "Custom",
            "lcsc": "Standard",
        }

        result = _parse_fields_argument("+all", available_fields, False, False)
        # Should include all fields in sorted order
        self.assertEqual(result, ["customfield", "lcsc", "reference", "value"])

    def test_parse_custom_fields(self):
        """Test parsing custom comma-separated field list"""
        from jbom.common.fields import parse_fields_argument as _parse_fields_argument

        available_fields = {
            "reference": "Standard",
            "quantity": "Standard",
            "value": "Standard",
            "lcsc": "Standard",
        }

        result = _parse_fields_argument(
            "reference,quantity,value,lcsc", available_fields, False, False
        )
        self.assertEqual(result, ["reference", "quantity", "value", "lcsc"])

    def test_parse_mixed_preset_and_custom(self):
        """Test mixing preset expansion with custom fields"""
        from jbom.common.fields import parse_fields_argument as _parse_fields_argument

        available_fields = {
            "reference": "Standard",
            "quantity": "Standard",
            "value": "Standard",
            "lcsc": "Standard",
            "customfield": "Custom",
            "datasheet": "Standard",
            "smd": "Standard",
            "description": "Standard",
            "footprint": "Standard",
            "i:package": "Inventory",
        }

        result = _parse_fields_argument(
            "+jlc,customfield", available_fields, False, False
        )
        self.assertIn("reference", result)
        self.assertIn("customfield", result)
        self.assertIn("fabricator_part_number", result)

    def test_invalid_preset_name(self):
        """Test error when using invalid preset name"""
        from jbom.common.fields import parse_fields_argument as _parse_fields_argument

        available_fields = {"reference": "Standard", "value": "Standard"}

        with self.assertRaises(ValueError) as context:
            _parse_fields_argument("+invalid", available_fields, False, False)

        self.assertIn("Unknown preset", str(context.exception))
        self.assertIn("invalid", str(context.exception))

    def test_invalid_field_name(self):
        """Test error when using invalid field name"""
        from jbom.common.fields import parse_fields_argument as _parse_fields_argument

        available_fields = {"reference": "Standard", "value": "Standard"}

        with self.assertRaises(ValueError) as context:
            _parse_fields_argument(
                "reference,InvalidField", available_fields, False, False
            )

        self.assertIn("Unknown field", str(context.exception))
        self.assertIn("InvalidField", str(context.exception))

    def test_deduplication(self):
        """Test that duplicate fields are removed"""
        from jbom.common.fields import parse_fields_argument as _parse_fields_argument

        available_fields = {
            "reference": "Standard",
            "value": "Standard",
        }

        result = _parse_fields_argument(
            "reference,value,reference", available_fields, False, False
        )
        # Should have only 2 items, not 3
        self.assertEqual(len([f for f in result if f == "reference"]), 1)

    def test_empty_fields_argument(self):
        """Test that empty --fields defaults to standard preset"""
        from jbom.common.fields import parse_fields_argument as _parse_fields_argument

        available_fields = {
            "reference": "Standard",
            "quantity": "Standard",
            "description": "Standard",
            "value": "Standard",
            "footprint": "Standard",
            "lcsc": "Standard",
            "datasheet": "Standard",
            "smd": "Standard",
        }

        result = _parse_fields_argument(None, available_fields, False, False)
        self.assertIn("reference", result)  # Now returns snake_case
        self.assertIsNotNone(result)
        self.assertGreater(len(result), 0)


class TestNormalizeFieldName(unittest.TestCase):
    """Test normalize_field_name function for case-insensitive field handling"""

    def test_snake_case_preserved(self):
        """Test that snake_case format is preserved as-is"""
        from jbom import normalize_field_name

        test_cases = [
            ("match_quality", "match_quality"),
            ("reference", "reference"),
            ("i:package", "i:package"),
            ("c:tolerance", "c:tolerance"),
            ("my_custom_field", "my_custom_field"),
        ]

        for input_val, expected in test_cases:
            with self.subTest(input_val=input_val):
                result = normalize_field_name(input_val)
                self.assertEqual(result, expected)

    def test_title_case_conversion(self):
        """Test that Title Case is converted to snake_case"""
        from jbom import normalize_field_name

        test_cases = [
            ("Match Quality", "match_quality"),
            ("Reference", "reference"),
            ("Package", "package"),
            ("Manufacturer PN", "manufacturer_pn"),
            ("I:Package", "i:package"),
            ("C:Tolerance", "c:tolerance"),
        ]

        for input_val, expected in test_cases:
            with self.subTest(input_val=input_val):
                result = normalize_field_name(input_val)
                self.assertEqual(result, expected)

    def test_camelcase_conversion(self):
        """Test that CamelCase is converted to snake_case"""
        from jbom import normalize_field_name

        test_cases = [
            ("MatchQuality", "match_quality"),
            ("MyCustomField", "my_custom_field"),
            ("IPackage", "ipackage"),  # Consecutive caps like XMLData -> xmldata
            ("DataValue", "data_value"),
        ]

        for input_val, expected in test_cases:
            with self.subTest(input_val=input_val):
                result = normalize_field_name(input_val)
                self.assertEqual(result, expected)

    def test_uppercase_conversion(self):
        """Test that UPPERCASE is converted to lowercase"""
        from jbom import normalize_field_name

        test_cases = [
            ("REFERENCE", "reference"),
            ("MATCH_QUALITY", "match_quality"),
            ("VALUE", "value"),
            ("I:PACKAGE", "i:package"),
            ("C:TOLERANCE", "c:tolerance"),
        ]

        for input_val, expected in test_cases:
            with self.subTest(input_val=input_val):
                result = normalize_field_name(input_val)
                self.assertEqual(result, expected)

    def test_prefix_handling(self):
        """Test that I: and C: prefixes are handled correctly"""
        from jbom import normalize_field_name

        test_cases = [
            ("I:Package", "i:package"),
            ("i:package", "i:package"),
            ("C:Tolerance", "c:tolerance"),
            ("c:tolerance", "c:tolerance"),
            ("I:Match Quality", "i:match_quality"),
            ("C:Custom Field", "c:custom_field"),
        ]

        for input_val, expected in test_cases:
            with self.subTest(input_val=input_val):
                result = normalize_field_name(input_val)
                self.assertEqual(result, expected)

    def test_whitespace_handling(self):
        """Test that extra whitespace is normalized"""
        from jbom import normalize_field_name

        test_cases = [
            ("Match  Quality", "match_quality"),  # Double space
            ("  reference  ", "reference"),  # Leading/trailing spaces
            ("Match   Quality", "match_quality"),  # Multiple spaces
            ("Package  Type", "package_type"),  # Space in middle
        ]

        for input_val, expected in test_cases:
            with self.subTest(input_val=input_val):
                result = normalize_field_name(input_val)
                self.assertEqual(result, expected)

    def test_whitespace_in_fields_conversion(self):
        """Test that various whitespace formats are handled"""
        from jbom import normalize_field_name

        # Note: normalize_field_name handles spaces but not embedded newlines
        # Newlines are typically cleaned by inventory loader before reaching normalize_field_name
        test_cases = [
            ("Match Quality", "match_quality"),  # Space between words
            ("Tol erance", "tol_erance"),  # Space in word
            ("Pack age", "pack_age"),  # Space in word
        ]

        for input_val, expected in test_cases:
            with self.subTest(input_val=input_val):
                result = normalize_field_name(input_val)
                self.assertEqual(result, expected)

    def test_hyphen_to_underscore(self):
        """Test that hyphens are converted to underscores"""
        from jbom import normalize_field_name

        test_cases = [
            ("match-quality", "match_quality"),
            ("ref-des", "ref_des"),
            ("custom-field-name", "custom_field_name"),
        ]

        for input_val, expected in test_cases:
            with self.subTest(input_val=input_val):
                result = normalize_field_name(input_val)
                self.assertEqual(result, expected)

    def test_multiple_underscores_collapsed(self):
        """Test that multiple underscores are collapsed to single"""
        from jbom import normalize_field_name

        test_cases = [
            ("match__quality", "match_quality"),
            ("field___name", "field_name"),
        ]

        for input_val, expected in test_cases:
            with self.subTest(input_val=input_val):
                result = normalize_field_name(input_val)
                self.assertEqual(result, expected)

    def test_idempotence(self):
        """Test that normalizing twice yields same result"""
        from jbom import normalize_field_name

        test_inputs = [
            "match_quality",
            "Match Quality",
            "MatchQuality",
            "i:package",
            "I:Package",
        ]

        for test_input in test_inputs:
            with self.subTest(test_input=test_input):
                first_norm = normalize_field_name(test_input)
                second_norm = normalize_field_name(first_norm)
                self.assertEqual(first_norm, second_norm)

    def test_empty_string(self):
        """Test edge case of empty string"""
        from jbom import normalize_field_name

        result = normalize_field_name("")
        self.assertEqual(result, "")

    def test_mixed_formats(self):
        """Test mixed format conversions"""
        from jbom import normalize_field_name

        test_cases = [
            ("Match_Quality", "match_quality"),  # Already snake_case
            ("MATCH_Quality", "match_quality"),  # Mixed case
            ("match Quality", "match_quality"),  # Space instead of underscore
            ("Match-Quality", "match_quality"),  # Hyphen instead of underscore
        ]

        for input_val, expected in test_cases:
            with self.subTest(input_val=input_val):
                result = normalize_field_name(input_val)
                self.assertEqual(result, expected)


class TestFieldToHeader(unittest.TestCase):
    """Test field_to_header function for CSV header generation"""

    def test_basic_snake_case_to_title_case(self):
        """Test conversion of snake_case to Title Case"""
        from jbom import field_to_header

        test_cases = [
            ("reference", "Reference"),
            ("match_quality", "Match Quality"),
            ("lcsc", "LCSC"),
            ("manufacturer", "Manufacturer"),
            ("mfgpn", "MFGPN"),
        ]

        for input_val, expected in test_cases:
            with self.subTest(input_val=input_val):
                result = field_to_header(input_val)
                self.assertEqual(result, expected)

    def test_prefixed_fields_no_space_after_prefix(self):
        """Test that prefixed fields have no space after prefix"""
        from jbom import field_to_header

        test_cases = [
            ("i:package", "I:Package"),  # NOT "I: Package"
            ("c:tolerance", "C:Tolerance"),  # NOT "C: Tolerance"
            ("i:voltage", "I:Voltage"),  # NOT "I: Voltage"
            ("c:custom_field", "C:Custom Field"),
        ]

        for input_val, expected in test_cases:
            with self.subTest(input_val=input_val):
                result = field_to_header(input_val)
                self.assertEqual(result, expected)
                # Verify no space after colon
                if ":" in result:
                    self.assertNotIn(": ", result)

    def test_multiword_fields(self):
        """Test multiword field name conversion"""
        from jbom import field_to_header

        test_cases = [
            ("match_quality", "Match Quality"),
            ("custom_field_name", "Custom Field Name"),
            ("temperature_coefficient", "Temperature Coefficient"),
            ("i:temperature_coefficient", "I:Temperature Coefficient"),
        ]

        for input_val, expected in test_cases:
            with self.subTest(input_val=input_val):
                result = field_to_header(input_val)
                self.assertEqual(result, expected)

    def test_empty_string(self):
        """Test edge case of empty string"""
        from jbom import field_to_header

        result = field_to_header("")
        self.assertEqual(result, "")

    def test_single_word_fields(self):
        """Test single-word field names"""
        from jbom import field_to_header

        test_cases = [
            ("reference", "Reference"),
            ("value", "Value"),
            ("datasheet", "Datasheet"),
            ("quantity", "Quantity"),
        ]

        for input_val, expected in test_cases:
            with self.subTest(input_val=input_val):
                result = field_to_header(input_val)
                self.assertEqual(result, expected)


class TestInventoryRawHeaderMatching(unittest.TestCase):
    """Test inventory field matching with raw CSV header variations"""

    def setUp(self):
        # Create inventory with intentionally varied column names
        self.temp_inv = tempfile.NamedTemporaryFile(
            mode="w", delete=False, suffix=".csv"
        )
        csv_content = [
            (
                "IPN,Name,Category,Package,Value,Tolerance,LCSC,Manufacturer,Mfg PN,"
                "Description,Priority"
            ),
            (
                "R001,330R 5%,RES,0603,330R,5%,C25231,UNI-ROYAL,0603WAJ0331T5E,"
                "330Ω resistor,1"
            ),
        ]
        self.temp_inv.write("\n".join(csv_content))
        self.temp_inv.close()

        self.matcher = InventoryMatcher(Path(self.temp_inv.name))
        self.bom_gen = BOMGenerator([], self.matcher)

    def tearDown(self):
        Path(self.temp_inv.name).unlink()

    def test_get_inventory_field_value_exact_match(self):
        """Test getting inventory field value with exact field name match"""
        item = self.matcher.inventory[0]

        # Direct field names from CSV
        value = self.bom_gen._get_inventory_field_value("value", item)
        self.assertEqual(value, "330R")

        value = self.bom_gen._get_inventory_field_value("tolerance", item)
        self.assertEqual(value, "5%")

    def test_get_inventory_field_value_with_spaces(self):
        """Test getting inventory field value when raw CSV has spaces"""
        item = self.matcher.inventory[0]

        # Raw CSV has "Mfg PN" with space, should still match normalized "mfg_pn"
        value = self.bom_gen._get_inventory_field_value("mfg_pn", item)
        self.assertEqual(value, "0603WAJ0331T5E")

    def test_has_inventory_field(self):
        """Test checking if inventory field exists"""
        item = self.matcher.inventory[0]

        # Should find normalized field names
        self.assertTrue(self.bom_gen._has_inventory_field("value", item))
        self.assertTrue(self.bom_gen._has_inventory_field("tolerance", item))
        self.assertTrue(self.bom_gen._has_inventory_field("mfg_pn", item))

        # Should not find non-existent fields
        self.assertFalse(self.bom_gen._has_inventory_field("nonexistent", item))

    def test_missing_inventory_field_returns_empty(self):
        """Test that missing fields return empty string"""
        item = self.matcher.inventory[0]

        value = self.bom_gen._get_inventory_field_value("nonexistent_field", item)
        self.assertEqual(value, "")


class TestCaseInsensitiveFieldInput(unittest.TestCase):
    """Test that field input accepts various case/format combinations"""

    def setUp(self):
        self.temp_inv = tempfile.NamedTemporaryFile(
            mode="w", delete=False, suffix=".csv"
        )
        csv_content = [
            "IPN,Category,Package,Value,Tolerance,LCSC,Priority",
            "R001,RES,0603,330R,5%,C25231,1",
        ]
        self.temp_inv.write("\n".join(csv_content))
        self.temp_inv.close()

        self.matcher = InventoryMatcher(Path(self.temp_inv.name))
        self.components = [
            Component(
                "R1",
                "Device:R",
                "330R",
                "PCM_SPCoast:0603-RES",
                properties={"Tolerance": "1%"},
            ),
        ]

    def tearDown(self):
        Path(self.temp_inv.name).unlink()

    def test_get_field_value_case_insensitive(self):
        """Test that _get_field_value works with various input formats"""
        bom_gen = BOMGenerator(self.matcher)
        bom_gen.components = self.components
        bom_entries, _, _ = bom_gen.generate_bom()

        entry = bom_entries[0]
        component = self.components[0]
        item = self.matcher.inventory[0]

        # All these formats should work and return same result
        test_cases = [
            "reference",  # snake_case
            "Reference",  # Title Case
            "REFERENCE",  # UPPERCASE
        ]

        expected = "R1"
        for field_input in test_cases:
            with self.subTest(field_input=field_input):
                result = bom_gen._get_field_value(field_input, entry, component, item)
                self.assertEqual(result, expected)

    def test_prefixed_field_case_insensitive(self):
        """Test that prefixed fields work case-insensitively"""
        bom_gen = BOMGenerator(self.matcher)
        bom_gen.components = self.components
        bom_entries, _, _ = bom_gen.generate_bom()

        entry = bom_entries[0]
        component = self.components[0]
        item = self.matcher.inventory[0]

        # All should return '5%' (from inventory tolerance)
        test_cases = [
            "i:tolerance",
            "I:Tolerance",
            "I:TOLERANCE",
        ]

        for field_input in test_cases:
            with self.subTest(field_input=field_input):
                result = bom_gen._get_field_value(field_input, entry, component, item)
                self.assertEqual(result, "5%")

    def test_component_property_case_insensitive(self):
        """Test that component property fields work case-insensitively"""
        bom_gen = BOMGenerator(self.matcher)
        bom_gen.components = self.components
        bom_entries, _, _ = bom_gen.generate_bom()

        entry = bom_entries[0]
        component = self.components[0]
        item = self.matcher.inventory[0]

        # All should return '1%' (from component tolerance)
        test_cases = [
            "c:tolerance",
            "C:Tolerance",
            "C:TOLERANCE",
        ]

        for field_input in test_cases:
            with self.subTest(field_input=field_input):
                result = bom_gen._get_field_value(field_input, entry, component, item)
                self.assertEqual(result, "1%")


if __name__ == "__main__":
    # Import the constants needed for skip conditions
    from jbom import EXCEL_SUPPORT, NUMBERS_SUPPORT

    # Run with verbose output
    unittest.main(verbosity=2)
