"""Excel document reader supporting .xlsx and .xls formats."""

import os
import openpyxl
import xlrd
from tabulate import tabulate


def get_excel_content(file_path: str) -> str:
    """
    Read content from an Excel file.
    
    Args:
        file_path: Path to the Excel file (.xlsx or .xls).
        
    Returns:
        Formatted content as a string.
    """
    if not os.path.isabs(file_path):
        file_path = os.path.abspath(file_path)

    if not os.path.exists(file_path):
        return f"Error: File not found at {file_path}"

    _, ext = os.path.splitext(file_path)
    ext = ext.lower()
    
    output = []

    try:
        if ext == '.xlsx':
            output = _read_xlsx(file_path)
        elif ext == '.xls':
            output = _read_xls(file_path)
        else:
            return f"Error: Unsupported file extension {ext}"
            
        return "\n".join(output)
                
    except Exception as e:
        return f"Error reading Excel file: {e}"


def _read_xlsx(file_path: str) -> list[str]:
    """Read .xlsx file using openpyxl."""
    output = []
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    
    for sheet_name in workbook.sheetnames:
        output.append(f"\n--- Sheet: {sheet_name} ---\n")
        sheet = workbook[sheet_name]
        data = []
        
        for row in sheet.iter_rows(values_only=True):
            cleaned_row = [str(cell) if cell is not None else "" for cell in row]
            data.append(cleaned_row)
        
        if data:
            try:
                table = tabulate(data, headers="firstrow", tablefmt="grid")
                output.append(table)
            except Exception:
                output.append(str(data))
        else:
            output.append("(Empty Sheet)")
            
    return output


def _read_xls(file_path: str) -> list[str]:
    """Read .xls file using xlrd."""
    output = []
    workbook = xlrd.open_workbook(file_path)
    
    for sheet_name in workbook.sheet_names():
        output.append(f"\n--- Sheet: {sheet_name} ---\n")
        sheet = workbook.sheet_by_name(sheet_name)
        data = []
        
        for row_idx in range(sheet.nrows):
            row_data = [str(cell) if cell is not None else "" for cell in sheet.row_values(row_idx)]
            data.append(row_data)
            
        if data:
            try:
                table = tabulate(data, headers="firstrow", tablefmt="grid")
                output.append(table)
            except Exception:
                output.append(str(data))
        else:
            output.append("(Empty Sheet)")
            
    return output
