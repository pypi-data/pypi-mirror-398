from __future__ import annotations

import json
import re
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, List, Literal, Optional, Tuple, Union

from ..models.control_catalog_model import CRControlCatalog
from ..models.attack_catalog_model import CRAttackCatalog
from ..models.control_relationships_model import CRControlRelationships
from ..models.attack_control_relationships_model import CRAttackControlRelationships


_WORKBOOK_FORMAT = "crml_xlsx_mapping"
_WORKBOOK_VERSION = "1.0"

_SHEET_META = "_meta"
_SHEET_CONTROL_CATALOGS = "control_catalogs"
_SHEET_ATTACK_CATALOGS = "attack_catalogs"
_SHEET_CONTROL_RELATIONSHIPS = "control_relationships"
_SHEET_ATTACK_CONTROL_RELATIONSHIPS = "attack_control_relationships"


# Common column labels/descriptions (Sonar: avoid duplicating literals)
_DOC_NAME_LABEL = "Document name"
_DOC_NAME_DESC = "Name of this CRML document (meta.name)."
_DOC_VERSION_LABEL = "Document version"
_DOC_VERSION_DESC = "Optional document version (meta.version)."
_DOC_DESCRIPTION_LABEL = "Document description"
_DOC_DESCRIPTION_DESC = "Optional description (meta.description)."
_DOC_TAGS_LABEL = "Document tags (JSON)"
_DOC_TAGS_DESC = "Optional tags array as JSON."
_DOC_TAGS_DESC_EXAMPLE = 'Optional tags array as JSON (e.g. ["community"]).'
_TAGS_JSON_LABEL = "Tags (JSON)"
_TAGS_JSON_DESC = "Optional tags array as JSON."


def _doc_meta_columns(*, tags_desc: str) -> List[Tuple[str, str, str]]:
    return [
        ("doc_name", _DOC_NAME_LABEL, _DOC_NAME_DESC),
        ("doc_version", _DOC_VERSION_LABEL, _DOC_VERSION_DESC),
        ("doc_description", _DOC_DESCRIPTION_LABEL, _DOC_DESCRIPTION_DESC),
        ("doc_tags_json", _DOC_TAGS_LABEL, tags_desc),
    ]


def _xlsx_header_row(ws) -> int:
    # Mapping sheets created by this exporter hide row 1 (machine keys) and
    # show row 2 as the human header.
    return 2 if ws.row_dimensions[1].hidden else 1


def _xlsx_header_index(ws, name: str) -> int:
    # Machine header row is always row 1.
    header = [c.value for c in ws[1]]
    try:
        return header.index(name) + 1
    except ValueError as e:
        raise KeyError(name) from e


def _xlsx_set_freeze_and_filter(ws, *, header_row: int, first_body_row: int, get_column_letter) -> None:
    ws.freeze_panes = f"A{first_body_row}"
    if ws.max_row >= header_row and ws.max_column >= 1:
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ws.max_column)}{ws.max_row}"


def _xlsx_style_header_row(ws, *, header_row: int, header_fill, header_font, header_align) -> None:
    for cell in ws[header_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align


def _xlsx_wrap_body_cells(ws, *, first_body_row: int, wrap_align) -> None:
    for row in ws.iter_rows(min_row=first_body_row, max_row=ws.max_row, values_only=False):
        for cell in row:
            v = cell.value
            if v is None:
                continue
            if isinstance(v, str) and ("{" in v or "[" in v or "\n" in v):
                cell.alignment = wrap_align


def _xlsx_apply_column_widths(ws, *, widths: dict[str, float], get_column_letter) -> None:
    for col_name, width in widths.items():
        try:
            idx = _xlsx_header_index(ws, col_name)
        except KeyError:
            continue
        ws.column_dimensions[get_column_letter(idx)].width = width


def _xlsx_style_sheet(
    ws,
    *,
    header_fill,
    header_font,
    header_align,
    wrap_align,
    get_column_letter,
    column_widths: Optional[Dict[str, float]] = None,
) -> None:
    hr = _xlsx_header_row(ws)
    first_body_row = hr + 1
    _xlsx_set_freeze_and_filter(
        ws, header_row=hr, first_body_row=first_body_row, get_column_letter=get_column_letter
    )
    _xlsx_style_header_row(
        ws,
        header_row=hr,
        header_fill=header_fill,
        header_font=header_font,
        header_align=header_align,
    )
    _xlsx_wrap_body_cells(ws, first_body_row=first_body_row, wrap_align=wrap_align)
    if column_widths:
        _xlsx_apply_column_widths(ws, widths=column_widths, get_column_letter=get_column_letter)


def _control_catalog_get_or_create_doc(
    *,
    out_by_doc: dict[tuple[str, str | None, str | None, str | None, str | None], dict[str, Any]],
    doc_name: str,
    doc_version: Optional[str],
    doc_description: Optional[str],
    doc_tags: Any,
    catalog_id: Optional[str],
    framework: str,
) -> dict[str, Any]:
    key = _doc_key(doc_name, doc_version, doc_description, doc_tags, catalog_id)
    container = out_by_doc.get(key)
    if container is None:
        container = {
            "crml_control_catalog": "1.0",
            "meta": {
                "name": doc_name,
                "version": doc_version,
                "description": doc_description,
                "tags": doc_tags,
            },
            "catalog": {
                "id": catalog_id,
                "framework": framework,
                "controls": [],
            },
        }
        out_by_doc[key] = container
        return container

    if container["catalog"].get("framework") != framework:
        raise ValueError(
            f"control_catalogs sheet has conflicting frameworks for doc {doc_name!r}: "
            f"{container['catalog']['framework']!r} vs {framework!r}"
        )
    return container


def _control_catalog_parse_ref_obj(*, control_id: Optional[str], data: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    ref_standard = _cell_str(data.get("ref_standard"))
    ref_control = _cell_str(data.get("ref_control"))
    ref_requirement = _cell_str(data.get("ref_requirement"))

    if ref_standard is None and ref_control is None and ref_requirement is None:
        return None
    if ref_standard is None or ref_control is None:
        raise ValueError(f"control_catalogs row has incomplete ref for control {control_id!r}")
    return {
        "standard": ref_standard,
        "control": ref_control,
        "requirement": ref_requirement,
    }


def _xlsx_add_list_validation(ws, *, column_name: str, allowed: list[str]) -> None:
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    try:
        col_idx = _xlsx_header_index(ws, column_name)
    except KeyError:
        return

    start_row = _xlsx_header_row(ws) + 1
    if ws.max_row < start_row:
        return

    # Escape double-quotes for Excel list literals by doubling them.
    safe_items = [s.replace('"', '""') for s in allowed]
    formula = '"' + ",".join(safe_items) + '"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(
        f"{get_column_letter(col_idx)}{start_row}:{get_column_letter(col_idx)}{ws.max_row}"
    )


def _xlsx_number_format(ws, *, column_name: str, fmt: str) -> None:
    try:
        col_idx = _xlsx_header_index(ws, column_name)
    except KeyError:
        return

    start_row = _xlsx_header_row(ws) + 1
    for r in range(start_row, ws.max_row + 1):
        cell = ws.cell(row=r, column=col_idx)
        if cell.value is None:
            continue
        cell.number_format = fmt


def _xlsx_module():
    try:
        import openpyxl  # type: ignore

        return openpyxl
    except Exception as e:  # pragma: no cover
        raise ImportError(
            "openpyxl is required for XLSX import/export: pip install 'crml-lang[xlsx]'"
        ) from e


def _to_json_cell(value: Any) -> Optional[str]:
    if value is None:
        return None
    return json.dumps(value, ensure_ascii=False, sort_keys=True)


def _from_json_cell(text: Any) -> Any:
    if text is None:
        return None
    if isinstance(text, str):
        s = text.strip()
        if s == "":
            return None
        return json.loads(s)
    raise TypeError(f"Expected JSON cell as string, got {type(text).__name__}")


def _cell_str(v: Any) -> Optional[str]:
    if v is None:
        return None
    s = str(v)
    return s if s != "" else None


def _cell_float(v: Any) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip()
        if s == "":
            return None
        return float(s)
    raise TypeError(f"Expected numeric cell, got {type(v).__name__}")


def _now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def _try_int(v: Any) -> Optional[int]:
    if v is None:
        return None
    try:
        return int(v)
    except Exception:
        return None


def _safe_filename(name: str) -> str:
    s = name.strip()
    if not s:
        return "document"

    s = re.sub(r"[^A-Za-z0-9._-]+", "-", s)
    s = re.sub(r"-+", "-", s).strip("-")
    return s or "document"


def _doc_key(
    doc_name: str,
    doc_version: Optional[str],
    doc_description: Optional[str],
    doc_tags: Any,
    pack_or_catalog_id: Optional[str],
) -> Tuple[str, Optional[str], Optional[str], Optional[str], Optional[str]]:
    return (
        doc_name,
        doc_version,
        doc_description,
        _to_json_cell(doc_tags),
        pack_or_catalog_id,
    )


def _read_sheet_rows(wb, sheet_name: str, *, header_rows: int) -> Tuple[List[str], List[Tuple[Any, ...]]]:
    if sheet_name not in wb.sheetnames:
        return ([], [])

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) <= header_rows:
        return ([], [])

    header = [str(c) for c in rows[0]]
    body = list(rows[header_rows:])
    return (header, body)


def _read_doc_meta(data: Dict[str, Any]) -> Tuple[Optional[str], Optional[str], Optional[str], Any]:
    doc_name = _cell_str(data.get("doc_name"))
    doc_version = _cell_str(data.get("doc_version"))
    doc_description = _cell_str(data.get("doc_description"))
    doc_tags = _from_json_cell(data.get("doc_tags_json"))
    return (doc_name, doc_version, doc_description, doc_tags)


def _get_or_create_container(
    *,
    out_by_doc: Dict[Tuple[str, Optional[str], Optional[str], Optional[str], Optional[str]], Dict[str, Any]],
    doc_type: str,
    doc_name: str,
    doc_version: Optional[str],
    doc_description: Optional[str],
    doc_tags: Any,
    payload_key: str,
    payload_id: Optional[str],
    framework: Optional[str],
    list_key: str,
    extra_payload: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    key = _doc_key(doc_name, doc_version, doc_description, doc_tags, payload_id)
    container = out_by_doc.get(key)
    if container is None:
        meta_obj: dict[str, Any] = {
            "name": doc_name,
            "version": doc_version,
            "description": doc_description,
            "tags": doc_tags,
        }

        payload_obj: dict[str, Any] = {
            "id": payload_id,
            list_key: [],
        }
        if framework is not None:
            payload_obj["framework"] = framework
        if extra_payload:
            payload_obj.update(extra_payload)

        container = {
            doc_type: "1.0",
            "meta": meta_obj,
            payload_key: payload_obj,
        }
        out_by_doc[key] = container
    else:
        if framework is not None:
            existing = container.get(payload_key, {}).get("framework")
            if existing is not None and existing != framework:
                raise ValueError(
                    f"{sheet_name_for_error(payload_key)} sheet has conflicting frameworks for doc {doc_name!r}: "
                    f"{existing!r} vs {framework!r}"
                )
    return container


def sheet_name_for_error(payload_key: str) -> str:
    # This keeps existing error wording stable-ish.
    return payload_key


def _group_last_or_new(rel_list: list[dict[str, Any]], *, key_field: str, key_value: str, targets_key: str) -> dict[str, Any]:
    if rel_list and rel_list[-1].get(key_field) == key_value:
        return rel_list[-1]
    grouped: dict[str, Any] = {key_field: key_value, targets_key: []}
    rel_list.append(grouped)
    return grouped


@dataclass(frozen=True)
class ImportedXlsx:
    control_catalogs: list[CRControlCatalog]
    attack_catalogs: list[CRAttackCatalog]
    control_relationships: list[CRControlRelationships]
    attack_control_relationships: list[CRAttackControlRelationships]


def export_xlsx(
    out_path: str,
    *,
    control_catalogs: Iterable[CRControlCatalog] = (),
    control_catalog_paths: Iterable[Union[str, Path]] = (),
    attack_catalogs: Iterable[CRAttackCatalog] = (),
    attack_catalog_paths: Iterable[Union[str, Path]] = (),
    control_relationships: Iterable[CRControlRelationships] = (),
    control_relationship_paths: Iterable[Union[str, Path]] = (),
    attack_control_relationships: Iterable[CRAttackControlRelationships] = (),
    attack_control_relationship_paths: Iterable[Union[str, Path]] = (),
) -> None:
    """Export CRML documents into a strict XLSX workbook.

    The workbook is versioned via the `_meta` sheet.
    """

    openpyxl = _xlsx_module()

    catalogs = list(control_catalogs) + _load_docs_from_paths(
        control_catalog_paths, CRControlCatalog
    )
    attack_catalogs_list = list(attack_catalogs) + _load_docs_from_paths(
        attack_catalog_paths, CRAttackCatalog
    )
    rels = list(control_relationships) + _load_docs_from_paths(
        control_relationship_paths, CRControlRelationships
    )
    attck_rels = list(attack_control_relationships) + _load_docs_from_paths(
        attack_control_relationship_paths, CRAttackControlRelationships
    )

    wb = openpyxl.Workbook()

    # Remove the default sheet.
    if wb.worksheets:
        wb.remove(wb.worksheets[0])

    _write_meta_sheet(wb)
    _write_control_catalogs_sheet(wb, catalogs)
    _write_attack_catalogs_sheet(wb, attack_catalogs_list)
    _write_control_relationships_sheet(wb, rels)
    _write_attack_control_relationships_sheet(wb, attck_rels)

    _apply_workbook_formatting(wb)

    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def _apply_workbook_formatting(wb) -> None:
    """Apply purely-presentational formatting to make sheets nicer to use.

    This intentionally does not change the workbook schema.
    """

    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor="F2F2F2")
    header_font = Font(bold=True)
    header_align = Alignment(vertical="top", wrap_text=True)
    wrap_align = Alignment(vertical="top", wrap_text=True)

    sheet_configs: list[tuple[str, dict[str, Any]]] = [
        (
            _SHEET_META,
            {
                "column_widths": {"format": 22, "version": 12, "created_at": 22, "header_rows": 12},
            },
        ),
        (
            _SHEET_CONTROL_CATALOGS,
            {
                "column_widths": {
                    "doc_name": 22,
                    "framework": 20,
                    "control_id": 18,
                    "title": 32,
                    "url": 34,
                    "tags_json": 22,
                    "defense_in_depth_layers_json": 22,
                },
            },
        ),
        (
            _SHEET_ATTACK_CATALOGS,
            {
                "column_widths": {
                    "doc_name": 22,
                    "framework": 26,
                    "attack_id": 18,
                    "title": 40,
                    "url": 34,
                    "tags_json": 22,
                },
            },
        ),
        (
            _SHEET_CONTROL_RELATIONSHIPS,
            {
                "column_widths": {
                    "doc_name": 22,
                    "source_id": 18,
                    "target_id": 18,
                    "relationship_type": 16,
                    "overlap_weight": 14,
                    "overlap_dimensions_json": 26,
                    "overlap_rationale": 34,
                    "confidence": 12,
                    "groupings_json": 26,
                    "references_json": 26,
                    "description": 34,
                },
                "list_validations": [
                    (
                        "relationship_type",
                        [
                            "overlaps_with",
                            "mitigates",
                            "supports",
                            "equivalent_to",
                            "parent_of",
                            "child_of",
                            "backstops",
                        ],
                    )
                ],
                "number_formats": [("overlap_weight", "0.00"), ("confidence", "0.00")],
            },
        ),
        (
            _SHEET_ATTACK_CONTROL_RELATIONSHIPS,
            {
                "column_widths": {
                    "doc_name": 22,
                    "attack_id": 18,
                    "control_id": 18,
                    "relationship_type": 16,
                    "strength": 12,
                    "confidence": 12,
                    "tags_json": 22,
                    "references_json": 26,
                    "description": 34,
                    "metadata_json": 26,
                },
                "list_validations": [
                    ("relationship_type", ["mitigated_by", "detectable_by", "respondable_by"])
                ],
                "number_formats": [("strength", "0.00"), ("confidence", "0.00")],
            },
        ),
    ]

    for sheet_name, cfg in sheet_configs:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        _xlsx_style_sheet(
            ws,
            header_fill=header_fill,
            header_font=header_font,
            header_align=header_align,
            wrap_align=wrap_align,
            get_column_letter=get_column_letter,
            column_widths=cfg.get("column_widths"),
        )

        for col_name, allowed in (cfg.get("list_validations") or []):
            _xlsx_add_list_validation(ws, column_name=str(col_name), allowed=list(allowed))

        for col_name, fmt in (cfg.get("number_formats") or []):
            _xlsx_number_format(ws, column_name=str(col_name), fmt=str(fmt))


def import_xlsx(source: Union[str, Path, Any]) -> ImportedXlsx:
    """Import CRML documents from an XLSX workbook created by `export_xlsx`.

    Args:
        source:
            Either a file path (str/Path) or an already-loaded openpyxl Workbook.
    """

    openpyxl = _xlsx_module()

    if isinstance(source, (str, Path)):
        wb = openpyxl.load_workbook(str(source))
    else:
        wb = source
    _validate_meta_sheet(wb)

    header_rows = _get_header_rows(wb)

    return ImportedXlsx(
        control_catalogs=_read_control_catalogs_sheet(wb, header_rows=header_rows),
        attack_catalogs=_read_attack_catalogs_sheet(wb, header_rows=header_rows),
        control_relationships=_read_control_relationships_sheet(wb, header_rows=header_rows),
        attack_control_relationships=_read_attack_control_relationships_sheet(wb, header_rows=header_rows),
    )


def _load_docs_from_paths(paths: Iterable[Union[str, Path]], model_cls) -> List[Any]:
    if not paths:
        return []

    from ..yamlio import load_yaml_mapping_from_path

    out = []
    for p in paths:
        data = load_yaml_mapping_from_path(str(p))
        out.append(model_cls.model_validate(data))
    return out


def _write_meta_sheet(wb) -> None:
    ws = wb.create_sheet(_SHEET_META)
    ws.append(["format", "version", "created_at", "header_rows"])
    ws.append([_WORKBOOK_FORMAT, _WORKBOOK_VERSION, _now_iso(), 2])


def _validate_meta_sheet(wb) -> None:
    if _SHEET_META not in wb.sheetnames:
        raise ValueError("XLSX workbook is missing required sheet '_meta'")

    ws = wb[_SHEET_META]
    rows = list(ws.iter_rows(min_row=1, max_row=2, values_only=True))
    if len(rows) < 2:
        raise ValueError("XLSX workbook '_meta' sheet is incomplete")

    header = list(rows[0] or [])
    values = list(rows[1] or [])

    try:
        fmt = values[header.index("format")]
        ver = values[header.index("version")]
    except Exception as e:
        raise ValueError("XLSX workbook '_meta' sheet is malformed") from e

    if fmt != _WORKBOOK_FORMAT:
        raise ValueError(f"Unsupported workbook format: {fmt!r}")
    if ver != _WORKBOOK_VERSION:
        raise ValueError(f"Unsupported workbook version: {ver!r}")


def _get_header_rows(wb) -> int:
    """Number of header rows in mapping sheets.

    Current format always uses 2 header rows:
    - Row 1 (hidden): machine keys
    - Row 2 (visible): human labels
    """

    ws = wb[_SHEET_META]
    rows = list(ws.iter_rows(min_row=1, max_row=2, values_only=True))
    if len(rows) < 2:
        return 1
    header = list(rows[0] or [])
    values = list(rows[1] or [])

    def _get(name: str) -> Any:
        try:
            return values[header.index(name)]
        except Exception:
            return None

    hr = _try_int(_get("header_rows"))
    if hr is None:
        # Treat missing value as current format.
        return 2

    if hr != 2:
        raise ValueError(
            "Unsupported workbook layout: expected _meta.header_rows == 2 for the current format"
        )
    return 2


def _write_human_header(ws, columns: list[tuple[str, str, str]]) -> None:
    """Write a two-row header.

    Row 1: machine keys (hidden)
    Row 2: human labels (visible) + comments with descriptions
    """

    from openpyxl.comments import Comment

    ws.append([c[0] for c in columns])
    ws.append([c[1] for c in columns])

    ws.row_dimensions[1].hidden = True

    for idx, (_, _, desc) in enumerate(columns, start=1):
        if not desc:
            continue
        ws.cell(row=2, column=idx).comment = Comment(desc, "crml")


def _write_control_catalogs_sheet(wb, docs: list[CRControlCatalog]) -> None:
    ws = wb.create_sheet(_SHEET_CONTROL_CATALOGS)
    _write_human_header(
        ws,
        _doc_meta_columns(tags_desc=_DOC_TAGS_DESC_EXAMPLE)
        + [
            ("catalog_id", "Catalog id", "Optional catalog identifier (catalog.id)."),
            ("framework", "Framework", "Framework label (e.g. CIS v8)."),
            ("control_id", "Control id", "Canonical control id (e.g. cisv8:4.2)."),
            ("ref_standard", "Ref standard", "Optional structured ref: standard."),
            ("ref_control", "Ref control", "Optional structured ref: control."),
            ("ref_requirement", "Ref requirement", "Optional structured ref: requirement."),
            ("title", "Title", "Optional short title."),
            ("url", "URL", "Optional URL."),
            ("tags_json", _TAGS_JSON_LABEL, _TAGS_JSON_DESC),
            (
                "defense_in_depth_layers_json",
                "Defense-in-depth layers (JSON)",
                "Optional array as JSON (e.g. [\"prevent\",\"detect\"]).",
            ),
        ],
    )

    for doc in docs:
        meta = doc.meta
        cat = doc.catalog
        for entry in cat.controls:
            ref_standard = ref_control = ref_requirement = None
            if entry.ref is not None:
                ref_standard = entry.ref.standard
                ref_control = entry.ref.control
                ref_requirement = entry.ref.requirement

            ws.append(
                [
                    meta.name,
                    meta.version,
                    meta.description,
                    _to_json_cell(meta.tags),
                    cat.id,
                    cat.framework,
                    entry.id,
                    ref_standard,
                    ref_control,
                    ref_requirement,
                    entry.title,
                    entry.url,
                    _to_json_cell(entry.tags),
                    _to_json_cell(entry.defense_in_depth_layers),
                ]
            )


def _write_attack_catalogs_sheet(wb, docs: list[CRAttackCatalog]) -> None:
    ws = wb.create_sheet(_SHEET_ATTACK_CATALOGS)
    _write_human_header(
        ws,
        _doc_meta_columns(tags_desc=_DOC_TAGS_DESC)
        + [
            ("catalog_id", "Catalog id", "Required catalog identifier / namespace (catalog.id)."),
            ("framework", "Framework", "Framework label (e.g. MITRE ATT&CK Enterprise)."),
            ("attack_id", "Attack id", "Canonical attack id (e.g. attck:T1059.003)."),
            (
                "kind",
                "Kind",
                "Required enum describing the entry kind (e.g. tactic, technique, sub-technique, phase, category, attack-pattern).",
            ),
            ("title", "Title", "Optional short title."),
            ("url", "URL", "Optional URL."),
            ("parent", "Parent", "Optional parent attack id (same namespace)."),
            ("tags_json", _TAGS_JSON_LABEL, _TAGS_JSON_DESC),
            (
                "phases_json",
                "Phases (JSON)",
                "Optional phases array as JSON (recommended list of phase-like ids in the same catalog, e.g. ATT&CK tactic ids).",
            ),
        ],
    )

    for doc in docs:
        meta = doc.meta
        cat = doc.catalog
        for entry in cat.attacks:
            ws.append(
                [
                    meta.name,
                    meta.version,
                    meta.description,
                    _to_json_cell(meta.tags),
                    cat.id,
                    cat.framework,
                    entry.id,
                    entry.kind,
                    entry.title,
                    entry.url,
                    entry.parent,
                    _to_json_cell(entry.tags),
                    _to_json_cell(entry.phases),
                ]
            )


def _read_control_catalogs_sheet(wb, *, header_rows: int) -> list[CRControlCatalog]:
    header, body_rows = _read_sheet_rows(wb, _SHEET_CONTROL_CATALOGS, header_rows=header_rows)
    if not header:
        return []
    out_by_doc: dict[tuple[str, str | None, str | None, str | None, str | None], dict[str, Any]] = {}

    for row in body_rows:
        data = dict(zip(header, row))

        doc_name, doc_version, doc_description, doc_tags = _read_doc_meta(data)
        if not doc_name:
            continue

        catalog_id = _cell_str(data.get("catalog_id"))
        framework = _cell_str(data.get("framework"))
        if not framework:
            raise ValueError(f"control_catalogs row missing framework for doc {doc_name!r}")

        container = _control_catalog_get_or_create_doc(
            out_by_doc=out_by_doc,
            doc_name=doc_name,
            doc_version=doc_version,
            doc_description=doc_description,
            doc_tags=doc_tags,
            catalog_id=catalog_id,
            framework=framework,
        )

        control_id = _cell_str(data.get("control_id"))
        ref_obj = _control_catalog_parse_ref_obj(control_id=control_id, data=data)

        container["catalog"]["controls"].append(
            {
                "id": control_id,
                "ref": ref_obj,
                "title": _cell_str(data.get("title")),
                "url": _cell_str(data.get("url")),
                "tags": _from_json_cell(data.get("tags_json")),
                "defense_in_depth_layers": _from_json_cell(
                    data.get("defense_in_depth_layers_json")
                ),
            }
        )

    return [CRControlCatalog.model_validate(d) for d in out_by_doc.values()]


def _read_attack_catalogs_sheet(wb, *, header_rows: int) -> list[CRAttackCatalog]:
    header, body_rows = _read_sheet_rows(wb, _SHEET_ATTACK_CATALOGS, header_rows=header_rows)
    if not header:
        return []
    out_by_doc: dict[tuple[str, str | None, str | None, str | None, str | None], dict[str, Any]] = {}

    for row in body_rows:
        data = dict(zip(header, row))

        doc_name, doc_version, doc_description, doc_tags = _read_doc_meta(data)
        if not doc_name:
            continue

        catalog_id = _cell_str(data.get("catalog_id"))
        framework = _cell_str(data.get("framework"))
        if not framework:
            raise ValueError(f"attack_catalogs row missing framework for doc {doc_name!r}")
        if not catalog_id:
            raise ValueError(f"attack_catalogs row missing catalog_id for doc {doc_name!r}")

        key = _doc_key(doc_name, doc_version, doc_description, doc_tags, catalog_id)
        container = out_by_doc.get(key)
        if container is None:
            container = {
                "crml_attack_catalog": "1.0",
                "meta": {
                    "name": doc_name,
                    "version": doc_version,
                    "description": doc_description,
                    "tags": doc_tags,
                },
                "catalog": {
                    "id": catalog_id,
                    "framework": framework,
                    "attacks": [],
                },
            }
            out_by_doc[key] = container
        else:
            if container["catalog"].get("framework") != framework:
                raise ValueError(
                    f"attack_catalogs sheet has conflicting frameworks for doc {doc_name!r}: "
                    f"{container['catalog']['framework']!r} vs {framework!r}"
                )

        container["catalog"]["attacks"].append(
            {
                "id": _cell_str(data.get("attack_id")),
                "kind": _cell_str(data.get("kind")),
                "title": _cell_str(data.get("title")),
                "url": _cell_str(data.get("url")),
                "parent": _cell_str(data.get("parent")),
                "tags": _from_json_cell(data.get("tags_json")),
                "phases": _from_json_cell(data.get("phases_json")),
            }
        )

    return [CRAttackCatalog.model_validate(d) for d in out_by_doc.values()]


def _write_control_relationships_sheet(wb, docs: list[CRControlRelationships]) -> None:
    ws = wb.create_sheet(_SHEET_CONTROL_RELATIONSHIPS)
    _write_human_header(
        ws,
        _doc_meta_columns(tags_desc=_DOC_TAGS_DESC)
        + [
            ("pack_id", "Pack id", "Optional relationship pack identifier (relationships.id)."),
            ("source_id", "Source control id", "Source control (scenario/threat-centric)."),
            ("target_id", "Target control id", "Target control (portfolio/implementation-centric)."),
            (
                "relationship_type",
                "Relationship type",
                "Optional enum: overlaps_with, mitigates, supports, equivalent_to, parent_of, child_of, backstops.",
            ),
            ("overlap_weight", "Overlap weight", "Required overlap weight in [0,1]."),
            (
                "overlap_dimensions_json",
                "Overlap dimensions (JSON)",
                "Optional dimension map as JSON (e.g. {\"coverage\":0.9}).",
            ),
            ("overlap_rationale", "Overlap rationale", "Optional rationale/explanation."),
            ("confidence", "Confidence", "Optional confidence in [0,1]."),
            ("groupings_json", "Groupings (JSON)", "Optional groupings array as JSON."),
            ("description", "Description", "Optional free-form description."),
            ("references_json", "References (JSON)", "Optional references array as JSON."),
        ],
    )

    def _row(meta, pack, rel, target) -> list[Any]:
        groupings_json = None
        if target.groupings:
            groupings_json = _to_json_cell([g.model_dump(exclude_none=True) for g in target.groupings])
        references_json = None
        if target.references:
            references_json = _to_json_cell([r.model_dump(exclude_none=True) for r in target.references])

        return [
            meta.name,
            meta.version,
            meta.description,
            _to_json_cell(meta.tags),
            pack.id,
            rel.source,
            target.target,
            target.relationship_type,
            target.overlap.weight,
            _to_json_cell(target.overlap.dimensions),
            target.overlap.rationale,
            target.confidence,
            groupings_json,
            target.description,
            references_json,
        ]

    for doc in docs:
        meta = doc.meta
        pack = doc.relationships
        for rel in pack.relationships:
            for target in rel.targets:
                ws.append(_row(meta, pack, rel, target))


def _read_control_relationships_sheet(
    wb, *, header_rows: int
) -> list[CRControlRelationships]:
    header, body_rows = _read_sheet_rows(wb, _SHEET_CONTROL_RELATIONSHIPS, header_rows=header_rows)
    if not header:
        return []
    out_by_doc: dict[tuple[str, str | None, str | None, str | None, str | None], dict[str, Any]] = {}

    for row in body_rows:
        data = dict(zip(header, row))

        doc_name, doc_version, doc_description, doc_tags = _read_doc_meta(data)
        if not doc_name:
            continue
        pack_id = _cell_str(data.get("pack_id"))

        key = _doc_key(doc_name, doc_version, doc_description, doc_tags, pack_id)
        container = out_by_doc.get(key)
        if container is None:
            container = {
                "crml_control_relationships": "1.0",
                "meta": {
                    "name": doc_name,
                    "version": doc_version,
                    "description": doc_description,
                    "tags": doc_tags,
                },
                "relationships": {
                    "id": pack_id,
                    "relationships": [],
                },
            }
            out_by_doc[key] = container

        source_id = _cell_str(data.get("source_id"))
        target_id = _cell_str(data.get("target_id"))
        if not source_id or not target_id:
            raise ValueError(f"control_relationships row missing source/target in doc {doc_name!r}")

        overlap_weight = _cell_float(data.get("overlap_weight"))
        if overlap_weight is None:
            raise ValueError(
                f"control_relationships row missing overlap_weight for {source_id!r} -> {target_id!r}"
            )

        relationship_type = _cell_str(data.get("relationship_type"))
        overlap_dimensions = _from_json_cell(data.get("overlap_dimensions_json"))
        overlap_rationale = _cell_str(data.get("overlap_rationale"))
        confidence = _cell_float(data.get("confidence"))
        groupings = _from_json_cell(data.get("groupings_json"))
        description = _cell_str(data.get("description"))
        references = _from_json_cell(data.get("references_json"))

        rel_list = container["relationships"]["relationships"]
        grouped = _group_last_or_new(rel_list, key_field="source", key_value=source_id, targets_key="targets")

        grouped["targets"].append(
            {
                "target": target_id,
                "relationship_type": relationship_type,
                "overlap": {
                    "weight": overlap_weight,
                    "dimensions": overlap_dimensions,
                    "rationale": overlap_rationale,
                },
                "confidence": confidence,
                "groupings": groupings,
                "description": description,
                "references": references,
            }
        )

    return [CRControlRelationships.model_validate(d) for d in out_by_doc.values()]


def _write_attack_control_relationships_sheet(
    wb, docs: list[CRAttackControlRelationships]
) -> None:
    ws = wb.create_sheet(_SHEET_ATTACK_CONTROL_RELATIONSHIPS)
    _write_human_header(
        ws,
        _doc_meta_columns(tags_desc=_DOC_TAGS_DESC)
        + [
            ("pack_id", "Pack id", "Optional relationship pack identifier (relationships.id)."),
            ("attack_id", "Attack id", "Attack pattern id (e.g. attck:T1059.003)."),
            ("control_id", "Control id", "Mapped control id (e.g. cap:edr)."),
            (
                "relationship_type",
                "Relationship type",
                "Enum: mitigated_by, detectable_by, respondable_by.",
            ),
            ("strength", "Strength", "Optional strength in [0,1]."),
            ("confidence", "Confidence", "Optional confidence in [0,1]."),
            ("description", "Description", "Optional free-form description."),
            ("tags_json", _TAGS_JSON_LABEL, _TAGS_JSON_DESC),
            ("references_json", "References (JSON)", "Optional references array as JSON."),
            ("metadata_json", "Pack metadata (JSON)", "Optional pack-level metadata map as JSON."),
        ],
    )

    for doc in docs:
        meta = doc.meta
        pack = doc.relationships
        for rel in pack.relationships:
            for tgt in rel.targets:
                ws.append(
                    [
                        meta.name,
                        meta.version,
                        meta.description,
                        _to_json_cell(meta.tags),
                        pack.id,
                        rel.attack,
                        tgt.control,
                        tgt.relationship_type,
                        tgt.strength,
                        tgt.confidence,
                        tgt.description,
                        _to_json_cell(tgt.tags),
                        _to_json_cell(
                            [r.model_dump(exclude_none=True) for r in (tgt.references or [])]
                        )
                        if tgt.references
                        else None,
                        _to_json_cell(pack.metadata),
                    ]
                )


def _read_attack_control_relationships_sheet(
    wb, *, header_rows: int
) -> list[CRAttackControlRelationships]:
    header, body_rows = _read_sheet_rows(wb, _SHEET_ATTACK_CONTROL_RELATIONSHIPS, header_rows=header_rows)
    if not header:
        return []
    out_by_doc: dict[tuple[str, str | None, str | None, str | None, str | None], dict[str, Any]] = {}

    for row in body_rows:
        data = dict(zip(header, row))

        doc_name, doc_version, doc_description, doc_tags = _read_doc_meta(data)
        if not doc_name:
            continue

        pack_id = _cell_str(data.get("pack_id"))
        metadata = _from_json_cell(data.get("metadata_json"))

        key = _doc_key(doc_name, doc_version, doc_description, doc_tags, pack_id)
        container = out_by_doc.get(key)
        if container is None:
            container = {
                "crml_attack_control_relationships": "1.0",
                "meta": {
                    "name": doc_name,
                    "version": doc_version,
                    "description": doc_description,
                    "tags": doc_tags,
                },
                "relationships": {
                    "id": pack_id,
                    "relationships": [],
                    "metadata": metadata,
                },
            }
            out_by_doc[key] = container
        else:
            if container["relationships"].get("metadata") != metadata:
                raise ValueError(
                    f"attack_control_relationships sheet has conflicting metadata for doc {doc_name!r}"
                )

        attack_id = _cell_str(data.get("attack_id"))
        control_id = _cell_str(data.get("control_id"))
        relationship_type = _cell_str(data.get("relationship_type"))
        if not attack_id or not control_id or not relationship_type:
            raise ValueError(
                f"attack_control_relationships row missing required fields in doc {doc_name!r}"
            )

        strength = _cell_float(data.get("strength"))
        confidence = _cell_float(data.get("confidence"))
        description = _cell_str(data.get("description"))
        tags = _from_json_cell(data.get("tags_json"))
        references = _from_json_cell(data.get("references_json"))

        rel_list = container["relationships"]["relationships"]
        grouped = _group_last_or_new(rel_list, key_field="attack", key_value=attack_id, targets_key="targets")

        grouped["targets"].append(
            {
                "control": control_id,
                "relationship_type": relationship_type,
                "strength": strength,
                "confidence": confidence,
                "description": description,
                "references": references,
                "tags": tags,
            }
        )

    return [CRAttackControlRelationships.model_validate(d) for d in out_by_doc.values()]


def write_imported_as_yaml(
    imported: ImportedXlsx, out_dir: str, *, overwrite: bool = False, sort_keys: bool = False
) -> list[str]:
    """Write imported documents to YAML files.

    Returns the list of file paths written.
    """

    from ..yamlio import dump_yaml_to_path

    out = []
    Path(out_dir).mkdir(parents=True, exist_ok=True)

    def _write(doc_type: str, doc_name: str, payload: dict[str, Any]) -> str:
        base = _safe_filename(doc_name)
        path = str(Path(out_dir) / f"{base}-{doc_type}.yaml")
        if not overwrite and Path(path).exists():
            raise FileExistsError(f"Refusing to overwrite existing file: {path}")
        dump_yaml_to_path(payload, path, sort_keys=sort_keys)
        return path

    for doc in imported.control_catalogs:
        out.append(_write("control-catalog", doc.meta.name, doc.model_dump(by_alias=True, exclude_none=True)))
    for doc in imported.attack_catalogs:
        out.append(_write("attack-catalog", doc.meta.name, doc.model_dump(by_alias=True, exclude_none=True)))
    for doc in imported.control_relationships:
        out.append(
            _write(
                "control-relationships",
                doc.meta.name,
                doc.model_dump(by_alias=True, exclude_none=True),
            )
        )
    for doc in imported.attack_control_relationships:
        out.append(
            _write(
                "attack-control-relationships",
                doc.meta.name,
                doc.model_dump(by_alias=True, exclude_none=True),
            )
        )

    return out
