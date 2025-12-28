from openpyxl import Workbook
from pathlib import Path
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment
from openpyxl.drawing.image import Image
from openpyxl.worksheet.table import Table, TableStyleInfo
from rich.progress import Progress

excel_progress = Progress()


class ExcelWriter:
    def __init__(self, condition_types, opt_metrics):
        self.condition_types = condition_types
        self.opt_metrics = opt_metrics

    def write_to_excel(self, output_df, batch_id, figure_output=[], figure_path=None, save_path=None, filetype="xlsx"):
        if filetype == "xlsx":
            # if not all(i in self.condition_types for i in figure_output):
            #     logger.warning("Figure output not in condition types, skipping...")
            #     return

            wb = Workbook()
            ws = self._create_worksheet(wb, batch_id)
            self._add_data_to_worksheet(ws, output_df)

            fixed_length_col = [chr(ord("A") + output_df.columns.get_loc(i)) for i in ["batch", "index", *self.opt_metrics]]
            self._auto_adjust_columns(ws, fixed_length_col)
            self._apply_table_style(ws, output_df)

            if figure_output != [] and figure_path:
                excel_progress.log("exporting with specific figures...", style="green")
                for figure_type in figure_output:
                    column_idx_letter = chr(ord("A") + output_df.columns.get_loc(figure_type))
                    if figure_type in self.condition_types:
                        self._process_figure(ws, figure_type, output_df, figure_path, column_idx_letter)
                    else:
                        excel_progress.log(f"Figure output '{figure_type}' not in condition types, skipping...", style="yellow")
            else:
                excel_progress.log("No figure output and path provided, exporting with names...", style="green")

            wb.save(save_path.with_suffix(".xlsx"))
        else:
            raise ValueError("Unknown filetype")

    def _create_worksheet(self, wb, batch_id):
        ws = wb.active
        ws.title = f"optimization in batch {batch_id}"
        return ws

    def _apply_cell_styles(self, cell, font, alignment):
        cell.font = font
        cell.alignment = alignment

    def _auto_adjust_columns(self, ws, number_col):
        max_width = 0
        for col in ws.columns:
            current_max = max(len(str(cell.value)) for cell in col if cell.value)
            max_width = current_max if current_max > max_width else max_width

        adjusted_width = (max_width + 2) * 1.2
        for col in ws.columns:
            column_letter = col[0].column_letter
            ws.column_dimensions[column_letter].width = adjusted_width

        for n_col in number_col:
            ws.column_dimensions[n_col].width = 16

    def _add_data_to_worksheet(self, ws, output_df):
        font = Font(name="Arial", size=18)
        alignment = Alignment(horizontal="center", vertical="center")

        for i, row in enumerate(dataframe_to_rows(output_df, index=False, header=True)):
            for j, value in enumerate(row):
                cell = ws.cell(row=i + 1, column=j + 1, value=value)
                self._apply_cell_styles(cell, font, alignment)

    def _process_figure(self, ws, figure_type, output_df, figure_path, column_idx_letter):
        ws.row_dimensions[1].height = 50
        for i in output_df.index:
            img_path = Path(figure_path) / Path(f"{figure_type}/{output_df.loc[i, figure_type]}.png")
            if not img_path.exists():
                excel_progress.log(f"{img_path} does not exist, skipping...", style="yellow")
                continue
            ws[f"{column_idx_letter}{i+2}"].value = None
            try:
                img = Image(img_path)
                img.width = int(img.width / 5)
                img.height = int(img.height / 5)

                ws.add_image(img, f"{column_idx_letter}{i+2}")
                ws.row_dimensions[i + 2].height = 0.0 if ws.row_dimensions[i + 2].height == None else ws.row_dimensions[i + 2].height
                ws.row_dimensions[i + 2].height = max(img.height * 0.8, ws.row_dimensions[i + 2].height)
                ws.column_dimensions[column_idx_letter].width = img.width * 0.2
            except Exception as e:
                excel_progress.log(f"Failed to add image for {figure_type} at row {i+2}: {str(e)}", style="red")

    def _apply_table_style(self, ws, output_df):
        max_row = len(output_df) + 1  # +1 for header
        max_col = len(output_df.columns)
        ref = f"A1:{chr(64 + max_col)}{max_row}"  # e.g. "A1:D10"

        table = Table(displayName="OptimizationTable", ref=ref)
        style = TableStyleInfo(
            name="TableStyleLight2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table.tableStyleInfo = style
        ws.add_table(table)
