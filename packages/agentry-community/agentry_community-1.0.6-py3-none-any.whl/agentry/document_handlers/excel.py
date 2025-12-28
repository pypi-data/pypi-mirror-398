from typing import Dict, Any
from .base import BaseDocumentHandler

class ExcelHandler(BaseDocumentHandler):
    """Handler for Excel spreadsheets (.xlsx) using openpyxl."""

    def __init__(self, file_path: str):
        super().__init__(file_path)
        self._wb = None
        self._text = ""
        self._metadata = {}

    def load(self) -> None:
        """Load and parse the XLSX file."""
        try:
            import openpyxl
        except ImportError:
            raise RuntimeError("openpyxl is not installed. Please install it via 'pip install openpyxl' to use ExcelHandler.")

        try:
            self._wb = openpyxl.load_workbook(self.file_path, data_only=True)
            
            text_parts = []
            for sheet_name in self._wb.sheetnames:
                sheet = self._wb[sheet_name]
                text_parts.append(f"--- Sheet: {sheet_name} ---")
                
                # Iterate rows
                for row in sheet.iter_rows(values_only=True):
                    # Filter out None values and convert to string
                    row_data = [str(cell) if cell is not None else "" for cell in row]
                    # Only add non-empty rows
                    if any(row_data):
                        text_parts.append("\t".join(row_data))
                
                text_parts.append("") # Spacer between sheets

            self._text = "\n".join(text_parts)
            
            # Extract properties
            props = self._wb.properties
            self._metadata = {
                "author": props.creator,
                "created": str(props.created),
                "modified": str(props.modified),
                "title": props.title,
                "subject": props.subject,
                "keywords": props.keywords,
                "category": props.category,
                "sheet_names": self._wb.sheetnames
            }
             # Remove None values
            self._metadata = {k: v for k, v in self._metadata.items() if v is not None}

        except Exception as e:
            raise RuntimeError(f"Failed to load Excel file {self.file_path}: {e}")

    def get_text(self) -> str:
        if self._wb is None:
            self.load()
        return self._text

    def get_metadata(self) -> Dict[str, Any]:
        if self._wb is None:
            self.load()
        return self._metadata
