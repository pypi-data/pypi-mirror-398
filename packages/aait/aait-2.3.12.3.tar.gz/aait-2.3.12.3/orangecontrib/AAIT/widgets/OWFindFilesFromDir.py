import Orange.data
from AnyQt.QtWidgets import QApplication, QLineEdit, QCheckBox,QSpinBox, QPushButton
from Orange.data import StringVariable, Table, Domain
from Orange.widgets.utils.signals import Input, Output
from Orange.widgets.settings import Setting
from datetime import datetime, timedelta
import openpyxl
import os
import sys
import shutil
from openpyxl import Workbook, load_workbook

if "site-packages/Orange/widgets" in os.path.dirname(os.path.abspath(__file__)).replace("\\", "/"):
    from Orange.widgets.orangecontrib.AAIT.utils import base_widget
    from Orange.widgets.orangecontrib.AAIT.utils.initialize_from_ini import apply_modification_from_python_file
else:
    from orangecontrib.AAIT.utils import base_widget
    from orangecontrib.AAIT.utils.initialize_from_ini import apply_modification_from_python_file


@apply_modification_from_python_file(filepath_original_widget=__file__)
class OWFileFromDir(base_widget.BaseListWidget):
    name = "Find Files From Dir"
    description = ("Search files by extension or no for all files in a directory or subdirectories.")
    category = "AAIT - TOOLBOX"
    icon = "icons/owfilesfromdir.svg"
    if "site-packages/Orange/widgets" in os.path.dirname(os.path.abspath(__file__)).replace("\\", "/"):
        icon = "icons_dev/owfilesfromdir.svg"
    gui = os.path.join(os.path.dirname(os.path.abspath(__file__)), "designer/owfindfilesfromdir.ui")
    want_control_area = False
    priority = 1060
    selected_column_name = Setting("input_dir")
    extension = Setting("")
    recursive = Setting("False")
    filter_per_date = Setting("False")
    time_filter= Setting("0-0-0-1")
    identify_sheets = Setting("False")

    class Inputs:
        data = Input("Data", Orange.data.Table)

    class Outputs:
        data = Output("Data", Orange.data.Table)


    @Inputs.data
    def set_path_table(self, in_data):
        self.data = in_data
        if in_data is None:
            self.Outputs.data.send(None)
            return
        if self.data:
            self.var_selector.add_variables(self.data.domain)
            self.var_selector.select_variable_by_name(self.selected_column_name)
        self.run()


    def __init__(self):
        super().__init__()
        # Qt Management
        self.setFixedWidth(500)
        self.setFixedHeight(620)
        if self.is_dash_int_castable(self.time_filter)!=0:
            self.time_filter = "0-0-0-1"

        self.edit_extension = self.findChild(QLineEdit, 'lineEdit')
        self.edit_extension.setPlaceholderText("Extension (.docx, .pdf, .xslx, .csv, .json ...)")
        self.edit_extension.setText(self.extension)
        self.edit_extension.editingFinished.connect(self.update_parameters)
        self.comboBox = self.findChild(QCheckBox, 'checkBox')
        self.checkBox_2 = self.findChild(QCheckBox, 'checkBox_2')
        self.spinBox_day = self.findChild(QSpinBox, 'spinBox_day')
        self.spinBox_hour = self.findChild(QSpinBox, 'spinBox_hour')
        self.spinBox_minute = self.findChild(QSpinBox, 'spinBox_minute')
        self.spinBox_second = self.findChild(QSpinBox, 'spinBox_second')
        self.checkBox_sheets = self.findChild(QCheckBox, 'checkBox_sheets')

        DD, JJ, SS, MM = self._parse_delta_4_numbers(self.time_filter)
        self.spinBox_day.setValue(int(DD))
        self.spinBox_hour.setValue(int(JJ))
        self.spinBox_minute.setValue(int(SS))
        self.spinBox_second.setValue(int(MM))
        # Data Management
        self.folderpath = None
        self.data = None
        self.temp_dir = None
        self.autorun = True
        self.post_initialized()

        if self.filter_per_date=="False":
            self.checkBox_2.setChecked(False)
            self.spinBox_day.setVisible(False)
            self.spinBox_hour.setVisible(False)
            self.spinBox_minute.setVisible(False)
            self.spinBox_second.setVisible(False)
        else:
            self.checkBox_2.setChecked(True)
            self.spinBox_day.setVisible(True)
            self.spinBox_hour.setVisible(True)
            self.spinBox_minute.setVisible(True)
            self.spinBox_second.setVisible(True)

        if self.recursive == "True":
            self.comboBox.setChecked(True)

        if self.identify_sheets == "True":
            self.checkBox_sheets.setChecked(True)
        else:
            self.checkBox_sheets.setChecked(False)

        self.comboBox.stateChanged.connect(self.on_checkbox_toggled)
        self.checkBox_2.stateChanged.connect(self.on_checkBox_2_toggled)
        self.spinBox_day.valueChanged.connect(self.on_value_changed)
        self.spinBox_hour.valueChanged.connect(self.on_value_changed)
        self.spinBox_minute.valueChanged.connect(self.on_value_changed)
        self.spinBox_second.valueChanged.connect(self.on_value_changed)
        self.checkBox_sheets.stateChanged.connect(self.on_checkBox_sheets_toggled)

        self.pushButton_run = self.findChild(QPushButton, 'pushButton_send')
        self.pushButton_run.clicked.connect(self.run)

    def on_value_changed(self):
        self.time_filter = str(self.spinBox_day.value()) + "-" + str(self.spinBox_hour.value()) + "-" + str(
            self.spinBox_minute.value()) + "-" + str(self.spinBox_second.value())

    def update_parameters(self):
        self.extension = (self.edit_extension.text() or "").strip()  # a jout de la gestion d'une zone vide
        if self.folderpath is not None:
            self.run()

    def on_checkbox_toggled(self, state):
        self.recursive = "True"
        if state == 0:
            self.recursive = "False"
        if self.folderpath is not None:
            self.run()

    def on_checkBox_2_toggled(self, state):
        self.filter_per_date = "True"
        if state == 0:
            self.filter_per_date = "False"

        if self.filter_per_date == "False":
            self.spinBox_day.setVisible(False)
            self.spinBox_hour.setVisible(False)
            self.spinBox_minute.setVisible(False)
            self.spinBox_second.setVisible(False)
        else:
            self.spinBox_day.setVisible(True)
            self.spinBox_hour.setVisible(True)
            self.spinBox_minute.setVisible(True)
            self.spinBox_second.setVisible(True)

        if self.folderpath is not None:
            self.run()

    def on_checkBox_sheets_toggled(self, state):
        self.identify_sheets = "True"
        if state == 0:
            self.identify_sheets = "False"
        if self.folderpath is not None:
            self.run()

    def split_excel_sheets(self, filepath):
        """
        Découpe un fichier Excel.
        Crée un dossier [NomFichier]_sheets dans le même répertoire que le fichier.
        Y dépose un fichier .xlsx par feuille.
        """
        new_paths = []
        try:
            source_dir = os.path.dirname(filepath)
            filename_no_ext = os.path.splitext(os.path.basename(filepath))[0]

            # Nom du dossier : ex "MonFichier_sheets"
            output_folder_name = f"{filename_no_ext}_sheets"
            output_folder_path = os.path.join(source_dir, output_folder_name)

            # Création du dossier (exist_ok=True évite de planter s'il existe déjà)
            os.makedirs(output_folder_path, exist_ok=True)

            # 2. Lire le fichier source avec Openpyxl
            wb = load_workbook(filepath, data_only=True)

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

                safe_sheet = "".join(c if c.isalnum() else "_" for c in sheet_name)
                target_filename = f"{safe_sheet}.xlsx"
                target_path = os.path.join(output_folder_path, target_filename).replace("\\", "/")

                new_wb = Workbook()
                new_ws = new_wb.active
                new_ws.title = sheet_name

                for row in ws.iter_rows():
                    new_ws.append([
                        "" if cell.value is None else str(cell.value)
                        for cell in row
                    ])

                new_wb.save(target_path)
                new_paths.append(target_path)

            wb.close()

        except Exception as e:
            # En cas d'échec (ex: fichier ouvert, permissions...), on log et on renvoie l'original
            print(f"Erreur lors de l'extraction des feuilles pour {filepath} : {e}")
            return [filepath]

        return new_paths

    def find_files(self):
        files_data = []
        suffixes = self.parse_extensions()
        excel_suffixes = (".xlsx", ".xlsm", ".xls")

        # Nettoyage du dossier temp précédent si nouvelle recherche
        # (Note : ce bloc est conservé tel quel, même si on utilise maintenant des dossiers locaux)
        if self.identify_sheets == "True" and self.temp_dir and os.path.exists(self.temp_dir):
            try:
                shutil.rmtree(self.temp_dir)
                self.temp_dir = None
            except:
                pass

        for i in range(len(self.folderpath)):
            base = self.folderpath[i]
            if self.recursive == "True":
                traversal = os.walk(base)
            else:
                try:
                    traversal = [(base, [], os.listdir(base))]
                except Exception:
                    continue

            for root, _, files in traversal:
                for file in files:
                    name = file.lower()
                    full_path = os.path.join(root, file).replace("\\", "/")

                    if suffixes is None or name.endswith(suffixes):
                        is_excel = any(name.endswith(ext) for ext in excel_suffixes)

                        if self.identify_sheets == "True" and is_excel:
                            filename_no_ext = os.path.splitext(file)[0]
                            expected_folder_name = f"{filename_no_ext}_sheets"
                            expected_folder_path = os.path.join(root, expected_folder_name)

                            if os.path.exists(expected_folder_path) and os.path.isdir(expected_folder_path):
                                found_sheets = False
                                try:
                                    existing_files = [
                                        os.path.join(expected_folder_path, f).replace("\\", "/")
                                        for f in os.listdir(expected_folder_path)
                                        if f.endswith(".xlsx") and not f.startswith("~$")
                                    ]
                                    if existing_files:
                                        for ex_path in existing_files:
                                            files_data.append([ex_path])
                                        found_sheets = True
                                except OSError:
                                    pass
                                # Si le dossier existait mais était vide
                                if not found_sheets:
                                    split_files = self.split_excel_sheets(full_path)
                                    for split_path in split_files:
                                        files_data.append([split_path])
                            # 3. Le dossier n'existe pas, on lance le découpage
                            else:
                                split_files = self.split_excel_sheets(full_path)
                                for split_path in split_files:
                                    files_data.append([split_path])
                        else:
                            files_data.append([full_path])
        return files_data

    def parse_extensions(self):
        """
        Convertit la saisie utilisateur en tuple de suffixes normalisés pour endswith.
        Exemples d'entrées valides :
          - ".pdf, .docx"
          - "pdf docx"
          - "csv"
          - "" (vide => aucune filtration, donc toutes extensions)
        """
        raw = (self.extension or "").strip().lower()
        if not raw:
            return None  # pas de filtre => tout passe

        # accepte virgules ou espaces multiples comme séparateurs
        parts = [p.strip() for chunk in raw.split(",") for p in chunk.split()]
        # normalise en ajoutant le point s'il manque, ignore les vides
        cleaned = []
        for p in parts:
            if not p:
                continue
            if not p.startswith("."):
                p = "." + p
            cleaned.append(p)

        if not cleaned:
            return None
        return tuple(set(cleaned))  # tuple unique pour endswith(...)

    def is_dash_int_castable(self,value: str) -> int:
        """
        Retourne 0 si tous les morceaux séparés par '-' sont castables en int, 1 sinon.
        Exemple :
          "1-0-0-1" -> 0
          "10-5-3-2" -> 0
          "1-a-0-1" -> 1
        """
        try:
            parts = value.split("-")
            if len(parts)!=4:
                return 1

            for p in parts:
                int(p)  # essaie de caster chaque partie
            return 0
        except (ValueError, TypeError):
            return 1

    def _parse_delta_4_numbers(self,delta_str):
        return(delta_str.split("-"))



    def _parse_delta(self,delta_str):
        """
        Accepte soit:
          - 'AA-MM-JJ-HH-mm-SS' (années, mois, jours, heures, minutes, secondes)
          - 'JJJJJ-HH-mm-SS'   (jours, heures, minutes, secondes)
        Les valeurs peuvent être non paddées (ex: '1-0-0-1').
        """
        parts = delta_str.split("-")
        try:
            nums = [int(p) for p in parts]
        except Exception as e:
            raise ValueError(
                "Delta invalide '{}'. Parties non entières.".format(delta_str)
            ) from e

        if len(nums) == 6:
            aa, mm, jj, hh, mi, ss = nums
            days = aa * 365 + mm * 30 + jj  # approximation mois=30j, année=365j
            return timedelta(days=days, hours=hh, minutes=mi, seconds=ss)
        elif len(nums) == 4:
            jj, hh, mi, ss = nums
            return timedelta(days=jj, hours=hh, minutes=mi, seconds=ss)
        else:
            raise ValueError(
                "Delta invalide '{}'. Formats acceptés: "
                "AA-MM-JJ-HH-mm-SS OU JJJJJ-HH-mm-SS.".format(delta_str)
            )

    def _created_or_modified_time(self,path):
        """
        Retourne l'instant (UTC) le plus récent entre création (si dispo) et modification.
        - Windows: st_ctime ~ création
        - Unix: st_ctime = change time; on utilise st_birthtime si disponible.
        """
        st = os.stat(path)
        mtime = st.st_mtime
        birth = getattr(st, "st_birthtime", None)
        best_epoch = max(mtime, birth) if birth is not None else mtime
        return datetime.utcfromtimestamp(best_epoch)

    def filter_files_newer_than_delta(self,
            files,
            delta_str,
            now=None,
            skip_missing=True,
    ):
        """
        Renvoie les fichiers créés/modifiés depuis moins que le delta donné.

        :param files: liste de chemins absolus OU chaîne avec chemins séparés.
        :param delta_str: 'AA-MM-JJ-HH-mm-SS' OU 'JJJJJ-HH-mm-SS' (sans padding requis).
        :param now: (optionnel) datetime de référence (UTC). Par défaut: datetime.utcnow().
        :param skip_missing: True -> ignore fichiers manquants, False -> lève FileNotFoundError.
        :return: liste de chemins satisfaisant (created/modified) >= now - delta
        """

        delta = self._parse_delta(delta_str)
        ref = now or datetime.utcnow()
        threshold = ref - delta
        results = []
        for pp in files:
            p=pp[0]

            if not os.path.isabs(p):
                # Chemins absolus attendus; on n'empêche pas pour autant.
                pass
            if not os.path.exists(p):
                if skip_missing:
                    continue
                raise FileNotFoundError(p)
            try:

                t = self._created_or_modified_time(p)
            except Exception:
                if not skip_missing:
                    raise
                continue
            if t >= threshold:
                results.append([p])
        return results

    def run(self):
        self.error("")
        self.warning("")
        if self.data is None:
            self.Outputs.data.send(None)
            return

        if not self.selected_column_name in self.data.domain:
            self.warning(f'Previously selected column "{self.selected_column_name}" does not exist in your data.')
            return

        self.folderpath = self.data.get_column(self.selected_column_name)

        try:
            files_data = self.find_files()

            if self.filter_per_date !="False":
                files_data=self.filter_files_newer_than_delta(files_data,self.time_filter)

            if len(files_data) == 0:
                self.Outputs.data.send(None)
                return
            X = [[] for _ in files_data]
            domain = Domain([], metas=[StringVariable("path")])
            table = Table.from_numpy(domain, X, metas=files_data)
            self.Outputs.data.send(table)
        except Exception as e:
            self.error(f"An error occurred: the provided file path may not be supported ({e})")
            return

    def post_initialized(self):
        pass

    def get_sheet_names_from_file(self, filepath: str):
        """
        Méthode simplifiée pour récupérer la liste des noms des feuilles
        en utilisant directement les bibliothèques sous-jacentes (openpyxl/xlrd).
        """
        # Récupérer l'extension en minuscules
        extension = os.path.splitext(filepath)[1].lower()
        sheets = []

        try:
            if extension in ('.xlsx', '.xlsm'):
                workbook = openpyxl.load_workbook(filepath, read_only=True, data_only=False)
                sheets = workbook.sheetnames
                workbook.close()
            else:
                print(f"INFO: Extension {extension} non supportée pour l'extraction de feuilles.")

        except ImportError as e:
            self.warning(
                f"Librairie manquante pour lire les feuilles : {e}. Assurez-vous d'avoir 'openpyxl'")
        except Exception as e:
            self.error(f"Impossible de lire les feuilles du fichier {os.path.basename(filepath)} : {e}")
        return sheets

if __name__ == "__main__":
    app = QApplication(sys.argv)
    my_widget = OWFileFromDir()
    my_widget.show()

    if hasattr(app, "exec"):
        app.exec()
    else:
        app.exec_()