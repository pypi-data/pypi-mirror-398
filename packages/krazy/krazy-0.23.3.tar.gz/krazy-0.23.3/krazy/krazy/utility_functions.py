#List of useful pandas dataframe commands

import pandas as pd
from openpyxl import load_workbook
import glob
import re
from io import StringIO
import os
import datetime
import csv
import pytesseract
from PIL import Image
import pytz
import time
import requests
from calendar import c
import numpy as np
from pathlib import Path
import chardet
from loguru import logger
import sys
from typing import List, Dict, Union, Optional

def files_identifier(
    folder_path: Union[str, Path],
    srch_pattern: str = '.*',
    excl_pattern: Optional[str] = None
) -> List[Path]:
    """
    Identifies files in a folder based on include/exclude regex patterns.

    Args:
        folder_path (str | Path): Directory path as a string or Path object.
        srch_pattern (str): Regex pattern to match files to include. Defaults to '.*' (all files).
        excl_pattern (str, optional): Regex pattern to exclude matching files. Defaults to None.

    Returns:
        List[Path]: List of file paths that match the criteria.
    """
    folder_path = Path(folder_path)

    if not folder_path.is_dir():
        raise ValueError(f"Provided path '{folder_path}' is not a valid directory.")

    files = [
        f for f in folder_path.iterdir()
        if f.is_file() and re.match(srch_pattern, f.name.lower())
    ]

    if excl_pattern:
        files = [f for f in files if not re.match(excl_pattern, f.name.lower())]

    return files


def detect_encoding(file_path: Path) -> str:
    """
    Detects the character encoding of a given file.

    Args:
        file_path (Path): Path to the file whose encoding is to be detected.

    Returns:
        str: Detected character encoding (e.g., 'utf-8', 'ISO-8859-1').
    """
    with open(file_path, 'rb') as f:
        result = chardet.detect(f.read())
    return result['encoding']


def detect_delimiter(file_path) -> List[str]:
    """
    Detects the delimiter and encoding of a text-based delimited file (e.g., CSV, TSV).

    Args:
        file_path (Path or str): Path to the file.

    Returns:
        List[str]: A list containing the detected delimiter and encoding.
                   Format: [delimiter, encoding]
    """
    encoding = detect_encoding(file_path)

    with open(file_path, 'r', encoding=encoding) as f:
        try:
            sample = f.read(5000)  # Read a small portion for detection
            sniffer = csv.Sniffer()
            detected_delimiter = sniffer.sniff(sample).delimiter
        except Exception:
            # Fallback strategy using first line analysis
            f.seek(0)
            first_line = f.readline()
            possible_delimiters = [',', ';', '\t', '|', ' ']
            delimiter_counts = {delim: first_line.count(delim) for delim in possible_delimiters}
            detected_delimiter = max(delimiter_counts, key=delimiter_counts.get)

    return [detected_delimiter, encoding]


def read_spreadsheet_single(
    full_path: Path,
    folder: Optional[str] = None,
    skip_rows: int = 0,
    low_memory: bool = True,
    dtypes: Optional[Dict[str, Union[str, type]]] = None
) -> Optional[pd.DataFrame]:
    """
    Reads a single spreadsheet file (CSV, XLSX, or XLS) into a pandas DataFrame.

    Args:
        full_path (Path): Full path to the file to be read.
        folder (str, optional): Source folder label to include as a column. Defaults to None.
        skip_rows (int, optional): Number of rows to skip from top of file. Defaults to 0.
        low_memory (bool, optional): Whether to use low_memory mode for CSVs. Defaults to True.
        dtypes (dict, optional): Dictionary specifying column data types. Defaults to None.

    Returns:
        pd.DataFrame or None: Parsed DataFrame with `folder` and `file_name` columns if successful, else None.
    """
    dtypes = dtypes or {}
    suffix = full_path.suffix.lower()

    try:
        if suffix == ".csv":
            delimiter, encoding = detect_delimiter(full_path)
            df = pd.read_csv(
                full_path,
                skiprows=skip_rows,
                low_memory=low_memory,
                dtype=dtypes,
                encoding=encoding,
                delimiter=delimiter
            )
        elif suffix == ".xlsx":
            df = pd.read_excel(full_path, skiprows=skip_rows, dtype=dtypes)
        elif suffix == ".xls":
            df = pd.read_excel(full_path, skiprows=skip_rows, dtype=dtypes, engine='xlrd')
        else:
            print(f"❌ Unsupported file extension: {full_path.name}")
            return None

        if folder:
            df["folder"] = folder
        df["file_name"] = full_path.name
        return df

    except Exception as e:
        print(f"❌ Failed to read file: {full_path.name} | Error: {e}")
        return None


def read_spreadsheet_multi(folder_path: Path, sheet_name: str=None, srch_pattern:str=None, excl_pattern:str=None, dtypes:dict=None, skip_rows:int=0, folder:str = None)->pd.DataFrame:
    # get files matching the pattern
    filenames = files_identifier(folder_path, srch_pattern, excl_pattern)
    df_consol = pd.DataFrame()
    # import file
    for item in filenames:
        df_temp = read_spreadsheet_single(item, low_memory=False, dtypes=dtypes, skip_rows=skip_rows)
        df_consol = pd.concat([df_consol, df_temp], ignore_index=True)
    
    # add folder name
    if folder is not None:
        df_consol['folder'] = folder
    
    return df_consol

def remove_dots(df_purge:pd.DataFrame, char_to_remove:str='.')->pd.DataFrame:
    '''Remove specified character from column names. Default is dot'''
    dot_cols = df_col_search(df_purge, f'{char_to_remove}')
    for col in dot_cols:
        renamed_col = col.replace(f'{char_to_remove}','')
        df_purge.rename(columns={col:renamed_col}, inplace=True)
    
    return df_purge

def df_find_index(dfdata,col_name):
    '''
    Finds index number of column in a dataframe
    '''
    colindex = dfdata.columns.get_loc(col_name) #finds index number of col
    return colindex

def df_seq(dfdata, colseq):
    '''
    Changes sequence of columns and returns dataframe
    '''
    dfdata = dfdata.reindex(columns=colseq) #changes sequence of columns

#example of creating dataframe using dictionery
def data_for_df():
    '''
    Returns a dataframe with sample data
    '''
    data = {
    "Duration":{
        "0":60,
        "1":60,
        "2":60,
        "3":45,
        "4":45,
        "5":60
    },
    "Pulse":{
        "0":110,
        "1":117,
        "2":103,
        "3":109,
        "4":117,
        "5":102
    },
    "Maxpulse":{
        "0":130,
        "1":145,
        "2":135,
        "3":175,
        "4":148,
        "5":127
    },
    "Calories":{
        "0":409,
        "1":479,
        "2":340,
        "3":282,
        "4":406,
        "5":300
    }
    }
    df = pd.DataFrame(data)
    return df

def df_to_sql(dfdata,table,colList,connection):
    # this inserts data from dataframe into database table
    # for this to work, columns and column sequence should be exactly in same manner as in database
    cur = connection.cursor()
    imFile = StringIO()
    dfdata.to_csv(imFile, index=False)
    imFile.seek(0)
    next(imFile)
    cur.copy_from(imFile,table, sep=',',columns=colList)
    imFile.close()
    connection.commit()
    cur.close()
    connection.close()

def df_compare(df1, df2, whichStr, colStr):
    if whichStr == '':
        comparisondf = df1.merge(df2, indicator = True, how = 'outer', on= colStr, validate = '1:1')
    else:
        comparisondf = df1.merge(df2, indicator = 'Exist', how = whichStr, left_on='InvNo', right_on='InvNo', validate = '1:1')
    return comparisondf

def df_group(dfdata, grColList, colGroupDict):
    dfdata = dfdata.groupby(grColList).agg(colGroupDict)
    return dfdata

def get_file_details(file_path):
    '''Gets file details for the given file. Created time, modified time, size in MB'''

    try:
        # creation date and time
        ct = os.path.getctime(file_path)
        if ct not in [None,'']:
            ct = time.ctime(ct)
            ct = datetime.datetime.strptime(ct,'%a %b %d %H:%M:%S %Y')
            ctstr = datetime.datetime.strftime(ct,'%d/%m/%Y')
        
        # modified date and time
        mt = os.path.getmtime(file_path)
        if mt not in [None,'']:
            mt = time.ctime(mt)
            mt = datetime.datetime.strptime(mt,'%a %b %d %H:%M:%S %Y')
            mtstr = datetime.datetime.strftime(mt,'%d/%m/%Y')

        # size in MB
        sz = os.path.getsize(file_path)
        if sz not in [None,'']:
            sz = round(sz/1024/1024,2)

    except:
        ct = None
        ctstr = None
        mt = None
        mtstr = None
        sz = None

    return [ct,mt,sz]

def get_full_file_list_df(dirName:str)->pd.DataFrame:
    ''' gets full file and directory tree under given directory '''

    dir_list = []
    dir_list.append(dirName)
    dffiles = pd.DataFrame({'Path':[], 'File':[], 'Creation':[], 'Modified': [], 'Size (MB)': []})

    for dir in dir_list:
        dir_content_list = os.listdir(dir)
        for entry in dir_content_list:
            full_path = os.path.join(dir,entry)
            if os.path.isdir(full_path):
                dir_list.append(full_path)
            else:
                file_details = get_file_details(full_path)
                dffiles.loc[len(dffiles)] = [dir, entry, file_details[0][0], file_details[1][0],
                    file_details[2]]
    return dffiles

def write_To_csv(header_tpl, data_tpl, filePath):
    outFile = open(filePath)
    writer = csv.writer(outFile)
    writer.writerow(header_tpl)
    writer.writerows(data_tpl)
    outFile.close()

def read_from_csv(filePath):
    inFile = open(filePath)
    reader = csv.reader(inFile)
    dataList = []
    for row in reader:
        dataList.append(row)
    inFile.close()
    return dataList

class date_converter:
    #takes string in dd/mm/yyyy or dd/mm/yy format, checks data validity and converts into
    # proper date in final_date attribute
    # this can accept either / or - as date separator
    '''following attributes are available:
    .final_date = gives date converted into proper date format
    .validity = True / False showing valid input
    .type = '/' or '-' or '' showing date separator used in input
    .manual_mapped_vals = values which were not auto mapped but mapped based on d/m/y sequence
    .manual_mapped_keys = keys which were not auto mapped but mapped based on d/m/y sequence            

    '''
    def __init__(self,date_string):
        self.date_string = date_string.strip(' ')
        self.__is_valid()
        if self.__is_valid:
            self.__convert_date()

    def __is_valid(self):

        # initializing variables
        slash_type = None
        hyphen_type = None
        first = ''
        second = ''
        third = ''

        # checking if / is used as date separator
        slash_type = self.date_string.count('/')
        if slash_type == 2:
            num_check = len(self.date_string.replace('/',''))
            first, second, third = self.date_string.split('/')

        # checking if - is used as date separator
        hyphen_type = self.date_string.count('-')
        if hyphen_type == 2:
            num_check = len(self.date_string.replace('-',''))
            first, second, third = self.date_string.split('-')

        #creating final validity
        self.validity = False
        
        if len(first) != 0 and len(second) != 0 and len(third) != 0:
            if (slash_type == 2 or hyphen_type ==2) and (4 <= num_check <= 8):
                if len(first) < 5 and len(second) < 5 and len(third) < 5:
                    if 2 < len(first) < 5:
                        if len(second)< 3 and len(third) < 3:
                            if int(second) < 32 and int(third) < 13:
                                self.validity = True
                            elif int(third) < 32 and int(second) < 13:
                                self.validity = True
                    elif 2 < len(second) < 5 :
                        if len(first) < 3 and len(third) < 3:
                            if int(first) < 32 and int(third) < 13:
                                self.validity = True
                            if int(third) < 32 and int(first) < 13:
                                self.validity == True
                    elif 2 < len(third) < 5:
                        if len(first) < 3 and len(second) < 3:
                            if int(first) < 32 and int(second) < 13:
                                self.validity = True
                            if int(second) < 32 and int(first) < 13:
                                self.validity = True

        if self.validity == True:
            if slash_type == 2:
                self.type = '/'
            if hyphen_type == 2:
                self.type = '-'
        else:
            self.type = ''


    def __convert_date(self):
        if self.validity == True:

            check = False
            if self.type == '/':
                first, second, third = self.date_string.split('/')
            else:
                first, second, third = self.date_string.split('-')
            first = int(first)
            second = int(second)
            third = int(third)
            
            # initializing working variables
            val_unmapped = [first, second, third]
            mapping = {'year': None, 'month': None, 'day': None}
            keys_unmapped = ['day','month','year']

            # getting year with 4 digits
            if len(str(first)) == 4 or first > 31:
                mapping.update(year = first,)
                val_unmapped.remove(first)
                keys_unmapped.remove('year')
            elif len(str(second))== 4 or second > 31:
                mapping.update(year = second,)
                val_unmapped.remove(second)
                keys_unmapped.remove('year')
            elif len(str(third))== 4 or third > 31:
                mapping.update(year = third,)
                val_unmapped.remove(third)
                keys_unmapped.remove('year')

            # getting day > 31            
            if 12 < first < 31:
                if first in val_unmapped:
                    mapping.update(day = first,)
                    val_unmapped.remove(first)
                    keys_unmapped.remove('day')
            elif 12 < second < 31:
                if second in val_unmapped:
                    mapping.update(day = second,)
                    val_unmapped.remove(second)
                    keys_unmapped.remove('day')
            elif 12 < third < 31:
                if third in val_unmapped:
                    mapping.update(day = third,)
                    val_unmapped.remove(third)
                    keys_unmapped.remove('day')
            
            # mapping remaining in entered sequence
            ind = 0
            for elem in keys_unmapped:
                tmp = {elem:val_unmapped[ind]}
                ind += 1
                mapping.update(tmp)
            
            if len(str(mapping['year'])) == 3:
                mapping['year'] = int(('2' + str(mapping['year'])))
            if len(str(mapping['year'])) == 2:
                mapping['year'] = int(('20' + str(mapping['year'])))
            if len(str(mapping['year'])) == 1:
                mapping['year'] = int(('200' + str(mapping['year'])))

            # creating date in date format
            self.final_date = datetime.datetime(mapping['year'],mapping['month'],mapping['day'])

            self.manual_mapped_vals = val_unmapped
            self.manual_mapped_keys = keys_unmapped
        else:
            self.final_date = None

def number_word_converter(amount:int) -> str:
    dict1 = {'0':'','1': ' One', '2': ' Two', '3': ' Three','4':' Four','5':' Five','6':' Six',
        '7':' Seven','8':' Eight','9':' Nine','10':' Ten', '11':' Eleven','12':' Twelve','13':' Thirteen',
        '14':' Forteen','15':' Fifteen','16':' Sixteen','17':' Seventeen','18':' Eighteen',
        '19':' Nineteen','20':' Twenty'}
    dict2 = {'0':'','1':' One','2':' Twenty','3':' Thirty','4':' Forty','5':' Fifty','6':' Sixty',
        '7':' Seventy','8':' Eighty','9':' Ninety'}
    dict_place = {0:'',1:'',2:' Hundred',3:' Thousand',4:'',5:' Lac',6:'',7:' Crore',8:'',
        9:' Hundred',10:'',11:' Thousand',12:'',13:' Lac'}
    amt_str = ''
    amount = str(amount)
    al = [x for x in reversed(amount)]
    
    max_digits = len(al) - 1

    for counter in range(max_digits+1):

        if counter == 2 and int(al[counter]) != 0:
            amt_str = dict_place[counter] + amt_str
        elif counter in [3,5,7,9]:
            if counter == max_digits:
                if int(al[counter]) != 0:
                    amt_str = dict_place[counter] + amt_str
            else:
                if (int(al[counter]) + int(al[counter+1])) != 0:
                    amt_str = dict_place[counter] + amt_str
        else:
            pass

        if counter in [0,3,5,7,9]:
            if counter == max_digits:
                amt_str = dict1[al[counter]] + amt_str
            elif al[counter+1] == '1':
                pass
            else:
                amt_str = dict1[al[counter]] + amt_str
        elif counter in [1,4,6,8,10]:
            if al[counter] == '1':
                val = al[counter] + al[(counter-1)]
                amt_str = dict1[val] + amt_str
            else:
                amt_str = dict2[al[counter]] + amt_str
        else:
            amt_str = dict1[al[counter]] + amt_str
    amt_str = amt_str.strip()
    return amt_str

def payment_ocr(img_path):
    '''
    ocr on google pay and paytm screenshots
    returns following:
    1. if successfule: [success, template, paid_to, paid_on, amount, upi_id]
    2. if image file not found: [FNF]
    3. if template not found: [TNF]
    '''
    # ocr on google pay and paytm screenshots
    # returns success, template, paid_to, paid_on, amount, upi_id if successful else returns None

    paid_to = None
    paid_on = None
    amount = None
    upi_id = None
    template = None
    success = False
    tz = pytz.timezone('Asia/Kolkata')

    if os.path.exists(img_path):

        # doing OCR and initializing text
        text = pytesseract.image_to_string(Image.open(img_path))
        data = []
        text_lower = text.lower()
        data = text.split('\n')
        data_lower = text_lower.split('\n')

        #checking gpay
        gpay = text_lower.find('g pay')
        gpay2 = -1
        if gpay > 0:
            gpay2 = text.lower().index('to')
        gpayv = text_lower.find('paid to')
        
        # checking paytm
        paytm = text_lower.find('paid successfully to')
        paytm2 = text_lower.find('paytm wallet')
        
        try:

            # processing if matched with gpay template v2
            if gpay != -1 and gpay2 == 0 and gpayv == -1:
                success = True
                template = 'Gpay v2' # used by google pay in 2021

                identifier = data_lower.index('upi transaction id')

                upi_id = data[identifier + 1]
                paid_to = data[identifier - 11].lower().replace('to ','').strip().title()
                if paid_to in [None,'']:
                    paid_to = data[identifier - 10].lower().replace('to ','').strip().title()

                paid_on = data[identifier - 4]
                paid_on = paid_on.replace('© Completed + ','').strip()
                paid_on = paid_on.replace('© Completed - ','').strip()
                paid_on = paid_on.replace('© Completed « ','').strip()
                paid_on = datetime.datetime.strptime(paid_on,'%d %b, %I:%M %p')
                paid_on = datetime.datetime(datetime.date.today().year, paid_on.month,paid_on.day,
                    paid_on.hour,paid_on.minute,tzinfo=tz)

                amount = int(float(data[identifier - 8].replace('=','').strip().replace(',','')))
            
            # processing if matched with gpay template v1
            elif gpay != -1 and gpayv != -1:
                success = True
                template = 'Gpay v1'

                identifier = data_lower.index('upi transaction id')

                upi_id = data[identifier + 1]
                paid_to = data[identifier - 7].title()

                paid_on = data[identifier - 2]
                paid_on = datetime.datetime.strptime(paid_on,'%d %b %Y %I:%M %p')
                paid_on = datetime.datetime(datetime.date.today().year, paid_on.month,paid_on.day,
                    paid_on.hour,paid_on.minute,tzinfo=tz)

                scrap = data[identifier - 3].split('= ')
                amount = int(float(scrap[-1].strip()))

            # processing if matched with paytm
            elif paytm != -1:
                success = True
                template = 'Paytm'

                identifier = data_lower.index('paid successfully to')

                upi_id = data[identifier + 6].split(':')[1].strip()

                paid_to = data[identifier + 2]
            
                amount = data[identifier + 4]
                amount = amount.replace('@','')
                amount = amount.replace('&','')
                amount = amount.replace('°','')
                amount = amount.strip()
                amount = int(float(amount[1:]))

                paid_on = data[identifier + 7]
                paid_on = datetime.datetime.strptime(paid_on.upper(),'%I:%M %p, %d %b %Y')
                paid_on = datetime.datetime(paid_on.year, paid_on.month,paid_on.day,
                    paid_on.hour,paid_on.minute,tzinfo=tz)
            
            elif paytm2 > -1:
                success = True
                template = 'Paytm v2' # paytm transaction screen

                identifier = data_lower.index('paytm wallet')

                upi_id = data[identifier + 2].split(':')[1].strip()

                paid_to = data[identifier - 3]
            
                amount = data[identifier - 9].strip()
                amount = int(float(amount))

                paid_on = data[identifier - 7].strip()
                paid_on = datetime.datetime.strptime(paid_on.upper(),'%d %b %Y, %I:%M %p')
                paid_on = datetime.datetime(paid_on.year, paid_on.month,paid_on.day,
                    paid_on.hour,paid_on.minute,tzinfo=tz)

            else:
                pass
            
            if success == True:
                return [success, template, paid_to, paid_on, amount, upi_id]
            else:
                return ['TNF', text]
        
        except:
            print(data)
            return ['TNF', text]

    else:
        return ['FNF']

def number_converter(str_num: str) -> int:
    '''Converts number to int'''

    if isinstance(str_num,int) or isinstance(str_num,float):
        return [True, int(str_num)]
    elif str_num in [None]:
        return [True, 0]
    elif isinstance(str_num,str):
        if str_num.replace('.','').replace('(','-').replace(')').strip().isdigit():
            str_num = int(float(str_num))
            return [True, str_num]
        elif str_num.strip() == '':
            return [True, 0]
        else:
            return [False]
    else:
        return [False]

class TimeKeeper():
    def __init__(self, supress:str) -> None:
        '''
        pass y flag to supresses printing by this class so that user can print as per his needs
        '''
        self.start_time = datetime.datetime.now()
        print(f'\nStarted at: {self.start_time}')
        self.logs = []
        self.logs.append(self.start_time)
        self.supress = supress
    
    def log(self) -> datetime:
        '''
        returns difference between last log and 2nd last log entry
        '''
        self.logs.append(datetime.datetime.now())
        diff = self.logs[-1] - self.logs[-2]
        if self.supress == 'y':
            pass
        else:
            print(f'\nLap runtime (sec): {diff}')
            print(f'Total runtime (sec): {self.logs[-1] - self.start_time}')
        
        return diff

    def get_log(self, log_idx:int) -> datetime:
        '''
        returns difference between given log and start time
        '''
        diff = self.logs[log_idx], self.start_time
        if self.supress == 'y':
            pass
        else:
            print(f'\nRuntime since start (sec): {diff}')
        
        return diff

    def end(self) -> datetime:
        self.end_time = datetime.datetime.now()
        self.logs.append(self.end_time)
        diff = self.end_time - self.start_time
        if self.supress == 'y':
            pass
        else:
            print(f'\nTotal runtime (sec): {diff}')
            print(f'End time: {self.end_time}')
        
        return diff

def progress_bar(iteration, total, prefix = '', suffix = '', decimals = 1, length = 100, fill = '█', printEnd = '\r'):
    """
    Call in a loop to create terminal progress bar
    @params:
        iteration   - Required  : current iteration (Int)
        total       - Required  : total iterations (Int)
        prefix      - Optional  : prefix string (Str)
        suffix      - Optional  : suffix string (Str)
        decimals    - Optional  : positive number of decimals in percent complete (Int)
        length      - Optional  : character length of bar (Int)
        fill        - Optional  : bar fill character (Str)
        printEnd    - Optional  : end character (e.g. "\r", "\r\n") (Str)
    """

    percent = ("{0:." + str(decimals) + "f}").format(100 * (iteration / float(total)))
    filledLength = int(length * iteration // total)
    bar = fill * filledLength + '-' * (length - filledLength)
    print(f'\r{prefix} |{bar}| {percent}% {suffix}', end = printEnd)
    # Print New Line on Complete
    if iteration == total: 
        print()


# Functions

# fixer API to get forex rates

def forex_get_latest(api_key, base_cur='USD', get_cur=''):
    '''
    Gets latest forex rates from api.apilayer.com i.e. fixer api
    @params
    api_key    -   required    :   api key
    base_cur   -   optional    :   base currency. USD by default.
    get_cur    -   optional    :   All currencies by default. Pass string like 'SGD,INR' if specific currencies are required.
    
    Returns response in json format
    '''

    url = "https://api.apilayer.com/fixer/latest?symbols={symbols}&base={base}"

    base = base_cur
    symbols = get_cur

    url = url.format(base=base, symbols=symbols)
    payload = {}
    headers= {
    "apikey": api_key
    }

    response = requests.request("GET", url, headers=headers, data = payload)

    status_code = response.status_code
    
    return response.json()


def forex_get_timeseries(api_key, start_date, end_date, base_cur='USD', get_cur=''):
    '''
    Gets forex rates for a timespan from api.apilayer.com i.e. fixer api
    @params
    api_key    -   required    :   api key
    start_date -   required    :   start date in YYYY-MM-DD format
    end_date   -   required    :   end date in YYYY-MM-DD format
    base_cur   -   optional    :   base currency. USD by default.
    get_cur    -   optional    :   All currencies by default. Pass string like 'SGD,INR' if specific currencies are required.

    Returns response in json
    '''

    url = f"https://api.apilayer.com/fixer/timeseries?start_date={start_date}&end_date={end_date}&base={base_cur}&symbols={get_cur}"

    base_cur = base_cur
    get_cur = get_cur
    start_date = start_date
    end_date = end_date
    
    url = f'{url}'

    payload = {}
    headers= {
      "apikey": api_key
    }  

    response = requests.request("GET", url, headers=headers, data = payload)

    status_code = response.status_code
    result = response.text

    return response.json()


def forex_get_symbols(api_key):
    '''
    Gets all currency symbols
    @params
    api_key    -   required    :   api key
    '''
    url = "https://api.apilayer.com/fixer/symbols"

    payload = {}
    headers= {
    "apikey": api_key
    }

    response = requests.request("GET", url, headers=headers, data = payload)

    status_code = response.status_code
    
    return response.json()


## functions from calculation utility brought here


def settings_importer(file_path: Path, sheet_index:int)->list:
    '''
    @params:
    file_path   - required  - path to the file
    sheet_index - required  - index of the sheet in the file
    columns required in the sheet:
    - import_cols
    - import_dtype (python data types)
    - rename
    - convert_int (y/n)
    - convert_float64 (y/n)
    - convert_date (y/n)

    returns following in a list:
    - import_cols as list
    - import_dtype as dict
    - col_rename as dic
    - date_cols as list
    - amt_cols as list
    - int_cols as list

    '''

    df_settings_file = pd.read_excel(file_path, sheet_name=sheet_index, engine='openpyxl')
    columns_in_setting = df_settings_file.columns
    check = all(item in ['import_cols', 'import_dtype', 'rename', 'convert_int', 'convert_float64', 'convert_date'] for item in columns_in_setting)

    if check:

        # picking rows till null value
        for ind, row in df_settings_file.iterrows():
            if pd.isna(row['import_cols']):
                break
        df_settings_file = df_settings_file[:ind]

        df_settings_file = df_settings_file.loc[~df_settings_file['import_cols'].isnull()]

        # list of columns to be imported

        import_cols = df_settings_file.loc[~df_settings_file['import_cols'].isnull()]['import_cols'].tolist()

        # importing column data types

        import_dtype = df_settings_file.loc[:,['import_cols','import_dtype']]
        import_dtype['import_dtype'] = import_dtype['import_dtype'].fillna('')
        import_dtype_dict = {}
        for ind, row in import_dtype.iterrows():
            if row['import_dtype']=='':
                import_dtype_dict[row['import_cols']] = 'str'
            else:
                import_dtype_dict[row['import_cols']] = row['import_dtype']

        # importing column rename dict

        df_col_rename = df_settings_file.loc[~df_settings_file['rename'].isnull(),['import_cols','rename']]
        col_rename_dict = {}
        for ind, row in df_col_rename.iterrows():
            col_rename_dict[row['import_cols']] = row['rename']

        # list of cols to be converted to date, amount, int

        date_cols = df_settings_file.loc[~df_settings_file['convert_date'].isnull()]['import_cols'].tolist()
        amt_cols = df_settings_file.loc[~df_settings_file['convert_float64'].isnull()]['import_cols'].tolist()
        int_cols = df_settings_file.loc[~df_settings_file['convert_int'].isnull()]['import_cols'].tolist()

        return {'import_cols':import_cols,
            'import_dtype':import_dtype_dict,
            'col_rename':col_rename_dict,
            'date_cols':date_cols,
            'amt_cols':amt_cols,
            'int_cols':int_cols}

    else:
        return False

def col_dtype_converter(df:pd.DataFrame, date_cols:list=False, day_first = False, amt_cols:list=False, code_cols:list=False) -> pd.DataFrame:
    '''
    converts date, amount and int columns from string to these dtypes. Date column can handle dates with - or /
    Removes -, NA, na
    '''

    if date_cols:
        for col in date_cols:
            if df[col].dtype == 'datetime64[ns]':
                pass
            else:
                print(f'Attempting to convert {col} to date!!')
                df[col] = df[col].fillna('')
                df[col] = df[col].replace('-','').replace('NA','').replace('na','')
                df[col] = df[col].astype(str).str.replace('-','/')
                df[col] = pd.to_datetime(df[col], errors='raise', dayfirst=day_first)

    if amt_cols:
        for col in amt_cols:
            if df[col].dtype == 'float64':
                pass
            else:
                print(f'Attempting to convert {col} to amount!!')
                df[col] = df[col].fillna('0')
                df[col] = df[col].replace('-', '0').replace('NA','0').replace('na','0')
                df[col] = df[col].str.replace(',','').replace(')','').replace('(','-').astype(float)

    if code_cols:
        for col in code_cols:
            if df[col].dtype == 'float64':
                print(f'Attenpting to convert {col} from {df[col].dtype} into int')
                df[col] = df[col].replace('-',0).replace('NA',0).replace('na',0)
                df[col].fillna(0, inplace=True)
                df[col] = df[col].astype(np.int64)
            elif df[col].dtype == 'object':
                print(f'Attenpting to convert {col} from {df[col].dtype} into int')
                df[col] = df[col].fillna('0')
                df[col] = df[col].replace('-','0').replace('NA','0').replace('na','0')
                df[col] = df[col].astype(np.int64)

    return df

def excel_writer(df:pd.DataFrame, output_file:Path, sheet_name, export_header:bool = False):
    '''
    Checks if file / sheet already exists. If yes, appends data to existing file / sheet else creates new
    file / sheet.
    '''
    if output_file.exists():
        # checking last used row before appending
        wb = load_workbook(output_file, read_only=True)
        sheets = wb.sheetnames
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            last_row = len(list(ws.rows))
        else:
            last_row=0
        wb.close()

        # appending to file in existing sheet
        with pd.ExcelWriter(output_file, mode='a', date_format='DD/MM/YYYY', engine='openpyxl', if_sheet_exists='overlay') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=last_row, header=export_header)
    else:
        # writing to a new file if it does not exist
        with pd.ExcelWriter(output_file, mode='w', date_format='DD/MM/YYYY', engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False, header=export_header)

def df_col_search(df:pd.DataFrame, srch_string: str):
    '''
    Searches given string in column header in lower case and returns list of columns where string is found.
    Case insensitive.
    '''
    col_list = []
    for col in df.columns:
        if srch_string in col.lower():
            col_list.append(col)
    return col_list

def dtype_to_df(df: pd.DataFrame)->pd.DataFrame:
    '''
    Accepts a dataframe and returns df containing column dtypes
    '''
    df_cols = pd.DataFrame(columns=['Type'], data=df.dtypes)
    df_cols.index.name = 'Col'
    df_cols.reset_index(inplace=True)
    return df_cols

def ageing_creator(df:pd.DataFrame, col:str, to_date:datetime.date=datetime.date.today(), bins:list=[0,30,60,90,5000], bins_labels:list=['0 to 30', '30 to 60','60 to 90', 'above 90'])->pd.Series:
    '''Creates following parameters and eturns dataframe with ageing:
    df = pandas dataframe
    col = column name of the column in dataframe. This should contain data in date format.
    bins = list of integers containing the bins
    bins_labels = labels to be used for bins    

    Returns binned series which can be directly added to the source dataframe
    '''

    if len(bins) != len(bins_labels)+1:
        return None

    diff = to_date - df[col]

    bins_converted = []
    for bin in bins:
        bins_converted.append(datetime.timedelta(bin))

    binned = pd.cut(diff, bins=bins_converted, include_lowest=True, right=True, labels=bins_labels)
    del diff

    return binned


class LoggerConfigurator:
    def __init__(self, logger_name: str, log_dir: str = "logs", log_level: str = "DEBUG"):
        self.logger_name = logger_name
        self.log_level = log_level

        # Resolve project root
        project_root = Path(__file__).resolve().parent
        log_path = Path(log_dir)

        # Resolve relative path to absolute based on project root
        self.log_dir = log_path if log_path.is_absolute() else project_root / log_path
        self.log_dir.mkdir(parents=True, exist_ok=True)

    def configure(self):
        logger.remove()
        timestamp = datetime.datetime.now().strftime("%Y_%m_%d_%H_%M_%S")
        log_file = self.log_dir / f"{self.logger_name}_{timestamp}.log"

        # File logging
        logger.add(log_file, rotation="10 MB", level=self.log_level, mode="a")

        # Console logging
        log_format = (
            "<green>{time:YYYY-MM-DD HH:mm:ss.SSS}</green> | "
            "<level>{level: <8}</level> | "
            "<cyan>{name}</cyan>:<yellow>{function}:{line}</yellow> - "
            "<level>{message}</level>"
        )
        logger.add(sys.stderr, level=self.log_level, format=log_format, colorize=True, backtrace=True, diagnose=True)

        logger.debug(f"Logger initialized at: {log_file}")
        return logger
