"""Excel evaluator for InstantGrade.

This module implements a minimal Evaluator class that mirrors the
interface of the Python notebook `Evaluator`. It can create an answer
key from a solution workbook and grade student .xlsx submissions found
in a folder. The implementation is intentionally lightweight and self
contained (no external services).
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple

try:
    import openpyxl
except Exception as e:  # pragma: no cover - environment dependent
    raise ImportError("openpyxl is required for the Excel evaluator: install with 'pip install openpyxl'") from e


def _check_value(cell_address: str, worksheet) -> Any:
    # return the raw value stored in the cell (data_only workbook used separately)
    try:
        return worksheet[cell_address].value
    except Exception:
        return None


def _extract_nested_functions(formula: str, main_functions: list) -> List[str]:
    if not isinstance(formula, str) or not formula.startswith("="):
        return []
    formula = formula[1:]
    nested_functions = set()
    buffer = ""
    stack = []
    for char in formula:
        if char == "(":
            if buffer:
                func_name = buffer.strip().upper()
                if func_name.isalpha():
                    nested_functions.add(func_name)
                stack.append(func_name)
                buffer = ""
        elif char == ")":
            buffer = ""
            if stack:
                stack.pop()
        else:
            buffer += char
    return list(nested_functions)


def create_answers_loop(column_name: str, row_number: int, number_of_question: int, worksheet, nearby_columns: int = 15) -> Dict[str, List[Any]]:
    from openpyxl.utils import column_index_from_string, get_column_letter

    result: Dict[str, List[Any]] = {}
    start_col_idx = column_index_from_string(column_name)

    for i in range(number_of_question):
        row_idx = row_number + i
        main_cell = f"{column_name}{row_idx}"
        answer_values = []

        for offset in range(nearby_columns):
            col_idx = start_col_idx + offset
            col_letter = get_column_letter(col_idx)
            cell_address = f"{col_letter}{row_idx}"

            try:
                cell_value = _check_value(cell_address, worksheet)
                if cell_value is not None and str(cell_value).strip() != "":
                    answer_values.append(cell_value)
                else:
                    break
            except Exception:
                break

        result[main_cell] = answer_values

    return result


def create_formula_answers_loop(column_name: str, row_number: int, number_of_question: int, worksheet, nearby_columns: int = 15) -> Dict[str, List[str]]:
    from openpyxl.utils import column_index_from_string, get_column_letter

    result: Dict[str, List[str]] = {}
    start_col_idx = column_index_from_string(column_name)

    for i in range(number_of_question):
        row_idx = row_number + i
        main_cell = f"{column_name}{row_idx}"
        nested_functions = set()

        for offset in range(nearby_columns):
            col_idx = start_col_idx + offset
            col_letter = get_column_letter(col_idx)
            cell_address = f"{col_letter}{row_idx}"

            try:
                cell_value = _check_value(cell_address, worksheet)
                if isinstance(cell_value, str) and cell_value.startswith("="):
                    root_func = cell_value[1:].split("(", 1)[0].strip().upper()
                    nested = _extract_nested_functions(cell_value, [root_func])
                    nested_functions.update(nested)
                else:
                    break
            except Exception:
                break

        result[main_cell] = list(nested_functions)

    return result


def create_answer_key(excel_file_path: str | Path, column_name: str, row_number: int, number_of_question: int, worksheet_name: Optional[str] = None, nearby_columns: int = 15) -> Tuple[Dict[str, List[str]], Dict[str, List[Any]]]:
    formula_wb = openpyxl.load_workbook(excel_file_path)
    if worksheet_name is None or worksheet_name not in formula_wb.sheetnames:
        worksheet_name = formula_wb.sheetnames[0]
    formula_ws = formula_wb[worksheet_name]

    value_wb = openpyxl.load_workbook(excel_file_path, data_only=True)
    if worksheet_name not in value_wb.sheetnames:
        worksheet_name = value_wb.sheetnames[0]
    value_ws = value_wb[worksheet_name]

    formula_answer_key = create_formula_answers_loop(column_name, row_number, number_of_question, formula_ws, nearby_columns)
    value_answer_key = create_answers_loop(column_name, row_number, number_of_question, value_ws, nearby_columns)

    return formula_answer_key, value_answer_key


def evaluate_excel_file(file_path: str | Path, column_name: str, row_number: int, number_of_question: int, values_answer_key: Dict[str, List[Any]], formula_answer_key: Dict[str, List[str]], worksheet_name: Optional[str] = None, student_name_cell: str = "B2", student_roll_no_cell: str = "B3") -> Optional[Dict[str, Any]]:
    file_path = str(file_path)
    try:
        formula_wb = openpyxl.load_workbook(file_path)
    except Exception:
        return None

    ws_name = worksheet_name if worksheet_name and worksheet_name in formula_wb.sheetnames else formula_wb.sheetnames[0]
    formula_ws = formula_wb[ws_name]
    value_wb = openpyxl.load_workbook(file_path, data_only=True)
    value_ws = value_wb[ws_name]

    try:
        name = _check_value(student_name_cell, formula_ws)
    except Exception:
        name = None
    try:
        roll_no = _check_value(student_roll_no_cell, formula_ws)
    except Exception:
        roll_no = None

    # Basic scoring counters
    result_data = {
        "Name": name,
        "Roll Number": roll_no,
        "Result": 0,
        "Error Sum": 0,
        "Value Sum": 0,
        "Formula Sum": 0,
    }

    # For each question cell in the answer key, check values and formula
    for cell_address in sorted(set(list(values_answer_key.keys()) + list(formula_answer_key.keys()))):
        # value check
        expected_values = values_answer_key.get(cell_address, []) or []
        student_value = _check_value(cell_address, value_ws)

        value_ok = False
        if expected_values:
            for ev in expected_values:
                try:
                    if (student_value == ev) or (float(student_value) == float(ev)):
                        value_ok = True
                        break
                except Exception:
                    if student_value == ev:
                        value_ok = True
                        break

        if value_ok:
            result_data["Value Sum"] += 1
            result_data["Result"] += 1
        # formula check
        expected_funcs = formula_answer_key.get(cell_address, []) or []
        # get student's formula (from formula workbook)
        student_formula = _check_value(cell_address, formula_ws)
        formula_ok = False
        if isinstance(student_formula, str) and student_formula.startswith("=") and expected_funcs:
            formula_upper = student_formula.upper()
            if any(f.upper() in formula_upper for f in expected_funcs):
                formula_ok = True

        if formula_ok:
            result_data["Formula Sum"] += 1
            result_data["Result"] += 1

    return result_data


class Evaluator:
    """Lightweight Excel Evaluator.

    Parameters mirror the Python Evaluator where useful. Additional keyword
    arguments recognized (defaults provided):
      - column_name: starting column for answers (default: 'B')
      - row_number: starting row for answers (default: 5)
      - number_of_question: number of questions (default: 10)
      - worksheet_name: optional worksheet name to use
    """

    def __init__(
        self,
        solution_file_path: str | Path,
        submission_folder_path: str | Path,
        use_docker: bool = False,
        parallel_workers: int = 1,
        log_path: str | Path = "./logs",
        log_level: str = "normal",
        column_name: str = "B",
        row_number: int = 5,
        number_of_question: int = 10,
        worksheet_name: Optional[str] = None,
        nearby_columns: int = 15,
        **kwargs,
    ):
        self.solution_path = Path(solution_file_path)
        self.submission_path = Path(submission_folder_path)
        self.column_name = column_name
        self.row_number = row_number
        self.number_of_question = number_of_question
        self.worksheet_name = worksheet_name
        self.nearby_columns = nearby_columns

    def run(self):
        # Build answer keys
        formula_key, value_key = create_answer_key(
            self.solution_path,
            self.column_name,
            self.row_number,
            self.number_of_question,
            worksheet_name=self.worksheet_name,
            nearby_columns=self.nearby_columns,
        )

        # Discover submissions
        if not self.submission_path.exists():
            raise FileNotFoundError(f"Submission path does not exist: {self.submission_path}")

        if self.submission_path.is_file():
            submission_files = [self.submission_path]
        else:
            submission_files = sorted([p for p in self.submission_path.iterdir() if p.suffix.lower() in (".xlsx", ".xls", ".xlsm")])

        executed_results: List[Dict[str, Any]] = []

        for sub in submission_files:
            try:
                data = evaluate_excel_file(
                    sub,
                    self.column_name,
                    self.row_number,
                    self.number_of_question,
                    value_key,
                    formula_key,
                    worksheet_name=self.worksheet_name,
                )
                if not data:
                    # Represent a failed execution
                    executed_results.append(
                        {
                            "student_path": sub,
                            "execution": {"success": False, "errors": ["Could not read file"], "student_meta": {"name": None, "roll_number": None}},
                            "results": [],
                        }
                    )
                    continue

                # Convert the aggregate data into a series of assertion-style rows
                results: List[Dict[str, Any]] = []
                # For each question cell, produce a value-assertion and formula-assertion if appropriate
                all_cells = sorted(set(list(value_key.keys()) + list(formula_key.keys())))
                for cell in all_cells:
                    expected_vals = value_key.get(cell, []) or []
                    expected_funcs = formula_key.get(cell, []) or []

                    # Value assertion
                    student_value = None
                    try:
                        wb_val = openpyxl.load_workbook(sub, data_only=True)
                        ws_val = wb_val[wb_val.sheetnames[0]]
                        student_value = _check_value(cell, ws_val)
                    except Exception:
                        student_value = None

                    val_status = "failed"
                    val_score = 0
                    if expected_vals:
                        for ev in expected_vals:
                            try:
                                if (student_value == ev) or (float(student_value) == float(ev)):
                                    val_status = "passed"
                                    val_score = 1
                                    break
                            except Exception:
                                if student_value == ev:
                                    val_status = "passed"
                                    val_score = 1
                                    break

                    # Only produce a value assertion if the answer key expects a value
                    if expected_vals:
                        results.append(
                            {
                                "question": cell,
                                "assertion": "value",
                                "status": val_status,
                                "score": val_score,
                                "error": None,
                                "description": "Value match",
                            }
                        )

                    # Formula assertion
                    if expected_funcs:
                        student_formula = None
                        try:
                            wb_formula = openpyxl.load_workbook(sub)
                            ws_formula = wb_formula[wb_formula.sheetnames[0]]
                            student_formula = _check_value(cell, ws_formula)
                        except Exception:
                            student_formula = None

                        form_status = "failed"
                        form_score = 0
                        if isinstance(student_formula, str) and student_formula.startswith("="):
                            form_upper = student_formula.upper()
                            if any(f.upper() in form_upper for f in expected_funcs):
                                form_status = "passed"
                                form_score = 1

                        results.append(
                            {
                                "question": cell,
                                "assertion": "formula",
                                "status": form_status,
                                "score": form_score,
                                "error": None,
                                "description": "Formula contains expected function(s)",
                            }
                        )

                executed_results.append(
                    {
                        "student_path": sub,
                        "execution": {"success": True, "errors": [], "student_meta": {"name": data.get("Name"), "roll_number": data.get("Roll Number")}},
                        "results": results,
                    }
                )

            except Exception as e:
                executed_results.append(
                    {
                        "student_path": sub,
                        "execution": {"success": False, "errors": [str(e)], "student_meta": {"name": None, "roll_number": None}},
                        "results": [],
                    }
                )

        # Compute total assertions per attempt based on the answer keys.
        all_cells = sorted(set(list(value_key.keys()) + list(formula_key.keys())))
        total_assertions = 0
        for cell in all_cells:
            if value_key.get(cell):
                total_assertions += 1
            if formula_key.get(cell):
                total_assertions += 1

        # Return a ReportingService to keep parity with the Python Evaluator
        # which returns a ReportingService. This ensures reporting has the
        # correct `total_assertions` (out-of marks) and formatting.
        try:
            from instantgrade.reporting.reporting_service import ReportingService

            report = ReportingService(
                executed_results=executed_results,
                logger=None,
                total_assertions=total_assertions,
            )
            return report
        except Exception:
            # If ReportingService cannot be imported for any reason, fall back
            # to returning the raw executed results (older behavior).
            return executed_results

    def to_html(self, path: str | Path):
        # The Evaluator returns executed_results from run(); use ReportingService for HTML
        from instantgrade.reporting.reporting_service import ReportingService
        executed = self.run()
        # If run() already returned a ReportingService, delegate to it.
        if isinstance(executed, ReportingService):
            return executed.to_html(path)

        # Otherwise wrap executed_results in a ReportingService and render.
        report = ReportingService(executed_results=executed)
        return report.to_html(path)

    def summary(self, all_results=None):
        # Build a quick summary compatible with the python Evaluator.summary
        executed = all_results if all_results is not None else self.run()
        total = len(executed)
        passed = sum(1 for r in executed if r.get("execution", {}).get("success", False))
        return {"total": total, "passed": passed, "failed": total - passed}
