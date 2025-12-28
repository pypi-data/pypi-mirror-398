#!/usr/bin/python3
# coding=utf-8
###################################################################
#           ____     _     _ __  __                 
#          / __/__  (_)___(_) /_/ /  ___  ___  ___ _
#         _\ \/ _ \/ / __/ / __/ /__/ _ \/ _ \/ _ `/
#        /___/ .__/_/_/ /_/\__/____/\___/_//_/\_, / 
#           /_/                              /___/  
# Copyright (c) 2024 Chongqing Spiritlong Technology Co., Ltd.
# All rights reserved.  
# @author	arthuryang
# @brief	基于池的数据库驱动MySQL
#
###################################################################  

import	MySQLdb	# pip install mysqlclient
from	MySQLdb			import OperationalError 
from	dbutils.pooled_db	import PooledDB
import	threading
import	logging
import	time
import	redis
import	pickle
import	decimal
import	json
import	datetime
import	openpyxl

# 装饰器：线程同步
def synchronized(func):
	func.__lock__ = threading.Lock()

	def synced_func(*args, **kwargs):
		with func.__lock__:
			return func(*args, **kwargs)
        
	return synced_func

# 基于池的数据库类
class MySQLDatabase(object):
	# 确保此类对每个数据库只有一个对象
	instance = {}

	@synchronized
	def __new__(cls, *args, **kwargs):
		# 根据传入的_database_config中host和db来实现每个数据库只有一个对象
		key 	= args[0]["host"] + args[0]["database"]
		if key not in cls.instance.keys():
			cls.instance[key] = super().__new__(cls)
		return cls.instance[key]

	## 类初始化
	#	_database_config	数据库配置参数，这是一个字典{user, password, host, port, database}，详细参数：https://mysqlclient.readthedocs.io/user_guide.html#mysqldb
	#	use_redis		None表示不使用redis，传入一个字典{host, port, index}，都可省略以使用默认值
	def __init__(self, _database_config, use_redis=None):
		self.database_config			= _database_config
		self.database_config["blocking"]	= True	# 连接池中如果没有可用连接，阻塞等待

		# 数据库名称
		self.database_name		= _database_config['database']

		# 默认字段
		self.default_field_names	= ['ID', 'CREATE_TIMESTAMP', 'UPDATE_TIMESTAMP', 'DELETE_TIMESTAMP']

		# 数值类型
		self.digital_field_type		= ["INT", "FLOAT", "DOUBLE", "DECIMAL"]

		# 日期类型
		self.timestamp_field_type	= ["TIMESTAMP", "TIMESTAMP ON UPDATE CURRENT_TIMESTAMP", "DATE", "DATETIME"]

		# 确保属性不会被重复初始化
		if not hasattr(self, 'connection_pool'):
			# 数据库连接池
			self.connection_pool	= PooledDB(MySQLdb, **self.database_config)
		
		# 日志初始化
		self.logger	= logging.getLogger(__name__)
		if not self.logger.handlers:
			# 避免重复添加handler
			self.logger.setLevel(logging.DEBUG)
			console_handler	= logging.StreamHandler()
			console_handler.setLevel(logging.DEBUG)
			console_handler.setFormatter(logging.Formatter('%(asctime)s %(message)s', datefmt='%Y-%m-%d %H:%M:%S'))
			self.logger.addHandler(console_handler)

		# 更新redis配置。数据库配置中'redis'保存的是一个字典{host, port, index}，指出了使用的redis主机、端口和数据库（序号）
		if use_redis is not None:
			self.redis_host		= use_redis['host'] 	if 'host'	in use_redis else 'localhost'
			self.redis_port		= use_redis['port'] 	if 'port'	in use_redis else '6379'
			self.redis_index	= use_redis['index']	if 'index'	in use_redis else '0'
			self.redis_password	= use_redis['password']	if 'password'	in use_redis else None

			if self.redis_password is None:
				self.redis		= redis.Redis(connection_pool=redis.ConnectionPool(
								host	= self.redis_host,
								port	= self.redis_port,
								db	= self.redis_index, 
								decode_responses=False))
			else:
				self.redis		= redis.Redis(connection_pool=redis.ConnectionPool(
								host		= self.redis_host,
								port		= self.redis_port,
								db		= self.redis_index,
								password	= self.redis_password,
								decode_responses=False))
		else:
			self.redis	= None

	## 对于值中的特殊字符进行替换
	def replace(self, string):
		if not string:
			return string
		string	= str(string)
		string	= string.replace("\\", "\\\\")	# 反斜杠换成两个反斜杠
		string	= string.replace("'", "\\'")	# 单引号转义
		string	= string.replace('"', '\\"')	# 双引号转义
		
		return string
	
	# ------------------------ 数据库基本操作 ------------------------
	# DDL 对数据库和表的操作	
	# 	CREATE DATABASE/TABLE
	# 	DROP DATABASE/TABLE
	# 	USE 
	# 	DESC
	#	ALTER TABLE ... ADD/DROP/CHANGE/MODIFY
	# DQL 查询
	# DML 插入/修改/删除，需要提交
	# DCL 授权
	# 操作接口：
	#	query_single	查询一条记录
	#	query		查询所有记录
	#	------------ 以下皆考虑redis缓存 ---------------
	#	query_redis	获取指定表的数据，可以指定需要的字段和ID
	#	insert		插入记录；若ID已经存在，则更新该记录
	#	update		更新记录，必须指定ID
	#	delete		删除记录（将DELET_TIMESTAMP设为当前时间，并非删除该记录）
	#	undelete	恢复删除的记录（将DELETE_TIMESTAMP设为NULL）

	## 执行SQL语句
	# 其他函数中应调用此函数来执行SQL语句
	# 私有方法，避免从外部直接调用
	#	SQL_string	要执行的SQL语句
	#	need_result	查询语句返回查询结果
	#	return_ID	当need_result==False时，返回插入语句的ID，多条语句将会返回第一个新纪录的ID；当need_result==True时，返回受影响的行数
	def _execute(self, SQL_string, need_result=False, return_ID=False):
		result	= None
		try:
			connection	= self.connection_pool.connection()
			cursor 		= connection.cursor()

			# 执行SQL语句
			cursor.execute(SQL_string)

			if need_result:
				columns = [desc[0] for desc in cursor.description]
				result = [dict(zip(columns, row)) for row in cursor.fetchall()]
			
			elif return_ID:
				# 获取第一个新插入记录的ID
				result	= cursor.lastrowid
			else:
				result	= cursor.rowcount
				
			cursor.close()
				
			if 'autocommit' not in self.database_config or not self.database_config['autocommit']:
				# 提交
				connection.commit()
			
			return result
		except OperationalError as ex:
			# 操作失败
			self.logger.error(str(ex))
			cursor.close()
			if 'autocommit' not in self.database_config or not self.database_config['autocommit']:
				# 回滚
				connection.rollback()
						
			return None
		except Exception as ex:
			cursor.close()
			if 'autocommit' not in self.database_config or not self.database_config['autocommit']:
				# 回滚
				connection.rollback()
			#self.logger.info(SQL_string)
			self.logger.error(str(ex))
				
			return result
	
	## 查询一条记录
	def query_single(self, SQL_string, *args):
		result	= self._execute(SQL_string+" LIMIT 1", need_result=True)
		return result[0] if result else None
		
	## 查询所有记录
	def query(self, SQL_string, *args):
		return self._execute(SQL_string, need_result=True)

	# ------------------------- 缓存查询和更新 -------------------
	# 每个记录用"database {table} {ID}"来记录，是一个HASH，field:value

	## 查询redis缓存：先尝试查询redis，若ID没有缓存，则从数据库中重新读取。注意写入和修改时已经同步到redis
	#	table		表名
	#	ID_list		ID列表，None或[]表示全部ID，可以是一个ID
	#	field_list	字段列表，[]表示全部字段，可以是一个字段的字符串
	#	valid 		是否查询有效数据（DELETE_TIMESTAMP为空）
	# 返回的
	def query_redis(self, table, ID_list=None, field_list=None, valid=True):
		# table总是全大写
		table	= table.upper()

		#  --------- 处理ID列表 ---------
		# ID_list可以是整数列表，也可以是单个整数，None或者[]表示全部ID
		if not ID_list:
			# 从数据库中查询全部ID
			SQL_string	= f"SELECT ID FROM `{table}`"
			if valid:
				SQL_string	+= f" WHERE DELETE_TIMESTAMP IS NULL"
			result	= self.query(SQL_string)
			if result:
				ID_list = [r['ID'] for r in result]
			else:
				# 没有ID可用
				return []
		elif isinstance(ID_list, str):
			# 支持ID_list是英文逗号分隔的ID字符串
			ID_list	= [ID for ID in ID_list.split(',')]
		elif isinstance(ID_list, int):
			# 单个数值ID处理成列表
			ID_list	= [ID_list]
		elif not isinstance(ID_list, list):
			# 不是list的其他情况
			ID_list	= []
		_ID_list	= []
		for ID in ID_list:
			try:
				# 尝试转换为整数
				_ID_list.append(int(ID))
			except:
				pass
		ID_list	= _ID_list
		
		#  --------- 处理要查询的字段 ---------
		# 标记对全部字段的查询
		all_fields	= True if not field_list else False

		if field_list is None:
			field_list	= []
		elif isinstance(field_list, str):
			# 支持字段列表是英文逗号分隔的字符串
			field_list	= [f.strip() for f in field_list.split(",")]

		# ID总是被查询
		if 'ID' not in field_list:
			# ID放在最前面
			field_list.insert(0, 'ID')
		
		# DELETE_TIMESTAMP总是被查询
		if 'DELETE_TIMESTAMP' not in field_list:
			# DELETE_TIMESTAMP放最后
			field_list.append('DELETE_TIMESTAMP')
		
		# 支持redis缓存且该表已经缓存，直接查找该表的数据
		# "database {database} {table}"是redis中的散列，key是该表的ID，对象是该记录全部数据的序列化
		redis_table	= f"database {self.database_name} {table}"
		redis_result	= []			
		if self.redis:
			for ID in ID_list:
				if self.redis.hexists(redis_table, ID):
					# 该ID已经缓存
					record	= pickle.loads(self.redis.hget(redis_table, ID))
					if (valid and record['DELETE_TIMESTAMP'] is None) or (not valid):
						# 按有效性设置来放入到结果
						redis_result.append(record)
						# 从ID_list中去掉已经查到结果的ID
						ID_list.remove(ID)
			if not ID_list:
				# 全部找齐，直接返回
				return redis_result
		
		# 查询数据库全部字段，查询结果将会缓存到redis
		SQL_string	= f"SELECT * FROM `{table}`"
		# ID列表是必须的条件
		conditions	= [f"ID IN ({','.join([str(ID) for ID in ID_list])})"]
		
		if valid:
			# 检查DELETE_TIMESTAMP
			conditions.append(f"DELETE_TIMESTAMP IS NULL")
		if conditions:
			# SQL语句加上WHERE
			SQL_string	+= f" WHERE {' AND '.join(conditions)}"
		result	= self.query(SQL_string)

		if all_fields and result:
			# 设置全部字段
			field_list	= list(result[0].keys())

		# 缓存到redis
		self.update_redis(table, records=result)
			
		# 取需要的字段
		query_result	= [{f:r[f] for f in field_list} for r in result]
		
		return query_result+redis_result
		
	## 插入记录，若ID已经存在，则更新该记录。本实现中，insert包含了update
	#	table	要插入的表
	#	records	记录的列表[{}]
	# 返回新增的ID列表
	def insert(self, table, records):
		if not records:
			self.logger.info("新增/更新0条记录")
			return None
		
		if isinstance(records, dict):
			records		= [records]

		if not isinstance(records, list):
			self.logger.info("记录要用列表")
			return None
		
		# 已有ID的列表
		SQL_string	= f"SELECT ID FROM `{table}`"
		result		= self.query(SQL_string)
		ID_all		= [int(i['ID']) for i in result]

		# 把没有ID的选出来
		records_insert	= []
		records_update	= []
		for r in records:
			if 'ID' in r and not r['ID']:
				# 有ID而为空，去掉之
				r.pop('ID')

			if 'ID' not in r:
				# 无ID
				records_insert.append(r)
			else:
				# 有ID
				if ID_all and int(r['ID']) in ID_all:
					# ID已经存在
					records_update.append(r)
				else:
					# ID不存在
					records_insert.append(r)
		
		# 确保返回值正确
		inserted_ID_list	= []

		# UPDATE		
		if records_update:
			self.update(table, records_update)		

		# INSERT
		if records_insert: 	
			# 所有出现了的字段名合并到一个列表
			field_names_set	= set()
			for r in records_insert:
				field_names_set	= field_names_set.union(set(r.keys()))
			field_names	= list(field_names_set)

			# 构建赋值的值列表：若该字段若没有，则尝试设置为NULL
			def _generate_SQL_string(records):
				value_strings	= []
				for r in records:
					# 将每个记录的值生成字符串添加到values，要进行转义处理
					values	= []
					for f in field_names:
						if f not in r or r[f] is None:
							values.append(f"NULL")
						elif isinstance(r[f], str):
							v	= r[f]
							values.append(f"'{self.replace(v)}'")
						elif isinstance(r[f], float) or isinstance(r[f], decimal.Decimal) or isinstance(r[f], int):
							# 直接转换成字符串
							v	= str(r[f])
							values.append(f"'{self.replace(v)}'")
						elif isinstance(r[f], datetime.datetime):
							# 日期时间处理为标准格式
							v	= f"{r[f]:%Y-%m-%d %H:%M:%S}"
							values.append(f"'{self.replace(v)}'")
						else:
							# 其他数据类型都按JSON来处理
							v	= json.dumps(r[f], ensure_ascii=False)
							values.append(f"'{self.replace(v)}'")
					value_strings.append(f"( {','.join(values)} )")			
				SQL_string	= f"INSERT INTO `{table}` (`{'`,`'.join(field_names)}`) VALUES {','.join(value_strings)}"
				return SQL_string

			# 执行insert插入有ID的记录并输出结果
			records_with_ID		= [r for r in records_insert if 'ID' in r and r['ID']] if 'ID' in field_names else []
			if records_with_ID:
				try:
					result	= self._execute(_generate_SQL_string(records_with_ID), return_ID=True)
					self.logger.info(f"表{table}新增{len(records_insert) if result else 0}条记录，last_ID={result}")
				except Exception as ex:
					self.logger.info(f"表{table}新增{len(records_insert)}条记录失败：{ex}")
					return []
			inserted_with_ID	= [r['ID'] for r in records_with_ID]
			
			# 执行insert插入无ID的记录并输出结果
			records_without_ID	= [r for r in records_insert if r not in records_with_ID]
			if records_without_ID:
				# 先查询下一个自增值
				auto_increment_SQL_string	= f"select auto_increment from information_schema.tables where table_schema='{self.database_name}' and table_name='{table}';"
				auto_increment_before		= self.query_single(auto_increment_SQL_string)['auto_increment']
				try:
					result	= self._execute(_generate_SQL_string(records_without_ID), return_ID=True)
					self.logger.info(f"表{table}新增{len(records_insert) if result else 0}条记录，last_ID={result}")
				except Exception as ex:
					self.logger.info(f"表{table}新增{len(records_insert)}条记录失败：{ex}")
					return []
				auto_increment_after	= self.query_single(auto_increment_SQL_string)['auto_increment']
			# 全都是自增的，必定是连续的
			records_without_ID	= list(range(auto_increment_before, auto_increment_after)) if records_without_ID else []

			# 更新redis，从第一个要插入的ID开始
			inserted_ID_list	= inserted_with_ID+records_without_ID
			self.update_redis(table, ID_list=inserted_ID_list)

		return inserted_ID_list

	## 更新记录
	# 	table	记录所在的表
	# 	records	一条或多条记录的列表，每个记录必须包含ID，更新字段可以不一致
	# 返回ID_list
	def update(self, table, records):
		if not records:
			self.logger.info(f"update {table}：没有需要更新的数据")
			return False
		
		if isinstance(records, dict):
			records	= [records]
			
		# 对每一条记录进行更新
		ID_list	= []
		for r in records:
			if 'ID' not in r:
				# 必须要包含ID字段
				continue
			ID	= int(r['ID'])
			
			field_value_pairs	= []
			for f in r:
				if f!='ID':
					# 注意值要用单引号包起来，且内容需要处理
					v	= 'NULL' if r[f] is None else f"'{self.replace(r[f])}'"
					field_value_pairs.append(f"`{f}`={v}")
			SQL_string = f"UPDATE `{table}` SET {','.join(field_value_pairs)} WHERE ID={ID}"
			n	= self._execute(SQL_string)
			if n==1:
				# UPDATE返回受影响的行数，指定了ID应该只有一行受影响，此时表示成功更新
				# 注意：如果该记录每个域的新值都与旧值相同，即没有改变，则不会返回1
				ID_list.append(ID)
				self.logger.info(f"表{table}更新ID={ID}")
		# 更新redis
		self.update_redis(table, ID_list=ID_list)

		self.logger.info(f"表{table}更新了{len(ID_list)}条记录")

		return ID_list
		
	## 删除记录（将DELET_TIMESTAMP设为当前时间，并非删除该记录）
	def delete(self, table, ID):
		SQL_string	= f"UPDATE `{table}` SET DELETE_TIMESTAMP=CURRENT_TIMESTAMP()  WHERE `ID`={ID}"
		self._execute(SQL_string)

		# 更新redis
		self.update_redis(table, ID_list=[ID])

	## 恢复删除的记录（将DELETE_TIMESTAMP设为NULL）
	def undelete(self, table, ID):
		SQL_string	= f"UPDATE `{table}` SET DELETE_TIMESTAMP=NULL  WHERE `ID`={ID}"
		self._execute(SQL_string)
		
		# 更新redis
		self.update_redis(table, ID_list=[ID])
	
	# ------------------------- 工具函数 -------------------------

	## 更新redis缓存。records, start_ID或ID_list要至少一种
	#	records		用于更新的记录数据
	#	start_ID	从数据库查询从此ID到最大值的记录以更新
	#	ID_list		从数据库查询ID列表以更新
	def update_redis(self, table, records=None, start_ID=None, ID_list=None):
		if not self.redis:
			return 
		
		redis_table	= f"database {self.database_name} {table}"
		
		# 更新现有记录
		if records:
			try:
				self.redis.hmset(redis_table, {r['ID']:pickle.dumps(r) for r in records})
			except Exception as ex:
				self.logger.info(f"update_redis: {ex}")				
			return
		# 使用max_ID查询
		if start_ID:
			SQL_string	= f"SELECT * FROM `{table}` WHERE ID>{str(start_ID)}"
		elif ID_list:
			SQL_string	= f"SELECT * FROM `{table}` WHERE ID in ({','.join([str(ID) for ID in ID_list])})"
		else:
			return
		# 缓存到redis：以ID（int类型）为key，value是pickle序列化的结果
		# 注意redis要设置为decode_responses=False，否则会UnicodeDecodeError
		records		= self.query(SQL_string)
		self.redis.hmset(redis_table, {r['ID']:pickle.dumps(r) for r in records})

	## 将列表或逗号分隔的字符串转换为SQL语句中的字段列表
	def fields_string(self, table_name, fields):
		if isinstance(fields, str):
			# 字符串转列表
			fields	= fields.split(',')

		if isinstance(fields, list):
			# 列表
			try:
				fields_string	= ",".join([f"`{table_name}`.`{f.strip()}`" for f in fields])
			except Exception:
				# 可能会失败，万一不是字符串列表呢
				return None
		else:
			# 本来就不是字符串或字段列表
			return None
		
		if '*' in fields_string:
			# 不支持通配符*
			return None

		return fields_string

	## excel单元格的值转换为字符串
	def cell_string(self, sheet, row, column):
		value	= sheet.cell(row, column).value
		return str(value) if value else ""

	## 字体：主要只使用了字体名称、大小、颜色和加粗。默认黑色11号Calibri，英文数字较美观
	def cell_font(self, name='Calibri', size=11, color='000000', bold=False):
		return openpyxl.styles.Font(
			name		= name,
			size		= size,
			bold		= bold,
			vertAlign	= None,
			underline	= None,
			strike		= False,
			color		= color,
		)

	## 背景填充。默认白色
	def cell_fill(self, color='FFFFFFFF'):
		return openpyxl.styles.PatternFill(
			fill_type	= 'solid',
			fgColor		= color,
		)

	## 填充excel单元格，可以指定格式
	def set_cell(self, sheet, i, j, value, number_format='General', font=None, fill=None, border=None, alignment=None):
		cell		= sheet.cell(row=i, column=j)
		cell.value	= value
	
		cell.number_format	= number_format
		cell.font		= font if font is not None else openpyxl.styles.Font(name='Calibi', size=11, color='000000')
		# 默认无填充
		cell.fill		= fill if fill is not None else openpyxl.styles.PatternFill(fill_type=None)
		# 默认左对齐
		cell.alignment		= alignment if alignment is not None else self.alignment_left
		# 默认无边界
		cell.border		= border #if border is not None else openpyxl.styles.Border()
	
	## 自动调整指定列的宽度
	#	column		列（从1开始）
	#	width		0自动
	#	max_width	最大宽度限制
	def adjust_column_width(self, sheet, column, width=0, max_width=100):
		# 自动计算最大宽度
		if width==0:
			width	= 1
			for row in range(1, sheet.max_row+1):
				# 对于中文字符和非中文字符单独计算宽度
				s	= self.cell_string(sheet, row, column)
				w	= 0
				for c in s:
					w	+= (1 if ord(c)<128 else 2)
				width	= max(width, w+1)

		# 列宽得有最大限制
		if width>max_width:
			width	= max_width

		sheet.column_dimensions[openpyxl.utils.get_column_letter(column)].width	= width
	
	## 自动调整所有列的宽度
	def adjust_all_column_width(self, sheet, max_width=100):
		for column in range(1, sheet.max_column+1):
			self.adjust_column_width(sheet, column, max_width=max_width)

	# ------------------------ 输入输出和数据库管理操作 ------------------------
	## 字段COLUMN信息转换为我们自定义的字段字典
	def convert_COLUMN_field(self, COLUMN_field):
		# 设置字段信息
		field		= self.blank_field_string()
		field['name']	= COLUMN_field['COLUMN_NAME']

		# 保存原始的类型字符串
		field['MySQL_Type']	= COLUMN_field['COLUMN_TYPE']

		type_string	= COLUMN_field['COLUMN_TYPE'].split('(')
		field['type']	= type_string[0].upper()
		
		if field['type']=="TIMESTAMP" and "on update current_timestamp()" in COLUMN_field['EXTRA']:
			field['type']	+= " ON UPDATE CURRENT_TIMESTAMP"
			
		if len(type_string)>1:
			field['length']	= type_string[1][:-1]
		
		field['null']	= True if COLUMN_field['IS_NULLABLE']=="YES" else False
		
		if COLUMN_field['COLUMN_DEFAULT']=="current_timestamp()":
			field['default']	= "CURRENT_TIMESTAMP"
		elif COLUMN_field['COLUMN_DEFAULT']=="current_timestamp()":
			field['default']	= COLUMN_field['COLUMN_DEFAULT']
			
		if COLUMN_field['COLUMN_COMMENT']!='None':
			field['comment']	= COLUMN_field['COLUMN_COMMENT']

		if COLUMN_field['COLUMN_KEY']=='UNI':
			field['unique']	= True
		
		return field

	## 获取字段信息
	def get_field(self, table, field_name):
		# 获取字段信息
		SQL_string	= f'''
			SELECT * 
			FROM information_schema.`COLUMNS` 
			WHERE `TABLE_SCHEMA`='{self.database_config['database']}' AND 
				`TABLE_NAME`='{table}' AND 
				`COLUMN_NAME`='{field_name}'
			'''
		r	= self.query_single(SQL_string)
		if r:
			return self.convert_COLUMN_field(r)
		else:
			return None
		

	## 获取表的结构
	def get_table(self, table):
		t	= self.query(f'''
			SELECT table_name, table_comment 
			FROM information_schema.TABLES 
			WHERE table_schema = '{self.database_config['database']}' AND 
				table_name='{table}'
			''')
		if len(t)<1:
			# 没找到
			return None
		
		t	= t[0]
		
		fields = self.query(f'''
			SELECT * 
			FROM information_schema.`COLUMNS` 
			WHERE `TABLE_SCHEMA`='{self.database_config['database']}' AND 
				`TABLE_NAME`='{table}'
			''')
		
		# 排序后依次加入
		t['fields']	= []
		for f in sorted(fields, key=lambda item: item['ORDINAL_POSITION']):
			t['fields'].append(self.convert_COLUMN_field(f))
		
		return t

	## 获取数据库的结构
	def get_schema(self):
		# 读取表
		tables	= self.query(f'''
			SELECT table_name, table_comment 
			FROM information_schema.TABLES 
			WHERE table_schema = '{self.database_config['database']}'
		''')
		
		if tables==None:
			# 操作失败
			return None
		
		fields = self.query(f'''
			SELECT * 
			FROM information_schema.`COLUMNS` 
			WHERE `TABLE_SCHEMA`='{self.database_config['database']}'
		''')
		
		for t in tables:
			# 选择该表字段
			fs	= [ f for f in fields if f['TABLE_NAME']==t['table_name'] ]
			# 排序后依次加入
			t['fields']	= []
			for f in sorted(fs, key=lambda item: item['ORDINAL_POSITION']):
				t['fields'].append(self.convert_COLUMN_field(f))

			# 唯一约束
			# 该表的键（除ID主键外）
			keys	= self.query(f"SELECT * FROM information_schema.`KEY_COLUMN_USAGE` WHERE `TABLE_SCHEMA`='{self.database_config['database']}' and `TABLE_NAME`='{t['table_name']}'")
			if len(keys)>1:
				# 注意ID总是主键，排除这种情况
				constraints	= {}
				for k in keys:
					if k['CONSTRAINT_NAME'] in constraints:
						constraints[k['CONSTRAINT_NAME']].append(k['COLUMN_NAME'])
					else:
						constraints[k['CONSTRAINT_NAME']]	= [k['COLUMN_NAME']]
				t['constraints']	= constraints
		# tables按表名排序
		tables.sort(key=lambda t:t['table_name'])
		return tables

	## 空的字段SQL字符串(MySQL)
	def blank_field_string(self):
		return {
			'name'			: None,		# 字段名
			'type'			: None,		# 字段类型INT/DOUBLE/DECIMAL/CHAR/TEXT/DATETIME
			'length'		: None,		# 字段类型长度DECIMAL(10,2)/CHAR(255)
			'default'		: "NULL",	# 默认值
			'null'			: False,	# 是否可以为空，默认不能必填
			'unique'		: False,	# 是否唯一键
			'comment'		: None,		# 注释
		}

	## 从字段信息字典生成字段SQL字符串
	#	field	字段信息字典（包含name,type,length,default,null,comment）
	# 返回None表示构造失败
	def generate_field_string(self, field):
		# 检查键是否存在
		if not all(k in field for k in self.blank_field_string()):
			return None

		if not field['name']:
			return None
			
		MySQL_Type	= field['type']
		if field['length']:
			# 类型可能附带长度，要加小括号附在类型后
			MySQL_Type	+= f"({field['length']})"
		elif field['type'] in ("CHAR", "VARCHAR"):
			# CHAR/VARCHAR只使用255长度
			MySQL_Type	+= "(255)"
		s	= f"`{field['name']}` {MySQL_Type}"
		
		# 排序规则
		if field['type'] in ["CHAR", "VARCHAR", "TEXT"]:
			s	+= " CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci "
		
		# default中是初值
		if field['default'] and field['default']!='NULL':
			if field['type'] in self.digital_field_type:
				s	+= f" DEFAULT {field['default']} "
			elif field['type']=='TIMESTAMP':
				s	+= f" DEFAULT current_timestamp() " if field['default']=='CURRENT_TIMESTAMP' else ""
			else:
				s	+= f" DEFAULT '{self.replace(field['default'])}' "
		
		# 是否可以为空
		if field['null']:
			s	+= " NULL "
		else:
			s	+= " NOT NULL "

		# 注释
		if field['comment']:
			s	+= f" COMMENT '{self.replace(field['comment'])}'"

		return s

	## 增加表
	#	name	表名
	#	fields	字段信息，是一个list，每个成员是一个记录该字段信息的字典（包含name,type,length,default,on_update,null,primary,auto_increment,comment）
	#	comment	表注释
	def new_table(self, name, fields, comment=None):
		fields_strings	= []
		# 增加默认的ID字段
		fields_strings.append("`ID` INT NOT NULL AUTO_INCREMENT COMMENT 'ID'")

		# 增加各字段
		for i in range(len(fields)):
			s	= self.generate_field_string(fields[i])
			if s:
				fields_strings.append(s)

		# 增加默认的时间戳字段
		fields_strings.append("`CREATE_TIMESTAMP` TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP COMMENT '创建时间戳'")
		fields_strings.append("`UPDATE_TIMESTAMP` TIMESTAMP ON UPDATE CURRENT_TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP COMMENT '更新时间戳'")
		fields_strings.append("`DELETE_TIMESTAMP` TIMESTAMP NULL DEFAULT NULL COMMENT '删除时间戳，NULL表示未删除'")
	
		fields_strings.append('PRIMARY KEY(`ID`)')
		SQL_string	= f"CREATE TABLE `{self.database_config['database']}`.`{name}` (	\
				{','.join(fields_strings)}					\
				)  CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci ENGINE=InnoDB" 
		
		# 表注释
		if comment:
			SQL_string	+=  f" COMMENT='{self.replace(comment)}'"
	
		self._execute(SQL_string)
	
	## 修改/增加字段
	#	field	要修改/增加的字段属性，只包含要修改的属性，name必须要有
	def update_field(self, table, field, first=False):
		# field中必须要有name
		if 'name' not in field:
			return None
		
		# 读取原有字段信息
		field_old	= self.get_field(table, field['name'])
		
		if field_old:
			# 该字段已经存在	
			SQL_string	= f"ALTER TABLE `{table}` CHANGE `{field['name']}` "
		else:	
			# 需要添加字段
			SQL_string	= f"ALTER TABLE `{table}` ADD "
			field_old	= self.blank_field_string()
			field_old['name']	= field['name']
			field_old['MySQL_Type']	= field['MySQL_Type']

		if 'type' in field:
			field_old['type']	= field['type']
			if 'length' in field:
				field_old['length']	= field['length']
			else:
				field_old['length']	= None
			# 实际上是使用这个来决定数据库字段类型的
			field_old['MySQL_Type']		= field['MySQL_Type']
		
		if 'default' in field:
			field_old['default']	= field['default']
		
		if 'null' in field:
			field_old['null']	= field['null']

		if 'comment' in field:
			field_old['comment']	= field['comment']
		
		SQL_string	+= self.generate_field_string(field_old)

		if first:
			# 排到首位去
			SQL_string	+= " FIRST "

		self._execute(SQL_string)

		# 唯一性要单独处理
		if 'unique' in field:
			if field_old['unique']!= field['unique']:
				if field['unique']:	
					# 增加唯一性
					self._execute(f"ALTER TABLE `{table}` ADD UNIQUE (`{field['name']}`)")
				else:	
					# 去掉唯一性
					self._execute(f"ALTER TABLE `{table}` DROP INDEX `{field['name']}`")

		# json类型修改排序规则
		if field['type'] in ["LONGTEXT", "JSON"]:
			self._execute(f"ALTER TABLE `{table}` CHANGE `{field['name']}` `{field['name']}` LONGTEXT CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci NULL DEFAULT NULL COMMENT '{field['comment']}'")

	## 列出/删除此数据库中所有的位于DELETE_TIMESTAMP之后的字段
	# 在导入excel文件后，将会把excel内没有的字段移动到DELETE_TIMESTAMP之后，调用此函数检查无误后，再调用此函数删除	
	def obsolete_fields(self, delete=False):
		tables	= self.get_schema()
		
		result	= []
		for t in tables:
			field_names	= []
			for f in t['fields']:
				field_names.append(f['name'])
			i	= field_names.index('DELETE_TIMESTAMP')
			for k in range(i+1, len(field_names)):
				result.append(f"{t['table_name']:20}{t['fields'][k]['name']:20}{t['fields'][k]['comment']}")
		return result
	
	# ************************************************************************
	# 				EXCEL 输入输出
	# ************************************************************************
	## 从excel导入结构
	def import_excel_schema(self, filename):
		# 读取文件
		try:
			book	= openpyxl.load_workbook(filename)
			sheet	= book['数据库表结构']			
		except Exception as ex:
			self.logger.error(str(ex))
			return None

		tables	= []
		for row in range(2, sheet.max_row+1):
			if not self.cell_string(sheet, row, 1):
				# 第一列为空：表名和注释
				tables.append({
					'table_name'	: self.cell_string(sheet, row, 2),
					'table_comment'	: self.cell_string(sheet, row, 3),
					'constraints'	: self.cell_string(sheet, row, 4),
					'fields'	: [],
				})
			else:
				type, length	= self.excel_type_string[self.cell_string(sheet, row, 4)]
				default	= self.cell_string(sheet, row, 5) 
				if not default:
					default	= "NULL"
				
				MySQL_Type	= type
				if length:
					# 类型可能附带长度，要加小括号附在类型后
					MySQL_Type	+= f"({length})"
				elif type in ("CHAR", "VARCHAR"):
					# CHAR/VARCHAR只使用255长度
					MySQL_Type	+= "(255)"
				
				tables[-1]['fields'].append({
					'name'		: self.cell_string(sheet, row, 2),	# 字段名
					'type'		: type,					# 字段类型INT/DOUBLE/DECIMAL/CHAR/TEXT/DATETIME
					'length'	: length,				# 字段类型长度DECIMAL(10,2)/CHAR(255)
					'default'	: default,				# 默认值
					'null'		: self.cell_string(sheet, row, 6)=='' ,	# 是否可以为空（只要有任何值都表示不能为空）
					'unique'	: False,				# 是否唯一键
					'comment'	: self.cell_string(sheet, row, 3),	# 注释
					'MySQL_Type'	: MySQL_Type,				# MySQL字段类型字符串
				})
		
		current_tables	= self.get_schema()
		
		if current_tables==None:
			# 读取失败
			self.logger.error("读取表结构失败")
			return None
		
		# 当前数据库表名列表
		current_table_names	= []
		for t in current_tables:
			current_table_names.append(t['table_name'])
		
		# Excel中的表名列表
		table_names		= []
		for t in tables:
			table_names.append(t['table_name'])

		# 删除excel中不存在的表
		for t in current_table_names:
			if t not in table_names:
				self._execute(f"DROP TABLE {t}")
		
		# 按excel逐个表处理
		for t in tables:
			try:
				i	= current_table_names.index(t['table_name'])
			except Exception:
				i	= -1

			if i==-1:
				# 表还不存在，创建表
				self.new_table(t['table_name'], t['fields'], t['table_comment'])
				
				self.logger.info(f"创建了表{t['table_name']}({t['table_comment']})")
			else:
				# 表已经存在，需要更新字段：只增加不删除
				# Excel中没有字段将在现有字段的备注上标注待删除并移动到末尾
				# 	注意current_tables和current_table_names是顺序一致的
				self.logger.info(f"更新表{t['table_name']}")
				
				# 表中必须有四个默认字段：ID, CREATE_TIMESTAMP, UPDATE_TIMESTAMP, DELETE_TIMESTAMP
				for field_name in ['ID', 'CREATE_TIMESTAMP', 'UPDATE_TIMESTAMP', 'DELETE_TIMESTAMP']:
					result	= self.get_field(t['table_name'], field_name)
					if not result:
						# 该字段不存在，需要创建
						if field_name=='ID':
							# ID必须存在
							pass
						elif field_name=='CREATE_TIMESTAMP':
							# 增加默认的时间戳字段CREATE_TIMESTAMP
							self._execute(f"ALTER TABLE `{t['table_name']}` ADD`CREATE_TIMESTAMP` TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP COMMENT '创建时间戳'")
						elif field_name=='UPDATE_TIMESTAMP':
							# 增加默认的时间戳字段UPDATE_TIMESTAMP
							self._execute(f"ALTER TABLE `{t['table_name']}` ADD`UPDATE_TIMESTAMP` TIMESTAMP ON UPDATE CURRENT_TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP COMMENT '更新时间戳'")
						elif field_name=='DELETE_TIMESTAMP':
							# 增加默认的时间戳字段DELETE_TIMESTAMP
							self._execute(f"ALTER TABLE `{t['table_name']}` ADD`DELETE_TIMESTAMP` TIMESTAMP NULL DEFAULT NULL COMMENT '删除时间戳，NULL表示未删除'")							
									
				# 重新排序：把ID+t['fields']+CREATE_TIMESTAMP+UPDATE_TIMESTAMP+DELETE_TIMESTAMP倒序放到开头
				# 排序过程中会使用修改的字段定义，增加excel中新的字段，在excel中没有的字段将会被移动到末尾
				# 如果表中没有这默认三字段，将会导致数据库操作返回错误
				action	= 'MODIFY' 
				self._execute(f"ALTER TABLE `{t['table_name']}` {action} `DELETE_TIMESTAMP` TIMESTAMP NULL DEFAULT NULL COMMENT '删除时间戳，NULL表示未删除' FIRST")
				self._execute(f"ALTER TABLE `{t['table_name']}` {action} `UPDATE_TIMESTAMP` TIMESTAMP ON UPDATE CURRENT_TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP COMMENT '更新时间戳' FIRST")
				self._execute(f"ALTER TABLE `{t['table_name']}` {action} `CREATE_TIMESTAMP` TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP COMMENT '创建时间戳' FIRST")
				for i in range(len(t['fields'])-1, -1, -1):
					f	= t['fields'][i]
					self.update_field(t['table_name'], f, first=True)
				self._execute(f"ALTER TABLE `{t['table_name']}` MODIFY `ID` INT(11) AUTO_INCREMENT FIRST")

				# 查询约束
				constraints	= self.query(f"SELECT * FROM information_schema.`TABLE_CONSTRAINTS` \
					WHERE `TABLE_SCHEMA`='{self.database_config['database']}' \
						AND `TABLE_NAME`='{t['table_name']}'\
						AND `CONSTRAINT_TYPE`='UNIQUE'	")
				# 去掉约束
				for c in constraints:
					self._execute(f"ALTER TABLE `{t['table_name']}` DROP INDEX `{c['CONSTRAINT_NAME']}`")
				# 更新表的注释
				self._execute(f"ALTER TABLE `{t['table_name']}`  COMMENT '{t['table_comment']}'")
			# 更新约束
			if t['constraints']!='':
				# 分号划分
				constraints	= t['constraints'].split(';')
				for c in constraints:
					name, columns	= c.split(':')
					columns	= columns.split(',')
					self._execute(f"ALTER TABLE `{t['table_name']}` ADD UNIQUE `{name}`(`{'`,`'.join(columns)}`)")
				
	## 导出结构到excel
	def export_excel_schema(self, filename=None):
		if not filename:
			# 使用默认文件名
			filename	= f"数据库{self.database_config['database']}设计.xlsx"

		book	= openpyxl.Workbook()
		
		# SHEET 数据库表
		sheet		= book.active
		sheet.title	= '数据库表结构'
			
		# 标题
		row	= 1
		self.set_cell(sheet, row, 1, '序号',	**self.cell_style_title)	
		self.set_cell(sheet, row, 2, '字段名',	**self.cell_style_title)	
		self.set_cell(sheet, row, 3, '说明',	**self.cell_style_title)	
		self.set_cell(sheet, row, 4, '类型',	**self.cell_style_title)	
		self.set_cell(sheet, row, 5, '默认值',	**self.cell_style_title)	
		self.set_cell(sheet, row, 6, '必填',	**self.cell_style_title)	
		self.set_cell(sheet, row, 7, '备注',	**self.cell_style_title)	

		# 获取表结构
		tables	= self.get_schema()
		row	= 2
		# 设置格式验证，产生下拉菜单
		validation	= openpyxl.worksheet.datavalidation.DataValidation(
			type		= 'list',
			formula1	='"'+','.join(self.excel_type_string.keys())+'"',
			allow_blank	= False,
		)
				
		if tables:
			# 逐个表生成
			for t in tables:
				# 表名和表注释
				self.set_cell(sheet, row, 2, t['table_name'], **self.cell_style_table_name)
				self.set_cell(sheet, row, 3, t['table_comment'], **self.cell_style_table_comment)
				# 约束
				if 'constraints' in t:
					constraints	= []
					for c in t['constraints']:
						if c != 'PRIMARY':
							constraints.append(f"{c}:{','.join(t['constraints'][c])}")

					self.set_cell(sheet, row, 4, ';'.join(constraints), **self.cell_style_table_comment)
				row	+= 1

				# 各字段
				for i in range(len(t['fields'])):
					f	= t['fields'][i]
					if f['name'] not in self.default_field_names:
						self.set_cell(sheet, row, 1, i, 		**self.cell_style_ID)
						self.set_cell(sheet, row, 2, f['name'],	**self.cell_style_default)
						self.set_cell(sheet, row, 3, f['comment'],	**self.cell_style_default)
						# 需要将MySQL中的字段名转换为Excel中的统一名称
						self.set_cell(sheet, row, 4, self.MySQL_type_string[f['type']],	**self.cell_style_default)
						
						# 此单元格需要验证
						validation.add(f'D{row}')

						if f['default']!="NULL":
							self.set_cell(sheet, row, 5, f['default'],	**self.cell_style_default)
						if f['null']==False:
							self.set_cell(sheet, row, 6, '●', **self.cell_style_ID)
						row	+= 1
			# 添加验证
			sheet.add_data_validation(validation)
		else:
			self.logger.info("没有数据库表")
			return

		# 自动调整列宽
		self.adjust_all_column_width(sheet)
		
		# # SHEET 数据库表
		sheet	= book.create_sheet('数据库表', 1)
		self.set_cell(sheet, 1, 1, '表名', **self.cell_style_title)
		self.set_cell(sheet, 1, 2, '说明', **self.cell_style_title)
		for i in range(len(tables)):
			# 填写excel自动计算，但这会让速度变慢
			#self.excel_cell_fill(sheet, i+2, 1, f'=INDEX(数据库表结构!$B:$B,SMALL(IF(数据库表结构!$A:$A="",ROW(数据库表结构!$A:$A),65536),ROW(A{i+2})))')
			#self.excel_cell_fill(sheet, i+2, 2, f'=INDEX(数据库表结构!$C:$C,SMALL(IF(数据库表结构!$A:$A="",ROW(数据库表结构!$A:$A),65536),ROW(A{i+2})))')
			# 直接填写表名
			self.set_cell(sheet, i+2, 1, tables[i]['table_name'], 		**self.cell_style_table_comment)
			self.set_cell(sheet, i+2, 2, tables[i]['table_comment'],	**self.cell_style_default)
		# # 自动调整列宽
		self.adjust_all_column_width(sheet)

		# SHEET 字段类型
		sheet	= book.create_sheet('字段类型', 1)
		self.set_cell(sheet, 1, 1, '字段类型', 	**self.cell_style_title)
		self.set_cell(sheet, 1, 2, 'MySQL类型', **self.cell_style_title)
		self.set_cell(sheet, 1, 3, '备注', 	**self.cell_style_title)
		row	= 2
		for k in self.excel_type_string:
			self.set_cell(sheet, row, 1, k,					**self.cell_style_default)
			self.set_cell(sheet, row, 2, self.excel_type_string[k][0],	**self.cell_style_default)
			self.set_cell(sheet, row, 3, self.excel_type_string[k][1],	**self.cell_style_default)
			row	+= 1
	
		# # 自动调整列宽
		self.adjust_all_column_width(sheet)

		# 保存
		book.save(filename)

	## 导出数据到excel
	# 每个SHEET是一个表，SHEET名称就是表名称，第一行总是字段名称
	#	table_names	表名列表[]
	def export_excel_data(self, table_names=None, filename=None):
		if not filename:
			# 使用默认文件名
			filename	= f"数据库{self.database_config['database']}数据 {time.strftime('%Y-%m-%d %H_%M_%S')}.xlsx"

		book	= openpyxl.Workbook()
		# 新建有一个sheet，删除之
		book.remove(book.active)

		if isinstance(table_names, str):
			table_names	= [table_names]

		if table_names:
			tables	= []
			for t in table_names:
				tt	= self.get_table(t)
				if tt:
					tables.append(tt)
		else:
			# 没有指定表名就导出全部表
			tables	= self.get_schema()
		for t in tables:
			sheet		= book.create_sheet(t['table_name'])
			field_names	= []
			i		= 1
			for f in t['fields']:
				# 标题行：字段名
				self.set_cell(sheet, 1, i, f['name'], **self.cell_style_title)
				comment				= openpyxl.comments.Comment(f['comment'], author="EXON")
				comment.width			= 200
				comment.hight			= 25				
				sheet.cell(1, i).comment	= comment
				i				+= 1
				field_names.append(f['name'])
			# 查询数据
			row	= 2

			result	= self.query(f"SELECT * FROM `{t['table_name']}`")

			for r in result:
				for i in range(len(field_names)):
					value		= r[field_names[i]]
					type		= t['fields'][i]['type']
					number_format	= 'General'
					if value!=None and type in self.digital_field_type[1:]:
						# 数值类型
						if type=="INT":
							value	= int(r[field_names[i]])
						else:
							value	= float(r[field_names[i]])
					elif type in self.timestamp_field_type:
						# 时间类
						number_format	= "yyyy-mm-dd hh:mm:ss"

					if value != None and value != 'NULL':
						self.set_cell(sheet, row, i+1, value, number_format=number_format)
					
				row	+= 1
			# 自动调整列宽
			self.adjust_all_column_width(sheet)
		# 保存
		book.save(filename)	

	## 从excel导入数据
	# 	filename	excel文件名
	#	overwrite	是否覆盖已有数据，默认False，将excel中的数据添加到末尾，若ID已经存在则更新数据；True则先清除该表
	def import_excel_data(self, filename, overwrite=False):
		# 读取文件
		try:
			book	= openpyxl.load_workbook(filename)
		except Exception as ex:
			self.logger.error(str(ex))
			return None
		
		for sheet_name in book.sheetnames:
			sheet	= book[sheet_name]
			# SHEET名称必须是表名称
			table	= sheet.title

			if overwrite:
				# 清除表中数据
				self._execute(f"TRUNCATE `{table}`")

			# 首行必须是字段名称
			field_names	= []
			for i in range(sheet.max_column):
				field_names.append(str(sheet.cell(1, i+1).value))
			# 插入，若ID存在则将更新该记录
			records	= []
			for r in range(2, sheet.max_row+1):
				record	= {}
				for k in range(len(field_names)):
					value	= sheet.cell(r, k+1).value
					record[field_names[k]]	= value if value!=None and value!='' else None
				records.append(record)
			self.logger.info(f"导入数据{len(records)}条到表{table}")
			self.insert(table, records)
		
	## ----------------------------------- 配置和定义 -----------------------------------	
	# excel中的字段名称，值是MySQL中对应的类型和长度
	excel_type_string	= {
		"STRING"	: ["CHAR", 	"255"],
		"INT"		: ["INT",	"11"],
		"JSON"		: ["JSON", 	None],		# 在 MariaDB 的实现中， JSON 类型为 longtext 类型的别名
		"DATETIME"	: ["DATETIME", 	None],
		"FLOAT"		: ["DECIMAL", 	"10,5"],	# FLOAT实际上是10.5的DECIMAL，而不是浮点数
		"DECIMAL"	: ["DECIMAL", 	"10,2"],
		"DOUBLE"	: ["DOUBLE", 	None],
		"BLOB"		: ["LONGBLOB", 	None],		# BLOB都是LONGBLOB
		"TEXT"		: ["TEXT", 	None],		# 文本只有小于255和TEXT两种
		"TIMESTAMP"	: ["TIMESTAMP", None],
	}

	# MySQL中读取的字段名称，值是excel中的对应名称
	MySQL_type_string	= {
		"CHAR"					: "STRING",
		"DATE"					: "DATETIME",
		"DATETIME"				: "DATETIME",
		"DOUBLE"				: "DOUBLE",
		"FLOAT"					: "DECIMAL",	# FLOAT实际上是10.5的DECIMAL，而不是浮点数
		"DECIMAL"				: "DECIMAL",
		"INT"					: "INT",
		"BLOB"					: "BLOB",	# BLOB都是LONGBLOB
		"LONGBLOB"				: "BLOB",	# BLOB都是LONGBLOB
		"LONGTEXT"				: "JSON",	# 在 MariaDB 的实现中， JSON 类型为 longtext 类型的别名
		"MEDIUMBLOB"				: "BLOB",
		"TEXT"					: "TEXT",
		"TIMESTAMP"				: "TIMESTAMP",
		"TIMESTAMP ON UPDATE CURRENT_TIMESTAMP"	: "TIMESTAMP",
		"VARCHAR"				: "STRING",
		"JSON"					: "JSON",
	}

	# 单元格对齐：只考虑三种基本对齐方式
	alignment_center	= openpyxl.styles.Alignment( horizontal	= 'center',	vertical = 'center',	wrap_text	= False)
	alignment_left		= openpyxl.styles.Alignment( horizontal	= 'left',	vertical = 'center',	wrap_text	= False)
	alignment_right		= openpyxl.styles.Alignment( horizontal	= 'right',	vertical = 'center',	wrap_text	= False)

	# 默认单元格样式
	cell_style_default	= {
		'font'		: openpyxl.styles.Font(name='Calibri', color='000000', size=11),
	}

	# 标题的单元格样式
	cell_style_title	= {
		'font'		: openpyxl.styles.Font(name='Calibri', color='FFFFFF', size=11),
		'fill'		: openpyxl.styles.PatternFill(fill_type='solid', fgColor='FF4169E1'),
		'alignment'	: alignment_center, 
	}

	# 表名的单元格样式
	cell_style_table_name	= {
		'font'		: openpyxl.styles.Font(name='Calibri', color='FFFFFF', size=11),
		'fill'		: openpyxl.styles.PatternFill(fill_type='solid', fgColor='FFFC760F'),
	}

	# 表注释单元格样式
	cell_style_table_comment	= {
		'font'		: openpyxl.styles.Font(name='Calibri', color='FC760F', size=11),
	}

	# ID单元格样式
	cell_style_ID		= {
		'font'		: openpyxl.styles.Font(name='Calibri', color='4169E1', size=11)
	}
	

# 调试/测试代码
if __name__ == '__main__':
	pass