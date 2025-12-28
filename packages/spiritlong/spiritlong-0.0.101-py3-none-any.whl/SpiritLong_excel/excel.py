#!/usr/bin/python3
# coding=utf-8
###################################################################
#           ____     _     _ __  __                 
#          / __/__  (_)___(_) /_/ /  ___  ___  ___ _
#         _\ \/ _ \/ / __/ / __/ /__/ _ \/ _ \/ _ `/
#        /___/ .__/_/_/ /_/\__/____/\___/_//_/\_, / 
#           /_/                              /___/  
# Copyright (c) 2024 Chongqing Spiritlong Technology Co., Ltd.
# All rights reserved.  
# @author	arthuryang
# @brief	excel工具集，只支持xlsx
#
###################################################################  

import	re
import	os
import	platform
import	datetime
import	chardet
import	openpyxl
import	openpyxl.styles
import	chardet
import	csv

################################### 格式/样式 ###################################
## 字体：主要只使用了字体名称、大小、颜色和加粗。默认黑色11号Calibri，英文数字较美观
def cell_font(name='Calibri', size=11, color='000000', bold=False):
	return openpyxl.styles.Font(
		name		= name,
		size		= size,
		bold		= bold,
		vertAlign	= None,
		underline	= None,
		strike		= False,
		color		= color,
	)

## 背景填充。默认白色
def cell_fill(color='FFFFFFFF'):
	return openpyxl.styles.PatternFill(
		fill_type	= 'solid',
		fgColor		= color,
	)

## 单元格边界
#	sides	字符串，可以是left/right/top/bottom中的一个或多个，英文逗号分隔，None或""表示四边全有
def cell_border(color='000000', style='thin', sides=None):
	border_sides	= {
			'left'		: None,
			'right'		: None,
			'top'		: None,
			'bottom'	: None,
	}
	side		= openpyxl.styles.Side(border_style=style, color=color)
	if not sides:
		# 默认四边都有
		sides	= 'left, right, top, bottom'
	
	sides	= sides.split(',')
	for s in sides:
		ss	= s.strip()
		if ss in border_sides:
			border_sides[ss]	= side

	return openpyxl.styles.Border(**border_sides)

# 单元格对齐：只考虑三种基本对齐方式
alignment_center	= openpyxl.styles.Alignment( horizontal	= 'center',	vertical = 'center',	wrap_text	= False)
alignment_left		= openpyxl.styles.Alignment( horizontal	= 'left',	vertical = 'center',	wrap_text	= False)
alignment_right		= openpyxl.styles.Alignment( horizontal	= 'right',	vertical = 'center',	wrap_text	= False)

# 标题样式
style_title	= {
	'font'		: cell_font(color='FFFFFF'),
	'fill'		: cell_fill(color='4169E1'),
	'border'	: None,
	'alignment'	: alignment_center,
}

################################### 操作 ###################################
## 打开xlsx文件，返回book对象或None，默认只读（会大幅提升速度），使用公式
#	filename	None表示创建
def open_xlsx(filename=None, read_only=True, data_only=False):
	file_exists	= os.path.exists(filename)
	try:
		if file_exists and filename:
			book	= openpyxl.load_workbook(filename, read_only=read_only, data_only=data_only)
		else:
			# 创建新的对象
			book	= openpyxl.Workbook()
			# 去掉默认的Sheet
			book.remove(book.active)
		return book
	except Exception as ex:
		print(str(ex))
		return None

## 关闭文件，指定了文件名则需要保存
def close_xlsx(book, filename=None):
	if filename:
		book.save(filename)
	# read_only模式下需要显式调用
	book.close()

## 设置单元格样式
def set_cell_format(sheet, i, j, number_format='General', font=None, fill=None, border=None, alignment=None):
	cell		= sheet.cell(row=i, column=j)

	# 如果是超过10位的数值长字符串，则强制设为文本类型
	if isinstance(cell.value, str) and len(cell.value)>=10:
		number_format	= '@'

	if isinstance(cell.value, datetime.datetime) and number_format=='General':
		# 日期时间类需要设置格式默认值
		number_format	= 'yyyy-mm-dd hh:mm:ss'

	cell.number_format	= number_format
	cell.font		= font if font is not None else cell_font()
	# 默认无填充
	cell.fill		= fill if fill is not None else openpyxl.styles.PatternFill(fill_type	= None)
	# 默认左对齐
	cell.alignment		= alignment if alignment is not None else alignment_left
	# 默认无边界
	cell.border		= border if border is not None else openpyxl.styles.Border()

## 设置单元格的值
def set_cell_value(sheet, i, j, value):
	cell		= sheet.cell(row=i, column=j)
	# 如果是超过10位的数值长字符串，则强制设为文本类型
	if isinstance(value, str) and len(value)>=10:
		cell.number_format	= '@'

	# 日期时间类需要设置格式默认值
	if isinstance(value, datetime.datetime) and cell.number_format=='General':
		cell.number_format	= 'yyyy-mm-dd hh:mm:ss'
	
	cell.value	= value

## 设置单元格的值和格式
def set_cell(sheet, i, j, value, number_format='General', font=None, fill=None, border=None, alignment=None):
	cell		= sheet.cell(row=i, column=j)
	cell.value	= value
	set_cell_format(sheet, i, j, number_format, font, fill, border, alignment)

## 读取数据，每个单元格一个数据
#	sheet		表格sheet对象
# 返回一个双重列表，最里面每个成员是一个单元格的值
def get_cells(sheet):
	cells	= []
	# 使用rows比逐个cell读取要快很多
	for row in sheet.rows:
		cells.append([c.value for c in row])
	return cells

## 读取带标题行的数据，每一行是一个记录，可以指定标题行位置
#	sheet		表格sheet对象
#	title_row	标题行所在行，从1开始
# 返回[{}]，列表成员是一个记录，每个记录是一个字典
def get_records_with_title(sheet, title_row=1):
	records	= []
	# 使用rows比逐个cell读取要快很多
	for line, row in enumerate(sheet.rows):
		if line+1<title_row:
			continue
		if line+1==title_row:
			# 标题行
			titles	= [c.value for c in row]
		else:
			records.append({titles[i]:c.value for i, c in enumerate(row)})

	return records

## 自动调整指定列的宽度
#	column		列（从1开始）
#	width		0自动
#	max_width	最大宽度限制
def adjust_column_width(sheet, column, width=0, max_width=100):
	# 自动计算最大宽度
	if width==0:
		width	= 1
		for row in range(1, sheet.max_row+1):
			# 对于中文字符和非中文字符单独计算宽度
			s	= sheet.cell(row, column).value
			s	= str(s) if s is not None else ""
			w	= 0
			for c in s:
				w	+= (1 if ord(c)<128 else 2)
			width	= max(width, w+1)

	# 列宽得有最大限制
	if width>max_width:
		width	= max_width

	sheet.column_dimensions[openpyxl.utils.get_column_letter(column)].width	= width

## 自动调整所有列的宽度
def adjust_all_column_width(sheet, max_width=100):
	for column in range(1, sheet.max_column+1):
		adjust_column_width(sheet, column, max_width=max_width)

## 将'A1'转换为（1,1)
def cell_coordinate(cell_string):
	column_letters, row	= openpyxl.utils.cell.coordinate_from_string(cell_string)
	column			= openpyxl.utils.cell.column_index_from_string(column_letters)
	return (row, column)

## 将（1，1）转换为'A1'
def cell_code(row, column):
	return f"{openpyxl.utils.cell.get_column_letter(column)}{row}"

## 将记录保存到sheet
#	 	sheet 		sheet对象
#	{}	sheet_data	记录数据：{titles:{key:列名}或[列名], records=[{key:value}]
# 	{}	style_title	标题样式：{font,fill,border,alignment,number_format}
# 	{}	style_text	内容样式：{font,fill,border,alignment,number_format}	
# sheets中的titles为[列名]时，表示key和列名一致。records中的key可以和列名不一致，此时必须要在titles中指定{列名:key}
# records可以为空，此时将创建一个空的文件作为模板
def records_to_sheet(sheet, sheet_data, style_title=style_title, style_text={}):
	# 判断是否有数据
	titles_valid	= 'titles' in sheet_data and sheet_data['titles']
	data_valid	= 'records' in sheet_data and sheet_data['records']
	# 字段名
	if not titles_valid:
		# 未提供则使用数据中出现了的字段名
		if data_valid:
			titles	= {k:k for r in sheet_data['records'] for k in r}
		else:
			# 既无标题又无数据
			titles	= []
	elif isinstance(sheet_data['titles'], list):
		# []则表示字段名和标题名相同
		titles	= {t:t for t in sheet_data['titles']}
	elif isinstance(sheet_data['titles'], dict):
		# {key:列名}
		titles	= sheet_data['titles']
	else:			
		print(f"sheet_data['titles']必须为标题列表、字典或None")
		return

	# 写入标题
	if titles:
		for i, t in enumerate(titles.values()):
			set_cell(sheet, 1, i+1, t, **style_title)
	
	# 写入内容
	if data_valid:
		for i, r in enumerate(sheet_data['records']):
			for j, field in enumerate(titles):
				if field in r.keys():
					set_cell(sheet, i+2, j+1, r[field], **style_text)	

	# 自动调整列宽
	if data_valid :
		adjust_all_column_width(sheet)
		
## 保存记录到excel文件
#	str 		filename 	文件名,带路径
#	{sheet_name:{}}	data		字典中每个键对应于一个sheet，其值是{titles:{key:列名}或[列名], records:[{key:value}]
# 	{}		style_title	标题样式,{font,fill,border,alignment,number_format}
# 	{}		style_text	内容样式,{font,fill,border,alignment,number_format}	
# sheets中的titles为[列名]时，表示key和列名一致。records中的key可以和列名不一致，此时必须要在titles中指定{列名:key}
# titles可以为空，此时将按records中的数据来生成列名
# records可以为空，此时将创建一个空的文件作为模板
def records_to_excel(filename, data, style_title=style_title, style_text={}):
	if not data:
		return
	
	book 	= openpyxl.Workbook()
	# 新建有一个sheet，删除之
	book.remove(book.active)
	
	for sheet_name in data:
		sheet	= book.create_sheet(sheet_name)
		records_to_sheet(sheet, data[sheet_name], style_title, style_text)
	# 保存文件，确保是xlsx后缀名
	if not filename.endswith(".xlsx"):
		filename	+= ".xlsx"
	book.save(filename)

## CSV文件转换成xlsx
def csv_save_as_xlsx(csv_file, xlsx_file=None, overwrite=False):
	filename, extesion	= os.path.splitext(csv_file)
	if extesion!='.csv':
		return
	if xlsx_file is None:
		xlsx_file	= filename+'.xlsx'

	if overwrite==False and os.path.exists(xlsx_file):
		# 若已经存在则不转换，除非被强制要求
		return

	# 先检查编码
	with open (csv_file,'rb') as f:
		encoding	= chardet.detect(f.read())["encoding"]
	# 确保用GBK，GB2312支持的汉字较少
	encoding	= 'GBK' if encoding=='GB2312' else encoding

	book	= openpyxl.Workbook()
	sheet	= book.active
	
	with open(csv_file, encoding=encoding) as f:
		reader	= csv.reader(f)
		for row in reader:
			sheet.append(row)
	book.save(xlsx_file)

if platform.system()=="Windows":
	import xlwings

## 将xls转换为xlsx文件
#	xls_file	要转换的xls
#	xlsx_file	若为None则和xls取相同文件名的xlsx文件
def xls_save_as_xlsx(xls_file, xlsx_file=None, overwrite=False):
	filename, extesion	= os.path.splitext(xls_file)
	if extesion!='.xls':
		return
	if xlsx_file is None:
		xlsx_file	= filename+'.xlsx'

	if overwrite==False and os.path.exists(xlsx_file):
		# 若已经存在则不转换，除非被强制要求
		return
	
	# 注意：只能在windows下运行
	if platform.system()!="Windows":
		print("此功能只能在windows平台使用")
	else:
		xlwings_app = xlwings.App(visible=False, add_book=False)
		work_book = xlwings_app.books.open(xls_file, read_only=True)
		work_book.save(xlsx_file)
		work_book.close()
		xlwings_app.quit()

		print(f"{xls_file}已经转换并保存到：{xlsx_file}")
	return

## 读取带不定长参数的函数列表，每一行的第一列若为空则表示接续前一行，这一组行一起是一个记录，从而在一个记录中可以放多组多个参数（每个参数的名称/说明等占一行）
#	sheet			xlsx sheet
#	parameter_fields	这是一个字典，每个key是参数组的名称，value是一个有两个元素的列表，记录了参数部分字段范围起始（从1开始，只有一个列则起始相同），这意味着可以有多个参数部分（比如输入参数和返回项）
# 返回函数列表，每个成员是一个表示函数的字典，字典中parameter_fields所指定的参数（key是参数组名称）列表
def get_records_with_parameters(sheet, parameter_groups):
	# 把记录按首列是否为空进行分组
	sheet_records	= get_records_with_title(sheet)
	# 第一行是标题
	titles		= {i:sheet.cell(1, i).value for i in range(1, sheet.max_column+1)}

	# 参数部分字段
	if not sheet_records:
		return None
	for name, start_end in parameter_groups.items():
		start, end	= start_end
		parameter_groups[name]	= ([titles[i] for i in titles if i>=start and i<=end])
	# 所有的参数字段
	parameter_titles	= []
	for t in parameter_groups.values():
		parameter_titles	+= t

	# 分组
	records	= []
	for i in range(len(sheet_records)):
		if sheet_records[i][titles[1]]:
			# 第一列不为空是新的开始
			record	= {t:sheet_records[i][t] for t in sheet_records[i] if t not in parameter_titles}
			for name, group_titles in parameter_groups.items():
				record[name]	= []
			records.append(record)
		# 添加参数
		for name, group_titles in parameter_groups.items():
			if sheet_records[i][group_titles[0]]:
				# 参数组的第一个不为空就要加入
				record[name].append({t:sheet_records[i][t] for t in sheet_records[i] if t in group_titles})

	return records

## 读取表/字段配置（首行是标题），一个表用多行来表示，这些行的第一行是表名称和说明，后面每行对应该表的一个字段。第一列是表序号，序号为空表示该行是表的字段描述。
#	sheet		xlsx sheet
# 返回表字典
def get_records_with_table_fields(sheet):
	# 把记录按首列是否为空进行分组
	sheet_records	= get_records_with_title(sheet)
	# 第一行是标题
	titles		= {i:sheet.cell(1, i).value for i in range(1, sheet.max_column+1)}

	# 分组
	table_lines	= []
	for i in range(len(sheet_records)):
		if sheet_records[i][titles[1]]:
			# 第一列不为空是一个表
			table_lines.append(i)
	# 为了计算方便
	table_lines.append(len(sheet_records))

	tables	= []
	for i, t in enumerate(table_lines):
		if i+1==len(table_lines):
			# 最后一个不是表
			break

		tables.append({
			'表名'	: sheet_records[t][titles[2]],
			# 注意字段是从本行的下一行开始的
			'字段'	: [{
				'字段名'	: sheet_records[f][titles[2]],
				'说明'		: sheet_records[f][titles[3]],
				'类型'		: sheet_records[f][titles[4]],
				'备注'		: sheet_records[f][titles[5]],
			} for f in range(t+1, table_lines[i+1])],
			'说明'	: sheet_records[t][titles[3]],
			'备注'	: sheet_records[t][titles[5]],
		})

	return tables

# 调试/测试代码
if __name__ == '__main__':
	pass