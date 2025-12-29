import argparse
import glob
import math
import multiprocessing
import os
import statistics
import sys
from datetime import datetime
from pathlib import Path
import time
import shutil
import gzip
import pandas as pd
from tqdm import tqdm
import subprocess
from joblib import Parallel, delayed
from Bio import SeqIO
import hashlib
from collections import defaultdict
import platform
from Bio.SeqIO import SeqRecord
from Bio.Seq import reverse_complement, Seq
import numpy as np
import plotly.graph_objects as go
import plotly.express as px
import subprocess, sys, os
from openpyxl import load_workbook

# project_folder = Path('/Volumes/Coruscant/APSCALE_projects/naturalis_dataset_apscale_nanopore')
# project_settings_files = '/Users/tillmacher/Documents/GitHub/apscale_nanopore/apscale_nanopore/default_settings.xlsx'
# settings_df = pd.read_excel(project_settings_files, sheet_name='Settings')
# demultiplexing_df = pd.read_excel(project_settings_files, sheet_name='Demultiplexing').fillna('')
# cpu_count = "auto"

def check_cpu_count(cpu_count):
    max_cpu = multiprocessing.cpu_count()
    if cpu_count == 'auto':
        return multiprocessing.cpu_count() - 2
    elif cpu_count >= max_cpu:
        return multiprocessing.cpu_count() - 2
    else:
        return cpu_count

def check_dependencies(tools=["cutadapt", "vsearch", "swarm", "blastn"]):
    missing = []
    for tool in tools:
        if shutil.which(tool) is None:
            missing.append(tool)

    if missing:
        print(f'{datetime.now().strftime("%H:%M:%S")} - Missing required tools:')
        for tool in missing:
            print(f" - {tool}")
        print(f'{datetime.now().strftime("%H:%M:%S")} - Please install the missing dependencies and ensure they are in your PATH.')
        sys.exit(1)
    else:
        print(f'{datetime.now().strftime("%H:%M:%S")} - All required tools are installed and accessible.')

def quality_report(project_folder, sub_folder):
    # Set parameters
    folder = project_folder / sub_folder / 'data'
    output_folder = project_folder / '8_nanopore_report' / 'data'

    # Get all .fastq or .fastq.gz files
    fastq_files = folder.glob('*.fastq')
    fastq_gz_files = folder.glob('*.fastq.gz')

    cpu_count = multiprocessing.cpu_count()-2

    def analyse_reads(fastq_file, output_folder, compressed):
        """Analyze read qualities and lengths and create summary plots."""

        # Extract name
        name = fastq_file.name.replace('.fastq.gz', '') if compressed else fastq_file.name.replace('.fastq', '')

        # Create output folder
        outdir = output_folder / name
        os.makedirs(outdir, exist_ok=True)

        # Choose appropriate open method
        open_func = gzip.open if compressed else open

        mean_qualities = []
        read_lengths = []

        with open_func(fastq_file, "rt") as handle:
            for record in SeqIO.parse(handle, "fastq"):
                phred_scores = record.letter_annotations["phred_quality"]
                if phred_scores:
                    mean_q = math.ceil(np.mean(phred_scores))
                else:
                    mean_q = 0
                mean_qualities.append(mean_q)
                read_lengths.append(len(record.seq))

        df = pd.DataFrame({
            "MeanQuality": mean_qualities,
            "ReadLength": read_lengths
        })

        # ----------------------------------------
        # Plot 1: Distribution of Mean Phred Scores
        # ----------------------------------------
        fig1 = go.Figure()
        fig1.add_trace(go.Histogram(
            x=df["MeanQuality"],
            xbins=dict(start=0, end=60, size=1),
            marker_color='steelblue',
            name='Mean Quality',
            opacity=0.9
        ))
        # Add background color zones
        fig1.add_vrect(x0=0, x1=20, fillcolor="red", opacity=0.1, layer="below", line_width=0)
        fig1.add_vrect(x0=20, x1=30, fillcolor="yellow", opacity=0.1, layer="below", line_width=0)
        fig1.add_vrect(x0=30, x1=60, fillcolor="green", opacity=0.05, layer="below", line_width=0)

        fig1.update_layout(
            title='Distribution of Mean Phred Scores',
            xaxis_title='Mean Phred Score',
            yaxis_title='Read Count',
            template='simple_white',
            height=500,
        )
        fig1.write_html(outdir / f'{name}_mean_phred_distribution.html')

        # ----------------------------------------
        # Plot 2: Distribution of Read Lengths
        # ----------------------------------------
        fig2 = go.Figure()
        fig2.add_trace(go.Histogram(
            x=df["ReadLength"],
            xbins=dict(start=0, end=max(read_lengths) + 1, size=5),
            marker_color='indianred',
            name='Read Length',
            opacity=0.85
        ))

        fig2.update_layout(
            title='Distribution of Read Lengths',
            xaxis_title='Read Length (bp)',
            yaxis_title='Read Count',
            template='simple_white',
            height=500
        )
        fig2.write_html(outdir / f'{name}_read_length_distribution.html')

        # ----------------------------------------
        # Plot 3: Mean Quality vs. Read Length
        # ----------------------------------------
        fig3 = go.Figure()
        fig3.add_trace(go.Scatter(
            x=df["ReadLength"],
            y=df["MeanQuality"],
            mode='markers',
            marker=dict(size=4, opacity=0.5, color=df["MeanQuality"], colorscale='algae', showscale=True),
            name='Length vs Quality'
        ))

        # Add background coloring for y-axis quality zones
        fig3.add_shape(type="rect", x0=0, x1=max(read_lengths), y0=0, y1=20, fillcolor="red", opacity=0.05,
                       layer="below", line_width=0)
        fig3.add_shape(type="rect", x0=0, x1=max(read_lengths), y0=20, y1=30, fillcolor="yellow", opacity=0.05,
                       layer="below", line_width=0)
        fig3.add_shape(type="rect", x0=0, x1=max(read_lengths), y0=30, y1=60, fillcolor="green", opacity=0.03,
                       layer="below", line_width=0)
        fig3.update_yaxes(rangemode='tozero')
        fig3.update_xaxes(rangemode='tozero')
        fig3.update_layout(
            title='Mean Quality vs. Read Length',
            xaxis_title='Read Length (bp)',
            yaxis_title='Mean Phred Quality',
            template='simple_white',
            height=500
        )
        fig3.write_html(outdir / f'{name}_quality_vs_length.html')

        # ----------------------------------------
        # Plot 4: Boxplot of Quality per Read Length Bin
        # ----------------------------------------
        df["LengthBin"] = pd.cut(df["ReadLength"], bins=range(0, max(read_lengths) + 50, 50))

        fig4 = go.Figure()
        for name_, group in df.groupby("LengthBin", observed=False):
            fig4.add_trace(go.Box(
                y=group["MeanQuality"],
                name=str(name_),
                boxpoints=False,
                marker_color='darkslateblue',
                line=dict(width=1)
            ))

        fig4.update_layout(
            title="Mean Quality by Read Length Bins (50 bp)",
            xaxis_title="Read Length Bin (bp)",
            yaxis_title="Mean Phred Score",
            template='simple_white',
            height=500
        )
        fig4.write_html(outdir / f'{name}_boxplot_quality_by_lengthbin.html')

        print(f'{datetime.now():%H:%M:%S} - Finished analysing "{name}".')

    # Run in parallel
    Parallel(n_jobs=cpu_count, backend='loky')(delayed(analyse_reads)(fastq_file, output_folder, False) for fastq_file in fastq_files)
    Parallel(n_jobs=cpu_count, backend='loky')(delayed(analyse_reads)(fastq_file, output_folder, True) for fastq_file in fastq_gz_files)

def is_file_still_writing(filepath, wait_time=1.0):
    initial_size = os.path.getsize(filepath)
    time.sleep(wait_time)
    current_size = os.path.getsize(filepath)
    return initial_size != current_size

def open_file(filepath):
    system = platform.system()
    if system == 'Windows':
        os.startfile(filepath)
    elif system == 'Darwin':  # macOS
        subprocess.run(['open', filepath])
    else:  # Linux and others
        subprocess.run(['xdg-open', filepath])

def create_project(project_folder, project_name):

    # Create subfolders
    sub_folders = ['1_raw_data', '2_index_demultiplexing', '3_primer_trimming', '4_quality_filtering', '5_clustering_denoising', '6_read_table', '7_taxonomic_assignment', '8_nanopore_report']

    for folder in sub_folders:
        folder_path = project_folder.joinpath(folder)
        os.makedirs(folder_path, exist_ok=True)
        data_folder_path = folder_path.joinpath('data')
        os.makedirs(data_folder_path, exist_ok=True)
        print(f'{datetime.now().strftime("%H:%M:%S")} - Created "{folder}" folder.')

    res = input('Is your data already demultiplexed (y/n): ')
    if res.upper() == 'Y':
        default_settings_file = Path(__file__).resolve().parent.joinpath('default_settings_nd.xlsx')
        project_settings_file = project_folder.joinpath(project_name + '_settings.xlsx')
        shutil.copyfile(default_settings_file, project_settings_file)

        print('')
        print(f'{datetime.now().strftime("%H:%M:%S")} - Please copy your samples to the 1_raw_data/data folder.')
        print(f'{datetime.now().strftime("%H:%M:%S")} - Then run the "import_raw_data" command:')
        print(f'          $ apscale_nanopore import_raw_data -p {project_name}_apscale_nanopore')
        print(f'{datetime.now().strftime("%H:%M:%S")} - Next, fill out the "Demultiplexing" sheet.')
        print('')

    else:
        default_settings_file = Path(__file__).resolve().parent.joinpath('default_settings.xlsx')
        project_settings_file = project_folder.joinpath(project_name + '_settings.xlsx')
        shutil.copyfile(default_settings_file, project_settings_file)

    print(f'{datetime.now().strftime("%H:%M:%S")} - Created settings file.')
    print(f'{datetime.now().strftime("%H:%M:%S")} - Copy your data into the "1_raw_data/data" folder.')
    print(f'{datetime.now().strftime("%H:%M:%S")} - Adjust the settings file.')
    res = input('Open in settings file Excel? (y/n): ')
    if res.upper() == 'Y':
        open_file(project_settings_file)
    print(f'{datetime.now().strftime("%H:%M:%S")} - Then run:')
    print(f'          $ apscale_nanopore run -p {project_name}_apscale_nanopore')
    print('')

def import_raw_data(project_folder, settings_file, demultiplexing_df):
    wb = load_workbook(settings_file)
    ws1 = wb["Demultiplexing"]

    raw_data_folder = project_folder.joinpath('1_raw_data', 'data')
    raw_data_files = glob.glob(str(raw_data_folder / '*.fastq*'))

    new_demultiplexing_df = pd.DataFrame(
        [['']*4 + [Path(f).name.replace('.gz', '').replace('.fastq', '')] + ['']*2
         for f in raw_data_files],
        columns=demultiplexing_df.columns
    )

    ws1.delete_rows(1, ws1.max_row)
    ws1.append(list(new_demultiplexing_df.columns))
    for row in new_demultiplexing_df.itertuples(index=False, name=None):
        ws1.append(row)
    wb.save(settings_file)

def apscale_nanopore_watch_folder(project_folder, settings_df, demultiplexing_df, steps, skip_demultiplexing):

    try:
        while True:
            # Define folders
            raw_data_folder = project_folder.joinpath('1_raw_data', 'data')
            raw_tmp_folder = project_folder.joinpath('1_raw_data', 'tmp')
            os.makedirs(raw_tmp_folder, exist_ok=True)

            # Scan for files
            print(f'{datetime.now().strftime("%H:%M:%S")} - Scanning for files...')
            main_files = [i for i in glob.glob(str(raw_data_folder.joinpath('*.fastq*')))]
            main_files = {Path(file).name:Path(file) for file in main_files}

            # Collect number of available CPUs
            cpu_count = settings_df[settings_df['Category'] == 'cpu count']['Variable'].values.tolist()[0]
            cpu_count = check_cpu_count(cpu_count)

            # Gzip files if required
            for name, file in main_files.items():
                suffix = file.suffix
                if suffix == ".fastq":
                    print(f'{datetime.now().strftime("%H:%M:%S")} - Zipping {name}...')
                    file_gz = Path(str(file) + '.gz')
                    with open(file, 'rb') as f_in:
                        with gzip.open(file_gz, 'wb') as f_out:
                            shutil.copyfileobj(f_in, f_out)
                    time.sleep(0.1)
                    os.remove(file)
                    main_files[name] = file_gz

            # Sleep if no files are present
            if len(main_files) == 0:
                print(f'{datetime.now().strftime("%H:%M:%S")} - Could not find any files to process! Waiting for new files...')
                time.sleep(2)

            # Analyse files if present
            else:

                print(f'{datetime.now().strftime("%H:%M:%S")} - Found {len(main_files)} file(s) to process!\n')

                # Analyse the files
                i = 0
                for name, main_file in main_files.items():
                    # Check if file is still being written
                    while is_file_still_writing(main_file):
                        print("Waiting for file to finish writing...")
                        time.sleep(1)

                    # Start processing of the file
                    name = name.replace('.fastq.gz', '')
                    main_file = Path(main_file)

                    if skip_demultiplexing == False and "Index demultiplexing" in steps:
                        print(f'{datetime.now().strftime("%H:%M:%S")} - Starting demultiplexing for: {name} ({i+1}/{len(main_files)})')
                        #=======# Index demultiplexing #=======#
                        cutadapt_index_demultiplexing(project_folder, main_file, settings_df, demultiplexing_df)
                        print(f'{datetime.now().strftime("%H:%M:%S")} - Finsihed demultiplexing for: {name} ({i+1}/{len(main_files)})')
                    elif skip_demultiplexing == True and "Index demultiplexing" in steps:
                        print(f'{datetime.now().strftime("%H:%M:%S")} - Skipping demultiplexing.')
                        print(f'{datetime.now().strftime("%H:%M:%S")} - Samples will be copied to the "2_index_demultiplexing" folder.')
                        # Still copy files to respective folder
                        files = glob.glob(str(project_folder.joinpath('1_raw_data', 'data', '*.fastq*')))
                        for file in files:
                            new_file = file.replace('1_raw_data', '2_index_demultiplexing')
                            shutil.copyfile(file, new_file)

                    # Move file to tmp folder
                    new_file = Path(str(raw_tmp_folder.joinpath(name)) + '.fastq.gz')
                    shutil.move(main_file, new_file)
                    print(f'{datetime.now().strftime("%H:%M:%S")} - Moved {name}...')
                    print('')
                    i += 1

                print(f'{datetime.now().strftime("%H:%M:%S")} - Starting raw-data processing for {len(main_files)} files.')
                print('')

                #=======# Primer trimming #=======#
                if "Primer trimming" in steps:
                    print(f'{datetime.now().strftime("%H:%M:%S")} - Starting cutadapt primer trimming...')
                    fastq_files = glob.glob(str(project_folder.joinpath('2_index_demultiplexing', 'data', '*.fastq*')))
                    Parallel(n_jobs=cpu_count, backend='loky')(delayed(cutadapt_primer_trimming)(project_folder, fastq_file, settings_df, demultiplexing_df, skip_demultiplexing) for fastq_file in fastq_files)
                    print(f'{datetime.now().strftime("%H:%M:%S")} - Finished cutadapt primer trimming!')
                    print('')

                #=======# Quality filtering #=======#
                if "Quality filtering" in steps:
                    print(f'{datetime.now().strftime("%H:%M:%S")} - Starting quality filtering...')
                    fastq_files = glob.glob(str(project_folder.joinpath('3_primer_trimming', 'data', '*.fastq.gz')))
                    Parallel(n_jobs=cpu_count, backend='loky')(delayed(python_quality_filtering)(project_folder, fastq_file, settings_df) for fastq_file in fastq_files)
                    print(f'{datetime.now().strftime("%H:%M:%S")} - Finished quality filtering!')
                    print('')

                #=======# Denoising #=======#
                if "Clustering/denoising" in steps:
                    print(f'{datetime.now().strftime("%H:%M:%S")} - Starting clustering/denoising...')
                    fasta_files = glob.glob(str(project_folder.joinpath('4_quality_filtering', 'data', '*.fasta')))
                    Parallel(n_jobs=cpu_count, backend='loky')(delayed(clustering_denoising)(project_folder, fasta_file, settings_df) for fasta_file in fasta_files)
                    print(f'{datetime.now().strftime("%H:%M:%S")} - Finished clustering/denoising...')
                    print('')

                #=======# Read table #=======#
                if "Read table" in steps:
                    print(f'{datetime.now().strftime("%H:%M:%S")} - Starting to build read table...')
                    create_read_table(project_folder, settings_df)
                    print(f'{datetime.now().strftime("%H:%M:%S")} - Finished building read table!')
                    print('')

                #=======# Taxonomic assignment #=======#
                if "Tax. assignment" in steps:
                    print(f'{datetime.now().strftime("%H:%M:%S")} - Starting taxonomic assignment...')
                    apscale_taxonomic_assignment(project_folder, settings_df)
                    print(f'{datetime.now().strftime("%H:%M:%S")} - Finished taxonomic assignment!')
                    print('')

                # =======# Create report #=======#
                print(f'{datetime.now().strftime("%H:%M:%S")} - Collecting results...')
                create_report(project_folder)
                print(f'{datetime.now().strftime("%H:%M:%S")} - Finished creating summary report!')
                print('')

                print(f'{datetime.now().strftime("%H:%M:%S")} - Finished raw-data processing for {len(main_files)}) files.')
                print('')

    except KeyboardInterrupt:
        print('Stopping apscale nanopore live processing.')

def apscale_nanopore(project_folder, settings_df, demultiplexing_df, steps, skip_demultiplexing):

    # Define folders
    raw_data_folder = project_folder.joinpath('1_raw_data', 'data')
    raw_tmp_folder = project_folder.joinpath('1_raw_data', 'tmp')
    os.makedirs(raw_tmp_folder, exist_ok=True)

    # Scan for files
    print(f'{datetime.now().strftime("%H:%M:%S")} - Scanning for files...')
    main_files = [i for i in glob.glob(str(raw_data_folder.joinpath('*.fastq*')))]
    main_files = {Path(file).name:Path(file) for file in main_files}
    print(f'{datetime.now().strftime("%H:%M:%S")} - Found {len(main_files)} fastq raw data files.')
    print('')

    # Collect number of available CPUs
    cpu_count = settings_df[settings_df['Category'] == 'cpu count']['Variable'].values.tolist()[0]
    cpu_count = check_cpu_count(cpu_count)

    # Gzip files if required
    for name, file in main_files.items():
        suffix = file.suffix
        if suffix == ".fastq":
            print(f'{datetime.now().strftime("%H:%M:%S")} - Zipping {name}...')
            file_gz = Path(str(file) + '.gz')
            with open(file, 'rb') as f_in:
                with gzip.open(file_gz, 'wb') as f_out:
                    shutil.copyfileobj(f_in, f_out)
            time.sleep(0.1)
            os.remove(file)
            main_files[name] = file_gz

    # Demultiplex the files
    if skip_demultiplexing == False and "Index demultiplexing" in steps:
        i = 0
        for name, main_file in main_files.items():
            # Start processing of the file
            name = name.replace('.fastq.gz', '')
            main_file = Path(main_file)
            print(f'{datetime.now().strftime("%H:%M:%S")} - Starting demultiplexing for: {name} ({i+1}/{len(main_files)})')
            #=======# Index demultiplexing #=======#
            print(f'{datetime.now().strftime("%H:%M:%S")} - Starting cutadapt index demultiplexing...')
            cutadapt_index_demultiplexing(project_folder, main_file, settings_df, demultiplexing_df)
            print(f'{datetime.now().strftime("%H:%M:%S")} - Finished cutadapt index demultiplexing!')
            print('')
            i += 1

    elif skip_demultiplexing == True and "Index demultiplexing" in steps:
            print(f'{datetime.now().strftime("%H:%M:%S")} - Skipping demultiplexing.')
            print(f'{datetime.now().strftime("%H:%M:%S")} - Samples will be copied to the "2_index_demultiplexing" folder.')
            print('')
            # Still copy files to respective folder
            files = glob.glob(str(project_folder.joinpath('1_raw_data', 'data', '*.fastq*')))
            for file in files:
                new_file = file.replace('1_raw_data', '2_index_demultiplexing')
                shutil.copyfile(file, new_file)

    #=======# Primer trimming #=======#
    if "Primer trimming" in steps:
        print(f'{datetime.now().strftime("%H:%M:%S")} - Starting cutadapt primer trimming...')
        fastq_files = glob.glob(str(project_folder.joinpath('2_index_demultiplexing', 'data', '*.fastq*')))
        Parallel(n_jobs=cpu_count, backend='loky')(delayed(cutadapt_primer_trimming)(project_folder, fastq_file, settings_df, demultiplexing_df, skip_demultiplexing) for fastq_file in fastq_files)
        print(f'{datetime.now().strftime("%H:%M:%S")} - Finished cutadapt primer trimming!')
        print('')

    #=======# Quality filtering #=======#
    if "Quality filtering" in steps:
        print(f'{datetime.now().strftime("%H:%M:%S")} - Starting quality filtering...')
        fastq_files = glob.glob(str(project_folder.joinpath('3_primer_trimming', 'data', '*.fastq.gz')))
        Parallel(n_jobs=cpu_count, backend='loky')(delayed(python_quality_filtering)(project_folder, fastq_file, settings_df, demultiplexing_df, skip_demultiplexing) for fastq_file in fastq_files)
        print(f'{datetime.now().strftime("%H:%M:%S")} - Finished vsearch quality filtering!')
        print('')

    #=======# Denoising #=======#
    if "Clustering/denoising" in steps:
        print(f'{datetime.now().strftime("%H:%M:%S")} - Starting clustering/denoising...')
        fasta_files = glob.glob(str(project_folder.joinpath('4_quality_filtering', 'data', '*.fasta')))
        Parallel(n_jobs=cpu_count, backend='loky')(delayed(clustering_denoising)(project_folder, fasta_file, settings_df) for fasta_file in fasta_files)
        print(f'{datetime.now().strftime("%H:%M:%S")} - Finished clustering/denoising...')
        print('')

    #=======# Read table #=======#
    if "Read table" in steps:
        print(f'{datetime.now().strftime("%H:%M:%S")} - Starting to build read table...')
        create_read_table(project_folder, settings_df)
        print(f'{datetime.now().strftime("%H:%M:%S")} - Finished building read table!')
        print('')

    #=======# Taxonomic assignment #=======#
    if "Tax. assignment" in steps:
        print(f'{datetime.now().strftime("%H:%M:%S")} - Starting taxonomic assignment...')
        apscale_taxonomic_assignment(project_folder, settings_df)
        print(f'{datetime.now().strftime("%H:%M:%S")} - Finished taxonomic assignment!')
        print('')

    # =======# Create report #=======#
    print(f'{datetime.now().strftime("%H:%M:%S")} - Collecting results...')
    create_report(project_folder)
    print(f'{datetime.now().strftime("%H:%M:%S")} - Finished creating summary report!')
    print('')

def cutadapt_index_demultiplexing(project_folder, main_file, settings_df, demultiplexing_df):

    # Preprare output files
    # main_file = Path("/Users/tillmacher/Desktop/APSCALE_projects/test_dataset_apscale_nanopore/1_raw_data/data/merged_nanopore_data.fastq.gz")
    input_file = main_file
    name = input_file.name.replace('.fastq.gz', '')
    output_folder_tmp = project_folder.joinpath('2_index_demultiplexing', 'tmp')
    output_folder_data = project_folder.joinpath('2_index_demultiplexing', 'data')
    output_file = output_folder_tmp.joinpath("{name}_fwd.fastq")

    # Create tmp folder
    tmp_folder = project_folder.joinpath('2_index_demultiplexing', 'tmp')
    os.makedirs(tmp_folder, exist_ok=True)

    # Create reverse complement of untrimmed
    output_file_rc = output_folder_tmp.joinpath("{name}_rc.fastq")
    untrimmed_folder = project_folder.joinpath('2_index_demultiplexing', 'untrimmed')
    os.makedirs(untrimmed_folder, exist_ok=True)
    untrimmed_fastq = untrimmed_folder.joinpath('untrimmed.fastq')
    untrimmed_rc_fastq = untrimmed_folder.joinpath('untrimmed_rc.fastq')

    # Collect required settings
    number_of_errors = settings_df[settings_df['Category'] == 'allowed errors index']['Variable'].values.tolist()[0]
    cpu_count = settings_df[settings_df['Category'] == 'cpu count']['Variable'].values.tolist()[0]
    cpu_count = check_cpu_count(cpu_count)

    ##======## Demultuplexing forward reads ##======##
    # Run cutadapt demultiplexing
    g_args = []
    for _, row in demultiplexing_df.iterrows():
        # Create forward sequence
        fwd_seq = row['Forward index 5-3']
        # Create reverse sequence
        rvs_seq = reverse_complement(row['Reverse index 5-3'])
        # Combine to search sequence
        search_seq = f'{fwd_seq}...{rvs_seq}'
        g_args.extend(['-g', search_seq])

    # Run cutadapt demultiplexing
    try:
        command = f"cutadapt -e {number_of_errors} {' '.join(g_args)} --cores {cpu_count} -o {output_file} --untrimmed-output {untrimmed_fastq} --report=minimal {input_file}"
        process = subprocess.Popen(command, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
        stdout, stderr = process.communicate()
        in_reads1 = int(stdout.split()[11])
        out_reads1 = int(stdout.split()[-3])
    except Exception as e:
        print("=== Cutadapt STDOUT ===")
        print(stdout if 'stdout' in locals() else "No STDOUT captured.")
        print("=== Cutadapt STDERR ===")
        print(stderr if 'stderr' in locals() else "No STDERR captured.")
        print("=== PYTHON ERROR ===")
        print(f"Error: {e}")
        sys.exit()

    print(f'{datetime.now().strftime("%H:%M:%S")} - Finished demultiplexing in 5\'-3\' orientation!')

    ##======## Demultuplexing RC reads ##======##
    # Vsearch reverse complement
    try:
        command = f"vsearch --fastx_revcomp {untrimmed_fastq} --fastqout {untrimmed_rc_fastq}"
        process = subprocess.Popen(command, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
        stdout, stderr = process.communicate()
    except Exception as e:
        print("=== Cutadapt STDOUT ===")
        print(stdout if 'stdout' in locals() else "No STDOUT captured.")
        print("=== Cutadapt STDERR ===")
        print(stderr if 'stderr' in locals() else "No STDERR captured.")
        print("=== PYTHON ERROR ===")
        print(f"Error: {e}")
        sys.exit()

    if untrimmed_fastq.exists():
        os.remove(untrimmed_fastq)

    # Run cutadapt again
    try:
        command = f"cutadapt -e {number_of_errors} {' '.join(g_args)} --cores {cpu_count} -o {output_file_rc} --discard-untrimmed --report=minimal {untrimmed_rc_fastq}"
        process = subprocess.Popen(command, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
        stdout, stderr = process.communicate()
        out_reads2 = int(stdout.split()[-3])
    except Exception as e:
        print("=== Cutadapt STDOUT ===")
        print(stdout if 'stdout' in locals() else "No STDOUT captured.")
        print("=== Cutadapt STDERR ===")
        print(stderr if 'stderr' in locals() else "No STDERR captured.")
        print("=== PYTHON ERROR ===")
        print(f"Error: {e}")
        sys.exit()

    print(f'{datetime.now().strftime("%H:%M:%S")} - Finished demultiplexing in 3\'-5\' orientation!')

    try:
        reads_perc = round((out_reads1 + out_reads2) / in_reads1 * 100, 2)
    except ZeroDivisionError:
        reads_perc = 0
    print(f'{datetime.now().strftime("%H:%M:%S")} - Finished demultiplexing of {name}: {in_reads1:,} -> {out_reads1 + out_reads2:,} reads ({reads_perc}% passed).')

    if untrimmed_folder.exists():
        shutil.rmtree(untrimmed_folder)

    ##======## Merge files ##======##
    # Collect all .fastq files (uncompressed)
    demultiplexed_fwd_files = sorted(glob.glob(str(output_folder_tmp.joinpath('*_fwd.fastq'))))
    demultiplexed_rc_files = sorted(glob.glob(str(output_folder_tmp.joinpath('*_rc.fastq'))))

    # Only continue if files were demultiplexed
    if len(demultiplexed_fwd_files) == 0:
        print(f'{datetime.now().strftime("%H:%M:%S")} - Error: Could not find any demultiplexed files!')
        return

    # Define merging function (uncompressed)
    def cat_files(output_folder_data, tmp_file_fwd, tmp_file_rc, bufsize=1024 * 1024):
        index = int(Path(tmp_file_fwd).name.replace('_fwd.fastq', '')) - 1
        sample_id = demultiplexing_df['ID'][index]
        output_file = output_folder_data.joinpath(f'{sample_id}.fastq')
        with open(output_file, "wb") as out:
            for fname in [tmp_file_fwd, tmp_file_rc]:
                with open(fname, "rb") as f:
                    while chunk := f.read(bufsize):
                        out.write(chunk)
        os.remove(tmp_file_fwd)
        os.remove(tmp_file_rc)
        print(f'{datetime.now().strftime("%H:%M:%S")} - Merged and saved {sample_id}!')

    [cat_files(output_folder_data, tmp_file_fwd, tmp_file_rc) for tmp_file_fwd, tmp_file_rc in zip(demultiplexed_fwd_files, demultiplexed_rc_files)]

def cutadapt_primer_trimming(project_folder, file, settings_df, demultiplexing_df, skip_demultiplexing):

    # Preprare output files
    # file = '/Volumes/Coruscant/APSCALE_projects/test_dataset_apscale_nanopore/2_index_demultiplexing/data/3_subset.fastq.gz'
    input_file = Path(file)
    name = input_file.name.replace('.fastq', '').replace('.gz', '')
    output_folder_data = project_folder.joinpath('3_primer_trimming', 'data')
    output_file = output_folder_data.joinpath(f"{name}_trimmed.fastq.gz")

    # Also save the untrimmed reads
    # Create reverse complement of untrimmed
    untrimmed_folder = project_folder.joinpath('3_primer_trimming', 'untrimmed')
    os.makedirs(untrimmed_folder, exist_ok=True)
    output_file_untrimmed = untrimmed_folder.joinpath(f"{name}_untrimmed.fastq")
    output_file_untrimmed_rc = untrimmed_folder.joinpath(f"{name}_untrimmed_rc.fastq")
    output_file_rc = output_folder_data.joinpath(f"{name}_trimmed_rc.fastq.gz")
    tmp_file = untrimmed_folder.joinpath(f"{name}_tmp.fastq.gz")

    # Collect required settings
    number_of_errors = settings_df[settings_df['Category'] == 'allowed errors primer']['Variable'].values.tolist()[0]

    # Run cutadapt demultiplexing
    # Create forward sequence
    # Check if index demultiplexing is required
    try:
        sub_df = demultiplexing_df[demultiplexing_df['ID'] == name]
        fwd_seq = sub_df['Forward primer 5-3'].values.tolist()[0]
        # Create reverse sequence
        rvs_seq_rc = reverse_complement(sub_df['Reverse primer 5-3'].values.tolist()[0])
        adapter = f'{fwd_seq}...{rvs_seq_rc}'
    except:
        print(f'{datetime.now().strftime("%H:%M:%S")} - Could not find "{name}" in the Demultiplexing sheet!')
        return

    ##======## Trimming of reads in 5'-3' orientation ##======##
    # Run cutadapt demultiplexing and primer trimming
    try:
        command = f"cutadapt -e {number_of_errors} -g {adapter} --cores 1 -o {output_file} --untrimmed-output {output_file_untrimmed} --report=minimal {input_file}"
        process = subprocess.Popen(command, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
        stdout, stderr = process.communicate()
        # You can now use `stdout` and `stderr` as variables
        in_reads1 = int(stdout.split()[11])
        out_reads1 = int(stdout.split()[-3])
    except Exception as e:
        print("=== Cutadapt STDOUT ===")
        print(stdout if 'stdout' in locals() else "No STDOUT captured.")
        print("=== Cutadapt STDERR ===")
        print(stderr if 'stderr' in locals() else "No STDERR captured.")
        print("=== PYTHON ERROR ===")
        print(f"Error: {e}")
        sys.exit()

    ##======## Trimming of reads in 3'-5' orientation ##======##
    # Vsearch reverse complement
    try:
        command = f"vsearch --fastx_revcomp {output_file_untrimmed} --fastqout {output_file_untrimmed_rc}"
        process = subprocess.Popen(command, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
        stdout, stderr = process.communicate()
    except Exception as e:
        print("=== Cutadapt STDOUT ===")
        print(stdout if 'stdout' in locals() else "No STDOUT captured.")
        print("=== Cutadapt STDERR ===")
        print(stderr if 'stderr' in locals() else "No STDERR captured.")
        print("=== PYTHON ERROR ===")
        print(f"Error: {e}")
        sys.exit()

    if output_file_untrimmed.exists():
        os.remove(output_file_untrimmed)

    # Run cutadapt primer trimming
    try:
        command = f"cutadapt -e {number_of_errors} -g {adapter} --cores 1 -o {output_file_rc} --discard-untrimmed --report=minimal {output_file_untrimmed_rc}"
        process = subprocess.Popen(command, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
        stdout, stderr = process.communicate()
        # You can now use `stdout` and `stderr` as variables
        in_reads2 = int(stdout.split()[11])
        out_reads2 = int(stdout.split()[-3])
    except Exception as e:
        print("=== Cutadapt STDOUT ===")
        print(stdout if 'stdout' in locals() else "No STDOUT captured.")
        print("=== Cutadapt STDERR ===")
        print(stderr if 'stderr' in locals() else "No STDERR captured.")
        print("=== PYTHON ERROR ===")
        print(f"Error: {e}")
        sys.exit()

    # Combine and overwrite
    with gzip.open(tmp_file, 'wb') as out_f:
        with gzip.open(output_file, 'rb') as f1:
            shutil.copyfileobj(f1, out_f)
        with gzip.open(output_file_rc, 'rb') as f2:
            shutil.copyfileobj(f2, out_f)

    # Replace original file with combined content
    os.replace(tmp_file, output_file)
    # Remove tmp files
    if output_file_untrimmed_rc.exists():
        os.remove(output_file_untrimmed_rc)
    if output_file_rc.exists():
        os.remove(output_file_rc)

    try:
        reads_perc = round((out_reads1 + out_reads2) / in_reads1 * 100, 2)
    except ZeroDivisionError:
        reads_perc = 0
    print(f'{datetime.now().strftime("%H:%M:%S")} - {name}: {in_reads1:,} -> {out_reads1 + out_reads2:,} reads ({reads_perc}% passed).')

def python_quality_filtering(project_folder, file, settings_df, demultiplexing_df, skip_demultiplexing):

    # Preprare output files
    # file = '/Users/tillmacher/Desktop/APSCALE_projects/test_dataset_apscale_nanopore/3_primer_trimming/data/Sample_3_trimmed.fastq.gz'
    input_file = Path(file)
    name = input_file.name.replace('.fastq.gz', '')
    output_folder_data = project_folder.joinpath('4_quality_filtering', 'data')
    # output files
    filtered_fasta = output_folder_data.joinpath(f'{name}_filtered.fasta')
    dereplicated_fasta = output_folder_data.joinpath(f'{name}_filtered_derep.fasta')

    # Collect required settings
    try:
        sub_df = demultiplexing_df[demultiplexing_df['ID'] == name.replace('_trimmed', '')]
        min_len = sub_df['Minimum length'].values.tolist()[0]
        max_len = sub_df['Maximum length'].values.tolist()[0]
        trunc_val = settings_df[settings_df['Category'] == 'minimum quality']['Variable'].values.tolist()[0]
    except:
        print(f'{datetime.now().strftime("%H:%M:%S")} - Could not find "{name}" in the Demultiplexing sheet!')
        return

    # Run python-based quality filtering
    reads_1 = 0
    total_reads = 0

    with gzip.open(input_file, "rt") as in_handle, open(filtered_fasta, "w") as out_handle:
        for record in SeqIO.parse(in_handle, "fastq"):
            total_reads += 1
            phred_scores = record.letter_annotations["phred_quality"]
            read_length = len(record.seq)
            if not phred_scores:
                continue
            if np.mean(phred_scores) >= int(trunc_val) and int(min_len) <= read_length <= int(max_len):
                fasta_record = SeqRecord(
                    Seq(str(record.seq)),
                    id=record.id,
                    description=""
                )
                SeqIO.write(fasta_record, out_handle, "fasta")
                reads_1 += 1

    # Run vsearch dereplication
    try:
        command = f"vsearch --threads 1 --derep_fulllength {filtered_fasta} --sizeout --relabel_sha1 --output {dereplicated_fasta}"
        process = subprocess.Popen(command, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
        stdout, stderr = process.communicate()
        reads_2 = stderr.split()[-15]
    except Exception as e:
        print("=== Cutadapt STDOUT ===")
        print(stdout if 'stdout' in locals() else "No STDOUT captured.")
        print("=== Cutadapt STDERR ===")
        print(stderr if 'stderr' in locals() else "No STDERR captured.")
        print("=== PYTHON ERROR ===")
        print(f"Error: {e}")
        sys.exit()

    try:
        reads_perc = round(reads_1 / total_reads * 100, 2)
    except ZeroDivisionError:
        reads_perc = 0
    print(f'{datetime.now().strftime("%H:%M:%S")} - {name.replace("_trimmed", "")}: {total_reads:,} -> {reads_1:,} reads ({reads_perc}% passed) -> {int(reads_2):,} (dereplication)')

    if filtered_fasta.exists():
        os.remove(filtered_fasta)

def clustering_denoising(project_folder, file, settings_df):

    # Preprare output files
    # file = '/Volumes/Coruscant/APSCALE_projects/naturalis_dataset_apscale_nanopore/4_quality_filtering/data/naturalis_sample_35_trimmed_filtered_derep.fasta'
    input_file = Path(file)
    name = input_file.name.replace('_trimmed_filtered_derep.fasta', '')
    output_folder_data = project_folder.joinpath('5_clustering_denoising', 'data')
    cluster_file = output_folder_data.joinpath(f'{name}_reads.fasta')

    # Collect required settings
    mode = settings_df[settings_df['Category'] == 'mode']['Variable'].values.tolist()[0]

    if mode == 'ESVs':
        representative = ''
    else:
        representative = settings_df[settings_df['Category'] == 'representative']['Variable'].values.tolist()[0]

    if mode == 'Swarms':
        d_value = settings_df[settings_df['Category'] == 'd']['Variable'].values.tolist()[0]
        # Run swarm denoising
        try:
            if representative == 'centroid':
                command = f"swarm -d {d_value} --threads 1 -z --seeds {cluster_file} {input_file}"
            elif representative == 'consensus':
                command = f"swarm -d {d_value} --threads 1 -z --consensus {cluster_file} {input_file}"
            process = subprocess.Popen(command, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
            stdout, stderr = process.communicate()
        except Exception as e:
            print("=== Cutadapt STDOUT ===")
            print(stdout if 'stdout' in locals() else "No STDOUT captured.")
            print("=== Cutadapt STDERR ===")
            print(stderr if 'stderr' in locals() else "No STDERR captured.")
            print("=== PYTHON ERROR ===")
            print(f"Error: {e}")
            sys.exit()

    elif mode == 'ESVs':
        try:
            alpha_value = settings_df[settings_df['Category'] == 'alpha']['Variable'].values.tolist()[0]
            # Run vsearch denoising
            command = f"vsearch --cluster_unoise {input_file} --unoise_alpha {alpha_value} --threads 1 --centroids {cluster_file} "
            process = subprocess.Popen(command, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
            stdout, stderr = process.communicate()
        except Exception as e:
            print("=== Cutadapt STDOUT ===")
            print(stdout if 'stdout' in locals() else "No STDOUT captured.")
            print("=== Cutadapt STDERR ===")
            print(stderr if 'stderr' in locals() else "No STDERR captured.")
            print("=== PYTHON ERROR ===")
            print(f"Error: {e}")
            sys.exit()

    elif mode == 'Swarm OTUs':
        try:
            d_value = settings_df[settings_df['Category'] == 'd']['Variable'].values.tolist()[0]
            representative = settings_df[settings_df['Category'] == 'representative']['Variable'].values.tolist()[0]
            # Run swarm denoising
            cluster_file_0 = output_folder_data.joinpath(f'{name}_swarms.fasta')
            # choose output
            if representative == 'centroid':
                command = f"swarm -d {d_value} --threads 1 -z --seeds {cluster_file_0} {input_file}"
            elif representative == 'consensus':
                command = f"swarm -d {d_value} --threads 1 -z --consensus {cluster_file_0} {input_file}"
            process = subprocess.Popen(command, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
            stdout, stderr = process.communicate()
        except Exception as e:
            print("=== Cutadapt STDOUT ===")
            print(stdout if 'stdout' in locals() else "No STDOUT captured.")
            print("=== Cutadapt STDERR ===")
            print(stderr if 'stderr' in locals() else "No STDERR captured.")
            print("=== PYTHON ERROR ===")
            print(f"Error: {e}")
            sys.exit()

        # Then clustering
        percid_value = settings_df[settings_df['Category'] == 'percid']['Variable'].values.tolist()[0]
        # Run vsearch clustering
        try:
            # choose output
            if representative == 'centroid':
                command = f"vsearch --cluster_size {cluster_file_0} --id {percid_value} --threads 1 --centroids {cluster_file} "
            elif representative == 'consensus':
                command = f"vsearch --cluster_size {cluster_file_0} --id {percid_value} --threads 1 --consout {cluster_file} "
            process = subprocess.Popen(command, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
            stdout, stderr = process.communicate()
        except Exception as e:
            print("=== Cutadapt STDOUT ===")
            print(stdout if 'stdout' in locals() else "No STDOUT captured.")
            print("=== Cutadapt STDERR ===")
            print(stderr if 'stderr' in locals() else "No STDERR captured.")
            print("=== PYTHON ERROR ===")
            print(f"Error: {e}")
            sys.exit()

        if cluster_file_0.exists():
            os.remove(cluster_file_0)
    else:
        # Initial denoising
        alpha_value = settings_df[settings_df['Category'] == 'alpha']['Variable'].values.tolist()[0]
        # Run vsearch denoising
        cluster_file_0 = output_folder_data.joinpath(f'{name}_denoise.fasta')
        try:
            command = f"vsearch --cluster_unoise {input_file} --unoise_alpha {alpha_value} --threads 1 --centroids {cluster_file_0} "
            process = subprocess.Popen(command, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
            stdout, stderr = process.communicate()
        except Exception as e:
            print("=== Cutadapt STDOUT ===")
            print(stdout if 'stdout' in locals() else "No STDOUT captured.")
            print("=== Cutadapt STDERR ===")
            print(stderr if 'stderr' in locals() else "No STDERR captured.")
            print("=== PYTHON ERROR ===")
            print(f"Error: {e}")
            sys.exit()

        # Then clustering
        percid_value = settings_df[settings_df['Category'] == 'percid']['Variable'].values.tolist()[0]
        representative = settings_df[settings_df['Category'] == 'representative']['Variable'].values.tolist()[0]

        # Run vsearch clustering
        try:
            # choose output
            if representative == 'centroid':
                command = f"vsearch --cluster_size {cluster_file_0} --id {percid_value} --threads 1 --centroids {cluster_file} "
            elif representative == 'consensus':
                command = f"vsearch --cluster_size {cluster_file_0} --id {percid_value} --threads 1 --consout {cluster_file} "
            process = subprocess.Popen(command, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
            stdout, stderr = process.communicate()
        except Exception as e:
            print("=== Cutadapt STDOUT ===")
            print(stdout if 'stdout' in locals() else "No STDOUT captured.")
            print("=== Cutadapt STDERR ===")
            print(stderr if 'stderr' in locals() else "No STDERR captured.")
            print("=== PYTHON ERROR ===")
            print(f"Error: {e}")
            sys.exit()

        if cluster_file_0.exists():
            os.remove(cluster_file_0)

    # Perform chimera detection
    nochimera_fasta = output_folder_data.joinpath(f'{name}_clusters_nochimera.fasta')
    try:
        command = f'vsearch --uchime3_denovo {cluster_file} --threads 1 --nonchimeras {nochimera_fasta}'
        process = subprocess.Popen(command, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
        stdout, stderr = process.communicate()
        nochimera_total = stderr.split()[-12]
        nochimera_perc = stderr.split()[-11]
    except Exception as e:
        print("=== Cutadapt STDOUT ===")
        print(stdout if 'stdout' in locals() else "No STDOUT captured.")
        print("=== Cutadapt STDERR ===")
        print(stderr if 'stderr' in locals() else "No STDERR captured.")
        print("=== PYTHON ERROR ===")
        print(f"Error: {e}")
        sys.exit()

    if nochimera_total == '0':
        print(f'{datetime.now().strftime("%H:%M:%S")} - {name}: Wrote 0 non-chimera {representative} {mode}.')
    else:
        print(f'{datetime.now().strftime("%H:%M:%S")} - {name}: Wrote {int(nochimera_total):,} {nochimera_perc} non-chimera {representative} {mode}.')
    if cluster_file.exists():
        os.remove(cluster_file)
    time.sleep(1)

def create_read_table(project_folder, settings_df):
    # Prepare output files
    swarm_files_path = project_folder.joinpath('5_clustering_denoising', 'data', '*_clusters_nochimera.fasta')
    swarm_files = [Path(i) for i in glob.glob(str(swarm_files_path))]
    data = defaultdict(lambda: defaultdict(int))  # nested dict: hash -> sample -> size
    seq_dict = {}  # hash -> sequence

    # Collect required settings
    min_reads = float(settings_df[settings_df['Category'] == 'minimum reads']['Variable'].values.tolist()[0])
    if min_reads < 1:
        print(f'Using relative threshold per sample: {min_reads}%')
        relative = True
    else:
        print(f'Using absoulte threshold per sample: {min_reads}')
        relative = False
    mode = settings_df[settings_df['Category'] == 'mode']['Variable'].values.tolist()[0]

    # Parse files
    for file in sorted(swarm_files):
        sample = file.name.replace('_clusters_nochimera.fasta', '')
        with open(file) as handle:
            for record in SeqIO.parse(handle, "fasta"):
                size = int(record.id.split(';')[1].replace('size=', ''))
                seq = str(record.seq)
                hash = hashlib.sha3_256(seq.encode("ascii")).hexdigest()
                data[hash][sample] += size
                seq_dict[hash] = seq

    # Create and prepare DataFrame
    df = pd.DataFrame.from_dict(data, orient='index').fillna(0).astype(int)
    df.insert(0, "Seq", df.index.map(seq_dict))

    # Filter low-abundance counts
    sample_cols = df.columns.difference(['Seq'])
    reads_before_dict = {i:df[i].sum() for i in sample_cols}
    total_reads_before = df[sample_cols].sum().sum()
    n_swarms_before = len(df)

    # apply read filter
    cutoff_dict = {}
    if relative:  # relative cutoff, expects fraction (e.g. 0.0001 for 0.01%)
        for col in sample_cols:
            total = df[col].sum()
            if total > 0:
                # compute absolute cutoff in reads for this sample
                abs_cutoff = int(total * min_reads)
                cutoff_dict[col] = abs_cutoff
                df[col] = df[col].where(df[col] >= abs_cutoff, 0)
    else:  # absolute cutoff
        abs_cutoff = int(min_reads)
        for col in sample_cols:
            cutoff_dict[col] = abs_cutoff
            df[col] = df[col].where(df[col] >= abs_cutoff, 0)

    # Filter low-abundance ESVs
    df.index.name = "ID"
    df.reset_index(inplace=True)
    df['sum'] = df[sample_cols].sum(axis=1)
    df = df[df['sum'] != 0]
    df['sum'] = df[sample_cols].sum(axis=1)
    df = df.sort_values('sum', ascending=False)

    # Print discard reads
    reads_after_dict = {i:df[i].sum() for i in sample_cols}
    for sample in sample_cols:
        r0 = reads_before_dict[sample]
        r1 = reads_after_dict[sample]
        perc = round(100 - (r0 - r1) / r0 * 100, 2)
        cutoff = cutoff_dict[sample]
        print(f'{datetime.now():%H:%M:%S} - {sample}: {r0:,} -> {r1:,} reads ({perc}%) (cutoff={cutoff})')

    # Calculate stats
    total_reads_after = df['sum'].sum()
    removed = total_reads_before - total_reads_after
    n_swarms = len(df)
    swarms_discarded = n_swarms_before - n_swarms
    print(f'{datetime.now():%H:%M:%S} - Final read table contains {n_swarms:,} {mode} accounting for {total_reads_after:,} reads.')
    print(f'{datetime.now():%H:%M:%S} - Discarded {swarms_discarded:,} {mode} accounting for {removed:,} reads (<= {min_reads} reads).')

    # Final clean-up
    df.drop(columns='sum', inplace=True)

    # insert empty files
    for file in sorted(swarm_files):
        sample = file.name.replace('_clusters_nochimera.fasta', '')
        if sample not in list(df.columns):
            df[sample] = [0] * len(df)

    # Collect name of the project
    project_name = project_folder.name.replace('_apscale_nanopore', '')

    # Write to files
    if df.shape[0] < 65000:
        excel_file = project_folder.joinpath('6_read_table', f'{project_name}_read_table.xlsx')
        df.to_excel(excel_file, index=False)
    parquet_file = project_folder.joinpath('6_read_table', f'{project_name}_read_table.parquet.snappy')
    df.to_parquet(parquet_file, compression='snappy')

    # Write sequences to fasta
    fasta_file = project_folder.joinpath('6_read_table', 'data', f'{project_name}_reads.fasta')
    with open(fasta_file, "w") as output_handle:
        for hash, seq in df[['ID', 'Seq']].values.tolist():
            record = SeqRecord(Seq(seq), id=hash, description='')
            SeqIO.write(record, output_handle, "fasta")

def apscale_taxonomic_assignment(project_folder, settings_df):
    # Define files
    project_name = project_folder.name.replace('_apscale_nanopore', '')
    fasta_file = project_folder.joinpath('6_read_table', 'data', f'{project_name}_reads.fasta')
    results_folder = project_folder.joinpath('7_taxonomic_assignment', project_name)

    # Collect variables
    run_blastn = settings_df[settings_df['Category'] == 'apscale blast']['Variable'].values.tolist()[0]
    blastn_db = settings_df[settings_df['Category'] == 'apscale db']['Variable'].values.tolist()[0]
    task = settings_df[settings_df['Category'] == 'task']['Variable'].values.tolist()[0]

    # Run apscale blast
    if run_blastn == 'Yes':
        try:
            shutil.rmtree(results_folder)
            os.makedirs(results_folder, exist_ok=True)
        except FileNotFoundError:
            pass
        try:
            command = f"apscale_blast -db {blastn_db} -q {fasta_file} -o {results_folder} -task {task}"
            process = subprocess.Popen(command, shell=True, text=True)
            process.wait()
        except Exception as e:
            print("=== Cutadapt STDOUT ===")
            print(stdout if 'stdout' in locals() else "No STDOUT captured.")
            print("=== Cutadapt STDERR ===")
            print(stderr if 'stderr' in locals() else "No STDERR captured.")
            print("=== PYTHON ERROR ===")
            print(f"Error: {e}")
            sys.exit()

def create_report(project_folder):
    # collect information
    project_name = project_folder.name.replace('_apscale_nanopore', '')
    read_table_file = project_folder.joinpath('6_read_table', f'{project_name}_read_table.parquet.snappy')
    taxonomy_table_file = project_folder.joinpath('7_taxonomic_assignment', project_name, f'{project_name}_taxonomy.xlsx')

    # create folder to store batch results
    report_folder = project_folder.joinpath('8_nanopore_report') / 'main'
    os.makedirs(report_folder, exist_ok=True)

    if read_table_file.exists() and taxonomy_table_file.exists():
        read_table_df = pd.read_parquet(read_table_file).fillna('')
        samples = read_table_df.columns.tolist()[2:]
        n_reads, n_groups = {}, {}
        for sample in samples:
            sub_df = read_table_df[sample]
            n_reads[sample] = sub_df.sum()
            n_groups[sample] = len(sub_df[sub_df != 0])

        # 1. Bar plot: Total reads per sample
        fig_reads = go.Figure()
        fig_reads.add_trace(go.Bar(
            x=list(n_reads.keys()),
            y=list(n_reads.values()),
            text=[f"{v:,}" for v in n_reads.values()],  # add labels with commas
            textposition="outside"
        ))
        fig_reads.update_layout(
            title="Total Reads per Sample",
            xaxis_title="Sample",
            yaxis_title="Total Reads",
            xaxis_tickangle=-45,
            bargap=0.4,
            height = 500,
            template="simple_white"
        )
        fig_reads.update_xaxes(dtick='linear')
        output_file = report_folder / Path(project_name + '_reads_per_sample.html')
        fig_reads.write_html(output_file)

        # 2. Bar plot: Number of detected groups per sample
        fig_groups = go.Figure()
        fig_groups.add_trace(go.Bar(
            x=list(n_groups.keys()),
            y=list(n_groups.values()),
            text=list(n_groups.values()),
            textposition="outside",
            marker_color="indianred"
        ))
        fig_groups.update_layout(
            title="Number of Groups per Sample",
            xaxis_title="Sample",
            yaxis_title="Detected Groups",
            xaxis_tickangle=-45,
            bargap=0.4,
            height=500,
            template="simple_white"
        )
        fig_groups.update_xaxes(dtick='linear')
        output_file = report_folder / Path(project_name + '_groups_per_sample.html')
        fig_groups.write_html(output_file)

        # 3. Scatter: Groups vs Reads (richness vs depth)
        fig_scatter = px.scatter(
            x=list(n_reads.values()),
            y=list(n_groups.values()),
            text=list(n_reads.keys()),
            labels={"x": "Total Reads", "y": "Detected Groups"},
            title="Groups vs. Sequencing Depth"
        )
        fig_scatter.update_traces(textposition="top center")
        fig_scatter.update_layout(template="simple_white", height=500)
        fig_scatter.update_yaxes(rangemode='tozero')
        fig_scatter.update_xaxes(rangemode='tozero')
        output_file = report_folder / Path(project_name + '_groups_vs_sequencing_depth.html')
        fig_scatter.write_html(output_file)

def main():
    """
    APSCALE nanopore suite
    Command-line tool to process nanopore sequence data.
    """

    # Introductory message with usage examples
    message = """
    APSCALE nanopore command line tool
    Example commands:
    $ apscale_nanopore create -p PATH/TO/PROJECT
    $ apscale_nanopore run -p PATH/TO/PROJECT
    $ apscale_nanopore qc -p PATH/TO/PROJECT
    $ apscale_nanopore import_raw_data -p PATH/TO/PROJECT
    $ apscale_nanopore gui

    """
    print(message)

    # Check dependencies
    check_dependencies()

    # Initialize main parser
    parser = argparse.ArgumentParser(description='APSCALE nanopore')
    subparsers = parser.add_subparsers(dest='command', required=True)

    # === Subparser: create ===
    create_parser = subparsers.add_parser('create', help='Create a new APSCALE nanopore project.')
    create_parser.add_argument('-p', '--project', type=str, required=True, help='Path to project.')

    # === Subparser: quality control (qc) ===
    create_parser = subparsers.add_parser('qc', help='Quality control for subfolder.')
    create_parser.add_argument('-p', '--project', type=str, required=True, help='Path to project.')

    # === Subparser: run ===
    run_parser = subparsers.add_parser('run', help='Run the APSCALE nanopore pipeline.')
    # General
    run_parser.add_argument('-p', '--project', type=str, required=True, help='Path to project.')
    run_parser.add_argument('-live', '--live_calling', action='store_true', help='Scan 1_raw_data for new batches.')
    run_parser.add_argument('-sd', action='store_true', help='Skip demultiplexing.')
    # Adjust parameters on the fly
    # allowed errors index
    run_parser.add_argument('-e1', type=str, help='Overwrite: allowed index demultiplexing errors.')
    # allowed errors primer
    run_parser.add_argument('-e2', type=str, help='Overwrite: allowed primer trimming errors.')
    # minimum length
    run_parser.add_argument('-minlen', type=str, help='Overwrite: minimum length.')
    # maximum length
    run_parser.add_argument('-maxlen', type=str, help='Overwrite: maximum length.')
    # minimum quality
    run_parser.add_argument('-minq', type=str, help='Overwrite: minimum quality value.')
    # mode
    run_parser.add_argument('-mode', type=str, help="Overwrite: Clustering/denoising mode.")
    # percid
    run_parser.add_argument('-percid', type=str, help="Overwrite: Vsearch clustering percid.")
    # alpha
    run_parser.add_argument('-alpha', type=str, help="Overwrite: Vsearch denoising alpha.")
    # d
    run_parser.add_argument('-d', type=str, help="Overwrite: swarm's d value.")
    # representative
    run_parser.add_argument('-rep', type=str, help="Overwrite: clustering representative value.")
    # minimum reads
    run_parser.add_argument('-minreads', type=str, help="Overwrite: Read filter threshold.")
    # STEPS
    run_parser.add_argument('-step', type=str, help="Select step to re-run individually. "
                                                    "1:Index demultiplexing, 2:Primer trimming, "
                                                    "3:Tag demultiplexing, 4:Quality filtering, "
                                                    "5:Clustering/denoising, 6:Read table, 7:Tax. assignment ")
    run_parser.add_argument('-steps', type=str, help="Select step from which to re-run all subsequent steps. "
                                                    "1:Index demultiplexing, 2:Primer trimming, "
                                                    "3:Tag demultiplexing, 4:Quality filtering, "
                                                    "5:Clustering/denoising, 6:Read table, 7:Tax. assignment ")
    # === Subparser: gui ===
    gui_parser = subparsers.add_parser('gui', help='Open Graphical User Interface.')

    # === Subparser: import_raw_data ===
    import_raw_data_parser = subparsers.add_parser('import_raw_data', help='Load all files from the raw data folder in the demultiplexing sheet.')
    import_raw_data_parser.add_argument('-p', '--project', type=str, required=True, help='Path to project.')

    # Parse arguments
    args = parser.parse_args()

    if args.command == 'gui':
        try:
            script_dir = os.path.dirname(os.path.abspath(__file__))
            script_path = os.path.join(script_dir, 'apscale_nanopore_gui.py')

            print('Press CTRL + C to close APSCALE-Nanopore in GUI mode!')
            subprocess.run(['streamlit', 'run', script_path, '--theme.base', 'dark', '--server.address', 'localhost'])

        except KeyboardInterrupt:
            print("Exiting...")
            sys.exit(1)

    # Create project
    if args.command == 'create':
        project_folder = Path(str(Path(args.project)) + '_apscale_nanopore')
        project_name = project_folder.name.replace('_apscale_nanopore', '')
        create_project(project_folder, project_name)

    ## Load samples
    if args.command == 'import_raw_data':
        project_folder = Path(args.project)
        project_name = project_folder.name.replace('_apscale_nanopore', '')
        settings_file = project_folder.joinpath(project_name + '_settings.xlsx')
        # Load settings dataframe
        if settings_file.exists():
            print('')
            print(f'{datetime.now().strftime("%H:%M:%S")} - Detected all .fastq files in the raw data folder and added them to the Demultiplexing sheet.')
            demultiplexing_df = pd.read_excel(settings_file, sheet_name='Demultiplexing')
            import_raw_data(project_folder, settings_file, demultiplexing_df)
            print(f'{datetime.now().strftime("%H:%M:%S")} - Update complete. Please review and adjust the Demultiplexing sheet as needed.')
            open_file(settings_file)
            print('')
        else:
            print(settings_file)
            print(f'{datetime.now().strftime("%H:%M:%S")} - Error: Cannot find settings file!')
            print('')

    if args.command == 'qc':
        print('Select for folder quality control: ')
        print('1: "1_raw_data/data"')
        print('2: "2_index_demultiplexing/data"')
        print('3: "3_primer_trimming/data"')
        res = input('Your choice: ')
        folder = {'1': "1_raw_data", '2': "2_index_demultiplexing", '3': "3_primer_trimming"}
        if res not in folder.keys():
            print('Please select a suitable folder!')
            return
        sub_folder = folder[res]
        project_folder = Path(args.project)
        quality_report(project_folder, sub_folder)

    # Run apscale
    if args.command == 'run':

        # Collect step information
        all_steps = {"1": "Index demultiplexing", "2": "Primer trimming",
                     "3": "Quality filtering", "4": "Clustering/denoising", "5": "Read table", "6": "Tax. assignment"}
        if args.step:
            steps = [all_steps[str(args.step)]]
        elif args.steps:
            steps = [all_steps[str(i)] for i in range(int(args.steps), 7)]
        else:
            steps = list(all_steps.values())
        if steps == []:
            print('Error: Please choose a suitable step index!')
            return

        # Define folders and files
        project_folder = Path(args.project)
        project_name = project_folder.name.replace('_apscale_nanopore', '')
        settings_file = project_folder.joinpath(project_name + '_settings.xlsx')

        # Load settings dataframe
        if settings_file.exists():
            settings_df = pd.read_excel(settings_file, sheet_name='Settings').fillna('')
            demultiplexing_df = pd.read_excel(settings_file, sheet_name='Demultiplexing').fillna('')

            # Check if argument require to be adjusted
            # allowed errors index
            if args.e1:
                index = settings_df[settings_df['Category'] == 'allowed errors index'].index[0]
                settings_df.loc[index, 'Variable'] = args.e1
                print(f'Adjusted value: Number of allowed index errors: {args.e1}')
            # allowed errors primer
            if args.e2:
                index = settings_df[settings_df['Category'] == 'allowed errors primer'].index[0]
                settings_df.loc[index, 'Variable'] = args.e2
                print(f'Adjusted value: Number of allowed primer errors: {args.e2}')
            # minimum length
            if args.minlen:
                index = settings_df[settings_df['Category'] == 'minimum length'].index[0]
                settings_df.loc[index, 'Variable'] = args.minlen
                print(f'Adjusted value: Minimum length: {args.minlen}')
            # maximum length
            if args.maxlen:
                index = settings_df[settings_df['Category'] == 'maximum length'].index[0]
                settings_df.loc[index, 'Variable'] = args.maxlen
                print(f'Adjusted value: Maximum length: {args.maxlen}')
            # minimum quality
            if args.minq:
                index = settings_df[settings_df['Category'] == 'minimum quality'].index[0]
                settings_df.loc[index, 'Variable'] = args.minq
                print(f'Adjusted value: Minimum mean PHRED quality: {args.minq}')
            # mode
            if args.mode:
                index = settings_df[settings_df['Category'] == 'mode'].index[0]
                settings_df.loc[index, 'Variable'] = args.mode
                print(f"Adjusted value: Clustering/denoising mode: {args.mode}")
            # percid
            if args.percid:
                index = settings_df[settings_df['Category'] == 'percid'].index[0]
                settings_df.loc[index, 'Variable'] = args.percid
                print(f"Adjusted value: Vsearch clustering percid value: {args.percid}")
            # alpha
            if args.alpha:
                index = settings_df[settings_df['Category'] == 'alpha'].index[0]
                settings_df.loc[index, 'Variable'] = args.alpha
                print(f"Adjusted value: Vsearch denoising alpha value: {args.alpha}")
            # d
            if args.d:
                index = settings_df[settings_df['Category'] == 'd'].index[0]
                settings_df.loc[index, 'Variable'] = args.d
                print(f"Adjusted value: Swarm's d value: {args.d}")
            # representative
            if args.rep:
                index = settings_df[settings_df['Category'] == 'representative'].index[0]
                settings_df.loc[index, 'Variable'] = args.rep
                print(f"Adjusted value: Clustering representative value: {args.rep}")
            # minimum reads
            if args.minreads:
                index = settings_df[settings_df['Category'] == 'minimum reads'].index[0]
                settings_df.loc[index, 'Variable'] = args.minreads
                print(f"Adjusted value: Read filter threshold: {args.minreads}")
            print('')

            # =======# Live processing #=======#
            if args.live_calling == True:
                apscale_nanopore_watch_folder(project_folder, settings_df, demultiplexing_df, steps, args.sd)
            else:
                apscale_nanopore(project_folder, settings_df, demultiplexing_df, steps, args.sd)

        else:
            print(settings_file)
            print('Error: Cannot find settings file!')

if __name__ == "__main__":
    main()