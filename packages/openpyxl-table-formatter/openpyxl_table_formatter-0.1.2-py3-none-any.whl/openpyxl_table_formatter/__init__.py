import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet


def main() -> None:
    pass


def add_formatted_table_to_worksheet(
    table: pd.DataFrame,
    worksheet: Worksheet | None,
    start_cell: str | None = None,
    index: bool = True,
    header: bool = True,
    style: str = "Medium 2",
    columns_formats: list[str] | None = None,
    columns_alignment: list[str] | None = None,
) -> None:
    """Appends a pandas DataFrame to file using one of default MS Office formattings.

    Parameters
    ----------
    table : pandas.DataFrame
        DataFrame to be added to the worksheet.
    worksheet : openpyxl.Worksheet
        Worksheet to use.
    worksheet : str or None, default: None
        Cell from which appending starts. If not provided - table will be appended starting in the first blank row of column A.
    index: bool, default: True
        If index of DataFrame should be included.
    header: bool, default: True
        If header of DataFrame should be included.
    style: str, default: "Medium 2"
        Style of table, from the Office 365 standard templates.
    columns_formats: list of str or None, default: None
        List of formats for each column, should follow Office 365 format standard, e.g. 'dd-mmm-yyyy'.
    columns_formats: list of str or None, default: None
            List of alignments for each column, should follow Office 365 standard, e.g. 'left', 'center', 'right'.
    """
    match style:
        case "Medium 1":
            header_fill_type = "solid"
            header_fill_color = "FF000000"
            header_font_color = "FFFFFFFF"
            odd_row_fill_type = "solid"
            odd_row_fill_color = "FFD9D9D9"
            even_row_fill_type = None
            even_row_fill_color = "00FFFFFF"
            border_color = "FF000000"
        case "Medium 2":
            header_fill_type = "solid"
            header_fill_color = "FF156082"
            header_font_color = "FFFFFFFF"
            odd_row_fill_type = "solid"
            odd_row_fill_color = "FFC0E6F5"
            even_row_fill_type = None
            even_row_fill_color = "00FFFFFF"
            border_color = "FF44B3E1"
        case "Medium 3":
            header_fill_type = "solid"
            header_fill_color = "FFE97123"
            header_font_color = "FFFFFFFF"
            odd_row_fill_type = "solid"
            odd_row_fill_color = "FFFBE2D5"
            even_row_fill_type = None
            even_row_fill_color = "00FFFFFF"
            border_color = "FFF1A983"
        case "Medium 4":
            header_fill_type = "solid"
            header_fill_color = "FF196B24"
            header_font_color = "FFFFFFFF"
            odd_row_fill_type = "solid"
            odd_row_fill_color = "FFC1F0C8"
            even_row_fill_type = None
            even_row_fill_color = "00FFFFFF"
            border_color = "FF47D359"
        case "Medium 5":
            header_fill_type = "solid"
            header_fill_color = "FF0F9ED5"
            header_font_color = "FFFFFFFF"
            odd_row_fill_type = "solid"
            odd_row_fill_color = "FFCAEDFB"
            even_row_fill_type = None
            even_row_fill_color = "00FFFFFF"
            border_color = "FF61CBF3"
        case "Medium 6":
            header_fill_type = "solid"
            header_fill_color = "FFA02B93"
            header_font_color = "FFFFFFFF"
            odd_row_fill_type = "solid"
            odd_row_fill_color = "FFF2CEEF"
            even_row_fill_type = None
            even_row_fill_color = "00FFFFFF"
            border_color = "FFD86DCD"
        case "Medium 7":
            header_fill_type = "solid"
            header_fill_color = "FF4EA72E"
            header_font_color = "FFFFFFFF"
            odd_row_fill_type = "solid"
            odd_row_fill_color = "FFDAF2D0"
            even_row_fill_type = None
            even_row_fill_color = "00FFFFFF"
            border_color = "FF8ED973"
        case "Custom Red":
            header_fill_type = "solid"
            header_fill_color = "FFED4C37"
            header_font_color = "FFFFFFFF"
            odd_row_fill_type = "solid"
            odd_row_fill_color = "FFFACEC8"
            even_row_fill_type = None
            even_row_fill_color = "00FFFFFF"
            border_color = "FFF06E5C"
        case _:
            header_fill_type = None
            header_fill_color = "00FFFFFF"
            header_font_color = "FF000000"
            odd_row_fill_type = None
            odd_row_fill_color = "00FFFFFF"
            even_row_fill_type = None
            even_row_fill_color = "00FFFFFF"
            border_color = "FF000000"
    if worksheet is not None:
        if start_cell is not None:
            start_row = worksheet[start_cell].row
            start_column = worksheet[start_cell].column
        else:
            max_h: int = worksheet.max_row
            if max_h == 1 and worksheet.cell(row=1, column=1).value is None:
                start_row = 1
                start_column = 1
            else:
                start_row = max_h + 1
                start_column = 1
        for row_count, dataframe_row in enumerate(
            dataframe_to_rows(table, index=index, header=header)
        ):
            if dataframe_row == [None]:
                continue
            if header and row_count == 0:
                for column_number, cell_value in enumerate(dataframe_row):
                    current_cell = worksheet.cell(
                        row=start_row, column=column_number + start_column
                    )
                    current_cell.value = cell_value
                    current_cell.font = Font(
                        name="Calibri", bold=True, color=header_font_color
                    )
                    current_cell.fill = PatternFill(
                        fill_type=header_fill_type,
                        start_color=header_fill_color,
                        end_color=header_fill_color,
                    )
                    if column_number == 0:
                        current_cell.border = Border(
                            top=Side(style="thin", color=border_color),
                            left=Side(style="thin", color=border_color),
                            bottom=Side(style="thin", color=border_color),
                        )
                    elif column_number == (len(dataframe_row) - 1):
                        current_cell.border = Border(
                            top=Side(style="thin", color=border_color),
                            right=Side(style="thin", color=border_color),
                            bottom=Side(style="thin", color=border_color),
                        )
                    else:
                        current_cell.border = Border(
                            top=Side(style="thin", color=border_color),
                            bottom=Side(style="thin", color=border_color),
                        )
            else:
                if header and index:
                    is_odd_row = row_count % 2 == 0
                    current_row = start_row + row_count - 1
                elif header and not index:
                    is_odd_row = row_count % 2 == 0
                    current_row = start_row + row_count
                elif not header and index:
                    is_odd_row = row_count % 2 == 1
                    current_row = start_row + row_count - 1
                else:
                    is_odd_row = row_count % 2 == 0
                    current_row = start_row + row_count
                for column_number, cell_value in enumerate(dataframe_row):
                    current_cell = worksheet.cell(
                        row=current_row, column=column_number + start_column
                    )
                    current_cell.value = cell_value
                    if (columns_formats is not None) and (
                        len(columns_formats) == len(dataframe_row)
                    ):
                        current_cell.number_format = columns_formats[column_number]
                    if (columns_alignment is not None) and (
                        len(columns_alignment) == len(dataframe_row)
                    ):
                        current_cell.alignment = Alignment(
                            horizontal=columns_alignment[column_number]
                        )
                    if column_number == 0:
                        current_cell.border = Border(
                            top=Side(style="thin", color=border_color),
                            left=Side(style="thin", color=border_color),
                            bottom=Side(style="thin", color=border_color),
                        )
                    elif column_number == (len(dataframe_row) - 1):
                        current_cell.border = Border(
                            top=Side(style="thin", color=border_color),
                            right=Side(style="thin", color=border_color),
                            bottom=Side(style="thin", color=border_color),
                        )
                    else:
                        current_cell.border = Border(
                            top=Side(style="thin", color=border_color),
                            bottom=Side(style="thin", color=border_color),
                        )
                    if is_odd_row:
                        current_cell.fill = PatternFill(
                            fill_type=odd_row_fill_type,
                            start_color=odd_row_fill_color,
                            end_color=odd_row_fill_color,
                        )
                    else:
                        current_cell.fill = PatternFill(
                            fill_type=even_row_fill_type,
                            start_color=even_row_fill_color,
                            end_color=even_row_fill_color,
                        )


if __name__ == "__main__":
    main()
