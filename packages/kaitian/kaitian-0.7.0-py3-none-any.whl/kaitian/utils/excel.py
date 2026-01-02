from __future__ import annotations

import logging
from colorsys import hls_to_rgb, rgb_to_hls
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any, Iterable, Literal

import openpyxl
import polars as pl
import xlsxwriter
import xlsxwriter.format
import xlsxwriter.worksheet
from openpyxl.styles.colors import COLOR_INDEX, Color
from openpyxl.utils import coordinate_to_tuple, get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.xml.functions import QName, fromstring
from typing_extensions import Self
from xlsxwriter.utility import xl_cell_to_rowcol

logger = logging.getLogger(__name__)


def rgb_tint(rgb_str: str, tint: float) -> str:
    """颜色亮度调整

    <0: 变暗, 可用于深化背景色
    >0: 变量, 可用于忽略不重要项目
    """
    if rgb_str.startswith("#"):
        rgb_str = rgb_str[1:]
    r = int(rgb_str[0:2], 16)
    g = int(rgb_str[2:4], 16)
    b = int(rgb_str[4:6], 16)
    h_, l_, s_ = rgb_to_hls(r / 255.0, g / 255.0, b / 255.0)
    if tint != 0:
        lum = l_ * 240
        if tint < 0:
            new_lum = lum * (1.0 + tint)
        else:
            new_lum = lum * (1.0 - tint) + (240 - 240 * (1.0 - tint))
        new_lum = max(0, min(240, new_lum))
        new_l = new_lum / 240
    else:
        new_l = l_
    new_r, new_g, new_b = hls_to_rgb(h_, new_l, s_)
    r_final = min(255, max(0, round(new_r * 255)))
    g_final = min(255, max(0, round(new_g * 255)))
    b_final = min(255, max(0, round(new_b * 255)))

    return f"#{r_final:02X}{g_final:02X}{b_final:02X}"


@dataclass
class CellFormat:
    align: str | None = None
    valign: str | None = None
    text_wrap: bool = False
    background_color: str | None = None
    num_format: str | None = None

    def to_writer(self) -> dict:
        fmt_all = {
            "align": self.align,
            "valign": self.valign,
            "text_wrap": self.text_wrap,
            "bg_color": self.background_color,
            "num_format": self.num_format,
        }
        return {k: v for k, v in fmt_all.items() if v is not None}

    @classmethod
    def load(cls, **kwargs) -> Self:
        return cls(**kwargs)


@dataclass
class FontFormat:
    """字体格式对象"""

    font_name: str | None
    font_size: float = 11
    font_color: str = "#000000"
    bold: bool = False
    italic: bool = False
    underline: Literal["single", "double", "single_full", "double_full"] | None = None

    def to_writer(self) -> dict:
        if self.underline is None:
            _underline = None
        elif self.underline == "single":
            _underline = 1
        elif self.underline == "double":
            _underline = 2
        elif self.underline == "single_full":
            _underline = 33
        elif self.underline == "double_full":
            _underline = 34
        else:
            logger.warning(f"underline {self.underline} is not supported")
            _underline = 1

        fmt_all = {
            "font_name": self.font_name,
            "font_size": self.font_size,
            "font_color": self.font_color,
            "bold": self.bold,
            "italic": self.italic,
            "underline": _underline,
        }
        return {k: v for k, v in fmt_all.items() if v is not None}

    @classmethod
    def load(cls, **kwargs) -> Self:
        return cls(**kwargs)


@dataclass
class BorderFormat:
    """边框格式对象"""

    direction: str | Literal["top", "bottom", "left", "right", "all"] = "all"
    style: str = "thin"
    color: str | None = "#000000"

    @staticmethod
    def _style2idx(style: str) -> int:
        if style == "thin":
            return 1
        elif style == "double":
            return 5
        elif style == "medium":
            return 2
        elif style == "mediumDashed":
            return 8
        else:
            return 1

    def to_writer(self) -> dict:
        if self.direction == "all":
            return {"border": self._style2idx(self.style), "border_color": self.color}
        else:
            return {
                f"{self.direction}": self._style2idx(self.style),
                f"{self.direction}_color": self.color,
            }

    @classmethod
    def load(cls, **kwargs) -> Self:
        return cls(**kwargs)


@dataclass
class ConditionalFormat:  # TODO
    """Sheet 格式对象

    条件格式:
    """

    cell_range: str | tuple[int, int, int, int]  # A1:B2 or (1, 1, 2, 2)
    ftype: Literal["data_bar", "color_scale"]
    fopt: dict | None = None

    def to_writer(self) -> dict:
        if isinstance(self.cell_range, str):
            cell_range = (
                *(xl_cell_to_rowcol(self.cell_range.split(":")[0])),
                *(xl_cell_to_rowcol(self.cell_range.split(":")[1])),
            )
        else:
            cell_range = self.cell_range

        fmt: dict[str, Any] = {
            "first_row": cell_range[0],
            "first_col": cell_range[1],
            "last_row": cell_range[2],
            "last_col": cell_range[3],
        }

        if self.ftype == "data_bar":
            fmt["options"] = {"type": "data_bar", **(self.fopt or {})}
        elif self.ftype == "color_scale":
            opt_name = list(map(lambda x: x.split("_")[0], list((self.fopt or {}).keys())))
            if "mid" in opt_name:
                fmt["options"] = {"type": "3_color_scale", **(self.fopt or {})}
            else:
                fmt["options"] = {"type": "2_color_scale", **(self.fopt or {})}

        return fmt


@dataclass
class DataBar:
    """数据条格式"""

    min_type: Literal["min", "num", "percent", "percentile", "formula"] | None = None
    max_type: Literal["max", "num", "percent", "percentile", "formula"] | None = None
    min_value: float | None = None
    max_value: float | None = None
    bar_color: str | None = None
    bar_only: bool = False
    bar_solid: bool = False
    bar_negative_color: str | None = None
    bar_border_color: str | None = None
    bar_negative_border_color: str | None = None
    bar_negative_color_same: bool = False
    bar_negative_border_color_same: bool = False
    bar_no_border: bool = False
    bar_direction: Literal["left", "right"] = "left"
    bar_axis_position: Literal["middle"] | None = None
    bar_axis_color: str | None = None
    data_bar_2010: bool = True

    def to_writer(self) -> dict:
        opt = {"type": "data_bar"}
        for k, v in asdict(self).items():
            if v is not None:
                opt[k] = v
        return opt


@dataclass
class ColorScale:
    """热力图"""

    min_type: Literal["min", "num", "percent", "percentile", "formula"] | None = None
    mid_type: Literal["min", "num", "percent", "percentile", "formula"] | None = None
    max_type: Literal["max", "num", "percent", "percentile", "formula"] | None = None
    min_value: float | None = None
    mid_value: float | None = None
    max_value: float | None = None
    min_color: str | None = None
    mid_color: str | None = None
    max_color: str | None = None

    def to_writer(self) -> dict:
        if self.mid_color is not None:
            opt = {"type": "3_color_scale"}
        else:
            opt = {"type": "2_color_scale"}
        for k, v in asdict(self).items():
            if v is not None:
                opt[k] = v
        return opt


class FormatParser:
    """格式解析器

    based on openpyxl

    - 背景色 / 主题色
    - 文本格式
    - 条件格式
    - 边框样式
    - 单元格尺寸
    """

    HLSMAX = 240
    XLMNS = "http://schemas.openxmlformats.org/drawingml/2006/main"

    def __init__(self, workbook: openpyxl.Workbook | Path | str) -> None:
        if isinstance(workbook, (Path, str)):
            self.workbook = openpyxl.load_workbook(Path(workbook), read_only=False)
        else:
            self.workbook = workbook

        self._parse_theme_color()
        self._logger = logging.getLogger(__name__)

    def _parse_theme_color(self) -> None:
        _color_scheme = (
            fromstring(self.workbook.loaded_theme)
            .find(QName(self.XLMNS, "themeElements").text)
            .findall(QName(self.XLMNS, "clrScheme").text)[0]
        )

        self.colors = []

        for _c in [
            "lt1",
            "dk1",
            "lt2",
            "dk2",
            "accent1",
            "accent2",
            "accent3",
            "accent4",
            "accent5",
            "accent6",
        ]:
            accent = _color_scheme.find(QName(self.XLMNS, _c).text)
            for i in list(accent):  # walk all child nodes, rather than assuming [0]
                if "window" in i.attrib["val"]:
                    self.colors.append(i.attrib["lastClr"])
                else:
                    self.colors.append(i.attrib["val"])

    def get_theme_color(self, theme: int = 0) -> str:
        """获取主题色"""
        return f"#{self.colors[theme % len(self.colors)]}"

    def get_indexed_color(self, index: int = 0) -> str:
        return f"#{COLOR_INDEX[index % len(COLOR_INDEX)][2:]}"

    def get_sheet(self, sheet: str | int) -> Worksheet:
        if isinstance(sheet, str):
            worksheet = self.workbook[sheet]
        elif isinstance(sheet, int):
            worksheet = self.workbook.worksheets[sheet]
        else:
            raise ValueError("sheet must be str or int")

        return worksheet

    def _openpyxl_color2rgb(self, color: Color, default: str | None = None) -> str:
        default_rgb = default or "#000000"

        if color.type == "rgb":
            if color.rgb == "00000000":
                return default_rgb
            return f"#{color.rgb[2:]}"
        elif color.type == "theme":
            return self.get_theme_color(color.theme)
        elif color.type == "indexed":
            return self.get_indexed_color(color.indexed)
        elif color.type == "auto":
            return default_rgb
        else:
            self._logger.warning(f"Unknown color type: {color.type}")
            return default_rgb

    def get_format(self, sheet: str | int | Worksheet, position: str | tuple[int, int]) -> dict:
        if isinstance(sheet, (str, int)):
            worksheet = self.get_sheet(sheet)
        elif isinstance(sheet, Worksheet):
            worksheet = sheet
        else:
            raise ValueError("sheet must be str or int or Worksheet")

        if isinstance(position, str):
            cell = worksheet.cell(*coordinate_to_tuple(position))
            col_s = get_column_letter(coordinate_to_tuple(position)[1])
            row_s = coordinate_to_tuple(position)[0]
        else:
            cell = worksheet.cell(*position)
            col_s = get_column_letter(position[1])
            row_s = position[0]

        fmt = {}
        # 字体 / 对齐 / 数值格式
        font = cell.font
        fmt["font"] = FontFormat.load(
            **{
                "font_name": font.name if font is not None else None,
                "font_size": font.size if font is not None else None,
                "font_color": self._openpyxl_color2rgb(font.color),
                "bold": font.bold if font is not None else None,
                "italic": font.italic if font is not None else None,
                "underline": font.underline if font is not None else None,
            }
        )
        fmt["cell"] = CellFormat.load(
            **{
                "align": cell.alignment.horizontal if cell.alignment is not None else None,
                "valign": cell.alignment.vertical if cell.alignment is not None else None,
                "text_wrap": cell.alignment.wrap_text if cell.alignment is not None else False,
                "num_format": cell.number_format,
                "background_color": self._openpyxl_color2rgb(cell.fill.fgColor, default="#FFFFFF")
                if cell.fill is not None
                else None,
            }
        )

        # 宽高
        col_dim = worksheet.column_dimensions[col_s]
        row_dim = worksheet.row_dimensions[row_s]
        if getattr(col_dim, "width") is not None:
            width = col_dim.width
        else:
            width = col_dim.style

        if getattr(row_dim, "height") is not None:
            height = row_dim.height
        else:
            height = row_dim.s

        fmt["width"] = round(width, 2)  # FIXME: 与WPS显示的字符宽度存在差异
        fmt["height"] = height

        # 边框
        borders = []
        for direction in ["left", "right", "top", "bottom"]:
            if getattr(cell.border, direction).style is not None:
                border = getattr(cell.border, direction)
                borders.append(
                    BorderFormat.load(
                        **{
                            "direction": direction,
                            "style": border.style,
                            "color": self._openpyxl_color2rgb(border.color, default="#000000"),
                        }
                    )
                )
        fmt["borders"] = borders

        # TODO: 条件格式
        fmt["conditional_formats"] = None

        return fmt


class ExcelWriter:
    """Excel读取器

    based on xlsxwriter

    Notes
    -----
    - xlsxwriter 的坐标默认以 0 开始, openpyxl 默认以 1 开始, 更符合使用情况, 需要转换
    """

    def __init__(self, file_path: str | Path) -> None:
        self.workbook = xlsxwriter.Workbook(Path(file_path), options={"nan_inf_to_errors": True})
        self._format_dict = {}
        self._logger = logging.getLogger(__name__)
        self._cursor_row = 1  # 当前行数
        self._cursor_col = 1  # 当前列数
        self._cursor_sheet: str | None = None  # 当前 Sheet
        self._cursor: dict[str, tuple[int, int]] = {}  # 当前 Sheet 的坐标

    def done(self) -> None:
        self.workbook.close()

    def _get_worksheet_or_create(self, sheet: str | None = None) -> xlsxwriter.worksheet.Worksheet:
        if sheet is None and self._cursor_sheet is None:
            worksheet = self.workbook.add_worksheet()
            assert worksheet.name is not None, "创建 Sheet 失败"
            self._cursor_sheet = worksheet.name
            self._cursor[worksheet.name] = (1, 1)
            self._logger.debug(f"新建默认 Sheet: {self._cursor_sheet}")
            self.set_default_format(sheet=self._cursor_sheet)
        elif isinstance(sheet, str) or self._cursor_sheet is not None:
            sheet = sheet or self._cursor_sheet
            assert sheet is not None, "sheet must be str"
            if sheet in self.workbook.sheetnames:
                self._logger.debug(f"使用已有的 Sheet: {sheet}")
                worksheet = self.workbook.get_worksheet_by_name(sheet)
                self._cursor_sheet = sheet
                if sheet not in self._cursor:  # 坐标缺失
                    self._cursor[sheet] = (1, 1)
            else:
                self._logger.debug(f"新建 Sheet: {sheet}")
                worksheet = self.workbook.add_worksheet(sheet)
                self._cursor[sheet] = (1, 1)
                self._cursor_sheet = sheet
                self.set_default_format(sheet=self._cursor_sheet)
        else:
            raise ValueError("sheet must be str or None")
        assert worksheet is not None
        return worksheet

    def add_format(
        self,
        name: str | None = None,
        cell: CellFormat | None = None,
        font: FontFormat | None = None,
        border: BorderFormat | None = None,
    ) -> Self:
        """预设格式

        可预设 str -> Format 的映射, 绘制接口均可直接使用, 也可设置默认格式,
        默认格式会通过 self.set_default_format 接口自动添加到所有的新 Sheet.
        """
        if name is None:
            _name = "default"
        else:
            _name = name

        if self._format_dict.get(_name) is None:
            self._format_dict[_name] = {}

        self._format_dict[_name]["cell"] = cell
        self._format_dict[_name]["font"] = font
        self._format_dict[_name]["border"] = border
        self._logger.debug(f"更新格式: {_name}")
        return self

    def _parse_format(
        self,
        cell_format: CellFormat | None = None,
        font_format: FontFormat | None = None,
        border_format: BorderFormat | list[BorderFormat] | None = None,
    ) -> dict:
        fmt = {}

        if cell_format is not None:
            fmt.update(cell_format.to_writer())

        if font_format is not None:
            fmt.update(font_format.to_writer())

        if isinstance(border_format, BorderFormat):
            border_format = [border_format]
            for bdfmt in border_format:
                fmt.update(bdfmt.to_writer())

        return fmt

    def get_format(
        self,
        name: str | None = None,
        cell_format: CellFormat | None = None,
        font_format: FontFormat | None = None,
        border_format: BorderFormat | list[BorderFormat] | None = None,
    ) -> xlsxwriter.format.Format:
        if name is not None:
            fmt_base = self._format_dict.get(name)
        else:
            fmt_base = self._format_dict.get("default")

        if fmt_base is not None:
            fmt = self._parse_format(fmt_base.get("cell"), fmt_base.get("font"), fmt_base.get("border"))
        else:
            fmt = {}

        fmt.update(self._parse_format(cell_format, font_format, border_format))
        return self.workbook.add_format(fmt)

    def fill_cell(
        self,
        value: Any,
        sheet: str | None = None,
        position: str | None = None,
        row: int | None = None,
        col: int | None = None,
        format_name: str | None = None,
        cell_format: CellFormat | None = None,
        font_format: FontFormat | None = None,
        border_format: BorderFormat | None = None,
    ) -> Self:
        """填充单元格"""
        worksheet = self._get_worksheet_or_create(sheet)
        assert worksheet.name is not None, "获取工作簿失败"
        _cursor_row, _cursor_col = self._cursor.get(worksheet.name, (1, 1))

        row = (row or _cursor_row) - 1  # 指定的坐标从 1 开始
        col = (col or _cursor_col) - 1

        if position is not None:
            row, col = xl_cell_to_rowcol(position)

        worksheet.write(row, col, value, self.get_format(format_name, cell_format, font_format, border_format))
        # 更新游标
        # cell |  -    |
        #   -  | cell* |
        self._cursor[worksheet.name] = (row + 2, col + 2)
        return self

    def fill_row(
        self,
        value: Iterable[Any],
        sheet: str | None = None,
        position: str | None = None,
        row: int | None = None,
        col: int | None = None,
        format_name: str | None = None,
        cell_format: CellFormat | None = None,
        font_format: FontFormat | None = None,
        border_format: BorderFormat | None = None,
    ) -> Self:
        """填充整行"""
        worksheet = self._get_worksheet_or_create(sheet)
        assert worksheet.name is not None, "获取工作簿失败"
        _cursor_row, _cursor_col = self._cursor.get(worksheet.name, (1, 1))

        row = (row or _cursor_row) - 1  # 指定的坐标从 1 开始
        col = (col or _cursor_col) - 1

        if position is not None:
            row, col = xl_cell_to_rowcol(position)

        worksheet.write_row(
            row, col, value, cell_format=self.get_format(format_name, cell_format, font_format, border_format)
        )
        # 更新游标
        # row  +----->
        # row* |
        self._cursor[worksheet.name] = (row + 2, col + 1)
        return self

    def fill_column(
        self,
        value: Iterable[Any],
        sheet: str | None = None,
        position: str | None = None,
        row: int | None = None,
        col: int | None = None,
        format_name: str | None = None,
        cell_format: CellFormat | None = None,
        font_format: FontFormat | None = None,
        border_format: BorderFormat | None = None,
    ) -> Self:
        """填充整列"""
        worksheet = self._get_worksheet_or_create(sheet)
        assert worksheet.name is not None, "获取工作簿失败"
        _cursor_row, _cursor_col = self._cursor.get(worksheet.name, (1, 1))

        row = (row or _cursor_row) - 1  # 指定的坐标从 1 开始
        col = (col or _cursor_col) - 1

        if position is not None:
            row, col = xl_cell_to_rowcol(position)

        worksheet.write_column(
            row, col, value, cell_format=self.get_format(format_name, cell_format, font_format, border_format)
        )
        # 更新游标
        # col | col* |
        # --- |  --  |
        self._cursor[worksheet.name] = (row + 1, col + 2)
        return self

    def fill_merge(
        self,
        value: Any,
        nrows: int,
        ncols: int,
        sheet: str | None = None,
        position: str | None = None,
        row: int | None = None,
        col: int | None = None,
        format_name: str | None = None,
        cell_format: CellFormat | None = None,
        font_format: FontFormat | None = None,
        border_format: BorderFormat | None = None,
    ) -> Self:
        """填充合并单元格"""
        worksheet = self._get_worksheet_or_create(sheet)
        assert worksheet.name is not None, "获取工作簿失败"
        _cursor_row, _cursor_col = self._cursor.get(worksheet.name, (1, 1))

        row = (row or _cursor_row) - 1  # 指定的坐标从 1 开始
        col = (col or _cursor_col) - 1

        if position is not None:
            row, col = xl_cell_to_rowcol(position)

        _ = worksheet.merge_range(
            first_row=row,
            first_col=col,
            last_row=row + nrows - 1,
            last_col=col + ncols - 1,
            data=value,
            cell_format=self.get_format(format_name, cell_format, font_format, border_format),
        )
        self._cursor[worksheet.name] = (row + nrows + 1, col + ncols + 1)
        return self

    def write_table(
        self,
        data: pl.DataFrame | list[pl.DataFrame],
        title: str | None = None,
        index: str | None = None,
        sheet: str | None = None,
        position: str | None = None,
        row: int | None = None,
        col: int | None = None,
        title_format: str | None = None,
        header_format: str | None = None,
        body_format: str | None = None,  # TODO: 增加 body_format : dict 类型, 用来区分不同列的样式
        index_format: str | None = None,
    ) -> Self:
        """写入表格

        支持写入单表和多表, 其中多表必须能够进行 pl.concat 合并, 进行的特殊操作仅有调整背景色使多个表明暗交错

        Parameters
        ----------
        data : pl.DataFrame | list[pl.DataFrame]
            需要绘制的表格
        title : str | None, default None
            标题是跨整个表头的合并单元格, 变量内容为标题名, 默认无
        index : str | None, default None
            索引列可新建可选择, 默认无
        """
        worksheet = self._get_worksheet_or_create(sheet)
        assert worksheet.name is not None, "获取工作簿失败"
        _cursor_row, _cursor_col = self._cursor.get(worksheet.name, (1, 1))

        row = row or _cursor_row
        col = col or _cursor_col

        if position is not None:
            row, col = xl_cell_to_rowcol(position)
            row += 1
            col += 1

        if isinstance(data, list):
            data_ = pl.concat(data)
            data_list = data
        else:
            data_ = data
            data_list = [data]

        if index is not None:
            if index in data_.columns:
                index_list = data_.get_column(index).to_list()
            else:
                index_list = list(range(1, data_.height + 1))
            features = [f for f in data_.columns if f != index]
            width_ = len(features) + 1
        else:
            features = data_.columns
            width_ = data_.width
            index_list = []

        # 绘制标题
        if title is not None:
            self.fill_merge(
                value=title,
                nrows=1,
                ncols=width_,
                row=row,
                col=col,
                sheet=worksheet.name,
                format_name=title_format,
            )
            self._logger.debug(f"绘制标题: ({row}, {col})")
            row += 1

        # 绘制 Index
        if index is not None:
            self.fill_cell(value=index, row=row, col=col, sheet=worksheet.name, format_name=header_format)
            self.fill_column(value=index_list, row=row + 1, col=col, sheet=worksheet.name, format_name=index_format)
            col += 1

        # 绘制表头
        self.fill_row(value=features, row=row, col=col, format_name=header_format)
        self._logger.debug(f"绘制表头: ({row}, {col})")
        row += 1

        # 绘制表体
        body_fmt = self.get_format(body_format)
        if body_fmt.bg_color is not None:
            bg_color = f"#{body_fmt.bg_color._rgb_hex_value()}"
        else:
            bg_color = "#FFFFFF"
        bg_color_shadow = rgb_tint(bg_color, tint=-0.05)
        cell_shadow = CellFormat(background_color=bg_color_shadow)
        col_from = col
        for idx, body in enumerate(data_list):
            for _, series in enumerate(body):
                _ = self.fill_column(
                    value=series.to_list(),
                    row=row,
                    col=col,
                    sheet=worksheet.name,
                    format_name=body_format,
                    cell_format=cell_shadow if idx % 2 != 0 else None,
                )
                col += 1
            col = col_from
            row = row + body.height

        self._cursor[worksheet.name] = (row + 2, col)

        return self

    def write_chart(self) -> Self:
        """写入图表

        将其他库生成的二进制图像流写入表格
        """
        raise NotImplementedError

    def set_height(self, height: float, sheet: str | None = None, row: int | list[int] | str | None = None) -> Self:
        """设置行高

        设置表格的行高, 支持不同类型的 row 来实现多行同时配置:

        1. int 单行行号处理
        2. Iterable[int] 多行行号处理
        3. str-range 多行处理: 1:2
        4. None 只处理当前所在行
        """
        worksheet = self._get_worksheet_or_create(sheet)
        assert worksheet.name is not None, "获取工作簿失败"
        _cursor_row, _ = self._cursor.get(worksheet.name, (1, 1))

        if isinstance(row, int):
            rows = [row]
        elif isinstance(row, list):
            rows = row
        elif isinstance(row, str) and ":" in row:
            r_from = int(row.split(":")[0])
            r_to = int(row.split(":")[1])
            rows = list(range(r_from, r_to + 1))
        elif row is None:
            rows = [_cursor_row]
        else:
            raise ValueError(f"row 参数错误: {row}")

        for r in rows:
            worksheet.set_row(r - 1, height)
        return self

    def set_width(self, width: float, sheet: str | None = None, col: int | list[int] | str | None = None) -> Self:
        """设置列宽

        设置表格的列宽, 支持不同类型的 col 来实现多列同时配置:

        1. int 单列序号处理
        2. Iterable[int] 多列序号处理
        3. Iterable[str] 多列标识处理
        4. str-range 多列处理: A:C
        5. str-single 单列处理: A
        6. None 只处理当前所在列
        """
        worksheet = self._get_worksheet_or_create(sheet)
        assert worksheet.name is not None, "获取工作簿失败"
        _, _cursor_col = self._cursor.get(worksheet.name, (1, 1))

        if isinstance(col, int):
            cols = [col]
        elif isinstance(col, list):
            cols = []
            for c in col:
                if isinstance(c, str):
                    cols.append(xl_cell_to_rowcol(f"{c}1")[1] + 1)
                else:
                    cols.append(c)
        elif isinstance(col, str) and ":" in col:
            c_from = col.split(":")[0]
            c_to = col.split(":")[1]
            _, c_from_idx = xl_cell_to_rowcol(f"{c_from}1")
            _, c_to_idx = xl_cell_to_rowcol(f"{c_to}1")
            cols = list(range(c_from_idx + 1, c_to_idx + 2))
        elif isinstance(col, str) and ":" not in col:
            _, c_ = xl_cell_to_rowcol(f"{col}1")
            cols = [c_ + 1]
        elif col is None:
            cols = [_cursor_col]
        else:
            raise ValueError(f"row 参数错误: {col}")

        for c in cols:
            worksheet.set_column(first_col=c - 1, last_col=c - 1, width=width)
        return self

    def set_default_format(
        self,
        sheet: str | None = None,
        format_name: str | None = None,
        cell_format: CellFormat | None = None,
        font_format: FontFormat | None = None,
        border_format: BorderFormat | None = None,
    ) -> Self:
        worksheet = self._get_worksheet_or_create(sheet)
        assert worksheet.name is not None, "获取工作簿失败"
        fmt = self.get_format(format_name, cell_format, font_format, border_format)
        worksheet.set_column(first_col=0, last_col=worksheet.xls_colmax - 1, cell_format=fmt)
        return self

    def add_databar(
        self,
        bar_format: DataBar,
        cell_range: str | None = None,
        first_row: int | None = None,
        first_col: int | None = None,
        last_row: int | None = None,
        last_col: int | None = None,
        sheet: str | None = None,
    ) -> Self:
        worksheet = self._get_worksheet_or_create(sheet)
        assert worksheet.name is not None, "获取工作簿失败"

        if cell_range is not None:
            corrd = (
                *(xl_cell_to_rowcol(cell_range.split(":")[0])),
                *(xl_cell_to_rowcol(cell_range.split(":")[1])),
            )
        else:
            assert first_row is not None and first_col is not None and last_row is not None and last_col is not None, (
                "请指定 cell_range 或 first_row, first_col, last_row, last_col 参数"
            )
            corrd = (first_row - 1, first_col - 1, last_row - 1, last_col - 1)

        worksheet.conditional_format(*corrd, bar_format.to_writer())
        return self

    def add_colorscale(
        self,
        color_format: ColorScale,
        cell_range: str | None = None,
        first_row: int | None = None,
        first_col: int | None = None,
        last_row: int | None = None,
        last_col: int | None = None,
        sheet: str | None = None,
    ) -> Self:
        worksheet = self._get_worksheet_or_create(sheet)
        assert worksheet.name is not None, "获取工作簿失败"

        if cell_range is not None:
            corrd = (
                *(xl_cell_to_rowcol(cell_range.split(":")[0])),
                *(xl_cell_to_rowcol(cell_range.split(":")[1])),
            )
        else:
            assert first_row is not None and first_col is not None and last_row is not None and last_col is not None, (
                "请指定 cell_range 或 first_row, first_col, last_row, last_col 参数"
            )
            corrd = (first_row - 1, first_col - 1, last_row - 1, last_col - 1)

        worksheet.conditional_format(*corrd, color_format.to_writer())
        return self

    # TODO: 这个只支持 range 写法
    # TODO: 所有的 writer 统一为 row, col, nrows, ncols 的写法, 其中 col 可以用字母
    def add_minichart(
        self,
        data_range: str,
        chart_type: Literal["column", "line"],
        data_sheet: str | None = None,
        sheet: str | None = None,
        position: str | None = None,
        row: int | None = None,
        col: int | None = None,
        high_point: bool = False,
        low_point: bool = False,
        series_color: str | None = None,
        min: float | None = None,
        max: float | None = None,
        nrows: int = 1,
        ncols: int = 1,
        content: str | None = None,
        format_name: str | None = None,
        cell_format: CellFormat | None = None,
        font_format: FontFormat | None = None,
        border_format: BorderFormat | None = None,
    ) -> Self:
        worksheet = self._get_worksheet_or_create(sheet)
        assert worksheet.name is not None, "获取工作簿失败"
        datasheet = self._get_worksheet_or_create(data_sheet)
        assert datasheet.name is not None, "获取数据簿失败"

        _cursor_row, _cursor_col = self._cursor.get(worksheet.name, (1, 1))

        _row = (row or _cursor_row) - 1  # 指定的坐标从 1 开始
        _col = (col or _cursor_col) - 1

        if position is not None:
            _row, _col = xl_cell_to_rowcol(position)

        options = {
            "range": f"{datasheet.name}!{data_range}",
            "type": chart_type,
            "high_point": high_point,
            "low_point": low_point,
            "min": min,
            "max": max,
            # First / Last Point
        }

        if series_color is not None:
            options["series_color"] = series_color

        self.fill_merge(
            value=content or "",
            nrows=nrows,
            ncols=ncols,
            sheet=sheet,
            position=position,
            row=row,
            col=col,
            format_name=format_name,
            cell_format=cell_format,
            font_format=font_format,
            border_format=border_format,
        )

        worksheet.add_sparkline(row=_row, col=_col, options=options)
        return self
