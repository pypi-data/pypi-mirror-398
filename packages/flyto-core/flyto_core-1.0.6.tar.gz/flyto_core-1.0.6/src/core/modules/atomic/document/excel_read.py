"""
Excel Read Module
Read data from Excel files (xlsx, xls)
"""
import logging
import os
from typing import Any, Dict, List, Optional

from ...registry import register_module


logger = logging.getLogger(__name__)


@register_module(
    module_id='excel.read',
    version='1.0.0',
    category='document',
    subcategory='excel',
    tags=['excel', 'spreadsheet', 'read', 'xlsx', 'data'],
    label='Read Excel',
    label_key='modules.excel.read.label',
    description='Read data from Excel files (xlsx, xls)',
    description_key='modules.excel.read.description',
    icon='Table',
    color='#217346',

    # Connection types
    input_types=['file_path'],
    output_types=['array', 'object'],
    can_connect_to=['array.*', 'data.*', 'database.*'],

    # Execution settings
    timeout=60,
    retryable=False,
    concurrent_safe=True,

    # Security settings
    requires_credentials=False,
    handles_sensitive_data=True,
    required_permissions=['file.read'],

    params_schema={
        'path': {
            'type': 'string',
            'label': 'Excel Path',
            'label_key': 'modules.excel.read.params.path.label',
            'description': 'Path to the Excel file',
            'description_key': 'modules.excel.read.params.path.description',
            'required': True,
            'placeholder': '/path/to/data.xlsx'
        },
        'sheet': {
            'type': 'string',
            'label': 'Sheet Name',
            'label_key': 'modules.excel.read.params.sheet.label',
            'description': 'Sheet name to read (default: first sheet)',
            'description_key': 'modules.excel.read.params.sheet.description',
            'required': False
        },
        'header_row': {
            'type': 'number',
            'label': 'Header Row',
            'label_key': 'modules.excel.read.params.header_row.label',
            'description': 'Row number for headers (1-based, 0 for no headers)',
            'description_key': 'modules.excel.read.params.header_row.description',
            'required': False,
            'default': 1
        },
        'range': {
            'type': 'string',
            'label': 'Cell Range',
            'label_key': 'modules.excel.read.params.range.label',
            'description': 'Cell range to read (e.g., "A1:D10")',
            'description_key': 'modules.excel.read.params.range.description',
            'required': False
        },
        'as_dict': {
            'type': 'boolean',
            'label': 'Return as Dict',
            'label_key': 'modules.excel.read.params.as_dict.label',
            'description': 'Return rows as dictionaries (using headers as keys)',
            'description_key': 'modules.excel.read.params.as_dict.description',
            'required': False,
            'default': True
        }
    },
    output_schema={
        'data': {
            'type': 'array',
            'description': 'Extracted data rows'
        },
        'headers': {
            'type': 'array',
            'description': 'Column headers'
        },
        'row_count': {
            'type': 'number',
            'description': 'Number of data rows'
        },
        'sheet_names': {
            'type': 'array',
            'description': 'All sheet names in the workbook'
        }
    },
    examples=[
        {
            'title': 'Read entire sheet',
            'title_key': 'modules.excel.read.examples.basic.title',
            'params': {
                'path': '/tmp/data.xlsx',
                'as_dict': True
            }
        }
    ],
    author='Flyto2 Team',
    license='MIT'
)
async def excel_read(context: Dict[str, Any]) -> Dict[str, Any]:
    """Read data from Excel file"""
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl is required for Excel reading. Install with: pip install openpyxl")

    params = context['params']
    path = params['path']
    sheet_name = params.get('sheet')
    header_row = params.get('header_row', 1)
    cell_range = params.get('range')
    as_dict = params.get('as_dict', True)

    # Validate file exists
    if not os.path.exists(path):
        raise FileNotFoundError(f"Excel file not found: {path}")

    # Open workbook
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet_names = wb.sheetnames

    # Select sheet
    if sheet_name:
        if sheet_name not in sheet_names:
            raise ValueError(f"Sheet not found: {sheet_name}")
        ws = wb[sheet_name]
    else:
        ws = wb.active

    # Determine range to read
    if cell_range:
        cells = ws[cell_range]
    else:
        cells = ws.iter_rows()

    # Read data
    all_rows: List[List[Any]] = []
    for row in cells:
        row_data = [cell.value for cell in row]
        all_rows.append(row_data)

    # Extract headers
    headers: List[str] = []
    data_rows: List[Any] = []

    if header_row > 0 and len(all_rows) >= header_row:
        headers = [str(h) if h else f"col_{i}" for i, h in enumerate(all_rows[header_row - 1])]
        data_rows = all_rows[header_row:]
    else:
        data_rows = all_rows

    # Convert to dicts if requested
    if as_dict and headers:
        data = []
        for row in data_rows:
            row_dict = {}
            for i, val in enumerate(row):
                key = headers[i] if i < len(headers) else f"col_{i}"
                row_dict[key] = val
            data.append(row_dict)
    else:
        data = data_rows

    wb.close()

    logger.info(f"Read Excel: {path} ({len(data)} rows)")

    return {
        'ok': True,
        'data': data,
        'headers': headers,
        'row_count': len(data),
        'sheet_names': sheet_names,
        'active_sheet': ws.title
    }
