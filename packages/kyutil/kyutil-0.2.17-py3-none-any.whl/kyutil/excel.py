# -*- coding: UTF-8 -*-
import os
from copy import copy

import logzero
import openpyxl
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Font, Alignment, NamedStyle, PatternFill
from openpyxl.utils import get_column_letter


class ExcelOp(object):
    """操作xls 表格工具类"""

    def __init__(self, file):
        self.file = file
        self.wb = load_workbook(self.file)
        self.ws = self.wb[self.wb.sheetnames[0]]  # 默认打开第一个表

    def get_all_sheet(self):
        """获取所有工作表名称"""
        return self.wb.sheetnames

    def get_row_clo_num(self):
        """获取表格的总行数和总列数"""
        rows = self.ws.max_row
        columns = self.ws.max_column
        return rows, columns

    def get_cell_value(self, row, column):
        """获取某个单元格的值"""
        cell_value = self.ws.cell(row=row, column=column).value
        return cell_value

    def get_col_value(self, column):
        """获取某列的所有值"""
        rows = self.ws.max_row
        column_data = []
        for i in range(1, rows + 1):
            cell_value = self.ws.cell(row=i, column=column).value
            column_data.append(cell_value)
        return column_data

    def get_row_value(self, row):
        """获取某行所有值"""
        columns = self.ws.max_column
        row_data = []
        for i in range(1, columns + 1):
            cell_value = self.ws.cell(row=row, column=i).value
            row_data.append(cell_value)
        return row_data

    def set_work_sheet(self, sheet_index=0):
        """设置当前工作表"""
        self.ws = self.wb[self.wb.sheetnames[sheet_index]]

    def set_cell_value(self, row, column, cell_value, commit=True):
        """设置某个单元格的值"""
        self.ws.cell(row=row, column=column).value = cell_value
        if commit:
            self.save_sheet()

    def set_row_value(self, content: "list or tuple", row_: int, commit=True):
        """设置某行所有的值"""
        if not isinstance(content, (list, tuple, str)):
            # 可迭代类型 及 参数无法写入
            raise ValueError("Parameter type error. Current row content cannot be set")

        if isinstance(content, str):
            content = [content, ]

        for i in range(0, len(content)):
            self.ws.cell(row=row_, column=i + 1, value=content[i])
        if commit:
            self.save_sheet()

    def save_sheet(self):
        """保存表变更值"""
        try:
            self.wb.save(self.file)
        except IOError as e:
            print("设置某行所有的值错误: %s" % e)
        finally:
            self.wb.close()

    @staticmethod
    def copy_cell_properties(source_cell, target_cell):
        """复制单元格及属性"""
        target_cell.data_type = copy(source_cell.data_type)
        target_cell.fill = copy(source_cell.fill)
        if source_cell.has_style:
            # 样式
            target_cell._style = copy(source_cell._style)
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = copy(source_cell.number_format)
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)

        if source_cell.hyperlink:
            # 超链接
            target_cell._hyperlink = copy(source_cell.hyperlink)

        if source_cell.comment:
            # 注解
            target_cell.comment = copy(source_cell.comment)


class CreateExcel:
    """CreateExcel"""

    def __init__(self, msg):
        # 创建五个表
        self.workbook = Workbook()
        self.workbook.active.title = "升级版本软件包"
        self.workbook.create_sheet("降低版本软件包")
        self.workbook.create_sheet("同版本软件包")
        self.workbook.create_sheet("删除软件包")
        self.workbook.create_sheet("新增软件包")
        self.workbook.create_sheet("文件目录比对")
        self.workbook.create_sheet("文件内容比对")

        # 填充
        self.blue_fill = PatternFill('solid', fgColor='4F81BD')

        # 顶部样式
        self.header_title = NamedStyle(name="header_title")
        self.header_title.font = Font(name=u"宋体", sz=14, bold=True)
        self.header_title.alignment = Alignment(horizontal='center', vertical="center")
        self.header_title.fill = self.blue_fill

        # 正文样式
        self.header_name = NamedStyle(name="header_name")
        self.header_name.font = Font(name=u"宋体", sz=14)
        self.header_name.alignment = Alignment(vertical="center", wrap_text=True)

        # 添加顶部信息
        for i in range(0, 7):
            self.__add_header__(i)
            self.workbook.active.cell(2, 1).value = msg
        self.workbook.active = 0

    def __add_header_up_down__(self, index):
        self.workbook.active = index
        self.workbook.active.merge_cells('A1:D1')
        self.workbook.active.merge_cells('A2:D2')
        self.workbook.active.merge_cells('A3:B3')
        self.workbook.active.merge_cells('C3:D3')
        self.workbook.active.cell(3, 1, value='SRPM').alignment = Alignment(horizontal='center', vertical='center')
        self.workbook.active.cell(3, 3, value='RPM').alignment = Alignment(horizontal='center', vertical='center')
        self.workbook.active.cell(1, 1).value = self.workbook.sheetnames[index]
        self.workbook.active.cell(1, 1).style = self.header_title
        self.workbook.active.cell(2, 1).style = self.header_name
        for i in range(1, 5):
            if (i % 2) == 0:
                self.workbook.active.cell(4, i, 'B').alignment = Alignment(horizontal='center', vertical='center')
            else:
                self.workbook.active.cell(4, i, 'A').alignment = Alignment(horizontal='center', vertical='center')
            self.workbook.active.column_dimensions[get_column_letter(i)].width = 70
            if i < 3:
                self.workbook.active.row_dimensions[i].height = 40

    def __add__header_add_del__(self, index):
        self.workbook.active = index
        self.workbook.active.merge_cells('A1:B1')
        self.workbook.active.merge_cells('A2:B2')
        self.workbook.active.cell(1, 1).value = self.workbook.sheetnames[index]
        self.workbook.active["A1"].style = self.header_title
        self.workbook.active["A2"].style = self.header_name
        for i in range(1, 3):
            self.workbook.active.column_dimensions[get_column_letter(i)].width = 70
            self.workbook.active.row_dimensions[i].height = 40
        self.workbook.active.cell(3, 1, value='SRPM').alignment = Alignment(horizontal='center', vertical='center')
        self.workbook.active.cell(3, 2, value='RPM').alignment = Alignment(horizontal='center', vertical='center')

    def __add_header_list__(self, index):
        self.workbook.active = index
        self.workbook.active.merge_cells('A1:C1')
        self.workbook.active.merge_cells('A2:C2')
        self.workbook.active.cell(1, 1).value = self.workbook.sheetnames[index]
        self.workbook.active["A1"].style = self.header_title
        self.workbook.active["A2"].style = self.header_name
        for i in range(1, 4):
            if i == 1:
                self.workbook.active.column_dimensions[get_column_letter(i)].width = 50
            else:
                self.workbook.active.column_dimensions[get_column_letter(i)].width = 35
            if i != 3:
                self.workbook.active.row_dimensions[i].height = 40
        self.workbook.active.cell(3, 1, value='文件名').alignment = Alignment(horizontal='center', vertical='center')
        self.workbook.active.cell(3, 2, value='A(md5)').alignment = Alignment(horizontal='center', vertical='center')
        self.workbook.active.cell(3, 3, value='B(md5)').alignment = Alignment(horizontal='center', vertical='center')

    def __add_header_content__(self, index):
        self.workbook.active = index
        self.workbook.active.merge_cells('A1:B1')
        self.workbook.active.merge_cells('A2:B2')
        self.workbook.active.cell(1, 1).value = self.workbook.sheetnames[index]
        self.workbook.active["A1"].style = self.header_title
        self.workbook.active["A2"].style = self.header_name
        self.workbook.active.column_dimensions['A'].width = 30
        self.workbook.active.row_dimensions[1].height = 40
        self.workbook.active.column_dimensions['B'].width = 100
        self.workbook.active.row_dimensions[2].height = 40
        self.workbook.active.cell(3, 1, value='文件名').alignment = Alignment(horizontal='center', vertical='center')
        self.workbook.active.cell(3, 2, value='差异内容').alignment = Alignment(horizontal='center', vertical='center')

    def __add_header__(self, index):
        if index < 0 or index > 7:
            return "参数错误"

        if 0 <= index < 2:
            self.__add_header_up_down__(index)
        elif 2 <= index <= 4:
            self.__add__header_add_del__(index)
        elif index == 5:
            self.__add_header_list__(index)
        else:
            self.__add_header_content__(index)

    def add_compare_info(self, index, info_a, info_b, compare_type):
        """
        添加源码包信息
        @param index: 表
        @param info_a:  添加信息a
        @param info_b: 添加信息b
        @param compare_type: 信息类型
        @return:
        """
        if index < 0 or index >= 2:
            return "参数有误"
        self.workbook.active = index
        if compare_type.lower() == "srpm":
            for i in range(0, len(info_a)):
                self.workbook.active.cell(i + 5, 1).value = info_a[i]
            for i in range(0, len(info_b)):
                self.workbook.active.cell(i + 5, 2).value = info_b[i]
        elif compare_type.lower() == 'rpm':
            for i in range(0, len(info_a)):
                self.workbook.active.cell(i + 5, 3).value = info_a[i]
            for i in range(0, len(info_b)):
                self.workbook.active.cell(i + 5, 4).value = info_b[i]

    def add_common(self, index, info, info_type):
        """
        @param index: 表
        @param info: 添加信息
        @param info_type:  添加信息类型
        @return:
        """
        if index <= 1 or index > 4:
            return "参数有误"
        self.workbook.active = index
        if info_type.lower() == 'srpm':
            for i in range(0, len(info)):
                self.workbook.active.cell(i + 4, 1).value = info[i]
        elif info_type.lower() == 'rpm':
            for i in range(0, len(info)):
                self.workbook.active.cell(i + 4, 2).value = info[i]
        self.workbook.active = 0

    def add_files(self, index, files, a_md5, b_md5):
        if index != 5:
            return "参数有误"
        self.workbook.active = index
        for _ in range(0, len(files)):
            self.workbook.active.cell(_ + 4, 1).value = files[_]
            self.workbook.active.cell(_ + 4, 2).value = a_md5[_]
            self.workbook.active.cell(_ + 4, 3).value = b_md5[_]
        self.workbook.active = 0

    def add_content_diff(self, index, files, contents, initrd_diff):
        if index != 6:
            return "参数有误"
        self.workbook.active = index
        for _ in range(0, len(files) + 1):
            if _ < len(files):
                self.workbook.active.cell(_ + 4, 1).value = files[_]
                self.workbook.active.cell(_ + 4, 2).value = contents[_]
            else:
                self.workbook.active.cell(_ + 4, 1).value = initrd_diff['fp']
                self.workbook.active.cell(_ + 4, 2).value = initrd_diff['diff_content']
            self.workbook.active.cell(_ + 4, 2).alignment = Alignment(wrap_text=True)
        self.workbook.active = 0

    def save(self, save_path):
        """
        保存表
        @param save_path: 保存路径
        @return:
        """
        if os.path.exists(save_path):
            print(save_path + " 此文件会被覆盖。------------")
        self.workbook.save(save_path)


class IntegratedJobOrder(ExcelOp):
    """集成工作单"""

    def __init__(self, file, logger_=logzero.logger):
        super().__init__(file=file)
        self._logger = logger_
        self.line_first_letter = "A"
        self.end_of_line_letters = "N"
        self.zen_info_row_num = 5  # 禅道信息在集成工作表的第五行之后追加
        self.other_instructions_row = 7
        self.other_instructions_column = 14
        self.other_instructions_are = "A7:N7"
        self.other_instructions_null_are = "A8:N8"

    def copy_worksheet_from_build01(self, build_ver):
        """根据build01 生成指定命名的其它工作表单"""
        ws2 = self.wb.copy_worksheet(self.ws)
        ws2.title = build_ver
        self.ws = ws2
        self.save_sheet()

    def write_row_info(self, info: list, sheet_index_=0, row_offset=1):
        """填写选中表的行信息"""
        self.set_work_sheet(sheet_index=sheet_index_)
        try:
            for i in range(0, len(info)):
                self.set_row_value(content=info[i], row_=row_offset + i, commit=False)
            return True
        except ValueError as e:
            self._logger.error("设置行内容错误，无法写入表格 %s" % e)
            return False
        except TypeError as e:
            self._logger.error("设置行内容错误，参数info [ %s ] 类型错误: %s" % info, e)
            return False

    def save_srpm_info(self, srpms: list):
        """源码包列表"""
        write_success = self.write_row_info(info=srpms, sheet_index_=-2, row_offset=2)
        if write_success:
            self.save_sheet()
        else:
            self._logger.error("[\" %s \"] 表写入失败!!" % self.wb.sheetnames[-2])

    def save_rpm_info(self, rpms: list):
        """二进制包列表"""
        write_success = self.write_row_info(info=rpms, sheet_index_=-1, row_offset=2)
        if write_success:
            self.save_sheet()
        else:
            self._logger.error("[\" %s \"] 表写入失败!!" % self.wb.sheetnames[-1])

    def handle_other_instructions_area(self, row_offset: int, area_idx: int = 0):
        """处理其它说明表格区域"""
        for c in range(1, self.other_instructions_column):
            # 其他说明：
            other_instructions_source = self.ws.cell(row=self.other_instructions_row + area_idx, column=c)
            other_instructions_target = self.ws.cell(row=self.other_instructions_row + area_idx + row_offset, column=c)
            # 复制样式样式
            self.copy_cell_properties(other_instructions_source, other_instructions_target)
            other_instructions_target.value = copy(other_instructions_source.value)
            # 复制行高
            self.ws.row_dimensions[self.other_instructions_row + area_idx + row_offset].height = self.ws.row_dimensions[
                self.other_instructions_row + area_idx].height

        merged_are_target = self.line_first_letter + str(self.other_instructions_row + area_idx + row_offset) + ":" + \
                            self.end_of_line_letters + str(self.other_instructions_row + area_idx + row_offset)

        self._logger.debug("待合并的其他说明单元格区域：[ %s ]" % merged_are_target)
        self.ws.merge_cells(merged_are_target)

        self.save_sheet()

    def change_build01_style(self, row_offset: int):
        """添加表格build01多行，同时保证格式"""
        self.handle_other_instructions_area(row_offset=row_offset, area_idx=0)  # 其他说明单元格
        self.handle_other_instructions_area(row_offset=row_offset, area_idx=1)  # 其他说明空白单元格
        self._logger.debug("处理其他说明区域完成，[ %s ] 下移: %s" % (self.other_instructions_are, row_offset))
        # 取消模板其它说明合并单元格
        self.ws.unmerge_cells(self.other_instructions_are)
        self.ws.unmerge_cells(self.other_instructions_null_are)
        # self._logger.debug("处理原其他说明区域完成，[ %s ] 取消合并" % self.other_instructions_are)

        for r in range(row_offset):
            r_idx = r + self.other_instructions_row
            self.ws.row_dimensions[r_idx].height = self.ws.row_dimensions[self.other_instructions_row - 1].height

            for c in range(1, self.other_instructions_column + 1):
                s_cell_other_ins = self.ws.cell(row=self.other_instructions_row - 1, column=1)
                t_cell_other_ins = self.ws.cell(row=r_idx, column=c)
                self.copy_cell_properties(s_cell_other_ins, t_cell_other_ins)
                t_cell_other_ins.value = ""

        self._logger.debug("处理偏移量明区域完成，[ %s ] -> [ A%s:N%s ] 设置单元格，初始化空值完成" % (
            self.other_instructions_are, str(self.other_instructions_row),
            str(self.other_instructions_row + row_offset)))

        self.save_sheet()

    def convertZenSrpmInfo(self, zen_info: list):
        """转换单个源码包的禅道信息为禅道地址，bug id, 修改内容， 修改人
        """
        if not zen_info:
            return '/', '/', '社区同步', '/'

        zen_urls = []
        bug_ids = []
        modify_contents = []
        modify_users = []
        for z_i in zen_info:
            zen_urls.append(z_i.get('zen_url'))
            bug_ids.append(z_i.get('bug_id'))
            modify_contents.append(z_i.get('modify_content'))
            modify_users.append(z_i.get('modify_user'))
        zen_url = "\n".join(zen_urls)  # 禅道地址
        bug_id = "\n".join(bug_ids)  # bugID
        modify_content = "\n".join(modify_contents)  # 修改内容
        modify_content = modify_content.replace('<p>', '').replace('</p>', '')  # 去除修改内容内的富文本标签
        modify_user = "\n".join(modify_users)  # 修改人

        return zen_url, bug_id, modify_content, modify_user

    def build_packages_data(self, tag_name: str, koji_ip: str, tag_srpms: list, iso_srpms: list):
        """
        工作单行数据转换
        :param tag_name:
        :param koji_ip:
        :param tag_srpms:
        [
            {
               "tag_id": 149,"tag_name": "v10-sp3","id": 25058,"build_id": 25058,
               "version": "0.18","release": "3.ky10","epoch": null,"state": 1,
               "completion_time": "2022-03-22 18:04:17.034721+08:00",
               "start_time": "2022-03-22 17:27:46.372950+08:00",
               "task_id": 304353,"creation_event_id": 386348,
               "creation_time": "2022-03-22 17:27:46.385286+08:00",
               "volume_id": 0,"volume_name": "DEFAULT","package_id": 489,
               "package_name": "jbig2dec","name": "jbig2dec","nvr": "jbig2dec-0.18-3.ky10",
               "owner_id": 56,"owner_name": "yangxudong"
            },
        ]
        :param iso_srpms:
        {
            "1213-123-123.src.rpm": [
               {
                   "bug_id": "62542",
                   "modify_content": "<p>test</p>",
                   "modify_user": "XXX",
                   "zen_url": "https://XXXXXXXXX/biz/task-view-62542.html"
               }
            ],
            "audit-3.0-5.se.08.ky10.src.rpm": [
               {
                   "bug_id": "11111111",
                   "modify_content": "<p>111111111111111111</p>",
                   "modify_user": "XXXX",
                   "zen_url": "https://XXXXXXXX/biz/bug-view-11111111.html"
               }
            ],
            "c.rpm": []
        }
        :return:
        """
        build_data = []

        # 转换iso_srpm内的包名为nvr
        if iso_srpms:
            for k in iso_srpms.keys():
                iso_srpms[k.strip('.src.rpm')] = iso_srpms.pop(k)

        for t_s in tag_srpms:
            # 行内容
            integration_flag = '是' if iso_srpms and t_s['nvr'] in iso_srpms.keys() else '否'
            # 一个源码包可能对应多个禅道信息
            if iso_srpms:
                zen_info = iso_srpms.get(t_s['nvr'])
                zen_url, bug_id, modify_content, modify_user = self.convertZenSrpmInfo(zen_info)
            else:
                zen_url, bug_id, modify_content, modify_user = '/', '/', '社区同步', '/'
            one_srpm_relation_data = [
                tag_srpms.index(t_s),  # 序号
                zen_url,  # 禅道地址
                bug_id,  # 需求内容/bug号(禅道)
                t_s['nvr'],  # 软件包 , 如： lmfit-8.2.2-1.p01.ky10
                modify_content,  # 修改内容
                modify_user,  # 研发人员
                '是',  # 自验证是否通过
                integration_flag,  # 是否集成
                '是',  # 是否已入库
                tag_name,  # 入库标签
                '',  # 升级至该版本及以上版本
                "https://" + koji_ip + '/koji/taskinfo?taskID=' +
                str(t_s['task_id']) if t_s['task_id'] else '',  # koji地址
                '',  # 软件包来源
                '',  # 备注
            ]
            build_data.append(one_srpm_relation_data)
        return build_data

    def build_second_row_data(self, product_name, build_ver, base_iso_ver, integr_time, inter_ip=''):
        # 集成工作单第二行数据内容
        self.ws['A2'].value = '项目名/项目ID：' + product_name
        self.ws['D2'].value = '集成版本号：' + build_ver
        self.ws['E2'].value = '基于哪个版本制作：' + base_iso_ver
        self.ws['K2'].value = '集成负责人：'
        self.ws['L2'].value = '集成时间：' + str(integr_time)
        self.ws['M2'].value = '集成用机IP：' + inter_ip



def is_merged_cell(cell, sheet):
    """
    检查单元格是否是合并单元格的一部分
    :param cell: 要检查的单元格
    :param sheet: 工作表对象
    :return: 如果是合并单元格的一部分，返回 True，否则返回 False
    """
    for merged_range in sheet.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return True
    return False


def is_top_left_merged_cell(cell, merged_range):
    """
    检查单元格是否是合并单元格的左上角单元格
    :param cell: 要检查的单元格
    :param merged_range: 合并单元格范围对象
    :return: 如果是合并单元格的左上角单元格，返回 True，否则返回 False
    """
    return cell.coordinate == str(merged_range.start_cell.column_letter) + str(merged_range.min_row)


def get_merged_cell_info(merged_range):
    """
    获取合并单元格的 rowspan 和 colspan
    :param merged_range: 合并单元格范围对象
    :return: 包含 rowspan 和 colspan 的元组
    """
    rowspan = merged_range.max_row - merged_range.min_row + 1
    colspan = merged_range.max_col - merged_range.min_col + 1
    return rowspan, colspan


def process_merged_cell(cell, merged_range, html, merged_cells):
    """
    处理合并单元格
    :param cell: 合并单元格的左上角单元格
    :param merged_range: 合并单元格范围对象
    :param html: 当前的 HTML 字符串
    :param merged_cells: 已处理的合并单元格列表
    :param get_column_letter: 获取列字母的函数
    :return: 更新后的 HTML 字符串
    """
    rowspan, colspan = get_merged_cell_info(merged_range)
    value = cell.value if cell.value is not None else ""
    html += f'<td rowspan="{rowspan}" colspan="{colspan}">{value}</td>'
    merged_cells.extend([f'{get_column_letter(col)}{row}' for row in
                         range(merged_range.min_row, merged_range.max_row + 1) for col in
                         range(merged_range.min_col, merged_range.max_col + 1)])
    return html


def process_normal_cell(cell, html):
    """
    处理普通单元格
    :param cell: 普通单元格
    :param html: 当前的 HTML 字符串
    :return: 更新后的 HTML 字符串
    """
    value = cell.value if cell.value is not None else ""
    value = value.replace('\n', '<br/>')
    html += f"<td>{value}</td>"
    return html


def deal_rows(row, html, sheet, merged_cells):
    """
    处理每行单元格
    """
    for cell in row:
        if is_merged_cell(cell, sheet):
            for merged_range in sheet.merged_cells.ranges:
                if is_top_left_merged_cell(cell, merged_range):
                    html = process_merged_cell(cell, merged_range, html, merged_cells)
                    break
        elif cell.coordinate not in merged_cells:
            html = process_normal_cell(cell, html)
    return html


def excel_to_html(file_path):
    # 加载 Excel 文件
    workbook = openpyxl.load_workbook(file_path)
    # 获取第一个工作表
    sheet = workbook.active

    # 开始构建 HTML 表格
    html = '<table>'

    # 用于跟踪已经处理过的合并单元格
    merged_cells = []

    # 遍历工作表的每一行
    for row in sheet.iter_rows():
        html += '<tr>'
        html = deal_rows(row, html, sheet, merged_cells)
        html += '</tr>'

    html += '</table>'
    return html
