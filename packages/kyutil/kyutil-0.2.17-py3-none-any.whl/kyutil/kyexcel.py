# -*- coding: UTF-8 -*-
import os
import re
import shutil
import time

import openpyxl
from openpyxl.cell import MergedCell
from openpyxl.styles import Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.cell_range import CellRange

from kyutil.file import ensure_dir

REGEX_ILLEGAL_CHARACTERS_RE = r'[\000-\010]|[\013-\014]|[\016-\037]'


class ExcelObj:
    def __init__(self, title=None, excel_template=None) -> None:
        self.ILLEGAL_CHARACTERS_RE = re.compile(REGEX_ILLEGAL_CHARACTERS_RE)
        self.title = title or [""]  # 一维列表
        self.data = []  # 二维列表
        self.file_name = ""
        self.excel_template = excel_template
        self.wb = self.get_save_wb_obj(excel_template)

    def get_save_wb_obj(self, excel_template=None):
        file_name = self.file_name = self.file_name or f'/tmp/repo_diff_{time.strftime("%H%M%S")}.xlsx'
        ensure_dir(file_name)
        excel_template = excel_template or self.excel_template
        if excel_template and os.path.exists(excel_template):
            shutil.copy(excel_template, file_name)
            return openpyxl.load_workbook(filename=file_name)
        else:
            return openpyxl.Workbook()

    def set_filename(self, filename):
        self.file_name = filename

    def append(self, data):
        self.data.append(data)

    @property
    def sheetnames(self):
        return self.wb.sheetnames

    def copy_sheet(self, sheet_name):
        if sheet_name in self.sheetnames:
            return
        ws = self.wb.create_sheet(sheet_name)
        ws = self.wb.copy_worksheet(ws)
        ws.title = sheet_name
        ws.freeze_panes = 'A3'

    def append_data_in_sheet(self, sheet_name, sheet_data):
        self.copy_sheet(sheet_name)
        for data_one in sheet_data:
            data_one = [self.ILLEGAL_CHARACTERS_RE.sub(r'', str(cell_one)) if cell_one is not None else "" for cell_one in data_one]
            self.wb[sheet_name].append(data_one)

    def set_sheet_info(self, sheet_name, sheet_info="", position="A1"):
        self.copy_sheet(sheet_name)
        if sheet_info != "":
            self.wb[sheet_name][position] = self.ILLEGAL_CHARACTERS_RE.sub(r'', sheet_info)

    def get_merged_cell_super_position(self, sheet_name, position):
        """检查指定单元格是否为合并单元格，如果是，则返回合并单元格的主位置；如果不是，返回当前位置
        eg. 存在合并单元格 A1:B1，输入B1，返回A1
        Args:
            sheet_name: 当前工作表对象
            position: 需要获取的单元格所在位置
        """
        cell = self.wb[sheet_name][position]
        if isinstance(cell, MergedCell):  # 判断该单元格是否为合并单元格
            tmp_range = CellRange(min_col=cell.column, min_row=cell.row, max_col=cell.column,
                                  max_row=cell.row)
            for cur_range in self.wb[sheet_name].merged_cells.ranges:
                if cur_range.issuperset(tmp_range):
                    return f"{get_column_letter(cur_range.min_col)}{cur_range.min_row}"
        else:
            return position

    def set_sheet_cell_value(self, sheet_name, position, value):
        """清空单元格原有元素，设置新值"""
        value = str(value)
        self.copy_sheet(sheet_name)
        position = self.get_merged_cell_super_position(sheet_name=sheet_name, position=position)  # 校验是否为合并单元格
        self.wb[sheet_name][position] = self.ILLEGAL_CHARACTERS_RE.sub(r'', value)

    def attach_sheet_cell_value(self, sheet_name, position, value):
        """不删除单元格中的原有元素，在后面追加新值"""
        value_a2 = self.get_cell_value(sheet_name=sheet_name, position=position)
        self.set_sheet_cell_value(sheet_name=sheet_name, position=position, value=f"{value_a2}{value}")

    def set_sheet_wraptext4sheet_all(self, sheet_name, start_line=2, ignore_line=None, start_row=0, ignore_row=None):
        """设置单元格自动换行
        Args:
            sheet_name: 表格名称
            start_line: 开始的行
            ignore_line: 忽略的行
            start_row: 开始的列
            ignore_row: 忽略的列
        """
        ignore_line = ignore_line or []
        ignore_row = ignore_row or []
        for idx, row in enumerate(self.wb[sheet_name]):
            if idx < start_line or idx in ignore_line:
                continue
            for idx_cell, cell in enumerate(row):
                if idx < start_row or idx in ignore_row:
                    continue
                cell.alignment = openpyxl.styles.Alignment(wrapText=True)

    def set_sheet_wraptext(self, sheet_name=None, sheet_name_list=None, position=None):
        if sheet_name:
            self.set_sheet_wraptext4sheet_all(sheet_name)
            return
        sheet_name_list = sheet_name_list or self.sheetnames
        if not sheet_name and not position:
            for sheet_name in sheet_name_list:
                self.set_sheet_wraptext4sheet_all(sheet_name)

    def set_cell_alignment(self, sheet_name: str, position: str, h=True, v=True) -> None:
        """position: 'A1'，h：横向居中，v：竖向居中"""
        if h and v:
            self.wb[sheet_name][position].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        elif h:
            self.wb[sheet_name][position].alignment = openpyxl.styles.Alignment(horizontal='center')
        elif v:
            self.wb[sheet_name][position].alignment = openpyxl.styles.Alignment(vertical='center')

    def set_border(self, sheet_name, area, style="thin"):
        """给指定区域添加边框线
        Args:
            sheet_name:
            area:
            style: 线段样式，可选参数：
                'dashDot','dashDotDot', 'dashed','dotted', 'double','hair', 'medium', 'mediumDashDot',
                'mediumDashDotDot', 'mediumDashed', 'slantDashDot', 'thick', 'thin'
        """
        cell_area = self.wb[sheet_name][area]
        for i in cell_area:
            for j in i:
                j.border = Border(right=Side(style=style), bottom=Side(style=style))

    def set_color_background(self, sheet_name, position: str, color: str, pattern_type: str = "solid"):
        """给指定位置设定背景颜色
        Args:
            sheet_name:
            position:指定单元格位置 e.g. E6
            color:指定要设定的颜色
            pattern_type:指定填充效果，可选参数：'darkDown', 'darkUp', 'lightDown', 'darkGrid', 'lightVertical',
               'solid', 'gray0625', 'darkHorizontal', 'lightGrid', 'lightTrellis', 'mediumGray', 'gray125',
               'darkGray', 'lightGray', 'lightUp', 'lightHorizontal', 'darkTrellis', 'darkVertical'
        """

        self.wb[sheet_name][position].fill = PatternFill(patternType=pattern_type, fgColor=color)

    def merge_cells(self, sheet_name: str, area: str, h=True, v=True) -> None:
        """position: 'A1:B2'"""
        self.wb[sheet_name].merge_cells(area)
        self.set_cell_alignment(sheet_name, area.split(':')[0], h, v)

    def unmerge_cells(self, sheet_name: str, range_string=None, start_row=None, start_column=None, end_row=None,
                      end_column=None):
        """取消合并单元格
        Args:
            sheet_name:表格名称
            range_string:存在合并单元格的区域
            start_row:
            start_column:
            end_row:
            end_column:
        """
        self.wb[sheet_name].unmerge_cells(
            range_string=range_string,
            start_row=start_row,
            start_column=start_column,
            end_row=end_row,
            end_column=end_column
        )

    def auto_merge_cells(self, sheet_name: str, cell_col: str, ignore_line=2, h=False, v=True) -> None:
        """position: 'A'"""
        cur_value = ""
        idx_start = ignore_line
        for line_idx in range(ignore_line + 1, self.wb[sheet_name].max_row + 1):
            if cur_value != self.wb[sheet_name][f'{cell_col}{str(line_idx)}'].value:
                idx_end = line_idx - 1
                if idx_start < idx_end:
                    self.merge_cells(sheet_name=sheet_name,
                                     area=f"{cell_col}{str(idx_start)}:{cell_col}{str(idx_end)}",
                                     h=h, v=v)
                idx_start = line_idx
                cur_value = self.wb[sheet_name][f'{cell_col}{str(line_idx)}'].value
        if idx_start < self.wb[sheet_name].max_row:
            self.merge_cells(sheet_name=sheet_name,  # 合并最后的几行
                             area=f"{cell_col}{str(idx_start)}:{cell_col}{str(self.wb[sheet_name].max_row)}",
                             h=h, v=v)

    def insert_rows(self, sheet_name: str, row_idx: int) -> None:
        """在指定行前面插入一行"""
        self.wb[sheet_name].insert_rows(row_idx)

    def insert_row_data(self, sheet_name: str, row_idx: int, data_line) -> None:
        """在指定行插入一行数据"""
        self.insert_rows(sheet_name=sheet_name, row_idx=row_idx)
        for idx_col, data_cell in enumerate(data_line):
            self.set_sheet_cell_value(sheet_name, position=f"{chr(ord('A') + idx_col)}{row_idx}", value=data_cell)

    def freeze_panes(self, sheet_name, line_num='3', col_num='A'):
        """冻结窗口"""
        self.copy_sheet(sheet_name)
        self.wb[sheet_name].freeze_panes = f'{col_num}{str(line_num)}'

    def set_sheet_title(self, sheet_name, title=None, line_num='2'):
        """
            根据参数title设定表格标题
                若title类型为list，根据line_num参数，在指定行添加
                若title类型为dict，k为表格中的cell的位置，v为要添加的值
        """
        if title is None:
            title = []
        self.copy_sheet(sheet_name)
        if isinstance(title, list):
            for i, v in enumerate(title):
                self.wb[sheet_name][f"{str(chr(65 + i))}{line_num}"] = v
        elif isinstance(title, dict):
            for position, value in title.items():
                self.set_sheet_cell_value(sheet_name, position, value)

    def set_sheet_filter(self, sheet_name, position=None, line_num='2'):
        width = self.wb[sheet_name].max_column
        height = self.wb[sheet_name].max_row
        position_end = chr(ord('A') - 1 + width) + str(height)
        position = position or f"A{line_num}:{position_end}"
        self.wb[sheet_name].auto_filter.ref = position

    def get_cell_value(self, sheet_name, position):
        """获取一个单元格中的数据值"""
        return self.wb[sheet_name][position].value

    def get_line_value(self, sheet_name, line_idx, wide=None):
        """获取一行单元格的数据值"""
        wide = wide or self.wb[sheet_name].max_column
        cell_col_list = [chr(ord('A') + i) for i in range(wide)]
        return [self.wb[sheet_name][f'{cell_col}{str(line_idx)}'].value for cell_col in cell_col_list]

    def get_col_value(self, sheet_name, col='B', idx_start=1, idx_end=None):
        """获取一列单元格的数据值"""
        idx_end = idx_end or self.wb[sheet_name].max_row
        return [self.wb[sheet_name][f'{col}{str(idx_line + 1)}'].value for idx_line in range(idx_start, idx_end)]

    def del_sheet(self, sheet_name_del="", sheet_name_keep=None):
        """按照名字删除表格
        Args:
            sheet_name_del:要删除的表格名称
            sheet_name_keep:要保留的表格名称（参数类型为字符串或列表）
        """
        if sheet_name_keep:
            if isinstance(sheet_name_keep, str):
                sheet_name_keep = [sheet_name_keep]
            if type(sheet_name_keep) not in [list, tuple, set]:
                return
            sheet_name_keep = list(sheet_name_keep)
            for sheet_name_one in self.sheetnames:
                if sheet_name_one not in sheet_name_keep:
                    del self.wb[sheet_name_one]

        elif sheet_name_del in self.sheetnames:
            del self.wb[sheet_name_del]

    def del_line(self, sheet_name, line_idx_start=None, line_idx_end=None):
        """删除指定行"""
        sheet_data = []
        line_idx_start = line_idx_start or 1
        line_idx_end = line_idx_end or self.wb[sheet_name].max_row
        for i in range(line_idx_end, line_idx_start, -1):
            sheet_data.append(self.get_line_value(sheet_name, i))
            self.wb[sheet_name].delete_rows(i)
        return sheet_data[::-1]

    def del_blank_cell(self, sheet_name, col_str):
        """删除指定列中的空行，下方数据上提"""
        col_list = col_str.split('-')
        cur_line_num = 3
        for i in range(3, self.wb[sheet_name].max_row + 1):
            is_blank_line = True
            for cell_col in col_list:
                if self.wb[sheet_name][f'{cell_col}{str(i)}'].value != "":  # 这一行不全为空
                    is_blank_line = False
                    break
            if not is_blank_line:
                for cell_col in col_list:  # 把这一行中的每个cell上移到指定位置
                    self.wb[sheet_name][f'{cell_col}{str(cur_line_num)}'] = self.wb[sheet_name][
                        f'{cell_col}{str(i)}'].value
                cur_line_num += 1
        for i in range(cur_line_num, self.wb[sheet_name].max_row + 1):
            for cell_col in col_list:  # 清空移完之后的无用数据
                self.wb[sheet_name][f'{cell_col}{str(i)}'] = ""

    @staticmethod
    def dir_init(file_path_all):
        file_path_all = os.path.dirname(file_path_all)
        if not os.path.exists(file_path_all):
            os.makedirs(file_path_all)

    def save(self, file_name=None):
        file_name = file_name or self.file_name
        file_name = self.file_name = file_name if file_name.endswith(".xlsx") else file_name + ".xlsx"
        self.dir_init(file_name)
        self.wb.save(file_name)
        self.wb.close()
        return file_name
