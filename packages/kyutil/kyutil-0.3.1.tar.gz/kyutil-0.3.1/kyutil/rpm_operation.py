# -*- coding: UTF-8 -*-
"""rpm_operation.py"""
import os
import uuid

import logzero
from openpyxl import load_workbook

from kyutil.base import sha256_file
from kyutil.rpms import read_rpm_header
from kyutil.shell import run_command


# 挂在iso到tmp目录下
def mount_iso(iso_path, logger=logzero.logger):
    """
    挂载iso
    @param iso_path: iso文件路径
    @param logger: 日志对象
    @return:
    """
    if not iso_path:
        return ""
    logger.info(f'iso_path:{iso_path}')
    name = os.path.basename(iso_path)
    mount_dir = "/mnt/" + name + uuid.uuid4().hex[:4]
    cmd = f'mount -o loop {iso_path} {mount_dir}'
    logger.info(f"挂载iso :{iso_path} -> {mount_dir}")

    # 判断是否已经挂载
    if os.path.ismount(mount_dir):
        run_command(f'umount -fl {mount_dir}', None)

    if not os.path.exists(mount_dir):
        os.makedirs(mount_dir)

    if os.path.exists(iso_path) and os.path.isdir(mount_dir):
        if not run_command(cmd, None):
            return ""
    else:
        logger.error(f'Mount failed. cmd is ：【{cmd}】')
        return ""
    return mount_dir


def get_rpms(dest_dir):
    """
    从挂载目录下获取属于ISO的所有rpm的绝对路径
    @param dest_dir: 挂载路径
    @return: []
    """
    l = []
    for path, dirs, files in os.walk(dest_dir):
        for file in files:
            if file.endswith(".rpm") and str(path).lower().find("packages-gcc") < 0:
                l.append(os.path.join(path, file))
    return l


def bytes_to_str(_b):
    """
    字节转换成字符串
    @param _b:
    @return: str
    """
    return _b.decode() if isinstance(_b, bytes) else _b


def get_rpm_info(rpm_path, logger=logzero.logger):
    """
    获取rpm的信息
    @param rpm_path:
    @return:
    """
    rpm_info = srpm_info = n = ""
    try:
        if rpm_path.endswith(".rpm"):
            header = read_rpm_header(rpm_path)
            tmp = "AppStream" if rpm_path.lower().find("appstream") > 0 else ""
            r_dir = "BaseOS" if rpm_path.lower().find("baseos") > 0 else tmp
            r = f'{r_dir}: ' if r_dir else ""
            if header:
                rpm_info = f'{r}{os.path.basename(rpm_path)}'
                if header["SourceRPM"]:
                    n = bytes_to_str(header["SourceRPM"])
                srpm_info = f'{r}{n}'
    except Exception as e:
        import traceback
        traceback.print_exc()
        logger.error(f"==Fail to get_rpm_info,err: {e}")
    return rpm_info, srpm_info


def unmount(mount_dir):
    """
    卸载
    @param mount_dir:
    @return:
    """
    if os.path.ismount(mount_dir):
        run_command(f'umount -fl {mount_dir}', None)
        os.removedirs(mount_dir)


def add_excel(path: str, rpm: list, srpm: list):
    """ add_excel(path"""
    w = load_workbook(path)
    w.active = 1
    for i in range(0, len(srpm)):
        w.active.cell(i + 2, 1).value = srpm[i]
    w.active = 2
    for j in range(0, len(rpm)):
        w.active.cell(j + 2, 1).value = rpm[j]
    w.active = 0
    w.save(path)


def get_hdr(rpm_path, logger=logzero.logger):
    """
    获取hdr信息
    Args:
        rpm_path:rpm地址
        logger:

    Returns:

    """
    import rpm

    ts = rpm.TransactionSet()
    ts.setVSFlags(-1)
    try:
        fdno = os.open(rpm_path, os.O_RDONLY)
        hdr = ts.hdrFromFdno(fdno)
        os.close(fdno)
    except Exception as e:
        logger.error(e)
    return hdr


def get_nvr(name):
    if name.count("-") < 2:
        return name, '', ''
    return name.rsplit("-", 2)


def get_rpm_info(file_path):
    """
    获取列表
    Args:
        file_path:
    Returns:
    """
    rpm_info = {}
    try:
        import rpm
        hdr = get_hdr(file_path)
        rpm_info.update({"name": str(hdr[rpm.RPMTAG_NAME])})
        rpm_info.update({"version": str(hdr[rpm.RPMTAG_VERSION])})
        rpm_info.update({"release": str(hdr[rpm.RPMTAG_R])})
        rpm_info.update({"sha256": sha256_file(file_path)})
        rpm_info.update({"package": os.path.basename(file_path)})
        rpm_info.update({"package_path": file_path})
    except Exception as e:
        print(f"no rpm package:{e}")
        n, v, r = get_nvr(os.path.basename(file_path))
        rpm_info.update({"name": n})
        rpm_info.update({"version": v})
        rpm_info.update({"release": r})
        rpm_info.update({"sha256": sha256_file(file_path)})
        rpm_info.update({"package": os.path.basename(file_path)})
        rpm_info.update({"package_path": file_path})
    return rpm_info


def read_rpm_file_content(file_path: str, target_file: str = './etc/.productinfo') -> str:
    """"""
    import rpmfile
    with rpmfile.open(file_path, 'rb') as rpm:
        fd = rpm.extractfile(target_file)
        content = fd.read().decode('utf-8')
        return content
