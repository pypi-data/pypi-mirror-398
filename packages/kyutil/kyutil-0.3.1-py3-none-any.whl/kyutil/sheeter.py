# -*- coding: UTF-8 -*-
import os
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


class SpreadSheet:
    def __init__(self, _sheet_dir_: str = None) -> None:
        self._current_sheet = None
        self.xlsx_suffix = '.xlsx'
        if _sheet_dir_ is None:
            self._workbook = Workbook()
        else:
            if not _sheet_dir_.endswith(self.xlsx_suffix):
                _sheet_dir_ += self.xlsx_suffix
            self._workbook = load_workbook(_sheet_dir_)

    def save_workbook(self, _filename_: str):
        _filename, *dir_suffix = _filename_.rsplit('.', 1)
        if dir_suffix and dir_suffix[0] != self.xlsx_suffix:
            _filename = _filename + '.' + dir_suffix[0]
        sheet_suffix = self.xlsx_suffix

        if os.path.exists(_filename + sheet_suffix):
            while True:
                _suffix_ = datetime.now().strftime("-%H-%M-%S")
                _tmp_name = _filename + _suffix_ + sheet_suffix
                if os.path.exists(_tmp_name):
                    continue
                self._workbook.save(_tmp_name)
                break
        else:
            self._workbook.save(_filename + sheet_suffix)
        self.reset_col(_filename + sheet_suffix)
        return _filename + sheet_suffix

    def active_this_sheet(self, _sheet_name_: str):
        """Active sheet

        Args:
            _sheet_name_ (str): The sheet you want to active
        """
        self._current_sheet = self._workbook.get_sheet_by_name(_sheet_name_)
        self._workbook.active = self._current_sheet

    def append(self, _sheet_name_: str, _content_: list):
        """Insert a list into current sheet

        Args:
            _sheet_name_ (str): The sheet you want to insert
            _content_ (list): The content you want to insert, forms like [a, b, c, d]
            _row_comb_index: (start, nums) : combine nums units after start
                                            eg. (3,2) combine 2 units after column 3
        """
        self.active_this_sheet(_sheet_name_)
        self._current_sheet.append(_content_)

    def merge_col(self, _sheet_name_: str, _cell_index: tuple):
        """merge cells

        Args:
            _sheet_name_ (str): sheet name
            _cell_index (tuple): demo: ((1,1), (3,1)) -> merge (A1:C1)
        """
        _cell_left_up_col, _cell_left_up_row = _cell_index[0]
        _cell_right_down_col, _cell_right_down_row = _cell_index[1]
        _cell_left = get_column_letter(_cell_left_up_col) + str(_cell_left_up_row)
        _cell_right = get_column_letter(_cell_right_down_col) + str(_cell_right_down_row)
        self.active_this_sheet(_sheet_name_)
        self._current_sheet.merge_cells(f'{_cell_left}:{_cell_right}')

    def add_sheet(self, _sheet_title_: str, _sheet_position_: int = None):
        self._workbook.create_sheet(_sheet_title_, _sheet_position_)

    def delete_rows(self, _sheet_name_: str, _row_number: tuple):
        self.active_this_sheet(_sheet_name_)
        self._current_sheet.delete_rows(*_row_number)

    def delete_cols(self, _sheet_name_: str, _row_number: tuple):
        self.active_this_sheet(_sheet_name_)
        self._current_sheet.delete_cols(*_row_number)

    def reset_col(self, _filename):
        wb = load_workbook(_filename)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for col in ws.columns:
                index = list(ws.columns).index(col)  # 列序号
                letter = get_column_letter(index + 1)  # 列字母
                ws.column_dimensions[letter].width = 60  # ws.max_row * 1.2，也就是列宽为最大长度*1.2
        wb.save(_filename)
